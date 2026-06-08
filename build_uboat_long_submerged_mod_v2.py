#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
UBOAT - Long Submerged 10x+ - Générateur de mod Data Sheets

Corrige la version précédente :
- Ajoute supportedGameVersions dans Manifest.json pour éviter l'avertissement "mod obsolète".
- Patche tous les General.xlsx trouvés, y compris Data Sheets/Realistic Travel/General.xlsx.
- Patche Entities.xlsx / Equipment de façon directe :
  * Accumulators / batteries : capacité multipliée.
  * Tous les paramètres EnergyUsage : consommation électrique divisée.
  * Ventilation / filtration : gain d'oxygène renforcé et coût électrique réduit.
- Génère des overrides minimaux, pas une copie complète des datasheets vanilla.

Prérequis :
    py -m pip install openpyxl

Commande conseillée :
    py build_uboat_long_submerged_mod_v2.py --uboat "C:\\Program Files (x86)\\Steam\\steamapps\\common\\UBOAT" --force --clear-cache

Important :
    Pour les changements Entities.xlsx, teste avec une NOUVELLE carrière.
"""

from __future__ import annotations

import argparse
import copy
import json
import math
import os
import re
import shutil
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


# =============================================================================
# CONFIG PAR DÉFAUT
# =============================================================================

MOD_FOLDER_NAME = "LongSubmerged10x"
MOD_DISPLAY_NAME = "Long Submerged 10x+"
MOD_VERSION = "1.1.0"
MOD_AUTHOR = "VotreNomOuVotreEquipe"

DEFAULT_GAME_VERSION = "2026.1 Patch 20"

# Profil par défaut :
# - Oxygène environ 15x plus durable.
#   Sur ton screenshot : 13h -> environ 8 jours si tout est bien appliqué.
# - Discipline/fatigue environ 15x plus lente pour rester proportionnel à l'air.
# - Batterie : capacité x10 + EnergyUsage x0.1 = autonomie électrique effective beaucoup plus longue.
DEFAULT_OXYGEN_FACTOR = 15.0
DEFAULT_DISCIPLINE_FACTOR = 15.0
DEFAULT_BATTERY_CAPACITY_FACTOR = 10.0
DEFAULT_EQUIPMENT_ENERGY_USAGE_FACTOR = 0.10
DEFAULT_VENTILATION_FACTOR = 10.0

# On évite de s'appuyer sur Energy Base Scale comme levier principal, car il ne semble
# pas produire l'effet visible attendu dans ton test. Il reste disponible en option CLI.
DEFAULT_PATCH_ENERGY_BASE_SCALE = False


# =============================================================================
# STRUCTURES
# =============================================================================

@dataclass(frozen=True)
class GeneralPatch:
    row_id: str
    multiplier: float
    expected_category: str | None
    reason: str


@dataclass
class PatchReport:
    changed_files: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    info: list[str] = field(default_factory=list)

    def file(self, path: Path) -> None:
        self.changed_files.append(str(path))

    def warn(self, message: str) -> None:
        self.warnings.append(message)

    def note(self, message: str) -> None:
        self.info.append(message)


@dataclass
class GeneralRowHit:
    patch: GeneralPatch
    source_row: int
    category_norm: str | None
    category_row: int | None
    value_cols: list[int]


# =============================================================================
# OUTILS GÉNÉRAUX
# =============================================================================

def norm_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").strip().lower()
    return re.sub(r"\s+", " ", text)


def norm_category(value: Any) -> str:
    text = norm_text(value)
    if not text:
        return ""
    if not text.startswith("/"):
        text = "/" + text
    return text


def contains_any(text: str, needles: Iterable[str]) -> bool:
    lowered = text.lower()
    return any(needle.lower() in lowered for needle in needles)


def safe_float(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None

    text = str(value).strip()
    if not text:
        return None

    if text.endswith("%"):
        text = text[:-1].strip()

    text = text.replace("\u00a0", "")
    text = text.replace(" ", "")
    text = text.replace(",", ".")

    try:
        number = float(text)
    except ValueError:
        return None

    return number if math.isfinite(number) else None


def scale_numeric_excel_value(original_value: Any, multiplier: float) -> float:
    parsed = safe_float(original_value)
    if parsed is None:
        raise ValueError(f"valeur non numérique : {original_value!r}")
    return parsed * multiplier


def copy_cell(src: Cell, dst: Cell) -> None:
    dst.value = src.value

    if src.has_style:
        dst._style = copy.copy(src._style)

    dst.number_format = src.number_format

    if src.font:
        dst.font = copy.copy(src.font)
    if src.fill:
        dst.fill = copy.copy(src.fill)
    if src.border:
        dst.border = copy.copy(src.border)
    if src.alignment:
        dst.alignment = copy.copy(src.alignment)
    if src.protection:
        dst.protection = copy.copy(src.protection)
    if src.comment:
        dst.comment = copy.copy(src.comment)


def copy_row(src_ws: Worksheet, dst_ws: Worksheet, src_row: int, dst_row: int) -> None:
    for col in range(1, src_ws.max_column + 1):
        copy_cell(src_ws.cell(src_row, col), dst_ws.cell(dst_row, col))

    if src_ws.row_dimensions[src_row].height:
        dst_ws.row_dimensions[dst_row].height = src_ws.row_dimensions[src_row].height


def copy_column_widths(src_ws: Worksheet, dst_ws: Worksheet) -> None:
    for key, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[key].width = dim.width


def get_sheet_case_insensitive(workbook: Workbook, wanted_name: str) -> Worksheet | None:
    wanted_norm = norm_text(wanted_name)
    for name in workbook.sheetnames:
        if norm_text(name) == wanted_norm:
            return workbook[name]
    return None


def default_local_uboat_dir() -> Path:
    user_profile = os.environ.get("USERPROFILE")
    if user_profile:
        return Path(user_profile) / "AppData" / "LocalLow" / "Deep Water Studio" / "UBOAT"
    return Path.cwd() / "UBOAT_LocalLow"


def default_local_mods_root() -> Path:
    return default_local_uboat_dir() / "Mods"


def resolve_data_sheets_dir(uboat_root: Path) -> Path:
    data_sheets = uboat_root / "UBOAT_Data" / "Data Sheets"
    if not data_sheets.exists():
        raise FileNotFoundError(
            f"Dossier introuvable : {data_sheets}\n"
            f"--uboat doit pointer vers le dossier d'installation UBOAT."
        )
    return data_sheets


def ensure_clean_directory(path: Path, force: bool) -> None:
    if path.exists():
        if not force:
            raise FileExistsError(
                f"Le dossier existe déjà : {path}\n"
                f"Relance avec --force pour le remplacer."
            )
        shutil.rmtree(path)
    path.mkdir(parents=True, exist_ok=True)


def find_datasheet_files(data_sheets_dir: Path, filename: str) -> list[Path]:
    files = []
    for path in data_sheets_dir.rglob(filename):
        if path.name.startswith("~$"):
            continue
        files.append(path)
    return sorted(files, key=lambda p: str(p.relative_to(data_sheets_dir)).lower())


def unique_preserve_order(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        key = value.strip().lower()
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(value.strip())
    return result


# =============================================================================
# GENERAL.XLSX
# =============================================================================

def build_general_patches(args: argparse.Namespace) -> list[GeneralPatch]:
    patches = [
        GeneralPatch(
            row_id="Oxygen Consumption Per Character",
            multiplier=1.0 / args.oxygen_factor,
            expected_category="/Resources",
            reason=f"Air environ {args.oxygen_factor:g}x plus durable.",
        ),
        GeneralPatch(
            row_id="Underwater Discipline Loss",
            multiplier=1.0 / args.discipline_factor,
            expected_category="/Discipline",
            reason=f"Discipline sous l'eau environ {args.discipline_factor:g}x moins punitive.",
        ),
        GeneralPatch(
            row_id="Fatigue - Per Day",
            multiplier=1.0 / args.discipline_factor,
            expected_category="/Discipline",
            reason=f"Fatigue longue durée environ {args.discipline_factor:g}x plus lente.",
        ),
        GeneralPatch(
            row_id="Fatigue - Max Penalty",
            multiplier=1.0 / args.discipline_factor,
            expected_category="/Discipline",
            reason=f"Pénalité max de fatigue réduite proportionnellement.",
        ),
    ]

    if args.patch_energy_base_scale:
        patches.append(
            GeneralPatch(
                row_id="Energy Base Scale",
                multiplier=args.battery_capacity_factor,
                expected_category="/Resources",
                reason=(
                    "Optionnel : réserve d'énergie globale augmentée. "
                    "Le mod patche déjà les Accumulators dans Entities.xlsx."
                ),
            )
        )

    return patches


def index_settings_sheet(ws: Worksheet) -> tuple[
    dict[str, int],
    dict[tuple[str, str], int],
    dict[str, tuple[int, str | None]],
]:
    """
    Retourne :
    - category_rows: catégorie -> ligne de catégorie
    - rows_by_category_and_id: (catégorie, id) -> ligne
    - rows_by_id: id -> (ligne, catégorie ou None)

    Supporte aussi les feuilles moins structurées, sans catégorie explicite.
    """
    category_rows: dict[str, int] = {}
    rows_by_category_and_id: dict[tuple[str, str], int] = {}
    rows_by_id: dict[str, tuple[int, str | None]] = {}

    current_category: str | None = None

    for row in range(1, ws.max_row + 1):
        first_raw = ws.cell(row, 1).value
        first = str(first_raw).strip() if first_raw is not None else ""

        if not first:
            continue

        if first.startswith("/"):
            current_category = norm_category(first)
            category_rows[current_category] = row
            continue

        row_id_norm = norm_text(first)
        if not row_id_norm:
            continue

        if current_category:
            rows_by_category_and_id[(current_category, row_id_norm)] = row

        rows_by_id.setdefault(row_id_norm, (row, current_category))

    return category_rows, rows_by_category_and_id, rows_by_id


def find_value_cols_for_row(ws: Worksheet, row: int) -> list[int]:
    """
    Je patche toutes les colonnes numériques de la ligne.
    C'est nécessaire pour /Discipline, qui a Normal, Hard et Very Hard.
    /Resources reste propre parce que seule sa colonne B contient une valeur.
    """
    value_cols: list[int] = []

    for col in range(2, ws.max_column + 1):
        if safe_float(ws.cell(row, col).value) is not None:
            value_cols.append(col)

    return value_cols


def find_general_hits(
    ws: Worksheet,
    patches: list[GeneralPatch],
    report: PatchReport,
    source_label: str,
) -> list[GeneralRowHit]:
    category_rows, rows_by_category_and_id, rows_by_id = index_settings_sheet(ws)
    hits: list[GeneralRowHit] = []

    for patch in patches:
        expected_category_norm = norm_category(patch.expected_category) if patch.expected_category else None
        row: int | None = None
        actual_category: str | None = None

        if expected_category_norm:
            row = rows_by_category_and_id.get((expected_category_norm, norm_text(patch.row_id)))
            if row is not None:
                actual_category = expected_category_norm

        if row is None:
            fallback = rows_by_id.get(norm_text(patch.row_id))
            if fallback:
                row, actual_category = fallback

        if row is None:
            report.warn(f"{source_label}: ligne introuvable dans Settings : {patch.row_id!r}")
            continue

        category_row = category_rows.get(actual_category or "") if actual_category else None
        value_cols = find_value_cols_for_row(ws, row)

        if not value_cols:
            report.warn(
                f"{source_label}: {patch.row_id!r} trouvé ligne {row}, "
                f"mais aucune colonne numérique n'a été trouvée."
            )
            continue

        hits.append(
            GeneralRowHit(
                patch=patch,
                source_row=row,
                category_norm=actual_category,
                category_row=category_row,
                value_cols=value_cols,
            )
        )

    return hits


def apply_general_hit_values(
    src_ws: Worksheet,
    out_ws: Worksheet,
    hit: GeneralRowHit,
    dest_row: int,
    report: PatchReport,
    source_label: str,
) -> None:
    changes: list[str] = []

    for value_col in hit.value_cols:
        original = src_ws.cell(hit.source_row, value_col).value
        new_value = scale_numeric_excel_value(original, hit.patch.multiplier)
        out_ws.cell(dest_row, value_col).value = new_value

        header = src_ws.cell(hit.category_row or 1, value_col).value
        header_text = str(header).strip() if header is not None else f"colonne {value_col}"
        changes.append(f"{header_text}: {original!r} -> {new_value!r}")

    report.note(
        f"{source_label} / Settings / {hit.patch.row_id}: "
        + "; ".join(changes)
        + f" | {hit.patch.reason}"
    )


def create_general_override(
    vanilla_general: Path,
    output_general: Path,
    patches: list[GeneralPatch],
    report: PatchReport,
    data_sheets_dir: Path,
) -> bool:
    source_label = str(vanilla_general.relative_to(data_sheets_dir))
    src_wb = load_workbook(vanilla_general, data_only=False)
    src_ws = get_sheet_case_insensitive(src_wb, "Settings")

    if src_ws is None:
        report.warn(f"{source_label}: onglet Settings introuvable, fichier ignoré.")
        return False

    hits = find_general_hits(src_ws, patches, report, source_label)
    if not hits:
        report.warn(f"{source_label}: aucune valeur General.xlsx patchée.")
        return False

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = src_ws.title
    copy_column_widths(src_ws, out_ws)

    dest_row = 1
    written_categories: set[str] = set()

    categorized_hits = [hit for hit in hits if hit.category_norm and hit.category_row]
    uncategorized_hits = [hit for hit in hits if not (hit.category_norm and hit.category_row)]

    for hit in sorted(categorized_hits, key=lambda h: (h.category_row or 0, h.source_row)):
        assert hit.category_norm is not None
        assert hit.category_row is not None

        if hit.category_norm not in written_categories:
            copy_row(src_ws, out_ws, hit.category_row, dest_row)
            dest_row += 1
            written_categories.add(hit.category_norm)

        copy_row(src_ws, out_ws, hit.source_row, dest_row)
        apply_general_hit_values(src_ws, out_ws, hit, dest_row, report, source_label)

        dest_row += 1

    if uncategorized_hits:
        if dest_row > 1:
            dest_row += 1

        first_row_text = " ".join(
            norm_text(src_ws.cell(1, col).value)
            for col in range(1, src_ws.max_column + 1)
        )

        if "value" in first_row_text or "id" in first_row_text:
            copy_row(src_ws, out_ws, 1, dest_row)
            dest_row += 1

        for hit in sorted(uncategorized_hits, key=lambda h: h.source_row):
            copy_row(src_ws, out_ws, hit.source_row, dest_row)
            apply_general_hit_values(src_ws, out_ws, hit, dest_row, report, source_label)

            dest_row += 1

    output_general.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_general)
    report.file(output_general)
    return True


# =============================================================================
# ENTITIES.XLSX
# =============================================================================

NUMBER_PATTERN = r"[-+]?\d+(?:[.,]\d+)?(?:[eE][-+]?\d+)?"
KEY_VALUE_PATTERN_TEMPLATE = (
    r"(?P<prefix>(?<![A-Za-z0-9_]){key}(?![A-Za-z0-9_])\s*=\s*)"
    r"(?P<number>" + NUMBER_PATTERN + r")(?P<percent>%?)"
)


def format_number(value: float, percent: str = "") -> str:
    return f"{value:.10g}" + percent


def scale_keyed_parameters(parameters: str, multipliers: dict[str, float]) -> tuple[str, list[str]]:
    result = parameters
    changes: list[str] = []

    for key in sorted(multipliers.keys(), key=len, reverse=True):
        multiplier = multipliers[key]
        pattern = re.compile(
            KEY_VALUE_PATTERN_TEMPLATE.format(key=re.escape(key)),
            flags=re.IGNORECASE,
        )

        def replace(match: re.Match[str]) -> str:
            raw_number = match.group("number")
            percent = match.group("percent") or ""
            parsed = safe_float(raw_number + percent)

            if parsed is None:
                return match.group(0)

            scaled = parsed * multiplier
            replacement_number = format_number(scaled, percent)
            changes.append(f"{key}: {raw_number}{percent} -> {replacement_number}")

            return match.group("prefix") + replacement_number

        result = pattern.sub(replace, result)

    return result, changes


def scale_all_numbers_in_parameter_string(parameters: str, multiplier: float) -> tuple[str, list[str]]:
    changes: list[str] = []
    pattern = re.compile(r"(?P<number>" + NUMBER_PATTERN + r")(?P<percent>%?)")

    def replace(match: re.Match[str]) -> str:
        raw_number = match.group("number")
        percent = match.group("percent") or ""
        parsed = safe_float(raw_number + percent)

        if parsed is None:
            return match.group(0)

        scaled = parsed * multiplier
        replacement_number = format_number(scaled, percent)
        changes.append(f"*fallback*: {raw_number}{percent} -> {replacement_number}")

        return replacement_number

    return pattern.sub(replace, parameters), changes


def find_header_and_parameters_col(ws: Worksheet) -> tuple[int | None, int | None]:
    for row in range(1, min(ws.max_row, 25) + 1):
        for col in range(1, ws.max_column + 1):
            if norm_text(ws.cell(row, col).value) == "parameters":
                return row, col

    if ws.max_column >= 16:
        return 1, 16

    return None, None


def row_values(ws: Worksheet, row: int) -> list[Any]:
    return [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]


def join_row_text(values: list[Any]) -> str:
    return " | ".join(str(value) for value in values if value is not None)


def is_accumulator_row(text: str) -> bool:
    lowered = text.lower()

    if "battery room" in lowered and "accumulator" not in lowered:
        return False

    return contains_any(
        lowered,
        (
            "accumulator",
            "accumulators",
            "akkumulator",
            "akkumulatoren",
        ),
    )


def is_ventilation_row(text: str, parameters: str) -> bool:
    lowered = text.lower()
    params = parameters.lower()

    return (
        "ventilation" in lowered
        or "potassium" in lowered
        or "oxygen" in lowered
        or "oxygengain" in params
        or "regenerationlimit" in params
    )


def patch_entity_parameters(
    parameters: str,
    row_text_full: str,
    args: argparse.Namespace,
) -> tuple[str, list[str]]:
    """
    Applique tous les patchs pertinents sur une cellule Parameters.
    """
    new_parameters = parameters
    all_changes: list[str] = []

    # 1) Batterie : ligne Accumulators.
    if is_accumulator_row(row_text_full):
        accumulator_keys = {
            "EnergyCapacity": args.battery_capacity_factor,
            "BatteryCapacity": args.battery_capacity_factor,
            "CapacityMultiplier": args.battery_capacity_factor,
            "CapacityScale": args.battery_capacity_factor,
            "Capacity": args.battery_capacity_factor,
            "MaxCapacity": args.battery_capacity_factor,
            "EnergyStorage": args.battery_capacity_factor,
            "Storage": args.battery_capacity_factor,
            "Value": args.battery_capacity_factor,
        }

        new_parameters, changes = scale_keyed_parameters(new_parameters, accumulator_keys)
        all_changes.extend(changes)

        # Fallback volontairement limité à la ligne Accumulators.
        if not changes:
            new_parameters, changes = scale_all_numbers_in_parameter_string(
                new_parameters,
                args.battery_capacity_factor,
            )

            if changes:
                all_changes.extend(changes)

    # 2) Consommation électrique générale.
    # C'est le patch qui doit changer directement l'affichage "Moteurs électriques -22.7/km".
    if "energyusage" in new_parameters.lower():
        new_parameters, changes = scale_keyed_parameters(
            new_parameters,
            {
                "EnergyUsage": args.energy_usage_factor,
            },
        )

        all_changes.extend(changes)

    # 3) Ventilation / filtration.
    if is_ventilation_row(row_text_full, new_parameters):
        ventilation_keys = {
            "OxygenGain": args.ventilation_factor,
            "RegenerationLimit": args.ventilation_factor,
            "AirQualityGain": args.ventilation_factor,
            "AirGain": args.ventilation_factor,
            "CO2Reduction": args.ventilation_factor,
            "CarbonDioxideReduction": args.ventilation_factor,
            "EnergyUsage": args.energy_usage_factor,
        }

        new_parameters, changes = scale_keyed_parameters(new_parameters, ventilation_keys)
        all_changes.extend(changes)

    unique_changes = unique_preserve_order(all_changes)

    return new_parameters, unique_changes


def create_entities_override(
    vanilla_entities: Path,
    output_entities: Path,
    report: PatchReport,
    data_sheets_dir: Path,
    args: argparse.Namespace,
) -> bool:
    source_label = str(vanilla_entities.relative_to(data_sheets_dir))
    src_wb = load_workbook(vanilla_entities, data_only=False)

    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    total_rows_changed = 0
    touched_accumulators = False
    touched_energy_usage = False
    touched_ventilation = False

    for sheet_name in src_wb.sheetnames:
        src_ws = src_wb[sheet_name]
        header_row, parameters_col = find_header_and_parameters_col(src_ws)

        if header_row is None or parameters_col is None:
            continue

        modified_rows: list[tuple[int, str, list[str], str]] = []

        for row_idx in range(header_row + 1, src_ws.max_row + 1):
            values = row_values(src_ws, row_idx)
            params_raw = src_ws.cell(row_idx, parameters_col).value

            if params_raw is None:
                continue

            parameters = str(params_raw)
            full_text = join_row_text(values)

            new_parameters, changes = patch_entity_parameters(parameters, full_text, args)

            if changes and new_parameters != parameters:
                modified_rows.append((row_idx, new_parameters, changes, full_text))

                if is_accumulator_row(full_text):
                    touched_accumulators = True

                if any("energyusage:" in change.lower() for change in changes):
                    touched_energy_usage = True

                if is_ventilation_row(full_text, new_parameters):
                    touched_ventilation = True

        if not modified_rows:
            continue

        out_ws = out_wb.create_sheet(title=sheet_name)
        copy_column_widths(src_ws, out_ws)

        dest_row = 1

        for src_header_row in range(1, header_row + 1):
            copy_row(src_ws, out_ws, src_header_row, dest_row)
            dest_row += 1

        for src_row_idx, new_parameters, changes, full_text in modified_rows:
            copy_row(src_ws, out_ws, src_row_idx, dest_row)
            out_ws.cell(dest_row, parameters_col).value = new_parameters

            row_id = src_ws.cell(src_row_idx, 1).value
            readable_name = src_ws.cell(src_row_idx, 2).value

            report.note(
                f"{source_label} / {sheet_name} / row={src_row_idx} "
                f"id={row_id!r} name={readable_name!r}: "
                + "; ".join(changes)
            )

            total_rows_changed += 1
            dest_row += 1

    if total_rows_changed == 0:
        report.warn(f"{source_label}: aucune ligne Entities patchée.")
        return False

    if not touched_accumulators:
        report.warn(
            f"{source_label}: aucune ligne Accumulators trouvée. "
            f"La capacité batterie ne sera peut-être pas multipliée sur cette version."
        )

    if not touched_energy_usage:
        report.warn(
            f"{source_label}: aucun EnergyUsage trouvé. "
            f"La consommation affichée des moteurs/appareils ne sera peut-être pas réduite."
        )

    if not touched_ventilation:
        report.warn(
            f"{source_label}: aucune ligne ventilation/oxygène trouvée. "
            f"Ce n'est pas bloquant si General.xlsx applique déjà l'oxygène."
        )

    output_entities.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_entities)
    report.file(output_entities)

    return True


# =============================================================================
# MANIFEST, README, CACHE
# =============================================================================

def make_supported_versions(game_version: str) -> list[str]:
    gv = game_version.strip()

    candidates = [
        gv,
        gv.lower(),
        gv.replace("Patch", "patch"),
        "2026.1",
    ]

    match = re.search(r"(\d{4}\.\d+)", gv)
    if match:
        candidates.append(match.group(1))

    return unique_preserve_order(candidates)


def write_manifest(mod_dir: Path, args: argparse.Namespace) -> None:
    supported_versions = make_supported_versions(args.game_version)

    manifest = {
        "name": MOD_DISPLAY_NAME,
        "version": MOD_VERSION,
        "description": (
            f"Longue immersion : oxygène x{args.oxygen_factor:g}, "
            f"discipline/fatigue x{args.discipline_factor:g}, "
            f"capacité accumulateurs x{args.battery_capacity_factor:g}, "
            f"EnergyUsage x{args.energy_usage_factor:g}. "
            "Mod Data Sheets sans DLL."
        ),
        "author": MOD_AUTHOR,

        # Ancien format conservé pour compatibilité.
        "minGameVersion": "2026.1",
        "maxGameVersion": "",

        # Format récent utilisé par le launcher pour éviter l'alerte obsolète.
        "supportedGameVersions": supported_versions,

        "assemblyName": "",
        "permissions": [],
        "steamFileId": 0,
    }

    (mod_dir / "Manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_readme(mod_dir: Path, args: argparse.Namespace) -> None:
    readme = f"""Long Submerged 10x+ v{MOD_VERSION}

Paramètres utilisés :
- Oxygène : consommation divisée par {args.oxygen_factor:g}
- Discipline/fatigue sous l'eau : pertes divisées par {args.discipline_factor:g}
- Accumulators / batterie : capacité multipliée par {args.battery_capacity_factor:g}
- Tous les EnergyUsage d'équipements : multipliés par {args.energy_usage_factor:g}
- Ventilation / filtration : gain multiplié par {args.ventilation_factor:g}

Installation :
1. Fermer UBOAT.
2. Générer le mod avec --force --clear-cache.
3. Activer le mod dans le launcher.
4. Confirmer que le mod affiche "active".
5. Lancer une nouvelle carrière pour valider les changements Entities.xlsx.

Notes :
- Si l'UI batterie affiche encore une autonomie vanilla, la sauvegarde garde probablement d'anciennes données Equipment.
- Si tu utilises plusieurs mods qui touchent Entities.xlsx ou General.xlsx, mets Long Submerged 10x+ après eux dans le launcher.
"""

    (mod_dir / "README_LongSubmerged10x.txt").write_text(readme, encoding="utf-8")


def clear_uboat_cache(local_uboat_dir: Path, report: PatchReport) -> None:
    for folder_name in ("Cache", "Data Sheets", "Temp"):
        folder = local_uboat_dir / folder_name

        if not folder.exists():
            continue

        for child in folder.iterdir():
            try:
                if child.is_dir():
                    shutil.rmtree(child)
                else:
                    child.unlink()
            except OSError as exc:
                report.warn(f"Impossible de supprimer {child}: {exc}")

        report.note(f"Cache vidé : {folder}")


# =============================================================================
# ARGUMENTS + MAIN
# =============================================================================

def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Génère le mod UBOAT Long Submerged 10x+ corrigé.",
    )

    parser.add_argument(
        "--uboat",
        type=Path,
        required=True,
        help=r"Chemin d'installation UBOAT, ex: C:\Program Files (x86)\Steam\steamapps\common\UBOAT",
    )

    parser.add_argument(
        "--out",
        type=Path,
        default=default_local_mods_root() / MOD_FOLDER_NAME,
        help=(
            "Dossier de sortie du mod. Par défaut : "
            "AppData\\LocalLow\\Deep Water Studio\\UBOAT\\Mods\\LongSubmerged10x"
        ),
    )

    parser.add_argument(
        "--force",
        action="store_true",
        help="Écrase le dossier de mod existant.",
    )

    parser.add_argument(
        "--clear-cache",
        action="store_true",
        help="Vide Cache, Data Sheets et Temp dans AppData\\LocalLow\\Deep Water Studio\\UBOAT.",
    )

    parser.add_argument(
        "--game-version",
        default=DEFAULT_GAME_VERSION,
        help='Version affichée par le launcher, ex: "2026.1 Patch 20".',
    )

    parser.add_argument(
        "--oxygen-factor",
        type=float,
        default=DEFAULT_OXYGEN_FACTOR,
        help="Facteur de durée d'oxygène. 10 = consommation divisée par 10. Défaut : 15.",
    )

    parser.add_argument(
        "--discipline-factor",
        type=float,
        default=DEFAULT_DISCIPLINE_FACTOR,
        help="Facteur de durée discipline/fatigue. 10 = pertes divisées par 10. Défaut : 15.",
    )

    parser.add_argument(
        "--battery-capacity-factor",
        type=float,
        default=DEFAULT_BATTERY_CAPACITY_FACTOR,
        help="Multiplicateur de capacité des Accumulators. Défaut : 10.",
    )

    parser.add_argument(
        "--energy-usage-factor",
        type=float,
        default=DEFAULT_EQUIPMENT_ENERGY_USAGE_FACTOR,
        help="Multiplicateur des EnergyUsage d'équipements. 0.1 = consommation divisée par 10. Défaut : 0.1.",
    )

    parser.add_argument(
        "--ventilation-factor",
        type=float,
        default=DEFAULT_VENTILATION_FACTOR,
        help="Multiplicateur des gains de ventilation/filtration. Défaut : 10.",
    )

    parser.add_argument(
        "--patch-energy-base-scale",
        action="store_true",
        default=DEFAULT_PATCH_ENERGY_BASE_SCALE,
        help="Patche aussi General.xlsx / Energy Base Scale. Désactivé par défaut.",
    )

    args = parser.parse_args(argv)

    for name in (
        "oxygen_factor",
        "discipline_factor",
        "battery_capacity_factor",
        "ventilation_factor",
    ):
        value = getattr(args, name)

        if value <= 0:
            raise ValueError(f"--{name.replace('_', '-')} doit être > 0.")

    if args.energy_usage_factor <= 0:
        raise ValueError("--energy-usage-factor doit être > 0.")

    return args


def main(argv: list[str]) -> int:
    try:
        args = parse_args(argv)
    except Exception as exc:
        print(f"Arguments invalides : {exc}", file=sys.stderr)
        return 2

    report = PatchReport()

    try:
        uboat_root = args.uboat.expanduser().resolve()
        data_sheets_dir = resolve_data_sheets_dir(uboat_root)

        out_mod_dir = args.out.expanduser().resolve()
        out_data_sheets = out_mod_dir / "Data Sheets"

        ensure_clean_directory(out_mod_dir, force=args.force)
        out_data_sheets.mkdir(parents=True, exist_ok=True)

        write_manifest(out_mod_dir, args)
        write_readme(out_mod_dir, args)

        general_patches = build_general_patches(args)
        general_files = find_datasheet_files(data_sheets_dir, "General.xlsx")
        entities_files = find_datasheet_files(data_sheets_dir, "Entities.xlsx")

        if not general_files:
            raise FileNotFoundError(f"Aucun General.xlsx trouvé sous {data_sheets_dir}")

        if not entities_files:
            raise FileNotFoundError(f"Aucun Entities.xlsx trouvé sous {data_sheets_dir}")

        general_changed = 0

        for vanilla_general in general_files:
            relative = vanilla_general.relative_to(data_sheets_dir)
            output_general = out_data_sheets / relative

            if create_general_override(
                vanilla_general,
                output_general,
                general_patches,
                report,
                data_sheets_dir,
            ):
                general_changed += 1

        entities_changed = 0

        for vanilla_entities in entities_files:
            relative = vanilla_entities.relative_to(data_sheets_dir)
            output_entities = out_data_sheets / relative

            if create_entities_override(
                vanilla_entities,
                output_entities,
                report,
                data_sheets_dir,
                args,
            ):
                entities_changed += 1

        if general_changed == 0:
            report.warn(
                "Aucun General.xlsx n'a été patché. "
                "L'oxygène/discipline ne changera probablement pas."
            )

        if entities_changed == 0:
            report.warn(
                "Aucun Entities.xlsx n'a été patché. "
                "Batterie/ventilation ne changeront probablement pas."
            )

        if args.clear_cache:
            clear_uboat_cache(default_local_uboat_dir(), report)

    except Exception as exc:
        print("\nERREUR : génération impossible.", file=sys.stderr)
        print(str(exc), file=sys.stderr)
        return 1

    print("\n=== Long Submerged 10x+ généré avec succès ===")
    print(f"Dossier du mod : {out_mod_dir}")
    print(f"Version déclarée : {args.game_version}")

    print("\nProfil appliqué :")
    print(f"  - Oxygène : durée x{args.oxygen_factor:g}")
    print(f"  - Discipline/fatigue : durée x{args.discipline_factor:g}")
    print(f"  - Accumulators : capacité x{args.battery_capacity_factor:g}")
    print(f"  - EnergyUsage équipements : x{args.energy_usage_factor:g}")
    print(f"  - Ventilation/filtration : gain x{args.ventilation_factor:g}")

    print("\nFichiers générés :")
    print(f"  - {out_mod_dir / 'Manifest.json'}")
    print(f"  - {out_mod_dir / 'README_LongSubmerged10x.txt'}")

    for file_path in report.changed_files:
        print(f"  - {file_path}")

    if report.info:
        print("\nChangements détectés :")

        for item in report.info:
            print(f"  - {item}")

    if report.warnings:
        print("\nAvertissements :")

        for warning in report.warnings:
            print(f"  - {warning}")

    print("\nÉtapes obligatoires :")
    print("  1. Ferme complètement UBOAT.")
    print("  2. Lance le script avec --force --clear-cache.")
    print("  3. Ouvre UBOAT Launcher > Mods.")
    print("  4. Active Long Submerged 10x+ et vérifie qu'il affiche bien 'active'.")
    print("  5. Démarre une NOUVELLE carrière pour tester batterie/ventilation.")
    print("  6. Dans le doute, teste sans autres mods qui touchent General.xlsx ou Entities.xlsx.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
