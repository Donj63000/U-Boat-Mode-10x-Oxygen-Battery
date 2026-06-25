#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
UBOAT - Long Submerged 10x+
Générateur de mod Data Sheets.

But de cette version :
- Garder la batterie longue durée qui fonctionne déjà.
- Corriger l'air / "qualité de l'air" pour viser environ 90 jours.
- Ne plus casser la ventilation : la ligne Ventilation reste vanilla par défaut.
- SurfaceSafe 1.4.7 : ne plus modifier l'oxygène ni la capacité d'air dans les Data Sheets.
- Appliquer l'oxygène long uniquement au runtime sur le drain négatif de respiration.

Prérequis :
    py -m pip install openpyxl

Commande conseillée :
    py build_uboat_long_submerged_mod.py --uboat "C:\\Program Files (x86)\\Steam\\steamapps\\common\\UBOAT" --force --clear-cache

Important :
    Teste l'air sur une NOUVELLE carrière après génération.
"""

from __future__ import annotations

import argparse
import copy
import json
import math
import os
import re
import shutil
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


# =============================================================================
# CONFIGURATION
# =============================================================================

MOD_FOLDER_NAME = "LongSubmerged10x"
MOD_DISPLAY_NAME = "Long Submerged 10x+"
MOD_VERSION = "1.4.16"
MOD_AUTHOR = "VotreNomOuVotreEquipe"
MOD_ASSEMBLY_NAME = "LongSubmerged10xPatch_1_4_16"

DEFAULT_GAME_VERSION = "2026.1 Patch 20"

# DonJ : le slider oxygene 100 vise environ 90 jours en conditions reelles.
DEFAULT_AIR_CAPACITY_FACTOR = 1.0

# On garde aussi la baisse de consommation par personnage, mais ce n'est plus le levier principal :
# certaines versions / configs semblent garder un minimum visible dans l'UI, par exemple "Équipage -4/min".
DEFAULT_OXYGEN_CONSUMPTION_FACTOR = 250.0

# Discipline/fatigue proportionnelles à l'immersion longue.
DEFAULT_DISCIPLINE_FACTOR = 15.0

# Batterie : le XLSX garde un fallback x0.1 restaurable, mais le runtime Mega Batterie coupe le drain.
DEFAULT_BATTERY_CAPACITY_FACTOR = 10.0
DEFAULT_EQUIPMENT_ENERGY_USAGE_FACTOR = 0.10

# Vitesse : seuls les deux derniers crans avant sont boostes.
DEFAULT_FAST_SPEED_FACTOR = 8.0
DEFAULT_FAST_SPEED_FUEL_FACTOR = 8.0
DEFAULT_FAST_SPEED_TOP_GEARS = 2
DEFAULT_PLAYER_SUBMARINE_MAX_SPEED = 45.0
DEFAULT_PATCH_PLAYER_SUBMARINE_SPEED_ROWS = False

# Mega torpilles : active par defaut, parce que le mod doit livrer le comportement demande.
# Je touche uniquement aux degats et aux effets d'explosion, pas a la vitesse ni a la portee.
DEFAULT_MEGA_TORPEDOES = True
DEFAULT_TORPEDO_DAMAGE_FACTOR = 10.0
DEFAULT_TORPEDO_CREW_DAMAGE_FACTOR = 10.0
DEFAULT_TORPEDO_EXPLOSION_RADIUS_FACTOR = 3.0
DEFAULT_TORPEDO_EXPLOSION_INTENSITY_FACTOR = 3.0
DEFAULT_PERFECT_TORPEDO_RELIABILITY = True
DEFAULT_TORPEDO_DUD_CHANCE = 0.0
DEFAULT_TORPEDO_MAGNETIC_FAILURE_CHANCE = 0.0
DEFAULT_TORPEDO_PREMATURE_MAGNETIC_CHANCE = 0.0

# IMPORTANT :
# On ne modifie plus la ventilation par défaut.
# Une ancienne tentative modifiait la ligne ventilation deux fois sur EnergyUsage et montait RegenerationLimit trop haut.
# Pour "la ventilation marche normalement", on ne copie plus cette ligne dans l'override.
DEFAULT_PATCH_VENTILATION = False

# Optionnel, désactivé par défaut.
# Si activé, seulement OxygenGain est augmenté et EnergyUsage est réduit UNE seule fois.
# RegenerationLimit reste vanilla pour éviter les comportements bizarres.
DEFAULT_VENTILATION_GAIN_FACTOR = 1.0
DEFAULT_PATCH_POTASSIUM = False
DEFAULT_POTASSIUM_DURATION_FACTOR = 1.0

# Patch optionnel de Energy Base Scale dans General.xlsx. Désactivé car la batterie marche via Entities.xlsx.
DEFAULT_PATCH_ENERGY_BASE_SCALE = False


# =============================================================================
# STRUCTURES
# =============================================================================

@dataclass(frozen=True)
class GeneralPatch:
    row_id: str
    multiplier: float
    expected_category: str | None
    reason: str


@dataclass
class PatchReport:
    changed_files: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    info: list[str] = field(default_factory=list)
    counters: dict[str, int] = field(default_factory=dict)

    def file(self, path: Path) -> None:
        path_text = str(path)
        if path_text not in self.changed_files:
            self.changed_files.append(path_text)

    def warn(self, message: str) -> None:
        self.warnings.append(message)

    def note(self, message: str) -> None:
        self.info.append(message)

    def inc(self, key: str, amount: int = 1) -> None:
        self.counters[key] = self.counters.get(key, 0) + amount


@dataclass
class GeneralRowHit:
    patch: GeneralPatch
    source_row: int
    category_norm: str | None
    category_row: int | None
    value_cols: list[int]


@dataclass
class ModifiedRow:
    source_row: int
    new_values_by_col: dict[int, Any]
    changes: list[str]
    readable_context: str


# =============================================================================
# OUTILS GÉNÉRAUX
# =============================================================================

def norm_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").strip().lower()
    return re.sub(r"\s+", " ", text)


def norm_category(value: Any) -> str:
    text = norm_text(value)
    if not text:
        return ""
    if not text.startswith("/"):
        text = "/" + text
    return text


def contains_any(text: str, needles: Iterable[str]) -> bool:
    lowered = text.lower()
    return any(needle.lower() in lowered for needle in needles)


def safe_float(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None

    text = str(value).strip()
    if not text:
        return None

    if text.endswith("%"):
        text = text[:-1].strip()

    text = text.replace("\u00a0", "")
    text = text.replace(" ", "")
    text = text.replace(",", ".")

    try:
        number = float(text)
    except ValueError:
        return None

    return number if math.isfinite(number) else None


def scale_numeric_excel_value(original_value: Any, multiplier: float) -> float:
    parsed = safe_float(original_value)
    if parsed is None:
        raise ValueError(f"valeur non numérique : {original_value!r}")
    return parsed * multiplier


def copy_cell(src: Cell, dst: Cell) -> None:
    dst.value = src.value

    if src.has_style:
        dst._style = copy.copy(src._style)

    dst.number_format = src.number_format

    if src.font:
        dst.font = copy.copy(src.font)
    if src.fill:
        dst.fill = copy.copy(src.fill)
    if src.border:
        dst.border = copy.copy(src.border)
    if src.alignment:
        dst.alignment = copy.copy(src.alignment)
    if src.protection:
        dst.protection = copy.copy(src.protection)
    if src.comment:
        dst.comment = copy.copy(src.comment)


def copy_row(src_ws: Worksheet, dst_ws: Worksheet, src_row: int, dst_row: int) -> None:
    for col in range(1, src_ws.max_column + 1):
        copy_cell(src_ws.cell(src_row, col), dst_ws.cell(dst_row, col))

    if src_ws.row_dimensions[src_row].height:
        dst_ws.row_dimensions[dst_row].height = src_ws.row_dimensions[src_row].height


def copy_column_widths(src_ws: Worksheet, dst_ws: Worksheet) -> None:
    for key, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[key].width = dim.width


def find_existing_override_row(dst_ws: Worksheet, source_row_id: Any) -> int | None:
    row_id = norm_text(source_row_id)
    if not row_id:
        return None

    for row in range(1, dst_ws.max_row + 1):
        if norm_text(dst_ws.cell(row, 1).value) == row_id:
            return row

    return None


def get_sheet_case_insensitive(workbook: Workbook, wanted_name: str) -> Worksheet | None:
    wanted = norm_text(wanted_name)
    for name in workbook.sheetnames:
        if norm_text(name) == wanted:
            return workbook[name]
    return None


def default_local_uboat_dir() -> Path:
    user_profile = os.environ.get("USERPROFILE")
    if user_profile:
        return Path(user_profile) / "AppData" / "LocalLow" / "Deep Water Studio" / "UBOAT"
    return Path.cwd() / "UBOAT_LocalLow"


def default_local_mods_root() -> Path:
    return default_local_uboat_dir() / "Mods"


def resolve_data_sheets_dir(uboat_root: Path) -> Path:
    data_sheets = uboat_root / "UBOAT_Data" / "Data Sheets"
    if not data_sheets.exists():
        raise FileNotFoundError(
            f"Dossier introuvable : {data_sheets}\n"
            f"--uboat doit pointer vers le dossier d'installation UBOAT."
        )
    return data_sheets


def ensure_clean_directory(path: Path, force: bool) -> None:
    if path.exists():
        if not force:
            raise FileExistsError(
                f"Le dossier existe déjà : {path}\n"
                f"Relance avec --force pour le remplacer."
            )
        shutil.rmtree(path)
    path.mkdir(parents=True, exist_ok=True)


def find_datasheet_files(data_sheets_dir: Path, filename: str | None = None) -> list[Path]:
    files: list[Path] = []
    for path in data_sheets_dir.rglob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        if filename is not None and path.name.lower() != filename.lower():
            continue
        files.append(path)
    return sorted(files, key=lambda p: str(p.relative_to(data_sheets_dir)).lower())


def find_type_ix_dlc_data_sheets_dir(uboat_root: Path) -> Path | None:
    data_sheets = (
        uboat_root
        / "UBOAT_Data"
        / "StreamingAssets"
        / "Packages"
        / "uboat.dlc.type-ix"
        / "Data Sheets"
    )
    return data_sheets if data_sheets.exists() else None


def unique_preserve_order(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []

    for value in values:
        key = value.strip().lower()
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(value.strip())

    return result


def row_values(ws: Worksheet, row: int) -> list[Any]:
    return [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]


def join_row_text(values: list[Any]) -> str:
    return " | ".join(str(value) for value in values if value is not None)


def readable_row_label(ws: Worksheet, row: int) -> str:
    values = row_values(ws, row)
    parts: list[str] = []
    for index in range(min(len(values), 6)):
        value = values[index]
        if value is not None and str(value).strip():
            parts.append(str(value).strip())
    return " | ".join(parts) if parts else f"row={row}"


def find_header_row(ws: Worksheet) -> int:
    """
    Cherche la première ligne d'en-tête probable.
    Dans les datasheets UBOAT, elle est très souvent en ligne 1.
    """
    best_row = 1
    best_score = -1

    for row in range(1, min(ws.max_row, 25) + 1):
        values = [norm_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        score = 0
        for value in values:
            if value in {"id", "name", "parameters", "type", "value", "category"}:
                score += 3
            elif value:
                score += 1

        if score > best_score:
            best_score = score
            best_row = row

    return best_row


def find_columns_by_header(ws: Worksheet, header_row: int) -> dict[str, list[int]]:
    headers: dict[str, list[int]] = {}
    for col in range(1, ws.max_column + 1):
        header = norm_text(ws.cell(header_row, col).value)
        if not header:
            continue
        headers.setdefault(header, []).append(col)
    return headers


# =============================================================================
# GENERAL.XLSX
# =============================================================================

def build_general_patches(args: argparse.Namespace) -> list[GeneralPatch]:
    patches = [
        GeneralPatch(
            row_id="Underwater Discipline Loss",
            multiplier=1.0 / args.discipline_factor,
            expected_category="/Discipline",
            reason=f"Discipline sous l'eau environ {args.discipline_factor:g}x moins punitive.",
        ),
        GeneralPatch(
            row_id="Fatigue - Per Day",
            multiplier=1.0 / args.discipline_factor,
            expected_category="/Discipline",
            reason=f"Fatigue longue durée environ {args.discipline_factor:g}x plus lente.",
        ),
        GeneralPatch(
            row_id="Fatigue - Max Penalty",
            multiplier=1.0 / args.discipline_factor,
            expected_category="/Discipline",
            reason="Pénalité maximale de fatigue réduite proportionnellement.",
        ),
    ]

    if args.patch_energy_base_scale:
        patches.append(
            GeneralPatch(
                row_id="Energy Base Scale",
                multiplier=args.battery_capacity_factor,
                expected_category="/Resources",
                reason="Optionnel : réserve d'énergie globale augmentée.",
            )
        )

    return patches


def index_settings_sheet(ws: Worksheet) -> tuple[
    dict[str, int],
    dict[tuple[str, str], int],
    dict[str, tuple[int, str | None]],
]:
    category_rows: dict[str, int] = {}
    rows_by_category_and_id: dict[tuple[str, str], int] = {}
    rows_by_id: dict[str, tuple[int, str | None]] = {}

    current_category: str | None = None

    for row in range(1, ws.max_row + 1):
        first_raw = ws.cell(row, 1).value
        first = str(first_raw).strip() if first_raw is not None else ""

        if not first:
            continue

        if first.startswith("/"):
            current_category = norm_category(first)
            category_rows[current_category] = row
            continue

        row_id_norm = norm_text(first)
        if not row_id_norm:
            continue

        if current_category:
            rows_by_category_and_id[(current_category, row_id_norm)] = row

        rows_by_id.setdefault(row_id_norm, (row, current_category))

    return category_rows, rows_by_category_and_id, rows_by_id


def find_value_cols_for_general_row(ws: Worksheet, row: int) -> list[int]:
    value_cols: list[int] = []

    for col in range(2, ws.max_column + 1):
        if safe_float(ws.cell(row, col).value) is not None:
            value_cols.append(col)

    return value_cols


def find_general_hits(
    ws: Worksheet,
    patches: list[GeneralPatch],
    report: PatchReport,
    source_label: str,
) -> list[GeneralRowHit]:
    category_rows, rows_by_category_and_id, rows_by_id = index_settings_sheet(ws)
    hits: list[GeneralRowHit] = []

    for patch in patches:
        expected_category_norm = norm_category(patch.expected_category) if patch.expected_category else None
        row: int | None = None
        actual_category: str | None = None

        if expected_category_norm:
            row = rows_by_category_and_id.get((expected_category_norm, norm_text(patch.row_id)))
            if row is not None:
                actual_category = expected_category_norm

        if row is None:
            fallback = rows_by_id.get(norm_text(patch.row_id))
            if fallback:
                row, actual_category = fallback

        if row is None:
            report.warn(f"{source_label}: ligne introuvable dans Settings : {patch.row_id!r}")
            continue

        value_cols = find_value_cols_for_general_row(ws, row)

        if not value_cols:
            report.warn(
                f"{source_label}: {patch.row_id!r} trouvé ligne {row}, "
                f"mais aucune colonne numérique n'a été trouvée."
            )
            continue

        hits.append(
            GeneralRowHit(
                patch=patch,
                source_row=row,
                category_norm=actual_category,
                category_row=category_rows.get(actual_category or "") if actual_category else None,
                value_cols=value_cols,
            )
        )

    return hits


def apply_general_hit_values(
    src_ws: Worksheet,
    out_ws: Worksheet,
    hit: GeneralRowHit,
    dest_row: int,
    report: PatchReport,
    source_label: str,
) -> None:
    changes: list[str] = []

    for value_col in hit.value_cols:
        original = src_ws.cell(hit.source_row, value_col).value
        new_value = scale_numeric_excel_value(original, hit.patch.multiplier)
        out_ws.cell(dest_row, value_col).value = new_value

        header = src_ws.cell(hit.category_row or 1, value_col).value
        header_text = str(header).strip() if header is not None else f"colonne {value_col}"
        changes.append(f"{header_text}: {original!r} -> {new_value!r}")

    report.note(
        f"{source_label} / Settings / {hit.patch.row_id}: "
        + "; ".join(changes)
        + f" | {hit.patch.reason}"
    )

def create_general_override(
    vanilla_general: Path,
    output_general: Path,
    patches: list[GeneralPatch],
    report: PatchReport,
    data_sheets_dir: Path,
) -> bool:
    source_label = str(vanilla_general.relative_to(data_sheets_dir))
    src_wb = load_workbook(vanilla_general, data_only=False)
    src_ws = get_sheet_case_insensitive(src_wb, "Settings")

    if src_ws is None:
        report.warn(f"{source_label}: onglet Settings introuvable, fichier ignoré.")
        return False

    hits = find_general_hits(src_ws, patches, report, source_label)
    if not hits:
        report.warn(f"{source_label}: aucune valeur General.xlsx patchée.")
        return False

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = src_ws.title
    copy_column_widths(src_ws, out_ws)

    dest_row = 1
    written_categories: set[str] = set()

    categorized_hits = [hit for hit in hits if hit.category_norm and hit.category_row]
    uncategorized_hits = [hit for hit in hits if not (hit.category_norm and hit.category_row)]

    for hit in sorted(categorized_hits, key=lambda h: (h.category_row or 0, h.source_row)):
        assert hit.category_norm is not None
        assert hit.category_row is not None

        if hit.category_norm not in written_categories:
            copy_row(src_ws, out_ws, hit.category_row, dest_row)
            dest_row += 1
            written_categories.add(hit.category_norm)

        copy_row(src_ws, out_ws, hit.source_row, dest_row)
        apply_general_hit_values(src_ws, out_ws, hit, dest_row, report, source_label)
        dest_row += 1

    if uncategorized_hits:
        if dest_row > 1:
            dest_row += 1

        first_row_text = " ".join(
            norm_text(src_ws.cell(1, col).value)
            for col in range(1, src_ws.max_column + 1)
        )

        if "value" in first_row_text or "id" in first_row_text:
            copy_row(src_ws, out_ws, 1, dest_row)
            dest_row += 1

        for hit in sorted(uncategorized_hits, key=lambda h: h.source_row):
            copy_row(src_ws, out_ws, hit.source_row, dest_row)
            apply_general_hit_values(src_ws, out_ws, hit, dest_row, report, source_label)
            dest_row += 1

    output_general.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_general)
    report.file(output_general)
    return True


# =============================================================================
# PATCH PARAMÈTRES TEXTUELS, EX: "EnergyUsage = 0.0002"
# =============================================================================

NUMBER_PATTERN = r"[-+]?\d+(?:[.,]\d+)?(?:[eE][-+]?\d+)?"
KEY_VALUE_PATTERN = re.compile(
    r"(?P<prefix>(?P<key>[A-Za-z][A-Za-z0-9_ ]*?)\s*(?P<op>[*+\-/]?=)\s*)"
    r"(?P<number>" + NUMBER_PATTERN + r")(?P<percent>%?)"
)


def format_number(value: float, percent: str = "") -> str:
    return f"{value:.10g}" + percent


def format_csharp_float(value: float) -> str:
    return f"{value:.10g}f"


def normalize_parameter_key(key: str) -> str:
    return re.sub(r"[^a-z0-9]", "", key.lower())


def replace_keyed_parameters(
    parameters: str,
    multiplier_by_key: dict[str, float],
    *,
    only_when: Callable[[str, re.Match[str]], bool] | None = None,
) -> tuple[str, list[str]]:
    """
    Remplace les valeurs dans une chaîne Parameters.

    multiplier_by_key utilise des clés normalisées :
      "EnergyUsage" -> "energyusage"
      "Air Quality Capacity" -> "airqualitycapacity"
    """
    normalized_multipliers = {
        normalize_parameter_key(key): value
        for key, value in multiplier_by_key.items()
    }

    changes: list[str] = []

    def replace(match: re.Match[str]) -> str:
        raw_key = match.group("key")
        normalized_key = normalize_parameter_key(raw_key)
        multiplier = normalized_multipliers.get(normalized_key)

        if multiplier is None:
            return match.group(0)

        if only_when is not None and not only_when(raw_key, match):
            return match.group(0)

        raw_number = match.group("number")
        percent = match.group("percent") or ""
        parsed = safe_float(raw_number + percent)

        if parsed is None:
            return match.group(0)

        scaled = parsed * multiplier
        replacement_number = format_number(scaled, percent)
        changes.append(f"{raw_key.strip()}: {raw_number}{percent} -> {replacement_number}")

        return match.group("prefix") + replacement_number

    new_parameters = KEY_VALUE_PATTERN.sub(replace, parameters)
    return new_parameters, changes


def find_parameters_col(ws: Worksheet, header_row: int) -> int | None:
    for col in range(1, ws.max_column + 1):
        if norm_text(ws.cell(header_row, col).value) == "parameters":
            return col

    # Fallback historique : colonne P.
    if ws.max_column >= 16:
        return 16

    return None


# =============================================================================
# DÉTECTION DES LIGNES À PATCHER
# =============================================================================

VENTILATION_WORDS = (
    "ventilation",
    "ventilator",
    "air recircul",
    "recirculating",
    "potassium",
    "absorber",
    "absorbent",
    "kali",
)

COMPRESSOR_WORDS = (
    "compressor",
    "compresseur",
    "compressed air",
    "compressedair",
    "druckluft",
)

ACCUMULATOR_WORDS = (
    "accumulator",
    "accumulators",
    "akkumulator",
    "akkumulatoren",
    "battery upgrade",
    "battery capacity",
)

PLAYER_SUBMARINE_WORDS = (
    "u-boat",
    "uboat",
    "u boat",
    "type ii",
    "type vii",
    "type viic",
    "type viib",
    "type ix",
    "submarine",
    "player ship",
    "player uboat",
)

AIR_CONTEXT_WORDS = (
    "air quality",
    "airquality",
    "air supply",
    "airsupply",
    "breathable",
    "atmosphere",
    "atmospheric",
    "oxygen capacity",
    "oxygen supply",
    "oxygen amount",
    "oxygen reserve",
    "starting air",
    "start air",
    "initial air",
)

AIR_BANNED_WORDS = (
    "compressed air",
    "compressedair",
    "compressor",
    "compresseur",
    "torpedo",
    "torpille",
    "diesel",
    "fuel",
    "emission",
    "exhaust",
    "noise",
    "visibility",
    "detection",
    "contrail",
)

CONSUMPTION_BANNED_WORDS = (
    "consumption",
    "consume",
    "usage",
    "use",
    "gain",
    "regeneration",
    "recharge",
    "efficiency",
    "noise",
    "rate",
    "persecond",
    "perminute",
    "percharacter",
    "per character",
)


# Clés exactes qui ressemblent à la CAPACITÉ de l'atmosphère, pas à sa consommation.
AIR_CAPACITY_KEYS = {
    "aircapacity",
    "airqualitycapacity",
    "airqualitymax",
    "airqualitymaximum",
    "maxairquality",
    "maxair",
    "airmax",
    "airamount",
    "airqualityamount",
    "airbaseamount",
    "baseairamount",
    "startingair",
    "startair",
    "initialair",
    "atmospherecapacity",
    "atmosphericcapacity",
    "atmospheremax",
    "atmosphereamount",
    "atmosphere",
    "oxygenmax",
    "maxoxygen",
    "oxygencapacity",
    "oxygenamount",
    "oxygenreserve",
    "oxygensupply",
    "oxygenstorage",
    "baseoxygen",
    "oxygenbase",
    "startingoxygen",
    "initialoxygen",
    "breathableair",
    "breathableaircapacity",
    "co2capacity",
    "carbondioxidecapacity",
}


ENERGY_USAGE_KEYS = {
    "energyusage",
    "electricityusage",
    "powerusage",
}


def patch_energy_usage_parameters(
    parameters: str,
    *,
    consumption_factor: float,
    recharge_factor: float,
) -> tuple[str, list[str], dict[str, int]]:
    """
    Je garde deux règles distinctes :
    - valeur positive : l'équipement consomme, donc je réduis la consommation ;
    - valeur négative : l'équipement produit/recharge, donc je laisse la recharge vanilla.
    """
    changes: list[str] = []
    counters = {
        "energy_usage_rows": 0,
        "energy_recharge_rows": 0,
    }

    def replace(match: re.Match[str]) -> str:
        raw_key = match.group("key")
        if normalize_parameter_key(raw_key) not in ENERGY_USAGE_KEYS:
            return match.group(0)

        raw_number = match.group("number")
        percent = match.group("percent") or ""
        parsed = safe_float(raw_number + percent)

        if parsed is None or parsed == 0:
            return match.group(0)

        if parsed < 0:
            return match.group(0)

        multiplier = consumption_factor
        counter_key = "energy_usage_rows"

        scaled = parsed * multiplier
        replacement_number = format_number(scaled, percent)
        changes.append(f"{raw_key.strip()}: {raw_number}{percent} -> {replacement_number}")
        counters[counter_key] = 1

        return match.group("prefix") + replacement_number

    new_parameters = KEY_VALUE_PATTERN.sub(replace, parameters)
    return new_parameters, changes, counters


ACCUMULATOR_CAPACITY_KEYS = {
    "energycapacity",
    "batterycapacity",
    "capacitymultiplier",
    "capacityscale",
    "capacity",
    "maxcapacity",
    "energycapacitygain",
    "batterycapacitygain",
    "energystorage",
    "storage",
    "value",
}


def is_ventilation_row(text: str) -> bool:
    lowered = text.lower()
    return contains_any(lowered, VENTILATION_WORDS) or "oxygengain" in lowered or "regenerationlimit" in lowered


def is_compressor_row(text: str) -> bool:
    return contains_any(text.lower(), COMPRESSOR_WORDS)


def is_accumulator_row(text: str) -> bool:
    lowered = text.lower()
    return contains_any(lowered, ACCUMULATOR_WORDS) and not is_compressor_row(lowered)


def is_player_submarine_context(text: str) -> bool:
    return contains_any(text.lower(), PLAYER_SUBMARINE_WORDS)


def has_air_context(text: str) -> bool:
    lowered = text.lower()

    if contains_any(lowered, AIR_BANNED_WORDS):
        return False

    if contains_any(lowered, AIR_CONTEXT_WORDS):
        return True

    # Fallback : ligne de sous-marin + paramètre exact "Oxygen" ou "Air".
    if is_player_submarine_context(lowered) and ("oxygen" in lowered or "air" in lowered):
        return True

    return False


def is_air_capacity_key(key: str, row_text_full: str, value: float) -> bool:
    normalized = normalize_parameter_key(key)

    if value <= 0:
        return False

    if normalized in AIR_CAPACITY_KEYS:
        return True

    # Fallback très contrôlé :
    # "Oxygen = 3200" ou "Air = 3200" peut être le compteur de base.
    # On ne l'accepte que sur une ligne clairement liée au bateau / atmosphère.
    if normalized in {"oxygen", "air"}:
        lowered = row_text_full.lower()
        if is_player_submarine_context(lowered) or contains_any(lowered, AIR_CONTEXT_WORDS):
            return True

    return False


def is_probable_air_capacity_row(ws: Worksheet, row: int, header_row: int, parameters_col: int | None) -> bool:
    values = row_values(ws, row)
    text = join_row_text(values)

    if not has_air_context(text):
        return False

    lowered = text.lower()
    if contains_any(lowered, CONSUMPTION_BANNED_WORDS):
        return False

    if parameters_col is not None and ws.cell(row, parameters_col).value:
        params = str(ws.cell(row, parameters_col).value)
        for match in KEY_VALUE_PATTERN.finditer(params):
            value = safe_float((match.group("number") or "") + (match.group("percent") or ""))
            if value is not None and is_air_capacity_key(match.group("key"), text, value):
                return True

    headers = find_columns_by_header(ws, header_row)
    header_text_by_col = {
        col: norm_text(ws.cell(header_row, col).value)
        for col in range(1, ws.max_column + 1)
    }

    for col in range(2, ws.max_column + 1):
        value = safe_float(ws.cell(row, col).value)
        if value is None or value <= 0:
            continue

        header_text = header_text_by_col.get(col, "")

        if contains_any(header_text, ("capacity", "amount", "maximum", "max", "initial", "start", "base", "value")):
            return True

        # Sans header utile, on n'accepte que de grandes valeurs typiques d'un compteur.
        if value >= 100:
            return True

    return False


# =============================================================================
# PATCH DES LIGNES NON-GENERAL
# =============================================================================

def patch_parameter_cell(
    parameters: str,
    row_text_full: str,
    args: argparse.Namespace,
) -> tuple[str, list[str], dict[str, int]]:
    """
    Patch une cellule Parameters.
    Retourne :
      new_parameters, changes, counters
    """
    counters: dict[str, int] = {}
    new_parameters = parameters
    changes: list[str] = []
    lowered_context = row_text_full.lower()

    # Batterie : je garde l'autonomie validée en jeu.
    if is_accumulator_row(row_text_full):
        accumulator_multipliers = {key: args.battery_capacity_factor for key in ACCUMULATOR_CAPACITY_KEYS}

        patched, local_changes = replace_keyed_parameters(new_parameters, accumulator_multipliers)
        if local_changes:
            new_parameters = patched
            changes.extend(local_changes)
            counters["battery_capacity_rows"] = 1

    # Conso/recharge électrique : je réduis les consommateurs et je garde les rechargeurs en vanilla.
    # Multiplier les EnergyUsage négatifs rendait les diesels trop rapides à recharger.
    if (
        not is_ventilation_row(row_text_full)
        and not is_compressor_row(row_text_full)
        and any(key in normalize_parameter_key(new_parameters) for key in ENERGY_USAGE_KEYS)
    ):
        patched, local_changes, local_counters = patch_energy_usage_parameters(
            new_parameters,
            consumption_factor=args.energy_usage_factor,
            recharge_factor=args.battery_capacity_factor,
        )
        if local_changes:
            new_parameters = patched
            changes.extend(local_changes)
            for key, value in local_counters.items():
                if value:
                    counters[key] = value

    # Torpilles : je les applique maintenant en runtime pour pouvoir les couper proprement dans le menu.
    # Les XLSX gardent les valeurs vanilla, le toggle Mega Torpilles ajoute/retire les multiplicateurs en jeu.

    # SurfaceSafe 1.4.7 :
    # Ne plus modifier les capacites Air/Oxygen/Atmosphere dans les Data Sheets.
    # La transition surface recharge l'air avec les valeurs vanilla ; l'oxygene long est applique uniquement au runtime.

    # Ventilation : désactivé par défaut pour revenir au fonctionnement vanilla.
    # Si tu actives explicitement --patch-ventilation, on ne touche PAS RegenerationLimit.
    if args.patch_ventilation and is_ventilation_row(row_text_full):
        ventilation_multipliers = {
            "OxygenGain": args.ventilation_gain_factor,
            "AirQualityGain": args.ventilation_gain_factor,
            "AirGain": args.ventilation_gain_factor,
        }

        if args.energy_usage_factor != 1.0:
            ventilation_multipliers["EnergyUsage"] = args.energy_usage_factor

        patched, local_changes = replace_keyed_parameters(new_parameters, ventilation_multipliers)
        if local_changes:
            new_parameters = patched
            changes.extend(local_changes)
            counters["ventilation_rows"] = 1

    # Potassium absorbers : désactivé par défaut.
    # Si activé, on allonge seulement les durées, pas la ventilation elle-même.
    if args.patch_potassium and contains_any(lowered_context, ("potassium", "absorber", "absorbent")):
        potassium_multipliers = {
            "Duration": args.potassium_duration_factor,
            "WorkTime": args.potassium_duration_factor,
            "WorkingTime": args.potassium_duration_factor,
            "BurnTime": args.potassium_duration_factor,
            "Time": args.potassium_duration_factor,
        }

        patched, local_changes = replace_keyed_parameters(new_parameters, potassium_multipliers)
        if local_changes:
            new_parameters = patched
            changes.extend(local_changes)
            counters["potassium_rows"] = 1

    return new_parameters, unique_preserve_order(changes), counters


def patch_simple_air_capacity_cells(
    ws: Worksheet,
    row: int,
    header_row: int,
    parameters_col: int | None,
    args: argparse.Namespace,
) -> tuple[dict[int, Any], list[str], dict[str, int]]:
    """
    Patch les cellules numériques d'une ligne qui représente clairement une réserve/capacité d'air,
    même si la ligne n'utilise pas de colonne Parameters.
    """
    # SurfaceSafe 1.4.7 :
    # Désactivé volontairement. Modifier directement les réserves/capacités d'air peut casser
    # la transition immersion -> surface quand UBOAT recharge l'atmosphère.
    return {}, [], {}

    if not is_probable_air_capacity_row(ws, row, header_row, parameters_col):
        return {}, [], {}

    values_by_col: dict[int, Any] = {}
    changes: list[str] = []
    header_text_by_col = {
        col: norm_text(ws.cell(header_row, col).value)
        for col in range(1, ws.max_column + 1)
    }

    for col in range(2, ws.max_column + 1):
        if parameters_col is not None and col == parameters_col:
            continue

        original = ws.cell(row, col).value
        parsed = safe_float(original)

        if parsed is None or parsed <= 0:
            continue

        header_text = header_text_by_col.get(col, "")
        accept = False

        if contains_any(header_text, ("capacity", "amount", "maximum", "max", "initial", "start", "base", "value")):
            accept = True

        if parsed >= 100:
            accept = True

        if not accept:
            continue

        new_value = parsed * args.air_capacity_factor
        values_by_col[col] = new_value

        label = header_text if header_text else f"colonne {col}"
        changes.append(f"{label}: {original!r} -> {new_value!r}")

    if not values_by_col:
        return {}, [], {}

    return values_by_col, changes, {"air_capacity_cell_rows": 1}


def patch_player_submarine_speed_cells(
    ws: Worksheet,
    row: int,
    header_row: int,
    args: argparse.Namespace,
) -> tuple[dict[int, Any], list[str], dict[str, int]]:
    # 1.4.8 : la vitesse doit etre controlee par le runtime F10.
    # On ne copie plus de lignes Types joueur dans Entities.xlsx, sinon x1 resterait deja modde.
    if not getattr(args, "patch_player_submarine_speed_rows", False):
        return {}, [], {}

    if norm_text(ws.title) != "types":
        return {}, [], {}

    row_name_raw = ws.cell(row, 1).value
    row_name = str(row_name_raw or "")
    category = norm_text(ws.cell(row, 2).value)

    if "(player)" not in row_name.lower() or category != "submarine":
        return {}, [], {}

    speed_col = None
    for col in range(1, ws.max_column + 1):
        header = norm_text(ws.cell(header_row, col).value)
        if "speed" in header and "km/h" in header:
            speed_col = col
            break

    if speed_col is None:
        return {}, [], {}

    original = ws.cell(row, speed_col).value
    parsed = safe_float(original)

    if parsed is None or parsed <= 0:
        return {}, [], {}

    target_speed = args.player_submarine_max_speed
    if parsed >= target_speed:
        return {}, [], {}

    return (
        {speed_col: target_speed},
        [f"Speed (km/h): {original!r} -> {target_speed!r}"],
        {"player_submarine_speed_rows": 1},
    )


def patch_sheet_generic(
    src_ws: Worksheet,
    source_label: str,
    sheet_name: str,
    report: PatchReport,
    args: argparse.Namespace,
) -> tuple[Worksheet | None, list[ModifiedRow]]:
    header_row = find_header_row(src_ws)
    parameters_col = find_parameters_col(src_ws, header_row)

    modified_rows_by_source: dict[int, ModifiedRow] = {}

    for row_idx in range(header_row + 1, src_ws.max_row + 1):
        values = row_values(src_ws, row_idx)
        full_text = join_row_text(values)
        readable = readable_row_label(src_ws, row_idx)

        new_values_by_col: dict[int, Any] = {}
        changes: list[str] = []
        counters: dict[str, int] = {}

        if parameters_col is not None:
            params_raw = src_ws.cell(row_idx, parameters_col).value
            if params_raw is not None:
                params = str(params_raw)
                new_params, param_changes, param_counters = patch_parameter_cell(params, full_text, args)

                if param_changes and new_params != params:
                    new_values_by_col[parameters_col] = new_params
                    changes.extend(param_changes)
                    for key, value in param_counters.items():
                        counters[key] = counters.get(key, 0) + value

        simple_values, simple_changes, simple_counters = patch_simple_air_capacity_cells(
            src_ws,
            row_idx,
            header_row,
            parameters_col,
            args,
        )

        if simple_changes:
            new_values_by_col.update(simple_values)
            changes.extend(simple_changes)
            for key, value in simple_counters.items():
                counters[key] = counters.get(key, 0) + value

        speed_values, speed_changes, speed_counters = patch_player_submarine_speed_cells(
            src_ws,
            row_idx,
            header_row,
            args,
        )

        if speed_changes:
            new_values_by_col.update(speed_values)
            changes.extend(speed_changes)
            for key, value in speed_counters.items():
                counters[key] = counters.get(key, 0) + value

        if new_values_by_col:
            modified_rows_by_source[row_idx] = ModifiedRow(
                source_row=row_idx,
                new_values_by_col=new_values_by_col,
                changes=unique_preserve_order(changes),
                readable_context=readable,
            )

            for key, value in counters.items():
                report.inc(key, value)

    return None, list(modified_rows_by_source.values())


def create_generic_xlsx_override(
    vanilla_xlsx: Path,
    output_xlsx: Path,
    report: PatchReport,
    source_root: Path,
    args: argparse.Namespace,
) -> bool:
    """
    Patch générique pour tous les fichiers XLSX sauf General.xlsx/Settings.
    Crée un override minimal avec seulement les feuilles/lignes modifiées.
    """
    source_label = str(vanilla_xlsx.relative_to(source_root))
    src_wb = load_workbook(vanilla_xlsx, data_only=False)

    if output_xlsx.exists():
        out_wb = load_workbook(output_xlsx, data_only=False)
    else:
        out_wb = Workbook()
        out_wb.remove(out_wb.active)

    total_rows_changed = 0

    for sheet_name in src_wb.sheetnames:
        src_ws = src_wb[sheet_name]

        _unused, modified_rows = patch_sheet_generic(src_ws, source_label, sheet_name, report, args)
        if not modified_rows:
            continue

        header_row = find_header_row(src_ws)

        out_ws = get_sheet_case_insensitive(out_wb, sheet_name)
        if out_ws is None:
            out_ws = out_wb.create_sheet(title=sheet_name)
            copy_column_widths(src_ws, out_ws)

            # Je copie les en-tetes une seule fois, puis j'ajoute les lignes modifiees.
            dest_row = 1
            for src_header_row in range(1, header_row + 1):
                copy_row(src_ws, out_ws, src_header_row, dest_row)
                dest_row += 1

        for modified in sorted(modified_rows, key=lambda item: item.source_row):
            source_row_id = src_ws.cell(modified.source_row, 1).value
            existing_row = find_existing_override_row(out_ws, source_row_id)
            dest_row = existing_row if existing_row is not None else out_ws.max_row + 1

            copy_row(src_ws, out_ws, modified.source_row, dest_row)

            for col, new_value in modified.new_values_by_col.items():
                out_ws.cell(dest_row, col).value = new_value

            report.note(
                f"{source_label} / {sheet_name} / {modified.readable_context}: "
                + "; ".join(modified.changes)
            )

            total_rows_changed += 1

    if total_rows_changed == 0:
        return False

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_xlsx)
    report.file(output_xlsx)
    return True


# =============================================================================
# MANIFEST / README / CACHE
# =============================================================================

def make_supported_versions(game_version: str) -> list[str]:
    gv = game_version.strip()
    candidates = [
        gv,
        gv.lower(),
        gv.replace("Patch", "patch"),
        "2026.1",
    ]

    match = re.search(r"(\d{4}\.\d+)", gv)
    if match:
        candidates.append(match.group(1))

    return unique_preserve_order(candidates)


def write_manifest(mod_dir: Path, args: argparse.Namespace) -> None:
    manifest = {
        "name": MOD_DISPLAY_NAME,
        "version": MOD_VERSION,
        "description": (
            "Immersion longue : oxygene long applique au runtime, recharge surface vanilla, "
            "aucune capacite d'air XLSX modifiee, "
            f"discipline x1/{args.discipline_factor:g}, "
            f"batterie x{args.battery_capacity_factor:g}, "
            f"consommation electrique runtime 0 par defaut, fallback x{args.energy_usage_factor:g}, "
            "recharge diesel vanilla, "
            f"{args.fast_speed_top_gears} crans rapides vitesse x{args.fast_speed_factor:g}, "
            f"carburant rapide x{args.fast_speed_fuel_factor:g}, "
            f"vitesse max joueur {args.player_submarine_max_speed:g} km/h, menu runtime F10. "
            "sliders runtime bornes par profil pour batterie, oxygene, SuperVitesse, torpilles et sonar. "
            + (
                f"mega torpilles : degats torpilles x{args.torpedo_damage_factor:g}, "
                f"effets visuels d'explosion bornes x{args.torpedo_explosion_radius_factor:g} pour stabilite. "
                if args.mega_torpedoes
                else "mega torpilles desactivees. "
            ) +
            (
                "fiabilite torpilles parfaite : DudChance et defaillances magnetiques a 0. "
                if args.perfect_torpedo_reliability
                else "fiabilite torpilles vanilla. "
            ) +
            "cette version garde AirCompressor et Ventilation vanilla pendant la remontée surface."
        ),
        "author": MOD_AUTHOR,
        "minGameVersion": "2026.1",
        "maxGameVersion": "",
        "supportedGameVersions": make_supported_versions(args.game_version),
        "assemblyName": MOD_ASSEMBLY_NAME,
        "permissions": ["Reflection"],
        "steamFileId": 0,
    }

    manifest["description"] = (
        "Immersion longue : oxygene runtime calibre 1=vanilla et 100=environ 90 jours, "
        "recharge surface vanilla, aucune capacite d'air XLSX modifiee, "
        f"discipline x1/{args.discipline_factor:g}, "
        "Mega Batterie rend la batterie infinie des que la case F10 est cochee, "
        "consommation electrique runtime restauree proprement, "
        "SuperVitesse reglable 1=x1 a 20=x20 sur les crans rapides, "
        "torpilles reglables 1=x1 a 10=x10, "
        "sonar hydrophone reglable 1=x1 a 10=x10, defaut x3, "
        "Blindage lourd F10 desactive par defaut avec degats joueur divisibles par 3 quand active, "
        "Super discrétion F10 desactivee par defaut avec bruit et detectabilite joueur divisibles par 3, "
        "DeepDive F10 active par defaut : ordres de profondeur x2, stress profondeur calcule sur profondeur /2, jusqu'a 600 m reels et crush a 700 m, "
        "Couleurs eclairage F10 personnalisables par listes predefinies, defaut Alarm orange ambre et SilentRun vert sous-marin, restauration vanilla quand decochees, "
        "menu runtime F10 debounce pour eviter les freezes de slider, bouton Appeler renforts avec fallback U-boats amis, "
        "AirCompressor et Ventilation vanilla."
    )

    (mod_dir / "Manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def write_runtime_patch_source(mod_dir: Path, args: argparse.Namespace, report: PatchReport) -> None:
    source_dir = mod_dir / "Source"
    source_dir.mkdir(parents=True, exist_ok=True)
    source_path = source_dir / "LongSubmergedRuntimePatch.cs"

    source = r'''using System;
using System.Collections.Generic;
using System.Reflection;
using System.Runtime.CompilerServices;
using System.Text;
using HarmonyLib;
using UBOAT.Game;
using UBOAT.Game.Core;
using UBOAT.Game.Core.AI;
using UBOAT.Game.Core.AI.GroupBehaviours;
using UBOAT.Game.Core.Data;
using UBOAT.Game.Scene.Characters;
using UBOAT.Game.Scene.Effects;
using UBOAT.Game.Scene.Entities;
using UBOAT.Game.Scene.Items;
using UBOAT.Game.Scene.Tasks;
using UBOAT.Game.Scene.Utilities;
using UBOAT.Game.Sandbox;
using UBOAT.Game.Sandbox.Missions;
using UBOAT.Game.UI.Notifications;
using UBOAT.Game.UI.Resources;
using UnityEngine;
using UnityEngine.EventSystems;
using UnityEngine.UI;

namespace LongSubmerged10x
{
    // DonJ : point d'entree du runtime UBOAT. Cette classe ne porte pas la logique gameplay elle-meme ;
    // elle charge les reglages, cree le menu F10, installe les hooks Harmony et lance une premiere passe runtime.
    public sealed class LongSubmergedRuntimePatchMod : IUserMod
    {
        private const string RuntimeVersion = "__MOD_VERSION__";

        public string Name
        {
            get { return "Long Submerged 10x+ AirFix"; }
        }

        public void OnLoaded()
        {
            try
            {
                // DonJ : je charge les reglages PlayerPrefs et je cree le menu avant Harmony.
                // Si un hook Harmony casse apres une mise a jour UBOAT, le menu et le heartbeat batterie existent quand meme.
                LongSubmergedRuntimeSettings.Load();
                LongSubmergedMenuController.Ensure();
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }

            // DonJ : je patche chaque hook un par un. Un patch rate ne doit jamais empecher la batterie,
            // l'oxygene, les torpilles ou le menu de continuer a fonctionner avec les autres hooks valides.
            LongSubmergedRuntimePatcher.PatchSafely(new Harmony("donj.longsubmerged10x.runtimefix1416"));

            try
            {
                // DonJ : premiere application directe. Elle couvre le cas ou des objets existent deja
                // avant que leurs hooks Awake/Start aient pu etre interceptes.
                LongSubmergedRuntimeApplier.ApplyAll("mod loaded");
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }

            Debug.Log("[LongSubmerged10x] Runtime patch loaded v" + RuntimeVersion + ". F10 ouvre le menu Long Submerged.");
        }
    }

    // DonJ : liste centralisee des hooks Harmony du mod. Garder cette liste explicite rend le chargement
    // robuste : on voit exactement quelles zones du jeu sont touchees et on peut ignorer un hook incompatible.
    internal static class LongSubmergedRuntimePatcher
    {
        private static readonly Type[] PatchTypes = new Type[]
        {
            typeof(PlayerShipAwakePatch),
            typeof(PlayerShipOnAfterDeserializePatch),
            typeof(PlayerShipUpdatePatch),
            typeof(ResourceUpdateAmountBatteryPatch),
            typeof(PlayerShipValidateTargetVelocityPatch),
            typeof(PlayerShipValidateOxygenBreathModifierPatch),
            typeof(DeepDivePlayerShipTargetDepthSetterPatch),
            typeof(DeepDiveHullCrushControllerDoUpdatePatch),
            typeof(DeepDivePlayerShipUpdateStressAndDisciplineGainPatch),
            typeof(PlayerShipSavesManagerOnLoadedPatch),
            typeof(PlayerShipCrewAddedPatch),
            typeof(PlayerShipCrewRemovedPatch),
            typeof(PlayerShipEngineAwakePatch),
            typeof(PlayerShipEngineOnAfterDeserializePatch),
            typeof(PlayerShipEngineSavesManagerOnLoadedPatch),
            typeof(AccumulatorsUpgradeStartPatch),
            typeof(DivingPlanesStationAwakePatch),
            typeof(DivingPlanesStationUpdateModifiersPatch),
            typeof(GyrocompassApplyModifiersPatch),
            typeof(TrimPumpOnEnablePatch),
            typeof(StoredTorpedoStartPatch),
            typeof(StoredTorpedoApplyWarmUpModifierPatch),
            typeof(TorpedoAwakePatch),
            typeof(TorpedoFixedUpdatePatch),
            typeof(TorpedoDetonatePatch),
            typeof(MegaSonarHydrophoneRefreshPatch),
            typeof(SuperStealthEntityUpdateDetectabilityPatch),
            typeof(SuperStealthAirCompressorOnEnablePatch),
            typeof(SuperStealthAirCompressorOnDisablePatch),
            typeof(SuperStealthVentilationOnEnablePatch),
            typeof(SuperStealthVentilationOnDisablePatch),
            typeof(SuperStealthPropellerPowerPatch),
            typeof(SuperStealthPropellerPowerMultiplierPatch),
            typeof(SuperStealthSnorkelUpdatePatch),
            typeof(SuperStealthPeriscopeUpdatePatch),
            typeof(HeavyArmorHullAddDamagePatch),
            typeof(HeavyArmorEquipmentAddDamagePatch),
            typeof(HeavyArmorEquipmentAddWaterDamagePatch),
            typeof(HeavyArmorPlayableCharacterAddDamagePatch),
            typeof(HeavyArmorDamageUtilityDoApplyDamagePatch),
            typeof(HeavyArmorDamageUtilityApplyDamageToComponentsPatch),
            typeof(HeavyArmorApplyWaterDamageToPlayerShipDoDamageTickPatch),
            typeof(HeavyArmorHullEffectsRendererApplyImpactPatch),
            typeof(ResourceGuiGetTooltipContentsPatch),
            typeof(ResourceGuiUpdateDisplayedValuePatch),
            typeof(InteriorLightingPlayerShipInteriorLightingAwakePatch),
            typeof(InteriorLightingPlayerShipInteriorLightingStartPatch),
            typeof(InteriorLightingPlayerShipInteriorLightingApplyPresetsPatch),
            typeof(InteriorLightingLightControllerUpdatePresetsPatch),
            typeof(InteriorLightingFillLightUpdatePresetsPatch)
        };

        public static void PatchSafely(Harmony harmony)
        {
            if (harmony == null)
                return;

            foreach (Type patchType in PatchTypes)
            {
                try
                {
                    harmony.CreateClassProcessor(patchType).Patch();
                    Debug.Log("[LongSubmerged10x] Harmony patch active: " + patchType.Name + ".");
                }
                catch (Exception ex)
                {
                    // DonJ : une seule methode renommee dans UBOAT ne doit plus neutraliser tout le mod.
                    Debug.LogWarning("[LongSubmerged10x] Harmony patch skipped: " + patchType.Name + " -> " + ex.GetType().Name + ": " + ex.Message);
                }
            }
        }
    }

    internal static class InteriorLightingColorPatcher
    {
        private static readonly InteriorLightingColorPreset[] LightingColorPresets =
            new InteriorLightingColorPreset[]
            {
                new InteriorLightingColorPreset("Orange ambre", new Color(1f, 0.55f, 0.12f, 1f)),
                new InteriorLightingColorPreset("Vert sous-marin", new Color(0.12f, 0.78f, 0.28f, 1f)),
                new InteriorLightingColorPreset("Rouge", new Color(0.95f, 0.12f, 0.10f, 1f)),
                new InteriorLightingColorPreset("Jaune", new Color(1f, 0.88f, 0.16f, 1f)),
                new InteriorLightingColorPreset("Vert", new Color(0.18f, 0.85f, 0.30f, 1f)),
                new InteriorLightingColorPreset("Bleu", new Color(0.16f, 0.42f, 1f, 1f)),
                new InteriorLightingColorPreset("Cyan", new Color(0.16f, 0.86f, 1f, 1f)),
                new InteriorLightingColorPreset("Turquoise", new Color(0.10f, 0.76f, 0.68f, 1f)),
                new InteriorLightingColorPreset("Violet", new Color(0.58f, 0.30f, 1f, 1f)),
                new InteriorLightingColorPreset("Rose", new Color(1f, 0.34f, 0.70f, 1f)),
                new InteriorLightingColorPreset("Blanc chaud", new Color(1f, 0.86f, 0.64f, 1f)),
                new InteriorLightingColorPreset("Blanc froid", new Color(0.78f, 0.90f, 1f, 1f))
            };

        private static readonly FieldInfo AlarmInteriorFogColorField =
            AccessTools.Field(typeof(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting), "alarmInteriorFogColor");

        private static readonly FieldInfo SilentRunInteriorFogColorField =
            AccessTools.Field(typeof(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting), "silentRunInteriorFogColor");

        private static readonly FieldInfo AlarmLightsColorMultiplierField =
            AccessTools.Field(typeof(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting), "alarmLightsColorMultiplier");

        private static readonly FieldInfo SilentRunLightsColorMultiplierField =
            AccessTools.Field(typeof(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting), "silentRunLightsColorMultiplier");

        private static readonly MethodInfo ApplyColorMultiplierMethod =
            AccessTools.Method(typeof(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting), "ApplyColorMultiplier");

        private static readonly HashSet<string> MissingMemberWarnings = new HashSet<string>();

        private static readonly ConditionalWeakTable<object, InteriorLightingObjectPatchData> ObjectColorPatches =
            new ConditionalWeakTable<object, InteriorLightingObjectPatchData>();

        private static bool refreshingInteriorLighting;

        public static bool IsEnabled()
        {
            return LongSubmergedRuntimeSettings.InteriorLightingColors;
        }

        public static int LightingColorPresetCount
        {
            get { return LightingColorPresets.Length; }
        }

        public static Color GetLightingColorPresetColor(int presetIndex)
        {
            return LightingColorPresets[ClampLightingColorPresetIndex(presetIndex)].Color;
        }

        public static string GetLightingColorPresetName(int presetIndex)
        {
            return LightingColorPresets[ClampLightingColorPresetIndex(presetIndex)].Name;
        }

        public static List<string> GetLightingColorPresetNames()
        {
            List<string> names = new List<string>();
            for (int index = 0; index < LightingColorPresets.Length; index++)
                names.Add(LightingColorPresets[index].Name);

            return names;
        }

        private static Color AlarmColor
        {
            get { return LongSubmergedRuntimeSettings.InteriorLightingAlarmColor; }
        }

        private static Color SilentRunColor
        {
            get { return LongSubmergedRuntimeSettings.InteriorLightingSilentRunColor; }
        }

        public static void ApplyAll(string reason)
        {
            try
            {
                foreach (UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting lighting in UnityEngine.Object.FindObjectsOfType<UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting>())
                    ApplyInteriorLighting(lighting, reason + ".PlayerShipInteriorLighting", true);

                foreach (UBOAT.Game.Scene.Effects.LightController controller in UnityEngine.Object.FindObjectsOfType<UBOAT.Game.Scene.Effects.LightController>())
                    ApplyLightController(controller, reason + ".LightController");

                foreach (UBOAT.Game.Scene.Effects.FillLight fillLight in UnityEngine.Object.FindObjectsOfType<UBOAT.Game.Scene.Effects.FillLight>())
                    ApplyFillLight(fillLight, reason + ".FillLight");
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }
        }

        public static void ApplyInteriorLighting(
            UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting lighting,
            string reason,
            bool refreshPresets
        )
        {
            if (lighting == null)
                return;

            try
            {
                if (!IsEnabled())
                {
                    RestoreInteriorLighting(lighting, reason, refreshPresets);
                    return;
                }

                ApplyPrivateInteriorColors(lighting);
                ApplyLightControllers(lighting, reason);
                ApplyColorMultiplier(lighting);

                if (refreshPresets && !refreshingInteriorLighting)
                {
                    refreshingInteriorLighting = true;
                    try
                    {
                        lighting.ApplyLightControllersPresets();
                    }
                    finally
                    {
                        refreshingInteriorLighting = false;
                    }
                }
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }
        }

        public static void ApplyLightController(
            UBOAT.Game.Scene.Effects.LightController controller,
            string reason
        )
        {
            if (controller == null)
                return;

            try
            {
                if (!IsEnabled())
                {
                    RestoreLightController(controller, reason);
                    return;
                }

                SetColorProperty(controller, "AlarmColor", controller.AlarmColor, AlarmColor);
                SetColorProperty(controller, "BlueColor", controller.BlueColor, SilentRunColor);
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }
        }

        public static void ApplyFillLight(
            UBOAT.Game.Scene.Effects.FillLight fillLight,
            string reason
        )
        {
            if (fillLight == null)
                return;

            try
            {
                if (!IsEnabled())
                {
                    RestoreFillLight(fillLight, reason);
                    return;
                }

                SetColorProperty(fillLight, "RedColor", fillLight.RedColor, AlarmColor);
                SetColorProperty(fillLight, "BlueColor", fillLight.BlueColor, SilentRunColor);
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }
        }

        private static void ApplyPrivateInteriorColors(
            UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting lighting
        )
        {
            SetColorField(lighting, AlarmInteriorFogColorField, "alarmInteriorFogColor", AlarmColor);
            SetColorField(lighting, SilentRunInteriorFogColorField, "silentRunInteriorFogColor", SilentRunColor);
            SetColorField(lighting, AlarmLightsColorMultiplierField, "alarmLightsColorMultiplier", Color.white);
            SetColorField(lighting, SilentRunLightsColorMultiplierField, "silentRunLightsColorMultiplier", Color.white);
        }

        private static void ApplyLightControllers(
            UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting lighting,
            string reason
        )
        {
            UBOAT.Game.Scene.Effects.LightController[] controllers =
                lighting.GetComponentsInChildren<UBOAT.Game.Scene.Effects.LightController>(true);
            for (int index = 0; index < controllers.Length; index++)
                ApplyLightController(controllers[index], reason);

            UBOAT.Game.Scene.Effects.FillLight[] fillLights =
                lighting.GetComponentsInChildren<UBOAT.Game.Scene.Effects.FillLight>(true);
            for (int index = 0; index < fillLights.Length; index++)
                ApplyFillLight(fillLights[index], reason);
        }

        private static void ApplyColorMultiplier(
            UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting lighting
        )
        {
            if (ApplyColorMultiplierMethod == null)
            {
                WarnMissingMember("ApplyColorMultiplier");
                return;
            }

            ApplyColorMultiplierMethod.Invoke(lighting, null);
        }

        private static void RestoreInteriorLighting(
            UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting lighting,
            string reason,
            bool refreshPresets
        )
        {
            RestoreColorField(lighting, AlarmInteriorFogColorField, "alarmInteriorFogColor");
            RestoreColorField(lighting, SilentRunInteriorFogColorField, "silentRunInteriorFogColor");
            RestoreColorField(lighting, AlarmLightsColorMultiplierField, "alarmLightsColorMultiplier");
            RestoreColorField(lighting, SilentRunLightsColorMultiplierField, "silentRunLightsColorMultiplier");
            ApplyLightControllers(lighting, reason);
            ApplyColorMultiplier(lighting);

            if (refreshPresets && !refreshingInteriorLighting)
            {
                refreshingInteriorLighting = true;
                try
                {
                    lighting.ApplyLightControllersPresets();
                }
                finally
                {
                    refreshingInteriorLighting = false;
                }
            }
        }

        private static void RestoreLightController(
            UBOAT.Game.Scene.Effects.LightController controller,
            string reason
        )
        {
            RestoreColorProperty(controller, "AlarmColor", controller.AlarmColor, delegate(Color value) { controller.AlarmColor = value; });
            RestoreColorProperty(controller, "BlueColor", controller.BlueColor, delegate(Color value) { controller.BlueColor = value; });
        }

        private static void RestoreFillLight(
            UBOAT.Game.Scene.Effects.FillLight fillLight,
            string reason
        )
        {
            RestoreColorProperty(fillLight, "RedColor", fillLight.RedColor, delegate(Color value) { fillLight.RedColor = value; });
            RestoreColorProperty(fillLight, "BlueColor", fillLight.BlueColor, delegate(Color value) { fillLight.BlueColor = value; });
        }

        private static void SetColorField(object target, FieldInfo field, string memberName, Color value)
        {
            if (field == null)
            {
                WarnMissingMember(memberName);
                return;
            }

            object current = field.GetValue(target);
            if (current is Color && ColorsEqual((Color)current, value))
                return;

            if (current is Color)
                RememberColorPatch(target, memberName, (Color)current, value);

            field.SetValue(target, value);
        }

        private static void SetColorProperty(object target, string memberName, Color current, Color value)
        {
            if (ColorsEqual(current, value))
                return;

            RememberColorPatch(target, memberName, current, value);
            if (target is UBOAT.Game.Scene.Effects.LightController)
            {
                UBOAT.Game.Scene.Effects.LightController controller =
                    (UBOAT.Game.Scene.Effects.LightController)target;
                if (memberName == "AlarmColor")
                    controller.AlarmColor = value;
                else if (memberName == "BlueColor")
                    controller.BlueColor = value;
            }
            else if (target is UBOAT.Game.Scene.Effects.FillLight)
            {
                UBOAT.Game.Scene.Effects.FillLight fillLight =
                    (UBOAT.Game.Scene.Effects.FillLight)target;
                if (memberName == "RedColor")
                    fillLight.RedColor = value;
                else if (memberName == "BlueColor")
                    fillLight.BlueColor = value;
            }
        }

        private static void RestoreColorField(object target, FieldInfo field, string memberName)
        {
            if (field == null)
            {
                WarnMissingMember(memberName);
                return;
            }

            object current = field.GetValue(target);
            if (!(current is Color))
                return;

            Color original;
            if (!TryConsumeColorPatch(target, memberName, (Color)current, out original))
                return;

            field.SetValue(target, original);
        }

        private static void RestoreColorProperty(
            object target,
            string memberName,
            Color current,
            Action<Color> setter
        )
        {
            Color original;
            if (!TryConsumeColorPatch(target, memberName, current, out original))
                return;

            setter(original);
        }

        private static void RememberColorPatch(object target, string memberName, Color original, Color patched)
        {
            InteriorLightingObjectPatchData data;
            if (!ObjectColorPatches.TryGetValue(target, out data))
            {
                data = new InteriorLightingObjectPatchData();
                ObjectColorPatches.Add(target, data);
            }

            InteriorLightingColorPatchValue stored;
            if (!data.Values.TryGetValue(memberName, out stored))
            {
                data.Values.Add(memberName, new InteriorLightingColorPatchValue(original, patched));
                return;
            }

            stored.PatchedValue = patched;
        }

        private static bool TryConsumeColorPatch(
            object target,
            string memberName,
            Color current,
            out Color original
        )
        {
            original = Color.clear;

            InteriorLightingObjectPatchData data;
            if (!ObjectColorPatches.TryGetValue(target, out data))
                return false;

            InteriorLightingColorPatchValue stored;
            if (!data.Values.TryGetValue(memberName, out stored))
                return false;

            if (!ColorsEqual(current, stored.PatchedValue))
                return false;

            data.Values.Remove(memberName);
            if (data.Values.Count == 0)
                ObjectColorPatches.Remove(target);

            original = stored.OriginalValue;
            return true;
        }

        private static bool ColorsEqual(Color left, Color right)
        {
            return Mathf.Abs(left.r - right.r) < 0.0001f
                && Mathf.Abs(left.g - right.g) < 0.0001f
                && Mathf.Abs(left.b - right.b) < 0.0001f
                && Mathf.Abs(left.a - right.a) < 0.0001f;
        }

        private static int ClampLightingColorPresetIndex(int presetIndex)
        {
            if (LightingColorPresets.Length == 0)
                return 0;

            return Mathf.Clamp(presetIndex, 0, LightingColorPresets.Length - 1);
        }

        private static void WarnMissingMember(string memberName)
        {
            if (!MissingMemberWarnings.Add(memberName))
                return;

            Debug.LogWarning("[LongSubmerged10x] Interior lighting color patch skipped missing member: " + memberName + ".");
        }
    }

    internal sealed class InteriorLightingColorPreset
    {
        public readonly string Name;
        public readonly Color Color;

        public InteriorLightingColorPreset(string name, Color color)
        {
            Name = string.IsNullOrEmpty(name) ? "Couleur" : name;
            Color = color;
        }
    }

    internal sealed class InteriorLightingObjectPatchData
    {
        public readonly Dictionary<string, InteriorLightingColorPatchValue> Values =
            new Dictionary<string, InteriorLightingColorPatchValue>();
    }

    internal sealed class InteriorLightingColorPatchValue
    {
        public readonly Color OriginalValue;
        public Color PatchedValue;

        public InteriorLightingColorPatchValue(Color originalValue, Color patchedValue)
        {
            OriginalValue = originalValue;
            PatchedValue = patchedValue;
        }
    }

    internal static class ReinforcementRuntimeController
    {
        private const float ReinforcementCooldownSeconds = 300f;
        private const float ReinforcementActiveTrackingSeconds = 900f;
        private const int RequiredPrimaryAirPatrolCalls = 2;
        private const int RequiredPrimaryWarshipCalls = 2;
        private const int DesiredFallbackUboatCount = 2;
        private const float FallbackMinimumPlayerDistance = 8f;
        private const float FallbackGroupClearance = 2.5f;
        private const float FallbackRallyDistance = 6f;
        private static readonly string[] FallbackSubmarineTypePriority = new string[]
        {
            "Type VIIC",
            "Type VIIB",
            "Type VIIC41",
            "Type IID",
            "Type IIB",
            "Type IIA"
        };

        private static readonly float[] FallbackSpawnDistances = new float[] { 10f, 12f, 14f, 16f };
        private static readonly float[] FallbackSpawnAngles = new float[] { 110f, -110f, 130f, -130f, 150f, -150f, 90f, -90f };

        private static readonly List<SandboxGroup> ActiveReinforcementGroups = new List<SandboxGroup>();
        private static readonly List<float> ActiveReinforcementGroupTrackedAt = new List<float>();
        private static readonly FieldInfo SandboxGroupWorldNavMeshField = AccessTools.Field(typeof(SandboxGroup), "worldNavMesh");

        private static bool reinforcementCallInProgress;
        private static float nextAllowedReinforcementCallTime;
        private static bool warnedMissingWorldNavMeshField;
        private static bool warnedWorldNavMeshValidationFailure;

        public static string GetStatusText()
        {
            CleanupActiveGroups();

            if (ActiveReinforcementGroups.Count > 0)
                return "Renforts deja actifs";

            float remainingSeconds = GetCooldownRemainingSeconds();
            if (remainingSeconds > 0f)
                return "Cooldown " + Mathf.CeilToInt(remainingSeconds) + "s";

            return "Pret";
        }

        public static string CallReinforcements(string reason)
        {
            CleanupActiveGroups();

            if (reinforcementCallInProgress)
            {
                Debug.LogWarning("[LongSubmerged10x] Reinforcement call skipped: already running.");
                return "Appel deja en cours";
            }

            if (ActiveReinforcementGroups.Count > 0)
            {
                Debug.LogWarning("[LongSubmerged10x] Reinforcement call skipped: active reinforcement groups still exist.");
                return "Renforts deja actifs";
            }

            float remainingSeconds = GetCooldownRemainingSeconds();
            if (remainingSeconds > 0f)
            {
                Debug.LogWarning("[LongSubmerged10x] Reinforcement call skipped: cooldown active for " + remainingSeconds + "s.");
                return "Cooldown " + Mathf.CeilToInt(remainingSeconds) + "s";
            }

            reinforcementCallInProgress = true;
            try
            {
                Debug.Log("[LongSubmerged10x] Reinforcement call requested: " + SafeReason(reason) + ".");

                PlayerShip playerShip = UnityEngine.Object.FindObjectOfType<PlayerShip>();
                if (playerShip == null)
                    return FailWithoutCooldown("Aucun sous-marin joueur", "player ship missing");

                SandboxGroup playerGroup = ResolvePlayerGroup(playerShip);
                if (playerGroup == null)
                    return FailWithoutCooldown("Groupe joueur introuvable", "player sandbox group missing");

                Country playerCountry = ResolvePlayerCountry(playerShip, playerGroup);
                if (playerCountry == null)
                    return FailWithoutCooldown("Pays joueur introuvable", "player country missing");

                Sandbox sandbox = ResolveSandbox();
                List<Country> friendlyCountries = BuildFriendlyCountries(sandbox, playerCountry);
                if (friendlyCountries.Count == 0)
                    return FailWithoutCooldown("Aucun pays ami", "no friendly country found");

                List<SandboxGroup> primaryGroups = new List<SandboxGroup>();
                int airGroups = SpawnFriendlyPatrols(
                    "LongSubmerged Air Reinforcement",
                    "Entities/Air Patrol",
                    "Air Patrol",
                    true,
                    RequiredPrimaryAirPatrolCalls,
                    friendlyCountries,
                    playerCountry,
                    playerGroup,
                    primaryGroups
                );
                int warshipGroups = SpawnFriendlyPatrols(
                    "LongSubmerged Warship Reinforcement",
                    "Entities/Warships",
                    "Warships",
                    false,
                    RequiredPrimaryWarshipCalls,
                    friendlyCountries,
                    playerCountry,
                    playerGroup,
                    primaryGroups
                );

                if (airGroups >= RequiredPrimaryAirPatrolCalls && warshipGroups >= RequiredPrimaryWarshipCalls)
                {
                    TrackCreatedGroups(primaryGroups);
                    StartReinforcementCooldown();
                    Debug.Log("[LongSubmerged10x] Reinforcement call spawned primary groups: air=" + airGroups + ", warships=" + warshipGroups + ".");
                    return "Renforts appeles";
                }

                DestroyCreatedGroups(primaryGroups, "primary incomplete");
                Debug.LogWarning("[LongSubmerged10x] Reinforcement call primary fallback: air=" + airGroups + ", warships=" + warshipGroups + ".");

                List<SandboxGroup> fallbackGroups = new List<SandboxGroup>();
                int submarineGroups = CreateManualFriendlyUboats(
                    sandbox,
                    friendlyCountries,
                    playerCountry,
                    playerGroup,
                    fallbackGroups
                );

                if (submarineGroups > 0)
                {
                    TrackCreatedGroups(fallbackGroups);
                    StartReinforcementCooldown();
                    Debug.Log("[LongSubmerged10x] Reinforcement call spawned manual fallback U-boats: submarines=" + submarineGroups + ".");
                    return submarineGroups == 1 ? "1 U-boat appele" : submarineGroups + " U-boats appeles";
                }

                DestroyCreatedGroups(fallbackGroups, "fallback failed");
                Debug.LogWarning("[LongSubmerged10x] Reinforcement call failed: no friendly U-boat fallback was available.");
                return "Aucun U-boat ami disponible";
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
                return "Erreur renforts";
            }
            finally
            {
                reinforcementCallInProgress = false;
            }
        }

        private static int SpawnFriendlyPatrols(
            string namePrefix,
            string displayNameKey,
            string patrolType,
            bool airborne,
            int requiredCalls,
            List<Country> friendlyCountries,
            Country playerCountry,
            SandboxGroup playerGroup,
            List<SandboxGroup> createdGroups
        )
        {
            int spawnedCount = 0;
            for (int attemptIndex = 0; attemptIndex < requiredCalls; attemptIndex++)
            {
                SandboxGroup group = SpawnOneFriendlyPatrol(
                    namePrefix + " " + (attemptIndex + 1),
                    displayNameKey,
                    patrolType,
                    airborne,
                    attemptIndex,
                    friendlyCountries,
                    playerCountry,
                    playerGroup
                );

                if (group == null)
                    continue;

                createdGroups.Add(group);
                spawnedCount++;
            }

            return spawnedCount;
        }

        private static SandboxGroup SpawnOneFriendlyPatrol(
            string groupName,
            string displayNameKey,
            string patrolType,
            bool airborne,
            int attemptIndex,
            List<Country> friendlyCountries,
            Country playerCountry,
            SandboxGroup playerGroup
        )
        {
            Vector2 preferredDirection = GetPreferredDirection(playerGroup, attemptIndex, airborne);
            for (int countryIndex = 0; countryIndex < friendlyCountries.Count; countryIndex++)
            {
                Country country = friendlyCountries[countryIndex];
                if (!IsFriendlyCountry(country, playerCountry))
                    continue;

                SandboxMobileGroup group = null;
                try
                {
                    group = MissionUtility.SpawnPatrol(
                        groupName,
                        new LocalizedString(displayNameKey),
                        patrolType,
                        false,
                        country,
                        preferredDirection,
                        airborne,
                        true
                    );
                }
                catch (Exception ex)
                {
                    Debug.LogWarning("[LongSubmerged10x] Reinforcement patrol spawn failed for " + patrolType + "/" + SafeCountryCode(country) + ": " + ex.GetType().Name + ": " + ex.Message);
                }

                if (group == null)
                    continue;

                if (!IsFriendlyGroup(group, playerCountry))
                {
                    Debug.LogWarning("[LongSubmerged10x] Reinforcement patrol rejected non-friendly group: " + patrolType + "/" + SafeCountryCode(group.Country) + ".");
                    DestroyCreatedGroup(group, "non-friendly group");
                    continue;
                }

                Debug.Log("[LongSubmerged10x] Reinforcement patrol spawned: " + patrolType + "/" + SafeCountryCode(group.Country) + " at " + group.Position + ".");
                return group;
            }

            Debug.LogWarning("[LongSubmerged10x] Reinforcement patrol unavailable for friendly countries: " + patrolType + ".");
            return null;
        }

        private static int CreateManualFriendlyUboats(
            Sandbox sandbox,
            List<Country> friendlyCountries,
            Country playerCountry,
            SandboxGroup playerGroup,
            List<SandboxGroup> createdGroups
        )
        {
            if (sandbox == null)
            {
                Debug.LogWarning("[LongSubmerged10x] Manual U-boat fallback skipped: sandbox missing.");
                return 0;
            }

            List<Country> fallbackCountries = BuildFallbackSubmarineCountries(sandbox, friendlyCountries, playerCountry);
            if (fallbackCountries.Count == 0)
            {
                Debug.LogWarning("[LongSubmerged10x] Manual U-boat fallback skipped: no friendly submarine country.");
                return 0;
            }

            int createdCount = 0;
            for (int reinforcementIndex = 0; reinforcementIndex < DesiredFallbackUboatCount; reinforcementIndex++)
            {
                SandboxMobileGroup group = CreateOneManualFriendlyUboat(sandbox, fallbackCountries, playerCountry, playerGroup, reinforcementIndex);
                if (group == null)
                    continue;

                createdGroups.Add(group);
                createdCount++;
            }

            return createdCount;
        }

        private static SandboxMobileGroup CreateOneManualFriendlyUboat(
            Sandbox sandbox,
            List<Country> fallbackCountries,
            Country playerCountry,
            SandboxGroup playerGroup,
            int reinforcementIndex
        )
        {
            Vector2 spawnPosition;
            if (!TryGetFallbackSpawnPosition(sandbox, playerGroup, reinforcementIndex, out spawnPosition))
            {
                Debug.LogWarning("[LongSubmerged10x] Manual U-boat fallback skipped: no clear horizon position.");
                return null;
            }

            Vector2 rallyPoint = GetFallbackRallyPoint(playerGroup, reinforcementIndex);
            for (int countryIndex = 0; countryIndex < fallbackCountries.Count; countryIndex++)
            {
                Country country = fallbackCountries[countryIndex];
                if (!IsFriendlyCountry(country, playerCountry))
                    continue;

                for (int typeIndex = 0; typeIndex < FallbackSubmarineTypePriority.Length; typeIndex++)
                {
                    SandboxMobileGroup group = TryCreateManualUboat(
                        sandbox,
                        country,
                        playerCountry,
                        playerGroup,
                        reinforcementIndex,
                        FallbackSubmarineTypePriority[typeIndex],
                        spawnPosition,
                        rallyPoint
                    );

                    if (group != null)
                        return group;
                }
            }

            Debug.LogWarning("[LongSubmerged10x] Manual U-boat fallback failed: no friendly submarine blueprint worked.");
            return null;
        }

        private static SandboxMobileGroup TryCreateManualUboat(
            Sandbox sandbox,
            Country country,
            Country playerCountry,
            SandboxGroup playerGroup,
            int reinforcementIndex,
            string submarineTypeName,
            Vector2 spawnPosition,
            Vector2 rallyPoint
        )
        {
            SandboxMobileGroup group = null;
            SandboxEntity entity = null;
            bool entityAttachedToGroup = false;
            try
            {
                group = SandboxGroup.Create<SandboxMobileGroup>(
                    "LongSubmerged U-boat Reinforcement " + (reinforcementIndex + 1),
                    "Submarine",
                    spawnPosition,
                    country
                );

                if (group == null)
                    return null;

                CharacterAI ai = EnsureCharacterAi(group);
                entity = SandboxEntity.Create(submarineTypeName, country);
                if (entity == null)
                {
                    DestroyCreatedGroup(group, "manual U-boat entity missing");
                    return null;
                }

                entity.Position = spawnPosition;
                entity.FormationPosition = Vector2.zero;
                entity.RandomizeSpawnPosition = false;
                entity.SpawnsInstantly = true;

                group.Position = spawnPosition;
                group.Up = GetDirectionTowards(spawnPosition, rallyPoint, ResolvePlayerForward(playerGroup));
                group.AddEntity(entity);
                entityAttachedToGroup = true;
                group.Velocity = Mathf.Max(0f, group.MaxVelocity * 0.55f);
                AddFallbackSailToBehaviour(ai, rallyPoint);

                sandbox.AddGroup(group);
                RefreshCreatedGroup(group);
                AddFallbackObservations(playerGroup, group);

                if (!IsFriendlyGroup(group, playerCountry))
                {
                    Debug.LogWarning("[LongSubmerged10x] Manual U-boat fallback rejected non-friendly group: " + SafeCountryCode(group.Country) + ".");
                    DestroyCreatedGroup(group, "manual non-friendly group");
                    return null;
                }

                Debug.Log("[LongSubmerged10x] Manual U-boat fallback spawned: type=" + submarineTypeName + ", country=" + SafeCountryCode(country) + ", position=" + group.Position + ".");
                return group;
            }
            catch (Exception ex)
            {
                Debug.LogWarning("[LongSubmerged10x] Manual U-boat fallback failed for " + submarineTypeName + "/" + SafeCountryCode(country) + ": " + ex.GetType().Name + ": " + ex.Message);
                if (group != null)
                    DestroyCreatedGroup(group, "manual U-boat exception");
                if (!entityAttachedToGroup && entity != null)
                    DestroyCreatedEntity(entity, "manual U-boat orphan");

                return null;
            }
        }

        private static List<Country> BuildFallbackSubmarineCountries(Sandbox sandbox, List<Country> friendlyCountries, Country playerCountry)
        {
            List<Country> countries = new List<Country>();
            AddFriendlyCountry(countries, FindCountryByCode(sandbox, "DE"), playerCountry);
            AddFriendlyCountry(countries, playerCountry, playerCountry);

            if (friendlyCountries != null)
            {
                for (int index = 0; index < friendlyCountries.Count; index++)
                    AddFriendlyCountry(countries, friendlyCountries[index], playerCountry);
            }

            return countries;
        }

        private static Country FindCountryByCode(Sandbox sandbox, string countryCode)
        {
            if (sandbox == null || sandbox.Countries == null || string.IsNullOrEmpty(countryCode))
                return null;

            Country[] countries = sandbox.Countries;
            for (int index = 0; index < countries.Length; index++)
            {
                Country country = countries[index];
                if (country != null && string.Equals(country.CountryCode, countryCode, StringComparison.OrdinalIgnoreCase))
                    return country;
            }

            return null;
        }

        private static bool TryGetFallbackSpawnPosition(Sandbox sandbox, SandboxGroup playerGroup, int reinforcementIndex, out Vector2 position)
        {
            position = Vector2.zero;
            if (playerGroup == null)
                return false;

            Vector2 origin = playerGroup.Position;
            Vector2 forward = ResolvePlayerForward(playerGroup);
            WorldNavMesh worldNavMesh = ResolveWorldNavMesh();
            int angleOffset = (reinforcementIndex * 2) % FallbackSpawnAngles.Length;
            for (int distanceIndex = 0; distanceIndex < FallbackSpawnDistances.Length; distanceIndex++)
            {
                for (int angleIndex = 0; angleIndex < FallbackSpawnAngles.Length; angleIndex++)
                {
                    float angle = FallbackSpawnAngles[(angleIndex + angleOffset) % FallbackSpawnAngles.Length];
                    float distance = FallbackSpawnDistances[distanceIndex];
                    Vector2 candidate = origin + Rotate(forward, angle) * distance;
                    candidate = SnapFallbackSpawnPosition(worldNavMesh, candidate);
                    if (!IsFallbackSpawnPositionClear(sandbox, playerGroup, worldNavMesh, candidate))
                        continue;

                    position = candidate;
                    return true;
                }
            }

            return false;
        }

        private static WorldNavMesh ResolveWorldNavMesh()
        {
            try
            {
                if (SandboxGroupWorldNavMeshField == null)
                {
                    if (!warnedMissingWorldNavMeshField)
                    {
                        warnedMissingWorldNavMeshField = true;
                        Debug.LogWarning("[LongSubmerged10x] Manual U-boat fallback navmesh check skipped: SandboxGroup.worldNavMesh field missing.");
                    }

                    return null;
                }

                return SandboxGroupWorldNavMeshField.GetValue(null) as WorldNavMesh;
            }
            catch (Exception ex)
            {
                WarnWorldNavMeshValidationSkipped(ex);
                return null;
            }
        }

        private static Vector2 SnapFallbackSpawnPosition(WorldNavMesh worldNavMesh, Vector2 position)
        {
            if (worldNavMesh == null)
                return position;

            try
            {
                return worldNavMesh.SnapWorld(position);
            }
            catch (Exception ex)
            {
                WarnWorldNavMeshValidationSkipped(ex);
                return position;
            }
        }

        private static bool IsFallbackSpawnPositionClear(Sandbox sandbox, SandboxGroup playerGroup, WorldNavMesh worldNavMesh, Vector2 position)
        {
            if (playerGroup != null)
            {
                Vector2 fromPlayer = position - playerGroup.Position;
                if (fromPlayer.sqrMagnitude < FallbackMinimumPlayerDistance * FallbackMinimumPlayerDistance)
                    return false;
            }

            if (!IsFallbackSpawnPositionOnNavMesh(worldNavMesh, playerGroup, position))
                return false;

            if (sandbox == null)
                return true;

            try
            {
                List<SandboxGroup> nearbyGroups = sandbox.GetGroupsInRange(position, FallbackGroupClearance, false);
                if (nearbyGroups == null)
                    return true;

                for (int index = 0; index < nearbyGroups.Count; index++)
                {
                    SandboxGroup group = nearbyGroups[index];
                    if (group != null && group != playerGroup && !ActiveReinforcementGroups.Contains(group))
                        return false;
                }
            }
            catch (Exception ex)
            {
                Debug.LogWarning("[LongSubmerged10x] Manual U-boat fallback group clearance check skipped: " + ex.GetType().Name + ": " + ex.Message);
            }

            return true;
        }

        private static bool IsFallbackSpawnPositionOnNavMesh(WorldNavMesh worldNavMesh, SandboxGroup playerGroup, Vector2 position)
        {
            if (worldNavMesh == null)
                return true;

            try
            {
                if (!worldNavMesh.IsOnNavMesh(position))
                    return false;

                if (playerGroup != null)
                {
                    Vector2 hit;
                    if (worldNavMesh.RaycastLandsNavMesh(position, playerGroup.Position, out hit))
                        return false;
                }
            }
            catch (Exception ex)
            {
                WarnWorldNavMeshValidationSkipped(ex);
                return true;
            }

            return true;
        }

        private static void WarnWorldNavMeshValidationSkipped(Exception ex)
        {
            if (warnedWorldNavMeshValidationFailure)
                return;

            warnedWorldNavMeshValidationFailure = true;
            Debug.LogWarning("[LongSubmerged10x] Manual U-boat fallback navmesh check skipped: " + ex.GetType().Name + ": " + ex.Message);
        }

        private static Vector2 GetFallbackRallyPoint(SandboxGroup playerGroup, int reinforcementIndex)
        {
            if (playerGroup == null)
                return Vector2.zero;

            Vector2 forward = ResolvePlayerForward(playerGroup);
            float rallyAngle = reinforcementIndex % 2 == 0 ? 70f : -70f;
            return playerGroup.Position + Rotate(forward, rallyAngle) * FallbackRallyDistance;
        }

        private static Vector2 ResolvePlayerForward(SandboxGroup playerGroup)
        {
            Vector2 forward = Vector2.up;
            if (playerGroup != null && playerGroup.Up.sqrMagnitude > 0.0001f)
                forward = playerGroup.Up.normalized;

            return forward;
        }

        private static Vector2 GetDirectionTowards(Vector2 from, Vector2 to, Vector2 fallback)
        {
            Vector2 direction = to - from;
            if (direction.sqrMagnitude <= 0.0001f)
                direction = fallback;

            if (direction.sqrMagnitude <= 0.0001f)
                direction = Vector2.up;

            return direction.normalized;
        }

        private static CharacterAI EnsureCharacterAi(SandboxMobileGroup group)
        {
            if (group == null)
                return null;

            try
            {
                CharacterAI ai = group.GetComponent<CharacterAI>();
                if (ai != null)
                    return ai;

                return group.gameObject.AddComponent<CharacterAI>();
            }
            catch (Exception ex)
            {
                Debug.LogWarning("[LongSubmerged10x] Manual U-boat fallback AI creation skipped: " + ex.GetType().Name + ": " + ex.Message);
                return null;
            }
        }

        private static void AddFallbackSailToBehaviour(CharacterAI ai, Vector2 rallyPoint)
        {
            if (ai == null)
                return;

            try
            {
                SailToBehaviour sailToBehaviour = new SailToBehaviour(ai, 1.5f, rallyPoint);
                sailToBehaviour.Flags = AIBehaviourFlags.OneShot;
                ai.AddBehaviour(sailToBehaviour);
            }
            catch (Exception ex)
            {
                Debug.LogWarning("[LongSubmerged10x] Manual U-boat fallback sail behaviour skipped: " + ex.GetType().Name + ": " + ex.Message);
            }
        }

        private static void RefreshCreatedGroup(SandboxMobileGroup group)
        {
            if (group == null)
                return;

            try
            {
                group.UpdateGroup();
                group.UpdateGroupLowFrequency(false);

                if (group.AI != null)
                {
                    for (int index = 0; index < 3; index++)
                        group.AI.UpdateAI(false);
                }
            }
            catch (Exception ex)
            {
                Debug.LogWarning("[LongSubmerged10x] Manual U-boat fallback refresh skipped: " + ex.GetType().Name + ": " + ex.Message);
            }
        }

        private static void AddFallbackObservations(SandboxGroup playerGroup, SandboxGroup reinforcementGroup)
        {
            if (playerGroup == null || reinforcementGroup == null)
                return;

            try
            {
                playerGroup.AddObservation(reinforcementGroup, GroupDetectionMethod.IndirectObservation);
                reinforcementGroup.AddObservation(playerGroup, GroupDetectionMethod.IndirectObservation);
            }
            catch (Exception ex)
            {
                Debug.LogWarning("[LongSubmerged10x] Manual U-boat fallback observations skipped: " + ex.GetType().Name + ": " + ex.Message);
            }
        }

        private static SandboxGroup ResolvePlayerGroup(PlayerShip playerShip)
        {
            if (playerShip == null)
                return null;

            if (playerShip.SandboxGroup != null)
                return playerShip.SandboxGroup;

            SandboxPlayerWolfpack wolfpack = playerShip.SandboxPlayerShip;
            return wolfpack as SandboxGroup;
        }

        private static Country ResolvePlayerCountry(PlayerShip playerShip, SandboxGroup playerGroup)
        {
            if (playerGroup != null && playerGroup.Country != null)
                return playerGroup.Country;

            if (playerShip != null && playerShip.SandboxEntity != null && playerShip.SandboxEntity.Country != null)
                return playerShip.SandboxEntity.Country;

            if (playerShip != null)
                return playerShip.Country;

            return null;
        }

        private static Sandbox ResolveSandbox()
        {
            if (Sandbox.Instance != null)
                return Sandbox.Instance;

            return UnityEngine.Object.FindObjectOfType<Sandbox>();
        }

        private static List<Country> BuildFriendlyCountries(Sandbox sandbox, Country playerCountry)
        {
            List<Country> friendlyCountries = new List<Country>();
            AddFriendlyCountry(friendlyCountries, playerCountry, playerCountry);

            if (sandbox != null && sandbox.Countries != null)
            {
                Country[] countries = sandbox.Countries;
                for (int index = 0; index < countries.Length; index++)
                    AddFriendlyCountry(friendlyCountries, countries[index], playerCountry);
            }

            return friendlyCountries;
        }

        private static void AddFriendlyCountry(List<Country> countries, Country candidate, Country playerCountry)
        {
            if (!IsFriendlyCountry(candidate, playerCountry))
                return;

            for (int index = 0; index < countries.Count; index++)
            {
                if (countries[index] == candidate)
                    return;
            }

            countries.Add(candidate);
        }

        private static bool IsFriendlyGroup(SandboxGroup group, Country playerCountry)
        {
            return group != null && IsFriendlyCountry(group.Country, playerCountry);
        }

        private static bool IsFriendlyCountry(Country candidate, Country playerCountry)
        {
            if (candidate == null || playerCountry == null)
                return false;

            if (candidate == playerCountry)
                return true;

            return playerCountry.GetRelationWith(candidate) == Country.Relation.Ally;
        }

        private static Vector2 GetPreferredDirection(SandboxGroup playerGroup, int attemptIndex, bool airborne)
        {
            Vector2 direction = Vector2.up;
            if (playerGroup != null && playerGroup.Up.sqrMagnitude > 0.0001f)
                direction = playerGroup.Up.normalized;

            float baseAngle = airborne ? 28f : -28f;
            float stepAngle = attemptIndex % 2 == 0 ? 18f : -18f;
            return Rotate(direction, baseAngle + stepAngle).normalized;
        }

        private static Vector2 Rotate(Vector2 vector, float degrees)
        {
            float radians = degrees * Mathf.Deg2Rad;
            float sin = Mathf.Sin(radians);
            float cos = Mathf.Cos(radians);
            return new Vector2(vector.x * cos - vector.y * sin, vector.x * sin + vector.y * cos);
        }

        private static void TrackCreatedGroups(List<SandboxGroup> groups)
        {
            for (int index = 0; index < groups.Count; index++)
            {
                SandboxGroup group = groups[index];
                if (group != null && !ActiveReinforcementGroups.Contains(group))
                {
                    ActiveReinforcementGroups.Add(group);
                    ActiveReinforcementGroupTrackedAt.Add(Time.unscaledTime);
                }
            }
        }

        private static void CleanupActiveGroups()
        {
            while (ActiveReinforcementGroupTrackedAt.Count < ActiveReinforcementGroups.Count)
                ActiveReinforcementGroupTrackedAt.Add(Time.unscaledTime);

            while (ActiveReinforcementGroupTrackedAt.Count > ActiveReinforcementGroups.Count)
                ActiveReinforcementGroupTrackedAt.RemoveAt(ActiveReinforcementGroupTrackedAt.Count - 1);

            for (int index = ActiveReinforcementGroups.Count - 1; index >= 0; index--)
            {
                SandboxGroup group = ActiveReinforcementGroups[index];
                float trackedAt = ActiveReinforcementGroupTrackedAt[index];
                bool trackingExpired = Time.unscaledTime - trackedAt >= ReinforcementActiveTrackingSeconds;
                if (group == null || trackingExpired)
                {
                    ActiveReinforcementGroups.RemoveAt(index);
                    ActiveReinforcementGroupTrackedAt.RemoveAt(index);
                    if (trackingExpired && group != null)
                        Debug.Log("[LongSubmerged10x] Reinforcement group tracking expired; cooldown now controls new calls.");
                }
            }
        }

        private static void DestroyCreatedGroups(List<SandboxGroup> groups, string reason)
        {
            for (int index = 0; index < groups.Count; index++)
                DestroyCreatedGroup(groups[index], reason);

            groups.Clear();
        }

        private static void DestroyCreatedGroup(SandboxGroup group, string reason)
        {
            if (group == null)
                return;

            try
            {
                group.DestroyGroup();
                Debug.Log("[LongSubmerged10x] Reinforcement group removed after " + reason + ".");
            }
            catch (Exception ex)
            {
                Debug.LogWarning("[LongSubmerged10x] Reinforcement group cleanup failed after " + reason + ": " + ex.GetType().Name + ": " + ex.Message);
            }
        }

        private static void DestroyCreatedEntity(SandboxEntity entity, string reason)
        {
            if (entity == null)
                return;

            try
            {
                entity.Destroy(false);
                Debug.Log("[LongSubmerged10x] Reinforcement entity removed after " + reason + ".");
            }
            catch (Exception ex)
            {
                Debug.LogWarning("[LongSubmerged10x] Reinforcement entity cleanup failed after " + reason + ": " + ex.GetType().Name + ": " + ex.Message);
            }
        }

        private static float GetCooldownRemainingSeconds()
        {
            return Mathf.Max(0f, nextAllowedReinforcementCallTime - Time.unscaledTime);
        }

        private static void StartReinforcementCooldown()
        {
            nextAllowedReinforcementCallTime = Time.unscaledTime + ReinforcementCooldownSeconds;
        }

        private static string FailWithoutCooldown(string uiMessage, string logReason)
        {
            Debug.LogWarning("[LongSubmerged10x] Reinforcement call failed: " + logReason + ".");
            return uiMessage;
        }

        private static string SafeCountryCode(Country country)
        {
            if (country == null)
                return "none";

            return string.IsNullOrEmpty(country.CountryCode) ? "unknown" : country.CountryCode;
        }

        private static string SafeReason(string reason)
        {
            return string.IsNullOrEmpty(reason) ? "unknown" : reason;
        }
    }

    internal static class LongSubmergedRuntimeSettings
    {
        private const string PrefPrefix = "LongSubmerged10x.";
        private const int RuntimeSettingsVersion = 19;

        public const float MinRuntimeFactor = 1f;
        public const float BatteryMaxFactor = 100f;
        public const float OxygenMaxFactor = 100f;
        public const float SpeedMaxFactor = 20f;
        public const float TorpedoMaxFactor = 10f;
        public const float SonarMaxFactor = 10f;
        public const float HeavyArmorDamageFactor = 3f;
        public const float SuperStealthFactor = 3f;

        // Compatibilite interne : les anciens blocs utilisaient MaxRuntimeFactor pour Batterie/Oxygene.
        // Les sliders vitesse, torpilles et sonar ont maintenant leurs propres bornes.
        public const float MaxRuntimeFactor = BatteryMaxFactor;

        private const bool DefaultMegaBattery = true;
        private const bool DefaultMegaOxygen = true;
        private const bool DefaultSuperSpeed = true;
        private const bool DefaultMegaTorpedoes = __DEFAULT_MEGA_TORPEDOES__;
        private const bool DefaultMegaSonar = true;
        private const bool DefaultHeavyArmor = false;
        private const bool DefaultSuperStealth = false;
        private const bool DefaultDeepDive = true;
        private const bool DefaultInteriorLightingColors = true;
        private const int DefaultInteriorLightingAlarmColorPresetIndex = 0;
        private const int DefaultInteriorLightingSilentRunColorPresetIndex = 1;

        // DonJ: readable in-game defaults. Mega Batterie now means a fully infinite battery.
        // The battery slider is kept only as a saved legacy value and no longer gates infinity.
        // Oxygen 100 is calibrated around 90 days, speed defaults to x8, torpedoes to x10,
        // and sonar defaults to x3 while remaining adjustable up to x10.
        private const float DefaultBatteryFactor = 10f;
        private const float DefaultOxygenFactor = 100f;
        private const float DefaultSpeedFactor = __FAST_SPEED_FACTOR__;
        private const float DefaultTorpedoFactor = __TORPEDO_DAMAGE_FACTOR__;
        private const float DefaultSonarFactor = 3f;

        public static bool MegaBattery = DefaultMegaBattery;
        public static bool MegaOxygen = DefaultMegaOxygen;
        public static bool SuperSpeed = DefaultSuperSpeed;
        public static bool MegaTorpedoes = DefaultMegaTorpedoes;
        public static bool MegaSonar = DefaultMegaSonar;
        public static bool HeavyArmor = DefaultHeavyArmor;
        public static bool SuperStealth = DefaultSuperStealth;
        public static bool DeepDive = DefaultDeepDive;
        public static bool InteriorLightingColors = DefaultInteriorLightingColors;
        public static int InteriorLightingAlarmColorPresetIndex = DefaultInteriorLightingAlarmColorPresetIndex;
        public static int InteriorLightingSilentRunColorPresetIndex = DefaultInteriorLightingSilentRunColorPresetIndex;
        public static float BatteryFactor = DefaultBatteryFactor;
        public static float OxygenFactor = DefaultOxygenFactor;
        public static float SpeedFactor = DefaultSpeedFactor;
        public static float TorpedoFactor = DefaultTorpedoFactor;
        public static float SonarFactor = DefaultSonarFactor;

        public static Color InteriorLightingAlarmColor
        {
            get { return InteriorLightingColorPatcher.GetLightingColorPresetColor(InteriorLightingAlarmColorPresetIndex); }
        }

        public static Color InteriorLightingSilentRunColor
        {
            get { return InteriorLightingColorPatcher.GetLightingColorPresetColor(InteriorLightingSilentRunColorPresetIndex); }
        }

        public static void Load()
        {
            int savedVersion = PlayerPrefs.GetInt(PrefPrefix + "RuntimeSettingsVersion", 0);

            MegaBattery = ReadBool("MegaBattery", DefaultMegaBattery);
            MegaOxygen = ReadBool("MegaOxygen", DefaultMegaOxygen);
            SuperSpeed = ReadBool("SuperSpeed", DefaultSuperSpeed);
            MegaTorpedoes = ReadBool("MegaTorpedoes", DefaultMegaTorpedoes);
            MegaSonar = ReadBool("MegaSonar", DefaultMegaSonar);
            HeavyArmor = ReadBool("HeavyArmor", DefaultHeavyArmor);
            SuperStealth = ReadBool("SuperStealth", DefaultSuperStealth);
            DeepDive = ReadBool("DeepDive", DefaultDeepDive);
            InteriorLightingColors = ReadBool("InteriorLightingColors", DefaultInteriorLightingColors);
            InteriorLightingAlarmColorPresetIndex = ReadInteriorLightingColorPresetIndex(
                "InteriorLightingAlarmColorPresetIndex",
                DefaultInteriorLightingAlarmColorPresetIndex
            );
            InteriorLightingSilentRunColorPresetIndex = ReadInteriorLightingColorPresetIndex(
                "InteriorLightingSilentRunColorPresetIndex",
                DefaultInteriorLightingSilentRunColorPresetIndex
            );

            BatteryFactor = ReadFactor("BatteryFactor", DefaultBatteryFactor, BatteryMaxFactor);
            OxygenFactor = ReadFactor("OxygenFactor", DefaultOxygenFactor, OxygenMaxFactor);
            SpeedFactor = ReadFactor("SpeedFactor", DefaultSpeedFactor, SpeedMaxFactor);
            TorpedoFactor = ReadFactor("TorpedoFactor", DefaultTorpedoFactor, TorpedoMaxFactor);
            SonarFactor = ReadFactor("SonarFactor", DefaultSonarFactor, SonarMaxFactor);

            if (savedVersion < RuntimeSettingsVersion)
            {
                // Keep existing runtime choices, but force Heavy Armor off once after its default changed.
                if (savedVersion < 16)
                    HeavyArmor = false;

                Save();
                Debug.Log("[LongSubmerged10x] Runtime settings migrated to v" + RuntimeSettingsVersion + ".");
            }
        }

        public static void Save()
        {
            BatteryFactor = ClampBatteryFactor(BatteryFactor);
            OxygenFactor = ClampOxygenFactor(OxygenFactor);
            SpeedFactor = ClampSpeedFactor(SpeedFactor);
            TorpedoFactor = ClampTorpedoFactor(TorpedoFactor);
            SonarFactor = ClampSonarFactor(SonarFactor);
            InteriorLightingAlarmColorPresetIndex =
                ClampInteriorLightingColorPresetIndex(InteriorLightingAlarmColorPresetIndex);
            InteriorLightingSilentRunColorPresetIndex =
                ClampInteriorLightingColorPresetIndex(InteriorLightingSilentRunColorPresetIndex);

            PlayerPrefs.SetInt(PrefPrefix + "MegaBattery", MegaBattery ? 1 : 0);
            PlayerPrefs.SetInt(PrefPrefix + "MegaOxygen", MegaOxygen ? 1 : 0);
            PlayerPrefs.SetInt(PrefPrefix + "SuperSpeed", SuperSpeed ? 1 : 0);
            PlayerPrefs.SetInt(PrefPrefix + "MegaTorpedoes", MegaTorpedoes ? 1 : 0);
            PlayerPrefs.SetInt(PrefPrefix + "MegaSonar", MegaSonar ? 1 : 0);
            PlayerPrefs.SetInt(PrefPrefix + "HeavyArmor", HeavyArmor ? 1 : 0);
            PlayerPrefs.SetInt(PrefPrefix + "SuperStealth", SuperStealth ? 1 : 0);
            PlayerPrefs.SetInt(PrefPrefix + "DeepDive", DeepDive ? 1 : 0);
            PlayerPrefs.SetInt(PrefPrefix + "InteriorLightingColors", InteriorLightingColors ? 1 : 0);
            PlayerPrefs.SetInt(PrefPrefix + "InteriorLightingAlarmColorPresetIndex", InteriorLightingAlarmColorPresetIndex);
            PlayerPrefs.SetInt(PrefPrefix + "InteriorLightingSilentRunColorPresetIndex", InteriorLightingSilentRunColorPresetIndex);
            PlayerPrefs.SetFloat(PrefPrefix + "BatteryFactor", BatteryFactor);
            PlayerPrefs.SetFloat(PrefPrefix + "OxygenFactor", OxygenFactor);
            PlayerPrefs.SetFloat(PrefPrefix + "SpeedFactor", SpeedFactor);
            PlayerPrefs.SetFloat(PrefPrefix + "TorpedoFactor", TorpedoFactor);
            PlayerPrefs.SetFloat(PrefPrefix + "SonarFactor", SonarFactor);
            PlayerPrefs.SetInt(PrefPrefix + "RuntimeSettingsVersion", RuntimeSettingsVersion);
            PlayerPrefs.Save();
        }

        public static void ResetToDefaults()
        {
            MegaBattery = DefaultMegaBattery;
            MegaOxygen = DefaultMegaOxygen;
            SuperSpeed = DefaultSuperSpeed;
            MegaTorpedoes = DefaultMegaTorpedoes;
            MegaSonar = DefaultMegaSonar;
            HeavyArmor = DefaultHeavyArmor;
            SuperStealth = DefaultSuperStealth;
            DeepDive = DefaultDeepDive;
            InteriorLightingColors = DefaultInteriorLightingColors;
            InteriorLightingAlarmColorPresetIndex = DefaultInteriorLightingAlarmColorPresetIndex;
            InteriorLightingSilentRunColorPresetIndex = DefaultInteriorLightingSilentRunColorPresetIndex;
            BatteryFactor = DefaultBatteryFactor;
            OxygenFactor = DefaultOxygenFactor;
            SpeedFactor = DefaultSpeedFactor;
            TorpedoFactor = DefaultTorpedoFactor;
            SonarFactor = DefaultSonarFactor;
        }

        public static float ClampFactor(float value)
        {
            return ClampFactor(value, MaxRuntimeFactor);
        }

        public static float ClampBatteryFactor(float value)
        {
            return ClampFactor(value, BatteryMaxFactor);
        }

        public static float ClampOxygenFactor(float value)
        {
            return ClampFactor(value, OxygenMaxFactor);
        }

        public static float ClampSpeedFactor(float value)
        {
            return ClampFactor(value, SpeedMaxFactor);
        }

        public static float ClampTorpedoFactor(float value)
        {
            return ClampFactor(value, TorpedoMaxFactor);
        }

        public static float ClampSonarFactor(float value)
        {
            return ClampFactor(value, SonarMaxFactor);
        }

        public static int ClampInteriorLightingColorPresetIndex(int value)
        {
            int presetCount = InteriorLightingColorPatcher.LightingColorPresetCount;
            if (presetCount <= 0)
                return 0;

            return Mathf.Clamp(value, 0, presetCount - 1);
        }

        public static float ClampFactor(float value, float maxValue)
        {
            if (float.IsNaN(value) || float.IsInfinity(value))
                return MinRuntimeFactor;

            return Mathf.Clamp(value, MinRuntimeFactor, Mathf.Max(MinRuntimeFactor, maxValue));
        }

        private static bool ReadBool(string key, bool fallback)
        {
            return PlayerPrefs.GetInt(PrefPrefix + key, fallback ? 1 : 0) != 0;
        }

        private static float ReadFactor(string key, float fallback, float maxValue)
        {
            return ClampFactor(PlayerPrefs.GetFloat(PrefPrefix + key, fallback), maxValue);
        }

        private static int ReadInteriorLightingColorPresetIndex(string key, int fallback)
        {
            int clampedFallback = ClampInteriorLightingColorPresetIndex(fallback);
            return ClampInteriorLightingColorPresetIndex(PlayerPrefs.GetInt(PrefPrefix + key, clampedFallback));
        }
    }

    // DonJ : vrai menu Unity UI en ScreenSpaceOverlay. Je n'utilise plus l'ancien rendu IMGUI,
    // car UBOAT pouvait figer ou masquer ce rendu. F10 ouvre/ferme, Escape ferme, et les changements s'appliquent en jeu.
    internal sealed class LongSubmergedMenuController : MonoBehaviour
    {
        private const KeyCode MenuKey = KeyCode.F10;
        private const int CanvasSortingOrder = 32000;
        private const float BatteryMaintenanceIntervalSeconds = 0.20f;
        private const float MegaSonarMaintenanceIntervalSeconds = 1.00f;
        private static LongSubmergedMenuController instance;
        private static Font cachedFont;

        private GameObject panelObject;
        private Toggle megaBatteryToggle;
        private Toggle megaOxygenToggle;
        private Toggle superSpeedToggle;
        private Toggle megaTorpedoesToggle;
        private Toggle megaSonarToggle;
        private Toggle heavyArmorToggle;
        private Toggle superStealthToggle;
        private Toggle deepDiveToggle;
        private Toggle interiorLightingToggle;
        private Slider batteryFactorSlider;
        private Slider oxygenFactorSlider;
        private Slider speedFactorSlider;
        private Slider torpedoFactorSlider;
        private Slider sonarFactorSlider;
        private Dropdown alarmColorDropdown;
        private Dropdown silentRunColorDropdown;
        private Image alarmColorSwatch;
        private Image silentRunColorSwatch;
        private Button callReinforcementsButton;
        private Text batteryFactorValueText;
        private Text oxygenFactorValueText;
        private Text speedFactorValueText;
        private Text torpedoFactorValueText;
        private Text sonarFactorValueText;
        private Text reinforcementsStatusText;
        private bool visible;
        private bool suppressToggleEvents;
        private float nextBatteryMaintenanceTime;
        private float nextMegaSonarMaintenanceTime;
        private string reinforcementStatusOverride;
        private float reinforcementStatusOverrideUntil;
        private bool cursorCaptured;
        private bool previousCursorVisible;
        private CursorLockMode previousCursorLockState;

        public static void Ensure()
        {
            if (instance != null)
            {
                instance.EnsureUi();
                return;
            }

            instance = UnityEngine.Object.FindObjectOfType<LongSubmergedMenuController>();
            if (instance != null)
            {
                instance.EnsureUi();
                return;
            }

            GameObject go = new GameObject("LongSubmerged10x Runtime Menu");
            UnityEngine.Object.DontDestroyOnLoad(go);
            instance = go.AddComponent<LongSubmergedMenuController>();
        }

        private void Awake()
        {
            instance = this;
            UnityEngine.Object.DontDestroyOnLoad(gameObject);
            EnsureUi();
        }

        private void OnDestroy()
        {
            RestoreCursorIfNeeded();

            if (instance == this)
                instance = null;
        }

        private void Update()
        {
            if (Input.GetKeyDown(MenuKey))
                SetVisible(!visible, "F10");

            if (visible && Input.GetKeyDown(KeyCode.Escape))
                SetVisible(false, "Escape");

            RunBatteryMaintenanceTick();
            RunMegaSonarMaintenanceTick();

            if (visible)
                RefreshReinforcementsStatus();
        }

        private void RunBatteryMaintenanceTick()
        {
            if (!LongSubmergedRuntimeSettings.MegaBattery)
                return;

            // DonJ : le heartbeat tourne meme menu ferme. UBOAT peut recalculer la batterie apres chargement,
            // changement d'equipement ou equipage ; je reapplique donc le mode nucleaire regulierement.
            float now = Time.unscaledTime;
            if (now < nextBatteryMaintenanceTime)
                return;

            nextBatteryMaintenanceTime = now + BatteryMaintenanceIntervalSeconds;
            LongSubmergedRuntimeApplier.MaintainBatteryRuntime("runtime heartbeat");
        }

        private void RunMegaSonarMaintenanceTick()
        {
            if (!LongSubmergedRuntimeSettings.MegaSonar)
                return;

            float now = Time.unscaledTime;
            if (now < nextMegaSonarMaintenanceTime)
                return;

            nextMegaSonarMaintenanceTime = now + MegaSonarMaintenanceIntervalSeconds;
            MegaSonarRuntimePatcher.ApplyAll("runtime sonar heartbeat");
        }

        private void SaveAndApplyCurrentControlsNow(string reason)
        {
            ReadControlStateIntoSettings();
            RefreshFactorLabels();
            RefreshColorPresetControls();
            LongSubmergedRuntimeSettings.Save();
            nextBatteryMaintenanceTime = 0f;
            nextMegaSonarMaintenanceTime = 0f;
            LongSubmergedRuntimeApplier.ApplyAll(string.IsNullOrEmpty(reason) ? "unity ui change" : reason);
        }

        private void EnsureUi()
        {
            if (panelObject != null)
                return;

            try
            {
                // DonJ : Canvas overlay avec ordre tres haut pour passer au-dessus de l'UI du jeu.
                Canvas canvas = gameObject.GetComponent<Canvas>();
                if (canvas == null)
                    canvas = gameObject.AddComponent<Canvas>();

                canvas.renderMode = RenderMode.ScreenSpaceOverlay;
                canvas.sortingOrder = CanvasSortingOrder;
                canvas.overrideSorting = true;

                CanvasScaler scaler = gameObject.GetComponent<CanvasScaler>();
                if (scaler == null)
                    scaler = gameObject.AddComponent<CanvasScaler>();

                scaler.uiScaleMode = CanvasScaler.ScaleMode.ScaleWithScreenSize;
                scaler.referenceResolution = new Vector2(1920f, 1080f);
                scaler.matchWidthOrHeight = 0.5f;

                if (gameObject.GetComponent<GraphicRaycaster>() == null)
                    gameObject.AddComponent<GraphicRaycaster>();

                EnsureEventSystem();
                BuildPanel();
                RefreshControlState();
                panelObject.SetActive(false);
                Debug.Log("[LongSubmerged10x] Runtime Unity UI menu controller ready on F10.");
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }
        }

        private void BuildPanel()
        {
            // DonJ : panneau compact de test runtime. Tous les controles modifient les valeurs sauvegardees
            // et rappellent ApplyAll pour voir le resultat directement dans la partie.
            panelObject = CreateUiObject("LongSubmerged10x Panel", transform);
            Image panelImage = panelObject.AddComponent<Image>();
            panelImage.color = new Color(0.04f, 0.05f, 0.06f, 0.96f);

            RectTransform panelRect = panelObject.GetComponent<RectTransform>();
            panelRect.anchorMin = new Vector2(0f, 1f);
            panelRect.anchorMax = new Vector2(0f, 1f);
            panelRect.pivot = new Vector2(0f, 1f);
            panelRect.anchoredPosition = new Vector2(28f, -82f);
            panelRect.sizeDelta = new Vector2(470f, 920f);

            CreateText(panelObject.transform, "Title", "Long Submerged 10x+", 20, FontStyle.Bold, new Vector2(18f, -16f), new Vector2(410f, 30f));
            CreateText(panelObject.transform, "Hint", "F10 ferme. Les reglages sont sauvegardes et appliques en partie.", 13, FontStyle.Normal, new Vector2(18f, -48f), new Vector2(430f, 24f));

            megaBatteryToggle = CreateToggle(panelObject.transform, "Mega Batterie", new Vector2(20f, -82f));
            batteryFactorSlider = CreateFactorSlider(panelObject.transform, "Batterie", LongSubmergedRuntimeSettings.BatteryMaxFactor, new Vector2(20f, -118f), out batteryFactorValueText);

            megaOxygenToggle = CreateToggle(panelObject.transform, "Mega Oxygene", new Vector2(20f, -158f));
            oxygenFactorSlider = CreateFactorSlider(panelObject.transform, "Oxygene", LongSubmergedRuntimeSettings.OxygenMaxFactor, new Vector2(20f, -194f), out oxygenFactorValueText);

            superSpeedToggle = CreateToggle(panelObject.transform, "SuperVitesse", new Vector2(20f, -234f));
            speedFactorSlider = CreateFactorSlider(panelObject.transform, "Vitesses rapides", LongSubmergedRuntimeSettings.SpeedMaxFactor, new Vector2(20f, -270f), out speedFactorValueText);

            megaTorpedoesToggle = CreateToggle(panelObject.transform, "Mega Torpilles", new Vector2(20f, -310f));
            torpedoFactorSlider = CreateFactorSlider(panelObject.transform, "Torpilles", LongSubmergedRuntimeSettings.TorpedoMaxFactor, new Vector2(20f, -346f), out torpedoFactorValueText);

            megaSonarToggle = CreateToggle(panelObject.transform, "Mega Sonar", new Vector2(20f, -386f));
            sonarFactorSlider = CreateFactorSlider(panelObject.transform, "Hydrophone portee", LongSubmergedRuntimeSettings.SonarMaxFactor, new Vector2(20f, -422f), out sonarFactorValueText);

            heavyArmorToggle = CreateToggle(panelObject.transform, "Blindage lourd x3", new Vector2(20f, -462f));
            superStealthToggle = CreateToggle(panelObject.transform, "Super discrétion", new Vector2(20f, -498f));

            deepDiveToggle = CreateToggle(panelObject.transform, "Plongée x2", new Vector2(20f, -534f));
            interiorLightingToggle = CreateToggle(panelObject.transform, "Couleurs eclairage", new Vector2(20f, -570f));
            alarmColorDropdown = CreateColorDropdown(panelObject.transform, "Alarme", new Vector2(20f, -606f), out alarmColorSwatch);
            silentRunColorDropdown = CreateColorDropdown(panelObject.transform, "Silencieux", new Vector2(20f, -650f), out silentRunColorSwatch);

            callReinforcementsButton = CreateButton(panelObject.transform, "Appeler renforts", new Vector2(20f, -704f), new Vector2(180f, 38f));
            callReinforcementsButton.onClick.AddListener(OnCallReinforcementsClicked);

            reinforcementsStatusText = CreateText(panelObject.transform, "Renforts Status", "Pret", 13, FontStyle.Normal, new Vector2(216f, -704f), new Vector2(230f, 38f));
            reinforcementsStatusText.alignment = TextAnchor.MiddleLeft;

            Button defaultsButton = CreateButton(panelObject.transform, "Par defaut", new Vector2(20f, -802f), new Vector2(140f, 38f));
            defaultsButton.onClick.AddListener(OnDefaultsClicked);

            Button refreshButton = CreateButton(panelObject.transform, "Reappliquer maintenant", new Vector2(176f, -802f), new Vector2(220f, 38f));
            refreshButton.onClick.AddListener(OnRefreshClicked);
        }

        private static void EnsureEventSystem()
        {
            if (UnityEngine.Object.FindObjectOfType<EventSystem>() != null)
                return;

            GameObject eventSystemObject = new GameObject("LongSubmerged10x EventSystem");
            UnityEngine.Object.DontDestroyOnLoad(eventSystemObject);
            eventSystemObject.AddComponent<EventSystem>();
            eventSystemObject.AddComponent<StandaloneInputModule>();
            Debug.Log("[LongSubmerged10x] Runtime menu created fallback EventSystem.");
        }

        private void SetVisible(bool value, string source)
        {
            EnsureUi();

            if (panelObject == null || visible == value)
                return;

            visible = value;
            panelObject.SetActive(visible);

            if (visible)
            {
                RefreshControlState();
                RefreshReinforcementsStatus();
                CaptureCursor();
            }
            else
            {
                RestoreCursorIfNeeded();
            }

            Debug.Log("[LongSubmerged10x] Runtime menu " + (visible ? "opened" : "closed") + " by " + source + ".");
        }

        private void OnToggleChanged(bool ignored)
        {
            if (suppressToggleEvents)
                return;

            SaveAndApplyCurrentControlsNow("unity ui toggle");
        }

        private void OnFactorSliderChanged(float ignored)
        {
            if (suppressToggleEvents)
                return;

            SaveAndApplyCurrentControlsNow("unity ui slider");
        }

        private void OnColorDropdownChanged(int ignored)
        {
            if (suppressToggleEvents)
                return;

            SaveAndApplyCurrentControlsNow("unity ui color dropdown");
        }

        private void ReadControlStateIntoSettings()
        {
            // UI changes must update the runtime state before every immediate apply.
            LongSubmergedRuntimeSettings.MegaBattery = megaBatteryToggle != null && megaBatteryToggle.isOn;
            LongSubmergedRuntimeSettings.MegaOxygen = megaOxygenToggle != null && megaOxygenToggle.isOn;
            LongSubmergedRuntimeSettings.SuperSpeed = superSpeedToggle != null && superSpeedToggle.isOn;
            LongSubmergedRuntimeSettings.MegaTorpedoes = megaTorpedoesToggle != null && megaTorpedoesToggle.isOn;
            LongSubmergedRuntimeSettings.MegaSonar = megaSonarToggle != null && megaSonarToggle.isOn;
            LongSubmergedRuntimeSettings.HeavyArmor = heavyArmorToggle != null && heavyArmorToggle.isOn;
            LongSubmergedRuntimeSettings.SuperStealth = superStealthToggle != null && superStealthToggle.isOn;
            LongSubmergedRuntimeSettings.DeepDive = deepDiveToggle != null && deepDiveToggle.isOn;
            LongSubmergedRuntimeSettings.InteriorLightingColors = interiorLightingToggle != null && interiorLightingToggle.isOn;
            LongSubmergedRuntimeSettings.InteriorLightingAlarmColorPresetIndex =
                ReadColorDropdownPresetIndex(alarmColorDropdown, LongSubmergedRuntimeSettings.InteriorLightingAlarmColorPresetIndex);
            LongSubmergedRuntimeSettings.InteriorLightingSilentRunColorPresetIndex =
                ReadColorDropdownPresetIndex(silentRunColorDropdown, LongSubmergedRuntimeSettings.InteriorLightingSilentRunColorPresetIndex);
            LongSubmergedRuntimeSettings.BatteryFactor = ReadSliderFactor(batteryFactorSlider, LongSubmergedRuntimeSettings.BatteryMaxFactor);
            LongSubmergedRuntimeSettings.OxygenFactor = ReadSliderFactor(oxygenFactorSlider, LongSubmergedRuntimeSettings.OxygenMaxFactor);
            LongSubmergedRuntimeSettings.SpeedFactor = ReadSliderFactor(speedFactorSlider, LongSubmergedRuntimeSettings.SpeedMaxFactor);
            LongSubmergedRuntimeSettings.TorpedoFactor = ReadSliderFactor(torpedoFactorSlider, LongSubmergedRuntimeSettings.TorpedoMaxFactor);
            LongSubmergedRuntimeSettings.SonarFactor = ReadSliderFactor(sonarFactorSlider, LongSubmergedRuntimeSettings.SonarMaxFactor);
        }

        private void OnDefaultsClicked()
        {
            LongSubmergedRuntimeSettings.ResetToDefaults();
            LongSubmergedRuntimeSettings.Save();
            RefreshControlState();
            nextBatteryMaintenanceTime = 0f;
            nextMegaSonarMaintenanceTime = 0f;
            LongSubmergedRuntimeApplier.ApplyAll("unity ui defaults");
        }

        private void OnRefreshClicked()
        {
            SaveAndApplyCurrentControlsNow("unity ui refresh");
        }

        private void OnCallReinforcementsClicked()
        {
            SaveAndApplyCurrentControlsNow("unity ui call reinforcements");
            SetReinforcementsStatusOverride("Appel...", 1f);
            string status = ReinforcementRuntimeController.CallReinforcements("unity ui call reinforcements");
            SetReinforcementsStatusOverride(status, 4f);
            RefreshReinforcementsStatus();
        }

        private void RefreshReinforcementsStatus()
        {
            string availabilityStatus = ReinforcementRuntimeController.GetStatusText();
            string displayStatus = availabilityStatus;

            if (!string.IsNullOrEmpty(reinforcementStatusOverride) && Time.unscaledTime < reinforcementStatusOverrideUntil)
                displayStatus = reinforcementStatusOverride;
            else
                reinforcementStatusOverride = null;

            SetReinforcementsStatus(displayStatus);

            if (callReinforcementsButton != null)
                callReinforcementsButton.interactable = availabilityStatus == "Pret";
        }

        private void SetReinforcementsStatusOverride(string status, float seconds)
        {
            reinforcementStatusOverride = string.IsNullOrEmpty(status) ? null : status;
            reinforcementStatusOverrideUntil = Time.unscaledTime + Mathf.Max(0f, seconds);
            SetReinforcementsStatus(status);
        }

        private void SetReinforcementsStatus(string status)
        {
            if (reinforcementsStatusText != null)
                reinforcementsStatusText.text = string.IsNullOrEmpty(status) ? "Pret" : status;
        }

        private void RefreshControlState()
        {
            suppressToggleEvents = true;

            try
            {
                if (megaBatteryToggle != null)
                    megaBatteryToggle.isOn = LongSubmergedRuntimeSettings.MegaBattery;

                if (megaOxygenToggle != null)
                    megaOxygenToggle.isOn = LongSubmergedRuntimeSettings.MegaOxygen;

                if (superSpeedToggle != null)
                    superSpeedToggle.isOn = LongSubmergedRuntimeSettings.SuperSpeed;

                if (megaTorpedoesToggle != null)
                    megaTorpedoesToggle.isOn = LongSubmergedRuntimeSettings.MegaTorpedoes;

                if (megaSonarToggle != null)
                    megaSonarToggle.isOn = LongSubmergedRuntimeSettings.MegaSonar;

                if (heavyArmorToggle != null)
                    heavyArmorToggle.isOn = LongSubmergedRuntimeSettings.HeavyArmor;

                if (superStealthToggle != null)
                    superStealthToggle.isOn = LongSubmergedRuntimeSettings.SuperStealth;

                if (deepDiveToggle != null)
                    deepDiveToggle.isOn = LongSubmergedRuntimeSettings.DeepDive;

                if (interiorLightingToggle != null)
                    interiorLightingToggle.isOn = LongSubmergedRuntimeSettings.InteriorLightingColors;

                SetSliderValue(batteryFactorSlider, LongSubmergedRuntimeSettings.BatteryFactor, LongSubmergedRuntimeSettings.BatteryMaxFactor);
                SetSliderValue(oxygenFactorSlider, LongSubmergedRuntimeSettings.OxygenFactor, LongSubmergedRuntimeSettings.OxygenMaxFactor);
                SetSliderValue(speedFactorSlider, LongSubmergedRuntimeSettings.SpeedFactor, LongSubmergedRuntimeSettings.SpeedMaxFactor);
                SetSliderValue(torpedoFactorSlider, LongSubmergedRuntimeSettings.TorpedoFactor, LongSubmergedRuntimeSettings.TorpedoMaxFactor);
                SetSliderValue(sonarFactorSlider, LongSubmergedRuntimeSettings.SonarFactor, LongSubmergedRuntimeSettings.SonarMaxFactor);
                RefreshFactorLabels();
                RefreshColorPresetControls();
            }
            finally
            {
                suppressToggleEvents = false;
            }
        }

        private void RefreshFactorLabels()
        {
            SetFactorLabel(
                batteryFactorValueText,
                batteryFactorSlider,
                LongSubmergedRuntimeSettings.BatteryMaxFactor,
                "x",
                LongSubmergedRuntimeSettings.MegaBattery ? "inf" : null
            );

            SetFactorLabel(
                oxygenFactorValueText,
                oxygenFactorSlider,
                LongSubmergedRuntimeSettings.OxygenMaxFactor,
                "x",
                oxygenFactorSlider != null && oxygenFactorSlider.value >= LongSubmergedRuntimeSettings.OxygenMaxFactor ? "90j" : null
            );

            SetFactorLabel(speedFactorValueText, speedFactorSlider, LongSubmergedRuntimeSettings.SpeedMaxFactor, "x", null);
            SetFactorLabel(torpedoFactorValueText, torpedoFactorSlider, LongSubmergedRuntimeSettings.TorpedoMaxFactor, "x", null);
            SetFactorLabel(sonarFactorValueText, sonarFactorSlider, LongSubmergedRuntimeSettings.SonarMaxFactor, "x", null);
        }

        private void RefreshColorPresetControls()
        {
            SetColorDropdownValue(alarmColorDropdown, LongSubmergedRuntimeSettings.InteriorLightingAlarmColorPresetIndex);
            SetColorDropdownValue(silentRunColorDropdown, LongSubmergedRuntimeSettings.InteriorLightingSilentRunColorPresetIndex);
            SetColorSwatch(alarmColorSwatch, LongSubmergedRuntimeSettings.InteriorLightingAlarmColor);
            SetColorSwatch(silentRunColorSwatch, LongSubmergedRuntimeSettings.InteriorLightingSilentRunColor);
        }

        private static void SetSliderValue(Slider slider, float value, float maxValue)
        {
            if (slider == null)
                return;

            slider.minValue = LongSubmergedRuntimeSettings.MinRuntimeFactor;
            slider.maxValue = maxValue;
            slider.wholeNumbers = true;
            slider.value = LongSubmergedRuntimeSettings.ClampFactor(value, maxValue);
        }

        private static float ReadSliderFactor(Slider slider, float maxValue)
        {
            return slider == null
                ? LongSubmergedRuntimeSettings.MinRuntimeFactor
                : LongSubmergedRuntimeSettings.ClampFactor(slider.value, maxValue);
        }

        private static int ReadColorDropdownPresetIndex(Dropdown dropdown, int fallback)
        {
            if (dropdown == null)
                return LongSubmergedRuntimeSettings.ClampInteriorLightingColorPresetIndex(fallback);

            return LongSubmergedRuntimeSettings.ClampInteriorLightingColorPresetIndex(dropdown.value);
        }

        private static void SetFactorLabel(Text text, Slider slider, float maxValue, string prefix, string suffixOverride)
        {
            if (text == null || slider == null)
                return;

            float value = LongSubmergedRuntimeSettings.ClampFactor(slider.value, maxValue);
            text.text = suffixOverride == null ? prefix + value.ToString("0") : suffixOverride;
        }

        private static void SetColorDropdownValue(Dropdown dropdown, int presetIndex)
        {
            if (dropdown == null)
                return;

            int clampedIndex = LongSubmergedRuntimeSettings.ClampInteriorLightingColorPresetIndex(presetIndex);
            dropdown.SetValueWithoutNotify(clampedIndex);
            dropdown.RefreshShownValue();
        }

        private static void SetColorSwatch(Image swatch, Color color)
        {
            if (swatch != null)
                swatch.color = color;
        }

        private void CaptureCursor()
        {
            if (cursorCaptured)
                return;

            previousCursorVisible = Cursor.visible;
            previousCursorLockState = Cursor.lockState;
            Cursor.visible = true;
            Cursor.lockState = CursorLockMode.None;
            cursorCaptured = true;
        }

        private void RestoreCursorIfNeeded()
        {
            if (!cursorCaptured)
                return;

            Cursor.visible = previousCursorVisible;
            Cursor.lockState = previousCursorLockState;
            cursorCaptured = false;
        }

        private Slider CreateFactorSlider(Transform parent, string label, float maxValue, Vector2 anchoredPosition, out Text valueText)
        {
            GameObject root = CreateUiObject(label + " Factor", parent);
            RectTransform rootRect = root.GetComponent<RectTransform>();
            rootRect.anchorMin = new Vector2(0f, 1f);
            rootRect.anchorMax = new Vector2(0f, 1f);
            rootRect.pivot = new Vector2(0f, 1f);
            rootRect.anchoredPosition = anchoredPosition;
            rootRect.sizeDelta = new Vector2(420f, 30f);

            Text labelText = CreateText(root.transform, "Label", label, 13, FontStyle.Bold, new Vector2(0f, -1f), new Vector2(116f, 24f));
            labelText.alignment = TextAnchor.MiddleLeft;

            valueText = CreateText(root.transform, "Value", "x1", 13, FontStyle.Bold, new Vector2(362f, -1f), new Vector2(58f, 24f));
            valueText.alignment = TextAnchor.MiddleRight;

            GameObject sliderObject = CreateUiObject("Slider", root.transform);
            RectTransform sliderRect = sliderObject.GetComponent<RectTransform>();
            sliderRect.anchorMin = new Vector2(0f, 0.5f);
            sliderRect.anchorMax = new Vector2(0f, 0.5f);
            sliderRect.pivot = new Vector2(0f, 0.5f);
            sliderRect.anchoredPosition = new Vector2(124f, -3f);
            sliderRect.sizeDelta = new Vector2(230f, 18f);

            Slider slider = sliderObject.AddComponent<Slider>();
            slider.minValue = LongSubmergedRuntimeSettings.MinRuntimeFactor;
            slider.maxValue = maxValue;
            slider.wholeNumbers = true;

            GameObject background = CreateUiObject("Background", sliderObject.transform);
            Image backgroundImage = background.AddComponent<Image>();
            backgroundImage.color = new Color(0.12f, 0.13f, 0.15f, 1f);
            RectTransform backgroundRect = background.GetComponent<RectTransform>();
            backgroundRect.anchorMin = new Vector2(0f, 0.5f);
            backgroundRect.anchorMax = new Vector2(1f, 0.5f);
            backgroundRect.pivot = new Vector2(0.5f, 0.5f);
            backgroundRect.anchoredPosition = Vector2.zero;
            backgroundRect.sizeDelta = new Vector2(0f, 6f);

            GameObject fillArea = CreateUiObject("Fill Area", sliderObject.transform);
            RectTransform fillAreaRect = fillArea.GetComponent<RectTransform>();
            fillAreaRect.anchorMin = new Vector2(0f, 0f);
            fillAreaRect.anchorMax = new Vector2(1f, 1f);
            fillAreaRect.offsetMin = new Vector2(5f, 0f);
            fillAreaRect.offsetMax = new Vector2(-5f, 0f);

            GameObject fill = CreateUiObject("Fill", fillArea.transform);
            Image fillImage = fill.AddComponent<Image>();
            fillImage.color = new Color(0.18f, 0.85f, 0.52f, 1f);
            RectTransform fillRect = fill.GetComponent<RectTransform>();
            fillRect.anchorMin = new Vector2(0f, 0.5f);
            fillRect.anchorMax = new Vector2(1f, 0.5f);
            fillRect.pivot = new Vector2(0f, 0.5f);
            fillRect.anchoredPosition = Vector2.zero;
            fillRect.sizeDelta = new Vector2(0f, 6f);

            GameObject handleArea = CreateUiObject("Handle Slide Area", sliderObject.transform);
            RectTransform handleAreaRect = handleArea.GetComponent<RectTransform>();
            handleAreaRect.anchorMin = Vector2.zero;
            handleAreaRect.anchorMax = Vector2.one;
            handleAreaRect.offsetMin = new Vector2(5f, 0f);
            handleAreaRect.offsetMax = new Vector2(-5f, 0f);

            GameObject handle = CreateUiObject("Handle", handleArea.transform);
            Image handleImage = handle.AddComponent<Image>();
            handleImage.color = new Color(0.92f, 0.95f, 0.98f, 1f);
            RectTransform handleRect = handle.GetComponent<RectTransform>();
            handleRect.sizeDelta = new Vector2(16f, 16f);

            slider.fillRect = fillRect;
            slider.handleRect = handleRect;
            slider.targetGraphic = handleImage;
            slider.onValueChanged.AddListener(OnFactorSliderChanged);

            return slider;
        }

        private Dropdown CreateColorDropdown(Transform parent, string label, Vector2 anchoredPosition, out Image swatch)
        {
            GameObject root = CreateUiObject(label + " Color Preset", parent);
            RectTransform rootRect = root.GetComponent<RectTransform>();
            rootRect.anchorMin = new Vector2(0f, 1f);
            rootRect.anchorMax = new Vector2(0f, 1f);
            rootRect.pivot = new Vector2(0f, 1f);
            rootRect.anchoredPosition = anchoredPosition;
            rootRect.sizeDelta = new Vector2(420f, 34f);

            Text labelText = CreateText(root.transform, "Label", label, 13, FontStyle.Bold, new Vector2(0f, -2f), new Vector2(92f, 28f));
            labelText.alignment = TextAnchor.MiddleLeft;

            GameObject swatchObject = CreateUiObject("Swatch", root.transform);
            swatch = swatchObject.AddComponent<Image>();
            swatch.color = Color.white;
            RectTransform swatchRect = swatchObject.GetComponent<RectTransform>();
            swatchRect.anchorMin = new Vector2(0f, 0.5f);
            swatchRect.anchorMax = new Vector2(0f, 0.5f);
            swatchRect.pivot = new Vector2(0f, 0.5f);
            swatchRect.anchoredPosition = new Vector2(96f, -2f);
            swatchRect.sizeDelta = new Vector2(24f, 24f);

            GameObject dropdownObject = CreateUiObject("Dropdown", root.transform);
            Image dropdownImage = dropdownObject.AddComponent<Image>();
            dropdownImage.color = new Color(0.12f, 0.13f, 0.15f, 1f);
            RectTransform dropdownRect = dropdownObject.GetComponent<RectTransform>();
            dropdownRect.anchorMin = new Vector2(0f, 0.5f);
            dropdownRect.anchorMax = new Vector2(0f, 0.5f);
            dropdownRect.pivot = new Vector2(0f, 0.5f);
            dropdownRect.anchoredPosition = new Vector2(130f, -2f);
            dropdownRect.sizeDelta = new Vector2(260f, 30f);

            Dropdown dropdown = dropdownObject.AddComponent<Dropdown>();
            dropdown.targetGraphic = dropdownImage;

            Text captionText = CreateText(dropdownObject.transform, "Caption", "", 13, FontStyle.Normal, new Vector2(10f, -1f), new Vector2(214f, 28f));
            captionText.alignment = TextAnchor.MiddleLeft;
            dropdown.captionText = captionText;

            Text arrowText = CreateText(dropdownObject.transform, "Arrow", "v", 14, FontStyle.Bold, new Vector2(232f, -1f), new Vector2(22f, 28f));
            arrowText.alignment = TextAnchor.MiddleCenter;

            Text itemText;
            dropdown.template = CreateDropdownTemplate(dropdownObject.transform, out itemText);
            dropdown.itemText = itemText;
            dropdown.ClearOptions();
            dropdown.AddOptions(InteriorLightingColorPatcher.GetLightingColorPresetNames());
            dropdown.onValueChanged.AddListener(OnColorDropdownChanged);
            dropdown.RefreshShownValue();

            return dropdown;
        }

        private RectTransform CreateDropdownTemplate(Transform parent, out Text itemText)
        {
            GameObject template = CreateUiObject("Template", parent);
            Image templateImage = template.AddComponent<Image>();
            templateImage.color = new Color(0.06f, 0.07f, 0.08f, 0.98f);
            RectTransform templateRect = template.GetComponent<RectTransform>();
            templateRect.anchorMin = new Vector2(0f, 1f);
            templateRect.anchorMax = new Vector2(1f, 1f);
            templateRect.pivot = new Vector2(0.5f, 0f);
            templateRect.anchoredPosition = new Vector2(0f, 32f);
            templateRect.sizeDelta = new Vector2(0f, 242f);

            Canvas templateCanvas = template.AddComponent<Canvas>();
            templateCanvas.overrideSorting = true;
            templateCanvas.sortingOrder = CanvasSortingOrder + 2;
            template.AddComponent<GraphicRaycaster>();

            ScrollRect scrollRect = template.AddComponent<ScrollRect>();
            scrollRect.horizontal = false;
            scrollRect.vertical = true;
            scrollRect.movementType = ScrollRect.MovementType.Clamped;

            GameObject viewport = CreateUiObject("Viewport", template.transform);
            Image viewportImage = viewport.AddComponent<Image>();
            viewportImage.color = new Color(1f, 1f, 1f, 0.02f);
            Mask viewportMask = viewport.AddComponent<Mask>();
            viewportMask.showMaskGraphic = false;
            RectTransform viewportRect = viewport.GetComponent<RectTransform>();
            viewportRect.anchorMin = Vector2.zero;
            viewportRect.anchorMax = Vector2.one;
            viewportRect.offsetMin = new Vector2(4f, 4f);
            viewportRect.offsetMax = new Vector2(-4f, -4f);

            GameObject content = CreateUiObject("Content", viewport.transform);
            RectTransform contentRect = content.GetComponent<RectTransform>();
            contentRect.anchorMin = new Vector2(0f, 1f);
            contentRect.anchorMax = new Vector2(1f, 1f);
            contentRect.pivot = new Vector2(0.5f, 1f);
            contentRect.anchoredPosition = Vector2.zero;
            contentRect.sizeDelta = new Vector2(0f, InteriorLightingColorPatcher.LightingColorPresetCount * 26f);

            VerticalLayoutGroup layout = content.AddComponent<VerticalLayoutGroup>();
            layout.childAlignment = TextAnchor.UpperLeft;
            layout.childControlHeight = false;
            layout.childControlWidth = true;
            layout.childForceExpandHeight = false;
            layout.childForceExpandWidth = true;

            ContentSizeFitter fitter = content.AddComponent<ContentSizeFitter>();
            fitter.verticalFit = ContentSizeFitter.FitMode.PreferredSize;

            GameObject item = CreateUiObject("Item", content.transform);
            RectTransform itemRect = item.GetComponent<RectTransform>();
            itemRect.anchorMin = new Vector2(0f, 1f);
            itemRect.anchorMax = new Vector2(1f, 1f);
            itemRect.pivot = new Vector2(0.5f, 1f);
            itemRect.sizeDelta = new Vector2(0f, 26f);

            Toggle itemToggle = item.AddComponent<Toggle>();
            Image itemBackground = item.AddComponent<Image>();
            itemBackground.color = new Color(0.10f, 0.11f, 0.13f, 1f);
            itemToggle.targetGraphic = itemBackground;

            GameObject checkmark = CreateUiObject("Item Checkmark", item.transform);
            Image checkmarkImage = checkmark.AddComponent<Image>();
            checkmarkImage.color = new Color(0.18f, 0.85f, 0.52f, 1f);
            RectTransform checkmarkRect = checkmark.GetComponent<RectTransform>();
            checkmarkRect.anchorMin = new Vector2(0f, 0.5f);
            checkmarkRect.anchorMax = new Vector2(0f, 0.5f);
            checkmarkRect.pivot = new Vector2(0f, 0.5f);
            checkmarkRect.anchoredPosition = new Vector2(8f, 0f);
            checkmarkRect.sizeDelta = new Vector2(10f, 10f);
            itemToggle.graphic = checkmarkImage;

            itemText = CreateText(item.transform, "Item Label", "Option", 13, FontStyle.Normal, new Vector2(26f, 0f), new Vector2(210f, 26f));
            itemText.alignment = TextAnchor.MiddleLeft;

            scrollRect.viewport = viewportRect;
            scrollRect.content = contentRect;
            template.SetActive(false);
            return templateRect;
        }

        private Toggle CreateToggle(Transform parent, string label, Vector2 anchoredPosition)
        {
            GameObject root = CreateUiObject(label + " Toggle", parent);
            RectTransform rootRect = root.GetComponent<RectTransform>();
            rootRect.anchorMin = new Vector2(0f, 1f);
            rootRect.anchorMax = new Vector2(0f, 1f);
            rootRect.pivot = new Vector2(0f, 1f);
            rootRect.anchoredPosition = anchoredPosition;
            rootRect.sizeDelta = new Vector2(330f, 30f);

            Toggle toggle = root.AddComponent<Toggle>();
            Image rowHitImage = root.AddComponent<Image>();
            rowHitImage.color = new Color(1f, 1f, 1f, 0f);
            rowHitImage.raycastTarget = true;

            GameObject box = CreateUiObject("Box", root.transform);
            Image boxImage = box.AddComponent<Image>();
            boxImage.color = new Color(0.16f, 0.18f, 0.2f, 1f);
            RectTransform boxRect = box.GetComponent<RectTransform>();
            boxRect.anchorMin = new Vector2(0f, 0.5f);
            boxRect.anchorMax = new Vector2(0f, 0.5f);
            boxRect.pivot = new Vector2(0f, 0.5f);
            boxRect.anchoredPosition = new Vector2(0f, 0f);
            boxRect.sizeDelta = new Vector2(24f, 24f);

            GameObject checkmark = CreateUiObject("Checkmark", box.transform);
            Image checkmarkImage = checkmark.AddComponent<Image>();
            checkmarkImage.color = new Color(0.18f, 0.85f, 0.52f, 1f);
            RectTransform checkRect = checkmark.GetComponent<RectTransform>();
            checkRect.anchorMin = new Vector2(0.5f, 0.5f);
            checkRect.anchorMax = new Vector2(0.5f, 0.5f);
            checkRect.pivot = new Vector2(0.5f, 0.5f);
            checkRect.anchoredPosition = Vector2.zero;
            checkRect.sizeDelta = new Vector2(14f, 14f);

            Text labelText = CreateText(root.transform, "Label", label, 16, FontStyle.Normal, new Vector2(34f, -2f), new Vector2(280f, 28f));
            labelText.alignment = TextAnchor.MiddleLeft;

            toggle.targetGraphic = boxImage;
            toggle.graphic = checkmarkImage;
            toggle.onValueChanged.AddListener(OnToggleChanged);

            return toggle;
        }

        private Button CreateButton(Transform parent, string label, Vector2 anchoredPosition, Vector2 size)
        {
            GameObject buttonObject = CreateUiObject(label + " Button", parent);
            Image image = buttonObject.AddComponent<Image>();
            image.color = new Color(0.13f, 0.26f, 0.42f, 1f);

            RectTransform rect = buttonObject.GetComponent<RectTransform>();
            rect.anchorMin = new Vector2(0f, 1f);
            rect.anchorMax = new Vector2(0f, 1f);
            rect.pivot = new Vector2(0f, 1f);
            rect.anchoredPosition = anchoredPosition;
            rect.sizeDelta = size;

            Button button = buttonObject.AddComponent<Button>();
            button.targetGraphic = image;

            Text text = CreateText(buttonObject.transform, "Label", label, 15, FontStyle.Bold, Vector2.zero, size);
            text.alignment = TextAnchor.MiddleCenter;
            RectTransform textRect = text.GetComponent<RectTransform>();
            textRect.anchorMin = Vector2.zero;
            textRect.anchorMax = Vector2.one;
            textRect.pivot = new Vector2(0.5f, 0.5f);
            textRect.anchoredPosition = Vector2.zero;
            textRect.sizeDelta = Vector2.zero;

            return button;
        }

        private static Text CreateText(Transform parent, string name, string value, int fontSize, FontStyle fontStyle, Vector2 anchoredPosition, Vector2 size)
        {
            GameObject textObject = CreateUiObject(name, parent);
            RectTransform rect = textObject.GetComponent<RectTransform>();
            rect.anchorMin = new Vector2(0f, 1f);
            rect.anchorMax = new Vector2(0f, 1f);
            rect.pivot = new Vector2(0f, 1f);
            rect.anchoredPosition = anchoredPosition;
            rect.sizeDelta = size;

            Text text = textObject.AddComponent<Text>();
            text.text = value;
            text.font = UiFont;
            text.fontSize = fontSize;
            text.fontStyle = fontStyle;
            text.color = Color.white;
            text.alignment = TextAnchor.UpperLeft;
            text.raycastTarget = false;
            return text;
        }

        private static GameObject CreateUiObject(string name, Transform parent)
        {
            GameObject go = new GameObject(name);
            go.transform.SetParent(parent, false);
            go.AddComponent<RectTransform>();
            return go;
        }

        private static Font UiFont
        {
            get
            {
                if (cachedFont == null)
                    cachedFont = Resources.GetBuiltinResource<Font>("Arial.ttf");

                return cachedFont;
            }
        }
    }

    // DonJ : coeur gameplay du mod. Cette classe applique les valeurs runtime sans reecrire les fichiers XLSX :
    // elle pose des modifiers sur les Parametres du jeu, garde la batterie pleine et ajuste torpilles/vitesse/oxygene.
    internal static class LongSubmergedRuntimeApplier
    {
        // DonJ : constantes du profil livre. Le joueur peut ensuite ajuster en F10 sans regenerer le mod.
        private const float OxygenRuntimeMaxFactor = __OXYGEN_CONSUMPTION_FACTOR__;
        private const float BatteryCapacityDataFactor = __BATTERY_CAPACITY_FACTOR__;
        private const float EnergyUsageDataFactor = __ENERGY_USAGE_FACTOR__;
        private const float BatteryCapacityVanillaRestoreScale = 1f / __BATTERY_CAPACITY_FACTOR__;
        private const float EnergyUsageVanillaRestoreScale = 1f / __ENERGY_USAGE_FACTOR__;
        private const float TorpedoDamageScale = __TORPEDO_DAMAGE_FACTOR__;
        private const float TorpedoCrewDamageScale = __TORPEDO_CREW_DAMAGE_FACTOR__;
        private const float TorpedoExplosionRadiusScale = __TORPEDO_EXPLOSION_RADIUS_FACTOR__;
        private const float TorpedoExplosionIntensityScale = __TORPEDO_EXPLOSION_INTENSITY_FACTOR__;
        private const bool PerfectTorpedoReliability = __PERFECT_TORPEDO_RELIABILITY__;
        private const float TorpedoGuidanceLeadSeconds = 4f;
        private const float TorpedoGuidanceMinimumDetonationDistance = 20f;
        private const float TorpedoGuidanceMaximumDetonationDistance = 80f;
        private const float TorpedoGuidanceDetonationRadiusRatio = 0.75f;
        private const string RuntimeScaleModifierName = "LongSubmerged10x Runtime Toggle";
        private const string RuntimeBatteryGainModifierName = "LongSubmerged10x Battery Gain Runtime";
        private const string RuntimeNuclearBatteryCapacityModifierName = "LongSubmerged10x Nuclear Battery Capacity Runtime";
        private const float NuclearBatteryCapacityFloor = 100000f;

        private static readonly FieldInfo OxygenBreathModifierField =
            AccessTools.Field(typeof(PlayerShip), "oxygenBreathModifier");

        private static readonly FieldInfo ResourcePlayerShipField =
            AccessTools.Field(typeof(Resource), "playerShip");

        private static readonly FieldInfo AirCompressorEnergyModifierField =
            AccessTools.Field(typeof(AirCompressor), "energyModifier");

        private static readonly FieldInfo GyrocompassEnergyGainModifierField =
            AccessTools.Field(typeof(Gyrocompass), "energyGainModifier");

        private static readonly FieldInfo TrimPumpEnergyGainModifierField =
            AccessTools.Field(typeof(TrimPump), "energyGainModifier");

        private static readonly FieldInfo VentilationEnergyModifierField =
            AccessTools.Field(typeof(Ventilation), "energyModifier");

        private static readonly FieldInfo ResourceGuiResourceField =
            AccessTools.Field(typeof(ResourceGUI), "resource");

        private static readonly FieldInfo DepletingResourceNotificationResourceField =
            AccessTools.Field(typeof(DepletingResourceNotification), "resource");

        private static readonly FieldInfo TorpedoHomingTargetField =
            AccessTools.Field(typeof(Torpedo), "homingTarget");

        private static readonly FieldInfo TorpedoRotatedField =
            AccessTools.Field(typeof(Torpedo), "rotated");

        private static readonly FieldInfo TorpedoSumOfAnglesField =
            AccessTools.Field(typeof(Torpedo), "sumOfAngles");

        private static readonly FieldInfo TorpedoHitEntityField =
            AccessTools.Field(typeof(Torpedo), "hitEntity");

        private static readonly FieldInfo TorpedoPassedDistanceField =
            AccessTools.Field(typeof(Torpedo), "passedDistance");

        private static readonly FieldInfo TorpedoArmDistanceField =
            AccessTools.Field(typeof(Torpedo), "armDistance");

        private static readonly MethodInfo TorpedoDoExplosionHitMethod =
            AccessTools.Method(typeof(Torpedo), "DoExplosionHit");

        private static readonly MethodInfo TorpedoDetonateMethod =
            AccessTools.Method(typeof(Torpedo), "Detonate", new Type[] { typeof(bool) });

        private static readonly ConditionalWeakTable<Parameter, ParameterScalePatchData> ParameterScaleData =
            new ConditionalWeakTable<Parameter, ParameterScalePatchData>();

        // DonJ : ConditionalWeakTable evite de garder en memoire des objets Unity detruits.
        // Chaque Parameter recoit un seul modifier DonJ, ensuite je change juste sa valeur.
        private static readonly ConditionalWeakTable<Parameter, ParameterDeltaPatchData> BatteryGainDeltaData =
            new ConditionalWeakTable<Parameter, ParameterDeltaPatchData>();

        private static readonly ConditionalWeakTable<Parameter, ParameterDeltaPatchData> NuclearBatteryCapacityDeltaData =
            new ConditionalWeakTable<Parameter, ParameterDeltaPatchData>();

        private static readonly ConditionalWeakTable<Torpedo, TorpedoGuidancePatchData> TorpedoGuidanceData =
            new ConditionalWeakTable<Torpedo, TorpedoGuidancePatchData>();

        private static readonly ConditionalWeakTable<Modifier, OxygenModifierPatchData> OxygenModifierData =
            new ConditionalWeakTable<Modifier, OxygenModifierPatchData>();

        private static readonly HashSet<int> InfiniteBatteryLoggedShipIds = new HashSet<int>();

        private static readonly HashSet<int> BatteryGainRuntimeLoggedResourceIds = new HashSet<int>();

        private static readonly HashSet<int> NuclearBatteryCapacityLoggedResourceIds = new HashSet<int>();

        private static readonly HashSet<int> BatteryTooltipRuntimeLoggedResourceIds = new HashSet<int>();

        // SurfaceSafe 1.4.7 :
        // Les callbacks de certains équipements peuvent être relancés pendant que l'on ajoute/modifie
        // leurs modifiers. Sans garde, un EnergyUsage_Changed déclenché par notre propre SetScale peut
        // entrer en récursion pendant la transition immersion -> surface.
        private static readonly HashSet<int> BatteryObjectApplicationGuardIds = new HashSet<int>();

        private static readonly Dictionary<Type, FieldInfo[]> GenericEnergyUsageFieldCache =
            new Dictionary<Type, FieldInfo[]>();

        private static readonly Dictionary<Type, FieldInfo[]> GenericParameterCollectionFieldCache =
            new Dictionary<Type, FieldInfo[]>();

        private static readonly Dictionary<Type, FieldInfo[]> GenericEnergyModifierFieldCache =
            new Dictionary<Type, FieldInfo[]>();

        public static void ApplyAll(string reason)
        {
            try
            {
                // DonJ : passe globale volontairement defensive. Elle resynchronise le menu,
                // le PlayerShip, les consommateurs batterie et toutes les torpilles visibles.
                LongSubmergedMenuController.Ensure();
                InteriorLightingColorPatcher.ApplyAll(reason + ".InteriorLighting");
                ApplyPlayerShip(UnityEngine.Object.FindObjectOfType<PlayerShip>(), reason);
                DeepDiveRuntimePatcher.ApplyAll(reason + ".DeepDive");
                ApplyBatteryConsumers(reason);
                MegaSonarRuntimePatcher.ApplyAll(reason + ".MegaSonar");

                foreach (StoredTorpedo item in UnityEngine.Object.FindObjectsOfType<StoredTorpedo>())
                    ApplyStoredTorpedo(item, reason + ".StoredTorpedo");

                foreach (Torpedo item in UnityEngine.Object.FindObjectsOfType<Torpedo>())
                    ApplyLaunchedTorpedo(item, reason + ".Torpedo");
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }
        }

        public static void MaintainBatteryRuntime(string reason)
        {
            try
            {
                // DonJ : tick leger appele toutes les 0.20s. Il ne rescane pas toute la scene,
                // il remet seulement la ressource batterie du sous-marin dans l'etat attendu.
                ApplyBatteryResource(UnityEngine.Object.FindObjectOfType<PlayerShip>(), reason);
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }
        }

        public static void ApplyBatteryConsumers(string reason)
        {
            try
            {
                // DonJ : UBOAT disperse les consommations electriques entre plusieurs composants.
                // Je traite les types connus puis je lance un scan generique pour les champs renommes ou caches.
                foreach (AccumulatorsUpgrade item in UnityEngine.Object.FindObjectsOfType<AccumulatorsUpgrade>())
                    ApplyBatteryObject(item, reason + ".AccumulatorsUpgrade");

                foreach (PlayerShipEngine item in UnityEngine.Object.FindObjectsOfType<PlayerShipEngine>())
                    ApplyBatteryObject(item, reason + ".PlayerShipEngine");

                foreach (DivingPlanesStation item in UnityEngine.Object.FindObjectsOfType<DivingPlanesStation>())
                    ApplyBatteryObject(item, reason + ".DivingPlanesStation");

                // SurfaceSafe 1.4.7 :
                // On ne touche plus AirCompressor ni Ventilation. Ces composants appartiennent au circuit
                // air/recharge de surface ; les modifier au moment où le bateau reprend l'air peut provoquer
                // une boucle EnergyUsage_Changed / modifier UI. La batterie infinie reste assurée par
                // ApplyBatteryResource(PlayerShip.Energy), donc il n'y a pas besoin de neutraliser leur coût.
                foreach (Gyrocompass item in UnityEngine.Object.FindObjectsOfType<Gyrocompass>())
                    ApplyBatteryObject(item, reason + ".Gyrocompass");

                foreach (TrimPump item in UnityEngine.Object.FindObjectsOfType<TrimPump>())
                    ApplyBatteryObject(item, reason + ".TrimPump");

                foreach (Equipment item in UnityEngine.Object.FindObjectsOfType<Equipment>())
                    ApplyBatteryEquipment(item, reason + ".Equipment");

                ApplyGenericBatteryConsumers(reason + ".Generic");
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }
        }

        public static void ApplyPlayerShip(PlayerShip ship, string reason)
        {
            LongSubmergedMenuController.Ensure();

            if (ship == null)
                return;

            OxygenBreathRecalculator.Recalculate(ship, reason);
            ApplyBatteryResource(ship, reason);
            EngineFastSpeedPatcher.PatchPlayerShip(ship, reason);
            DeepDiveRuntimePatcher.ApplyPlayerShip(ship, reason);
            SuperStealthRuntimePatcher.ApplyPlayerShip(ship, reason);
        }

        public static void ApplyBatteryResource(PlayerShip ship, string reason)
        {
            if (ship == null)
                return;

            ApplyBatteryRuntimeToResource(ship.Energy, reason);
        }

        public static void MaintainInfiniteBatteryCharge(PlayerShip ship, string reason)
        {
            if (ship == null)
                return;

            ApplyBatteryRuntimeToResource(ship.Energy, reason);
        }

        private static void ApplyBatteryRuntimeToResource(Resource energy, string reason)
        {
            if (energy == null)
                return;

            // DonJ : pipeline batterie unique. Capacite nucleaire, gain/drain et remplissage passent ici,
            // ce qui evite d'avoir plusieurs comportements batterie qui divergent.
            ApplyNuclearBatteryCapacityOverride(energy, reason);
            ApplyBatteryGainModifiers(energy, reason);

            if (IsInfiniteBatteryRuntimeActive())
                FillBatteryToCapacity(energy, reason);
            else
                ClampBatteryAmountToCapacity(energy);
        }

        private static void ApplyNuclearBatteryCapacityOverride(Resource energy, string reason)
        {
            if (energy == null || energy.Capacity == null)
                return;

            // DonJ: Mega Batterie does not merely reduce consumption; it adds a huge
            // capacity so the UI and gameplay see a nuclear battery immediately.
            float baseCapacity = energy.Capacity.GetValueExcludingModifier(RuntimeNuclearBatteryCapacityModifierName);
            float targetCapacity = baseCapacity;

            if (IsInfiniteBatteryRuntimeActive())
                targetCapacity = Math.Max(baseCapacity, NuclearBatteryCapacityFloor);

            float delta = targetCapacity - baseCapacity;
            SetDelta(
                energy.Capacity,
                NuclearBatteryCapacityDeltaData,
                RuntimeNuclearBatteryCapacityModifierName,
                delta
            );

            if (IsInfiniteBatteryRuntimeActive())
            {
                int resourceId = RuntimeHelpers.GetHashCode(energy);
                if (NuclearBatteryCapacityLoggedResourceIds.Add(resourceId))
                    Debug.Log("[LongSubmerged10x] Mega Batterie nuclear capacity active after " + reason + ".");
            }
        }

        private static void ClampBatteryAmountToCapacity(Resource energy)
        {
            if (energy == null)
                return;

            double capacity = GetResourceCapacity(energy);
            if (!IsUsableResourceValue(capacity) || capacity <= 0.0)
                return;

            if (energy.Amount > capacity)
                energy.Amount = capacity;
            else if (energy.Amount < 0.0)
                energy.Amount = 0.0;
        }

        public static bool TryMaintainBatteryResource(Resource resource, string reason)
        {
            try
            {
                if (!IsPlayerShipEnergyResource(resource))
                    return false;

                ApplyBatteryRuntimeToResource(resource, reason);
                return true;
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
                return false;
            }
        }

        public static bool TryFreezeInfiniteBatteryResource(Resource resource, string reason)
        {
            try
            {
                if (!IsInfiniteBatteryRuntimeActive())
                    return false;

                if (!IsPlayerShipEnergyResource(resource))
                    return false;

                double capacity = GetResourceCapacity(resource);
                if (!IsUsableResourceValue(capacity) || capacity <= 0.0)
                    return false;

                ApplyBatteryRuntimeToResource(resource, reason);
                return true;
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
                return false;
            }
        }

        public static Resource GetResourceFromGui(ResourceGUI gui)
        {
            try
            {
                return gui != null && ResourceGuiResourceField != null
                    ? ResourceGuiResourceField.GetValue(gui) as Resource
                    : null;
            }
            catch
            {
                return null;
            }
        }

        public static Resource GetResourceFromDepletingNotification(DepletingResourceNotification notification)
        {
            try
            {
                return notification != null && DepletingResourceNotificationResourceField != null
                    ? DepletingResourceNotificationResourceField.GetValue(notification) as Resource
                    : null;
            }
            catch
            {
                return null;
            }
        }

        public static bool ShouldSuppressBatteryDepletionUi(Resource resource, string reason)
        {
            if (!IsInfiniteBatteryRuntimeActive())
                return false;

            if (!TryMaintainBatteryResource(resource, reason))
                return false;

            int resourceId = RuntimeHelpers.GetHashCode(resource);
            if (BatteryTooltipRuntimeLoggedResourceIds.Add(resourceId))
                Debug.Log("[LongSubmerged10x] Mega Batterie depletion UI guard active after " + reason + ".");

            return true;
        }

        public static string BuildInfiniteBatteryTooltip(Resource resource)
        {
            // DonJ : ne pas appeler resource.PrintInfo ici.
            // PrintInfo repasse par les calculs vanilla de duree/recharge et peut toucher des etats UI
            // sensibles pendant la transition immersion -> surface. Le tooltip batterie infinie reste donc statique.
            if (resource == null)
                return "Mega Batterie : batterie infinie active.";

            StringBuilder builder = new StringBuilder();
            builder.AppendLine("Capacite de la batterie 100 %");
            builder.AppendLine("Mega Batterie : batterie nucleaire active.");
            builder.AppendLine("La batterie est maintenue au maximum par Long Submerged 10x+.");
            builder.AppendLine("Decoche Mega Batterie dans F10 pour revenir a la batterie vanilla.");
            return builder.ToString();
        }

        private static void FillBatteryToCapacity(Resource energy, string reason)
        {
            if (energy == null)
                return;

            double capacity = GetResourceCapacity(energy);
            if (!IsUsableResourceValue(capacity) || capacity <= 0.0)
                return;

            if (Math.Abs(energy.Amount - capacity) > 0.0001)
            {
                // DonJ : je garde la batterie au maximum avec le setter Amount pour forcer aussi le refresh UI.
                energy.Amount = capacity;

                int resourceId = RuntimeHelpers.GetHashCode(energy);
                if (InfiniteBatteryLoggedShipIds.Add(resourceId))
                    Debug.Log("[LongSubmerged10x] Mega Batterie infinite hold active after " + reason + ".");
            }
        }

        private static void ApplyBatteryGainModifiers(Resource energy, string reason)
        {
            if (energy == null)
                return;

            float factor = GetEffectiveBatteryGainFactor();
            ApplyBatteryGainParameter(energy.Gain, factor);
            ApplyBatteryGainParameter(energy.GainSandboxTimeScale, factor);

            if (factor >= LongSubmergedRuntimeSettings.BatteryMaxFactor - 0.0001f)
            {
                int resourceId = RuntimeHelpers.GetHashCode(energy);
                if (BatteryGainRuntimeLoggedResourceIds.Add(resourceId))
                    Debug.Log("[LongSubmerged10x] Mega Batterie infinite gain guard active after " + reason + ".");
            }
        }

        private static float GetEffectiveBatteryGainFactor()
        {
            if (!LongSubmergedRuntimeSettings.MegaBattery)
                return LongSubmergedRuntimeSettings.MinRuntimeFactor;

            return LongSubmergedRuntimeSettings.BatteryMaxFactor;
        }

        private static void ApplyBatteryGainParameter(Parameter parameter, float factor)
        {
            if (parameter == null)
                return;

            float baseValue = parameter.GetValueExcludingModifier(RuntimeBatteryGainModifierName);
            float desiredValue = baseValue;

            // DonJ: Mega Batterie is the single switch for infinity, so every negative
            // battery gain is neutralized without depending on the legacy slider value.
            if (factor >= LongSubmergedRuntimeSettings.BatteryMaxFactor - 0.0001f && baseValue < 0f)
                desiredValue = 0f;

            SetDelta(
                parameter,
                BatteryGainDeltaData,
                RuntimeBatteryGainModifierName,
                desiredValue - baseValue
            );
        }

        private static void SetDelta(
            Parameter parameter,
            ConditionalWeakTable<Parameter, ParameterDeltaPatchData> table,
            string modifierName,
            float delta
        )
        {
            if (parameter == null || table == null)
                return;

            ParameterDeltaPatchData data;
            if (!table.TryGetValue(parameter, out data))
            {
                data = new ParameterDeltaPatchData(parameter.AddDeltaModifier(modifierName, false));
                table.Add(parameter, data);
            }

            if (data.DeltaModifier == null)
                return;

            if (Math.Abs(data.DeltaModifier.Value - delta) > 0.000001f)
                data.DeltaModifier.Value = delta;
        }

        private static bool IsPlayerShipEnergyResource(Resource resource)
        {
            if (resource == null)
                return false;

            PlayerShip owner = null;
            if (ResourcePlayerShipField != null)
                owner = ResourcePlayerShipField.GetValue(resource) as PlayerShip;

            if (owner == null)
                owner = UnityEngine.Object.FindObjectOfType<PlayerShip>();

            // DonJ : securite anti-faux-positif. Si je trouve le PlayerShip, je n'accepte que sa vraie ressource Energy.
            // Le fallback par nom sert seulement quand UBOAT ne donne pas encore le lien owner.
            if (owner != null)
                return object.ReferenceEquals(owner.Energy, resource);

            return IsEnergyResourceName(resource.Name);
        }

        private static bool IsEnergyResourceName(string name)
        {
            return !string.IsNullOrEmpty(name)
                && (name.Equals("Energy", StringComparison.OrdinalIgnoreCase)
                    || name.IndexOf("Battery", StringComparison.OrdinalIgnoreCase) >= 0
                    || name.IndexOf("Batterie", StringComparison.OrdinalIgnoreCase) >= 0);
        }

        private static double GetResourceCapacity(Resource resource)
        {
            if (resource == null || resource.Capacity == null)
                return double.NaN;

            return resource.Capacity.Value;
        }

        private static bool IsUsableResourceValue(double value)
        {
            return !double.IsNaN(value) && !double.IsInfinity(value);
        }

        public static void ApplyOxygenBreathModifier(PlayerShip ship, string reason)
        {
            if (ship == null || OxygenBreathModifierField == null)
                return;

            try
            {
                Modifier oxygenModifier = OxygenBreathModifierField.GetValue(ship) as Modifier;
                if (oxygenModifier == null)
                    return;

                float currentValue = oxygenModifier.Value;
                if (!IsFinite(currentValue))
                    return;

                OxygenModifierPatchData data;
                if (!OxygenModifierData.TryGetValue(oxygenModifier, out data))
                {
                    data = new OxygenModifierPatchData(currentValue);
                    OxygenModifierData.Add(oxygenModifier, data);
                }

                float factor = GetEffectiveOxygenRuntimeFactor();

                // Surface and recharge states use zero or positive values; keep them vanilla.
                if (!LongSubmergedRuntimeSettings.MegaOxygen || factor <= 1.0001f || currentValue >= 0f)
                {
                    if (data.LastAppliedFactor > 1.0001f
                        && Math.Abs(currentValue - data.LastPatchedValue) <= 0.000001f)
                    {
                        oxygenModifier.Value = data.OriginalValue;
                    }
                    else
                    {
                        data.OriginalValue = currentValue;
                    }

                    data.LastAppliedFactor = 1f;
                    data.LastPatchedValue = oxygenModifier.Value;
                    return;
                }

                if (data.LastAppliedFactor <= 1.0001f
                    || Math.Abs(currentValue - data.LastPatchedValue) > 0.000001f)
                {
                    data.OriginalValue = currentValue;
                }

                float desiredValue = data.OriginalValue / factor;
                if (Math.Abs(oxygenModifier.Value - desiredValue) > 0.000000000001f)
                    oxygenModifier.Value = desiredValue;

                data.LastAppliedFactor = factor;
                data.LastPatchedValue = desiredValue;
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }
        }

        public static void ApplyBatteryObject(object target, string reason)
        {
            if (target == null || IsSurfaceAirRuntimeObject(target))
                return;

            if (!TryEnterBatteryObjectApplication(target))
                return;

            try
            {
                Equipment equipment = target as Equipment;
                if (equipment != null)
                    ApplyBatteryEquipment(equipment, reason);

                ApplyBatteryCapacityParameter(GetParameterField(target, "energyCapacityGain"));
                Parameter energyUsage = GetParameterField(target, "energyUsage");
                ApplyEnergyUsageParameter(energyUsage);
                ApplyDirectEnergyGainModifier(target, energyUsage, reason);
            }
            finally
            {
                ExitBatteryObjectApplication(target);
            }
        }

        public static void ApplyBatteryEquipment(Equipment equipment, string reason)
        {
            if (equipment == null || equipment.Parameters == null || IsSurfaceAirRuntimeObject(equipment))
                return;

            ApplyBatteryCapacityParameter(GetParameter(equipment.Parameters, "EnergyCapacityGain"));
            ApplyEnergyUsageParameter(GetParameter(equipment.Parameters, "EnergyUsage"));
        }

        public static void ApplyStoredTorpedo(StoredTorpedo storedTorpedo, string reason)
        {
            if (storedTorpedo == null)
                return;

            float reliabilityScale = IsMegaTorpedoRuntimeActive() && PerfectTorpedoReliability ? 0f : 1f;
            SetScale(storedTorpedo.DudChance, reliabilityScale);
        }

        public static void ApplyLaunchedTorpedo(Torpedo torpedo, string reason)
        {
            if (torpedo == null)
                return;

            if (torpedo.Parameters != null)
            {
                // DonJ : les torpilles sont reglees au runtime. A 1 elles redeviennent vanilla ;
                // a 10 elles utilisent le profil mega par defaut ; a 100 elles deviennent extremes.
                float torpedoFactor = GetEffectiveTorpedoFactor();
                float damageScale = torpedoFactor;
                float crewDamageScale = torpedoFactor;
                // DonJ : les degats peuvent rester x10, mais les effets visuels/particules sont bornes.
                // Des rayons/intensites x10 creent trop de surfaces feu/fumee et peuvent declencher
                // un crash natif Unity/particules pendant les phases surface + alarme.
                float radiusScale = Mathf.Min(torpedoFactor, 3f);
                float intensityScale = Mathf.Min(torpedoFactor, 3f);
                float reliabilityScale = IsMegaTorpedoRuntimeActive() && PerfectTorpedoReliability ? 0f : 1f;

                SetScale(GetParameter(torpedo.Parameters, "Damage"), damageScale);
                SetScale(GetParameter(torpedo.Parameters, "CrewDamage"), crewDamageScale);
                SetScale(GetParameter(torpedo.Parameters, "DamageRadius"), radiusScale);
                SetScale(GetParameter(torpedo.Parameters, "DamageEffectsRadius"), radiusScale);
                SetScale(GetParameter(torpedo.Parameters, "DamageEffectsIntensity"), intensityScale);
                SetScale(GetParameter(torpedo.Parameters, "MagneticExplosionOnArm"), reliabilityScale);
                SetScale(GetParameter(torpedo.Parameters, "MagneticExplosionAfterArm"), reliabilityScale);
                SetScale(GetParameter(torpedo.Parameters, "MagneticExplosionFail"), reliabilityScale);
            }

            // DonJ : stabilite surface/alarme.
            // Je garde les degats/fiabilite des Mega Torpilles, mais je desactive le guidage runtime
            // et la detonation forcee. Ces deux actions s'executaient en FixedUpdate avec des valeurs NaN
            // et pouvaient laisser une torpille/detonation dans un etat fragile pendant l'alarme.
            RestoreLockedTargetGuidance(torpedo);
        }

        private static void ApplyLockedTargetGuidance(Torpedo torpedo, string reason)
        {
            Entity target = torpedo.TargetEntity;
            if (target == null)
            {
                RestoreLockedTargetGuidance(torpedo);
                return;
            }

            TorpedoGuidancePatchData data = GetTorpedoGuidanceData(torpedo);
            if (!data.HasOriginalValues)
            {
                data.OriginalGyroAngle = torpedo.GyroAngle;
                data.OriginalTargetPosition = torpedo.TargetPosition;
                data.OriginalTargetPositionForReports = torpedo.TargetPositionForReports;
                data.HasOriginalValues = true;
            }

            Vector3 targetPoint = PredictLockedTargetPoint(torpedo, target);
            if (!IsFinite(targetPoint))
                return;

            // DonJ : je transforme le tir verrouille en visee cartesienne dynamique.
            // L'objectif est qu'une torpille tiree sur une cible correctement verrouillee corrige son angle pendant le vol.
            torpedo.GyroAngle = float.NaN;
            torpedo.TargetPosition = targetPoint;
            torpedo.TargetPositionForReports = targetPoint;
            data.GuidanceApplied = true;

            ResetCartesianTurnLimiter(torpedo);
            ApplyHomingPropeller(torpedo, target);
            TryForceLockedTargetDetonation(torpedo, target);

            if (!data.GuidanceLogged)
            {
                Debug.Log("[LongSubmerged10x] Mega torpedo locked-target guidance active after " + reason + ".");
                data.GuidanceLogged = true;
            }
        }

        private static void RestoreLockedTargetGuidance(Torpedo torpedo)
        {
            TorpedoGuidancePatchData data;
            if (!TorpedoGuidanceData.TryGetValue(torpedo, out data) || !data.HasOriginalValues || !data.GuidanceApplied)
                return;

            try
            {
                torpedo.GyroAngle = data.OriginalGyroAngle;
                torpedo.TargetPosition = data.OriginalTargetPosition;
                torpedo.TargetPositionForReports = data.OriginalTargetPositionForReports;

                if (TorpedoHomingTargetField != null)
                    TorpedoHomingTargetField.SetValue(torpedo, null);
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }

            data.GuidanceApplied = false;
        }

        private static Vector3 PredictLockedTargetPoint(Torpedo torpedo, Entity target)
        {
            Vector3 targetPoint = target.transform.position;
            Ship targetShip = target as Ship;
            if (targetShip != null && targetShip.RigidBody != null)
                targetPoint += targetShip.RigidBody.velocity * TorpedoGuidanceLeadSeconds;

            Vector3 torpedoPosition = torpedo.transform.position;
            targetPoint.y = torpedoPosition.y;
            return targetPoint;
        }

        private static void ResetCartesianTurnLimiter(Torpedo torpedo)
        {
            try
            {
                if (TorpedoRotatedField != null)
                    TorpedoRotatedField.SetValue(torpedo, false);

                if (TorpedoSumOfAnglesField != null)
                    TorpedoSumOfAnglesField.SetValue(torpedo, 0f);
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }
        }

        private static void ApplyHomingPropeller(Torpedo torpedo, Entity target)
        {
            if (TorpedoHomingTargetField == null)
                return;

            Ship targetShip = target as Ship;
            if (targetShip == null)
                return;

            Propeller[] propellers = targetShip.Propellers;
            if (propellers == null || propellers.Length == 0)
                return;

            for (int i = 0; i < propellers.Length; i++)
            {
                if (propellers[i] == null)
                    continue;

                TorpedoHomingTargetField.SetValue(torpedo, propellers[i]);
                return;
            }
        }

        private static void TryForceLockedTargetDetonation(Torpedo torpedo, Entity target)
        {
            if (target == null || torpedo.Detonated || TorpedoDoExplosionHitMethod == null || TorpedoDetonateMethod == null)
                return;

            TorpedoGuidancePatchData data = GetTorpedoGuidanceData(torpedo);
            if (data.ForcingDetonation)
                return;

            if (!IsTorpedoArmedForAssist(torpedo))
                return;

            Vector3 torpedoPosition = torpedo.transform.position;
            Vector3 targetPosition = target.transform.position;
            Vector2 delta = new Vector2(torpedoPosition.x - targetPosition.x, torpedoPosition.z - targetPosition.z);
            float detonationDistance = GetAssistDetonationDistance(torpedo);

            if (delta.sqrMagnitude > detonationDistance * detonationDistance)
                return;

            try
            {
                data.ForcingDetonation = true;

                if (TorpedoHitEntityField != null)
                    TorpedoHitEntityField.SetValue(torpedo, target);

                TorpedoDoExplosionHitMethod.Invoke(torpedo, new object[] { target });
                TorpedoDetonateMethod.Invoke(torpedo, new object[] { true });
                Debug.Log("[LongSubmerged10x] Mega torpedo forced locked-target detonation inside " + detonationDistance.ToString("0.0") + "m.");
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }
            finally
            {
                data.ForcingDetonation = false;
            }
        }

        private static bool IsTorpedoArmedForAssist(Torpedo torpedo)
        {
            if (TorpedoPassedDistanceField == null || TorpedoArmDistanceField == null)
                return true;

            try
            {
                float passedDistance = (float)TorpedoPassedDistanceField.GetValue(torpedo);
                Parameter armDistance = TorpedoArmDistanceField.GetValue(torpedo) as Parameter;
                return armDistance == null || passedDistance >= armDistance.Value;
            }
            catch
            {
                return true;
            }
        }

        private static float GetAssistDetonationDistance(Torpedo torpedo)
        {
            Parameter damageRadius = torpedo.Parameters == null ? null : GetParameter(torpedo.Parameters, "DamageRadius");
            float scaledDamageRadius = damageRadius == null ? 0f : damageRadius.Value * GetEffectiveTorpedoFactor();
            // DonJ : detonateur de secours proche cible. Il reste borne pour ne pas exploser trop loin,
            // mais suit le rayon mega afin de fiabiliser les impacts verrouilles.
            return Mathf.Clamp(
                scaledDamageRadius * TorpedoGuidanceDetonationRadiusRatio,
                TorpedoGuidanceMinimumDetonationDistance,
                TorpedoGuidanceMaximumDetonationDistance
            );
        }

        private static bool IsMegaTorpedoRuntimeActive()
        {
            return LongSubmergedRuntimeSettings.MegaTorpedoes && GetEffectiveTorpedoFactor() > 1.0001f;
        }

        private static float GetEffectiveTorpedoFactor()
        {
            return LongSubmergedRuntimeSettings.MegaTorpedoes
                ? LongSubmergedRuntimeSettings.ClampTorpedoFactor(LongSubmergedRuntimeSettings.TorpedoFactor)
                : 1f;
        }

        private static float GetEffectiveOxygenRuntimeFactor()
        {
            if (!LongSubmergedRuntimeSettings.MegaOxygen)
                return 1f;

            float sliderValue = LongSubmergedRuntimeSettings.ClampOxygenFactor(LongSubmergedRuntimeSettings.OxygenFactor);
            if (sliderValue <= LongSubmergedRuntimeSettings.MinRuntimeFactor)
                return 1f;

            float normalized = (sliderValue - LongSubmergedRuntimeSettings.MinRuntimeFactor)
                / (LongSubmergedRuntimeSettings.OxygenMaxFactor - LongSubmergedRuntimeSettings.MinRuntimeFactor);

            float maxFactor = Mathf.Max(1f, OxygenRuntimeMaxFactor);
            return 1f + normalized * (maxFactor - 1f);
        }

        private static float GetEffectiveBatteryCapacityScale()
        {
            if (!LongSubmergedRuntimeSettings.MegaBattery)
                return BatteryCapacityVanillaRestoreScale;

            float factor = LongSubmergedRuntimeSettings.ClampBatteryFactor(LongSubmergedRuntimeSettings.BatteryFactor);
            return factor / BatteryCapacityDataFactor;
        }

        private static float GetEffectiveBatteryEnergyUsageScale()
        {
            // DonJ: Mega Batterie is now fully infinite as soon as the toggle is on.
            // With the toggle off, restore the XLSX fallback x0.1 back to vanilla.
            if (IsInfiniteBatteryRuntimeActive())
                return 0f;

            return EnergyUsageVanillaRestoreScale;
        }

        private static bool IsInfiniteBatteryRuntimeActive()
        {
            return LongSubmergedRuntimeSettings.MegaBattery;
        }

        private static bool IsFinite(Vector3 value)
        {
            return IsFinite(value.x) && IsFinite(value.y) && IsFinite(value.z);
        }

        private static bool IsFinite(float value)
        {
            return !float.IsNaN(value) && !float.IsInfinity(value);
        }

        private static TorpedoGuidancePatchData GetTorpedoGuidanceData(Torpedo torpedo)
        {
            TorpedoGuidancePatchData data;
            if (!TorpedoGuidanceData.TryGetValue(torpedo, out data))
            {
                data = new TorpedoGuidancePatchData();
                TorpedoGuidanceData.Add(torpedo, data);
            }

            return data;
        }

        private static void ApplyBatteryCapacityParameter(Parameter parameter)
        {
            if (parameter == null)
                return;

            SetScale(
                parameter,
                GetEffectiveBatteryCapacityScale()
            );
        }

        private static void ApplyEnergyUsageParameter(Parameter parameter)
        {
            if (parameter == null)
                return;

            // DonJ : ne pas tester parameter.Value ici : en mode infini mon scale vaut 0.
            // Quand le joueur redescend le slider, je dois pouvoir restaurer le drain vanilla.
            float baseValue = parameter.GetValueExcludingModifier(RuntimeScaleModifierName);
            if (baseValue <= 0f)
                return;

            SetScale(
                parameter,
                GetEffectiveBatteryEnergyUsageScale()
            );
        }

        private static void ApplyDirectEnergyGainModifier(object target, Parameter energyUsage, string reason)
        {
            if (target == null || energyUsage == null || IsSurfaceAirRuntimeObject(target))
                return;

            if (target is AirCompressor)
            {
                ApplyDirectEnergyGainModifierField(target, AirCompressorEnergyModifierField, energyUsage);
                return;
            }

            if (target is Gyrocompass)
            {
                ApplyDirectEnergyGainModifierField(target, GyrocompassEnergyGainModifierField, energyUsage);
                return;
            }

            if (target is TrimPump)
            {
                ApplyDirectEnergyGainModifierField(target, TrimPumpEnergyGainModifierField, energyUsage);
                return;
            }

            if (target is Ventilation)
                ApplyDirectEnergyGainModifierField(target, VentilationEnergyModifierField, energyUsage);
        }

        private static void ApplyDirectEnergyGainModifierField(object target, FieldInfo modifierField, Parameter energyUsage)
        {
            if (modifierField == null || energyUsage == null)
                return;

            Modifier modifier = modifierField.GetValue(target) as Modifier;
            if (modifier == null)
                return;

            float usage = energyUsage.Value;
            if (usage < 0f)
                return;

            float desiredGain = -usage;
            if (Math.Abs(modifier.Value - desiredGain) > 0.0001f)
                modifier.Value = desiredGain;
        }

        private static void ApplyGenericBatteryConsumers(string reason)
        {
            // DonJ : filet de securite. Si UBOAT renomme un composant electrique,
            // je cherche quand meme les champs Parameter nommes EnergyUsage dans tous les MonoBehaviour.
            MonoBehaviour[] behaviours = UnityEngine.Object.FindObjectsOfType<MonoBehaviour>();
            foreach (MonoBehaviour behaviour in behaviours)
            {
                if (behaviour == null || behaviour is LongSubmergedMenuController)
                    continue;

                ApplyGenericBatteryConsumer(behaviour, reason);
            }
        }

        private static void ApplyGenericBatteryConsumer(object target, string reason)
        {
            if (target == null || IsSurfaceAirRuntimeObject(target))
                return;

            if (!TryEnterBatteryObjectApplication(target))
                return;

            try
            {
                Type type = target.GetType();

                foreach (FieldInfo field in GetGenericEnergyUsageFields(type))
                {
                    Parameter energyUsage = GetParameterFromField(target, field);
                    if (energyUsage == null)
                        continue;

                    ApplyEnergyUsageParameter(energyUsage);
                    ApplyDirectEnergyGainModifier(target, energyUsage, reason);
                    ApplyGenericEnergyModifierFields(target, energyUsage);
                }

                foreach (FieldInfo field in GetGenericParameterCollectionFields(type))
                {
                    ParameterCollection parameters = GetParameterCollectionFromField(target, field);
                    if (parameters == null)
                        continue;

                    ApplyBatteryCapacityParameter(GetParameter(parameters, "EnergyCapacityGain"));
                    Parameter energyUsage = GetParameter(parameters, "EnergyUsage");
                    ApplyEnergyUsageParameter(energyUsage);
                    ApplyDirectEnergyGainModifier(target, energyUsage, reason);
                    ApplyGenericEnergyModifierFields(target, energyUsage);
                }
            }
            finally
            {
                ExitBatteryObjectApplication(target);
            }
        }

        private static FieldInfo[] GetGenericEnergyUsageFields(Type type)
        {
            FieldInfo[] cached;
            if (GenericEnergyUsageFieldCache.TryGetValue(type, out cached))
                return cached;

            List<FieldInfo> fields = new List<FieldInfo>();
            CollectFields(type, fields, typeof(Parameter), true);
            cached = fields.ToArray();
            GenericEnergyUsageFieldCache[type] = cached;
            return cached;
        }

        private static FieldInfo[] GetGenericParameterCollectionFields(Type type)
        {
            FieldInfo[] cached;
            if (GenericParameterCollectionFieldCache.TryGetValue(type, out cached))
                return cached;

            List<FieldInfo> fields = new List<FieldInfo>();
            CollectFields(type, fields, typeof(ParameterCollection), false);
            cached = fields.ToArray();
            GenericParameterCollectionFieldCache[type] = cached;
            return cached;
        }

        private static FieldInfo[] GetGenericEnergyModifierFields(Type type)
        {
            FieldInfo[] cached;
            if (GenericEnergyModifierFieldCache.TryGetValue(type, out cached))
                return cached;

            List<FieldInfo> fields = new List<FieldInfo>();
            CollectFields(type, fields, typeof(Modifier), false);
            cached = fields.ToArray();
            GenericEnergyModifierFieldCache[type] = cached;
            return cached;
        }

        private static void CollectFields(Type type, List<FieldInfo> fields, Type requiredFieldType, bool energyUsageNameOnly)
        {
            for (Type current = type; current != null && current != typeof(object); current = current.BaseType)
            {
                FieldInfo[] declaredFields = current.GetFields(BindingFlags.Instance | BindingFlags.Public | BindingFlags.NonPublic | BindingFlags.DeclaredOnly);
                foreach (FieldInfo field in declaredFields)
                {
                    if (field == null || !requiredFieldType.IsAssignableFrom(field.FieldType))
                        continue;

                    if (energyUsageNameOnly && !IsEnergyUsageMemberName(field.Name))
                        continue;

                    if (!energyUsageNameOnly && requiredFieldType == typeof(Modifier) && !IsEnergyModifierMemberName(field.Name))
                        continue;

                    fields.Add(field);
                }
            }
        }

        private static bool IsEnergyUsageMemberName(string name)
        {
            return !string.IsNullOrEmpty(name)
                && name.IndexOf("EnergyUsage", StringComparison.OrdinalIgnoreCase) >= 0;
        }

        private static bool IsEnergyModifierMemberName(string name)
        {
            return !string.IsNullOrEmpty(name)
                && name.IndexOf("Energy", StringComparison.OrdinalIgnoreCase) >= 0
                && name.IndexOf("Modifier", StringComparison.OrdinalIgnoreCase) >= 0;
        }

        private static bool IsSurfaceAirRuntimeObject(object target)
        {
            if (target == null)
                return false;

            // AirCompressor et Ventilation sont volontairement laissés vanilla.
            // Ils s'activent autour du retour en surface et peuvent recalculer EnergyUsage en cascade.
            if (target is AirCompressor || target is Ventilation)
                return true;

            Type type = target.GetType();
            if (type != null && IsSurfaceAirName(type.Name))
                return true;

            UnityEngine.Object unityObject = target as UnityEngine.Object;
            if (unityObject != null && IsSurfaceAirName(unityObject.name))
                return true;

            return false;
        }

        private static bool IsSurfaceAirName(string name)
        {
            if (string.IsNullOrEmpty(name))
                return false;

            return name.IndexOf("AirCompressor", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Ventilation", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Atmosphere", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Oxygen", StringComparison.OrdinalIgnoreCase) >= 0;
        }

        private static bool TryEnterBatteryObjectApplication(object target)
        {
            if (target == null)
                return false;

            return BatteryObjectApplicationGuardIds.Add(RuntimeHelpers.GetHashCode(target));
        }

        private static void ExitBatteryObjectApplication(object target)
        {
            if (target == null)
                return;

            BatteryObjectApplicationGuardIds.Remove(RuntimeHelpers.GetHashCode(target));
        }

        private static Parameter GetParameterFromField(object target, FieldInfo field)
        {
            try
            {
                return field == null ? null : field.GetValue(target) as Parameter;
            }
            catch
            {
                return null;
            }
        }

        private static ParameterCollection GetParameterCollectionFromField(object target, FieldInfo field)
        {
            try
            {
                return field == null ? null : field.GetValue(target) as ParameterCollection;
            }
            catch
            {
                return null;
            }
        }

        private static void ApplyGenericEnergyModifierFields(object target, Parameter energyUsage)
        {
            if (target == null || energyUsage == null || IsSurfaceAirRuntimeObject(target))
                return;

            float usage = energyUsage.Value;
            if (usage < 0f)
                return;

            float desiredGain = -usage;
            foreach (FieldInfo field in GetGenericEnergyModifierFields(target.GetType()))
            {
                try
                {
                    Modifier modifier = field.GetValue(target) as Modifier;
                    if (modifier != null && Math.Abs(modifier.Value - desiredGain) > 0.0001f)
                        modifier.Value = desiredGain;
                }
                catch
                {
                }
            }
        }

        private static Parameter GetParameter(ParameterCollection parameters, string key)
        {
            try
            {
                return parameters.GetParameter(key);
            }
            catch
            {
                return null;
            }
        }

        private static Parameter GetParameterField(object target, string fieldName)
        {
            try
            {
                FieldInfo field = AccessTools.Field(target.GetType(), fieldName);
                return field == null ? null : field.GetValue(target) as Parameter;
            }
            catch
            {
                return null;
            }
        }

        private static void SetScale(Parameter parameter, float scale)
        {
            if (parameter == null)
                return;

            ParameterScalePatchData data;
            if (!ParameterScaleData.TryGetValue(parameter, out data))
            {
                data = new ParameterScalePatchData(parameter.AddScaleModifier(RuntimeScaleModifierName, false));
                ParameterScaleData.Add(parameter, data);
            }

            if (data.ScaleModifier == null)
                return;

            if (Math.Abs(data.ScaleModifier.Value - scale) > 0.0001f)
                data.ScaleModifier.Value = scale;
        }
    }

    internal static class MegaSonarRuntimePatcher
    {
        private const string MegaSonarScaleModifierName = "LongSubmerged10x Mega Sonar Runtime";

        private static readonly string[] HydrophoneParameterKeys = new string[]
        {
            "HydrophoneRange",
            "GroupHydrophoneRange",
            "DirectHydrophoneRange",
            "HydrophoneDirectRange",
            "HydrophoneDetectionRange",
            "NoiseHydrophoneRange",
            "HydrophoneNoiseRange",
            "ListeningRange",
            "PassiveSonarRange"
        };

        private static readonly string[] DirectRefreshMethodNames = new string[]
        {
            "Awake",
            "Start",
            "OnEnable",
            "OnAfterDeserialize",
            "SavesManagerOnLoaded",
            "Update",
            "FixedUpdate",
            "UpdateModifiers",
            "ApplyModifiers",
            "Refresh",
            "Recalculate",
            "Validate"
        };

        private static readonly ConditionalWeakTable<Parameter, ParameterScalePatchData> SonarScaleData =
            new ConditionalWeakTable<Parameter, ParameterScalePatchData>();

        private static readonly ConditionalWeakTable<object, MegaSonarObjectPatchData> ObjectPatchData =
            new ConditionalWeakTable<object, MegaSonarObjectPatchData>();

        private static readonly Dictionary<Type, FieldInfo[]> ParameterFieldCache =
            new Dictionary<Type, FieldInfo[]>();

        private static readonly Dictionary<Type, FieldInfo[]> ParameterCollectionFieldCache =
            new Dictionary<Type, FieldInfo[]>();

        private static readonly Dictionary<Type, FieldInfo[]> FloatFieldCache =
            new Dictionary<Type, FieldInfo[]>();

        private static readonly Dictionary<Type, PropertyInfo[]> FloatPropertyCache =
            new Dictionary<Type, PropertyInfo[]>();

        private static readonly HashSet<int> ApplicationGuardIds = new HashSet<int>();
        private static readonly HashSet<string> TargetMethodLogIds = new HashSet<string>();

        public static void ApplyAll(string reason)
        {
            try
            {
                Equipment[] equipmentItems = UnityEngine.Object.FindObjectsOfType<Equipment>();
                foreach (Equipment equipment in equipmentItems)
                    ApplyEquipment(equipment, reason + ".Equipment");

                MonoBehaviour[] behaviours = UnityEngine.Object.FindObjectsOfType<MonoBehaviour>();
                foreach (MonoBehaviour behaviour in behaviours)
                {
                    if (behaviour == null || behaviour is LongSubmergedMenuController || behaviour is Equipment)
                        continue;

                    if (!IsPotentialHydrophoneObject(behaviour))
                        continue;

                    ApplyObject(behaviour, reason + ".HydrophoneObject");
                }
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }
        }

        public static void ApplyObject(object target, string reason)
        {
            if (target == null || target is LongSubmergedMenuController)
                return;

            if (!TryEnter(target))
                return;

            try
            {
                Equipment equipment = target as Equipment;
                if (equipment != null)
                {
                    ApplyEquipment(equipment, reason);
                    return;
                }

                bool ownerLooksHydrophone = IsPotentialHydrophoneObject(target);
                if (!ownerLooksHydrophone)
                    return;

                Type type = target.GetType();
                ApplyParameterFields(target, type, ownerLooksHydrophone);
                ApplyParameterCollections(target, type);
                ApplyDirectFloatMembers(target, type, ownerLooksHydrophone);
            }
            catch (Exception ex)
            {
                Debug.LogWarning("[LongSubmerged10x] Mega Sonar skipped on " + SafeObjectName(target) + " -> " + ex.GetType().Name + ": " + ex.Message);
            }
            finally
            {
                Exit(target);
            }
        }

        public static IEnumerable<MethodBase> FindHydrophoneTargetMethods()
        {
            HashSet<string> emitted = new HashSet<string>();
            Assembly[] assemblies = AppDomain.CurrentDomain.GetAssemblies();

            foreach (Assembly assembly in assemblies)
            {
                Type[] types = GetTypesSafely(assembly);
                if (types == null)
                    continue;

                foreach (Type type in types)
                {
                    if (!IsPotentialHydrophoneType(type))
                        continue;

                    MethodInfo[] methods;
                    try
                    {
                        methods = type.GetMethods(BindingFlags.Instance | BindingFlags.Public | BindingFlags.NonPublic | BindingFlags.DeclaredOnly);
                    }
                    catch
                    {
                        continue;
                    }

                    foreach (MethodInfo method in methods)
                    {
                        if (method == null || method.IsAbstract || method.ContainsGenericParameters)
                            continue;

                        if (!IsHydrophoneRefreshMethod(method.Name))
                            continue;

                        string id = type.FullName + "::" + method.Name + "#" + method.GetParameters().Length;
                        if (!emitted.Add(id))
                            continue;

                        if (TargetMethodLogIds.Add(id))
                            Debug.Log("[LongSubmerged10x] Mega Sonar will refresh after " + id + ".");

                        yield return method;
                    }
                }
            }
        }

        private static void ApplyEquipment(Equipment equipment, string reason)
        {
            if (equipment == null || equipment.Parameters == null)
                return;

            bool hasHydrophoneRanges = ApplyParameterCollection(equipment.Parameters);
            bool equipmentLooksHydrophone = hasHydrophoneRanges || IsHydrophoneName(equipment.name);

            if (equipmentLooksHydrophone)
                ApplyDirectFloatMembers(equipment, equipment.GetType(), true);
        }

        private static void ApplyParameterFields(object target, Type type, bool ownerLooksHydrophone)
        {
            foreach (FieldInfo field in GetParameterFields(type))
            {
                if (!IsHydrophoneRangeMemberName(field.Name, ownerLooksHydrophone))
                    continue;

                try
                {
                    Parameter parameter = field.GetValue(target) as Parameter;
                    ApplyParameterScale(parameter);
                }
                catch
                {
                }
            }
        }

        private static void ApplyParameterCollections(object target, Type type)
        {
            foreach (FieldInfo field in GetParameterCollectionFields(type))
            {
                try
                {
                    ParameterCollection parameters = field.GetValue(target) as ParameterCollection;
                    ApplyParameterCollection(parameters);
                }
                catch
                {
                }
            }
        }

        private static bool ApplyParameterCollection(ParameterCollection parameters)
        {
            if (parameters == null)
                return false;

            bool foundAny = false;
            foreach (string key in HydrophoneParameterKeys)
            {
                Parameter parameter = GetParameter(parameters, key);
                if (parameter == null)
                    continue;

                ApplyParameterScale(parameter);
                foundAny = true;
            }

            return foundAny;
        }

        private static void ApplyParameterScale(Parameter parameter)
        {
            if (parameter == null)
                return;

            ParameterScalePatchData data;
            if (!SonarScaleData.TryGetValue(parameter, out data))
            {
                data = new ParameterScalePatchData(parameter.AddScaleModifier(MegaSonarScaleModifierName, false));
                SonarScaleData.Add(parameter, data);
            }

            if (data.ScaleModifier == null)
                return;

            float scale = GetEffectiveSonarFactor();
            if (Math.Abs(data.ScaleModifier.Value - scale) > 0.0001f)
                data.ScaleModifier.Value = scale;
        }

        private static void ApplyDirectFloatMembers(object target, Type type, bool ownerLooksHydrophone)
        {
            foreach (FieldInfo field in GetFloatFields(type))
            {
                if (!IsHydrophoneRangeMemberName(field.Name, ownerLooksHydrophone))
                    continue;

                ApplyFloatField(target, field);
            }

            foreach (PropertyInfo property in GetFloatProperties(type))
            {
                if (!IsHydrophoneRangeMemberName(property.Name, ownerLooksHydrophone))
                    continue;

                ApplyFloatProperty(target, property);
            }
        }

        private static void ApplyFloatField(object target, FieldInfo field)
        {
            try
            {
                object rawValue = field.GetValue(target);
                if (!(rawValue is float))
                    return;

                float currentValue = (float)rawValue;
                float desiredValue;
                if (!TryGetDesiredFloatValue(target, GetMemberKey(field), currentValue, out desiredValue))
                    return;

                if (Math.Abs(currentValue - desiredValue) > GetFloatTolerance(desiredValue))
                    field.SetValue(target, desiredValue);
            }
            catch
            {
            }
        }

        private static void ApplyFloatProperty(object target, PropertyInfo property)
        {
            try
            {
                object rawValue = property.GetValue(target, null);
                if (!(rawValue is float))
                    return;

                float currentValue = (float)rawValue;
                float desiredValue;
                if (!TryGetDesiredFloatValue(target, GetMemberKey(property), currentValue, out desiredValue))
                    return;

                if (Math.Abs(currentValue - desiredValue) > GetFloatTolerance(desiredValue))
                    property.SetValue(target, desiredValue, null);
            }
            catch
            {
            }
        }

        private static bool TryGetDesiredFloatValue(object target, string memberKey, float currentValue, out float desiredValue)
        {
            desiredValue = currentValue;

            if (!IsFinite(currentValue) || currentValue <= 0f)
                return false;

            float factor = GetEffectiveSonarFactor();
            MegaSonarObjectPatchData objectData = GetObjectPatchData(target);

            MegaSonarFloatMemberPatchData memberData;
            if (!objectData.FloatMembers.TryGetValue(memberKey, out memberData))
            {
                memberData = new MegaSonarFloatMemberPatchData(currentValue, 1f, currentValue);
                objectData.FloatMembers.Add(memberKey, memberData);
            }

            if (factor <= 1.0001f)
            {
                if (memberData.LastAppliedFactor > 1.0001f
                    && Math.Abs(currentValue - memberData.LastPatchedValue) <= GetFloatTolerance(memberData.LastPatchedValue))
                {
                    desiredValue = memberData.OriginalValue;
                }
                else
                {
                    memberData.OriginalValue = currentValue;
                    desiredValue = currentValue;
                }

                memberData.LastAppliedFactor = 1f;
                memberData.LastPatchedValue = desiredValue;
                return IsFinite(desiredValue) && desiredValue > 0f;
            }

            if (memberData.LastAppliedFactor <= 1.0001f
                || Math.Abs(currentValue - memberData.LastPatchedValue) > GetFloatTolerance(memberData.LastPatchedValue))
            {
                memberData.OriginalValue = currentValue;
            }

            desiredValue = memberData.OriginalValue * factor;
            memberData.LastAppliedFactor = factor;
            memberData.LastPatchedValue = desiredValue;
            return IsFinite(desiredValue) && desiredValue > 0f;
        }

        private static MegaSonarObjectPatchData GetObjectPatchData(object target)
        {
            MegaSonarObjectPatchData data;
            if (!ObjectPatchData.TryGetValue(target, out data))
            {
                data = new MegaSonarObjectPatchData();
                ObjectPatchData.Add(target, data);
            }

            return data;
        }

        private static float GetEffectiveSonarFactor()
        {
            if (!LongSubmergedRuntimeSettings.MegaSonar)
                return 1f;

            return LongSubmergedRuntimeSettings.ClampSonarFactor(LongSubmergedRuntimeSettings.SonarFactor);
        }

        private static FieldInfo[] GetParameterFields(Type type)
        {
            FieldInfo[] cached;
            if (ParameterFieldCache.TryGetValue(type, out cached))
                return cached;

            List<FieldInfo> fields = new List<FieldInfo>();
            CollectFields(type, fields, typeof(Parameter));
            cached = fields.ToArray();
            ParameterFieldCache[type] = cached;
            return cached;
        }

        private static FieldInfo[] GetParameterCollectionFields(Type type)
        {
            FieldInfo[] cached;
            if (ParameterCollectionFieldCache.TryGetValue(type, out cached))
                return cached;

            List<FieldInfo> fields = new List<FieldInfo>();
            CollectFields(type, fields, typeof(ParameterCollection));
            cached = fields.ToArray();
            ParameterCollectionFieldCache[type] = cached;
            return cached;
        }

        private static FieldInfo[] GetFloatFields(Type type)
        {
            FieldInfo[] cached;
            if (FloatFieldCache.TryGetValue(type, out cached))
                return cached;

            List<FieldInfo> fields = new List<FieldInfo>();

            for (Type current = type; current != null && current != typeof(object); current = current.BaseType)
            {
                FieldInfo[] declaredFields = current.GetFields(BindingFlags.Instance | BindingFlags.Public | BindingFlags.NonPublic | BindingFlags.DeclaredOnly);
                foreach (FieldInfo field in declaredFields)
                {
                    if (field == null || field.FieldType != typeof(float) || field.IsInitOnly || field.IsLiteral)
                        continue;

                    fields.Add(field);
                }
            }

            cached = fields.ToArray();
            FloatFieldCache[type] = cached;
            return cached;
        }

        private static PropertyInfo[] GetFloatProperties(Type type)
        {
            PropertyInfo[] cached;
            if (FloatPropertyCache.TryGetValue(type, out cached))
                return cached;

            List<PropertyInfo> properties = new List<PropertyInfo>();

            for (Type current = type; current != null && current != typeof(object); current = current.BaseType)
            {
                PropertyInfo[] declaredProperties = current.GetProperties(BindingFlags.Instance | BindingFlags.Public | BindingFlags.NonPublic | BindingFlags.DeclaredOnly);
                foreach (PropertyInfo property in declaredProperties)
                {
                    if (property == null || property.PropertyType != typeof(float))
                        continue;

                    if (property.GetIndexParameters().Length != 0)
                        continue;

                    if (property.GetGetMethod(true) == null || property.GetSetMethod(true) == null)
                        continue;

                    properties.Add(property);
                }
            }

            cached = properties.ToArray();
            FloatPropertyCache[type] = cached;
            return cached;
        }

        private static void CollectFields(Type type, List<FieldInfo> fields, Type requiredFieldType)
        {
            for (Type current = type; current != null && current != typeof(object); current = current.BaseType)
            {
                FieldInfo[] declaredFields = current.GetFields(BindingFlags.Instance | BindingFlags.Public | BindingFlags.NonPublic | BindingFlags.DeclaredOnly);
                foreach (FieldInfo field in declaredFields)
                {
                    if (field == null || !requiredFieldType.IsAssignableFrom(field.FieldType))
                        continue;

                    fields.Add(field);
                }
            }
        }

        private static bool IsPotentialHydrophoneObject(object target)
        {
            if (target == null)
                return false;

            Type type = target.GetType();
            if (IsPotentialHydrophoneType(type))
                return true;

            UnityEngine.Object unityObject = target as UnityEngine.Object;
            if (unityObject != null && IsHydrophoneName(unityObject.name))
                return true;

            Equipment equipment = target as Equipment;
            if (equipment != null && equipment.Parameters != null)
                return HasHydrophoneParameter(equipment.Parameters);

            return false;
        }

        private static bool IsPotentialHydrophoneType(Type type)
        {
            if (type == null || type.IsAbstract || type.ContainsGenericParameters)
                return false;

            if (type.Namespace != null && type.Namespace.IndexOf("LongSubmerged10x", StringComparison.OrdinalIgnoreCase) >= 0)
                return false;

            string name = type.FullName;
            if (string.IsNullOrEmpty(name))
                name = type.Name;

            return IsHydrophoneName(name);
        }

        private static bool IsHydrophoneName(string name)
        {
            if (string.IsNullOrEmpty(name))
                return false;

            return name.IndexOf("Hydrophone", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Horch", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Listening", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("PassiveSonar", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Gruppenhorch", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Balkon", StringComparison.OrdinalIgnoreCase) >= 0
                || string.Equals(name.Trim(), "GHG", StringComparison.OrdinalIgnoreCase)
                || string.Equals(name.Trim(), "KDB", StringComparison.OrdinalIgnoreCase);
        }

        private static bool HasHydrophoneParameter(ParameterCollection parameters)
        {
            if (parameters == null)
                return false;

            foreach (string key in HydrophoneParameterKeys)
            {
                if (GetParameter(parameters, key) != null)
                    return true;
            }

            return false;
        }

        private static bool IsHydrophoneRefreshMethod(string methodName)
        {
            if (string.IsNullOrEmpty(methodName))
                return false;

            if (methodName.StartsWith("get_", StringComparison.Ordinal) || methodName.StartsWith("set_", StringComparison.Ordinal))
                return false;

            if (methodName.StartsWith("add_", StringComparison.Ordinal) || methodName.StartsWith("remove_", StringComparison.Ordinal))
                return false;

            foreach (string directName in DirectRefreshMethodNames)
            {
                if (string.Equals(methodName, directName, StringComparison.Ordinal))
                    return true;
            }

            bool looksLikeRefresh =
                methodName.IndexOf("Update", StringComparison.OrdinalIgnoreCase) >= 0
                || methodName.IndexOf("Apply", StringComparison.OrdinalIgnoreCase) >= 0
                || methodName.IndexOf("Refresh", StringComparison.OrdinalIgnoreCase) >= 0
                || methodName.IndexOf("Recalculate", StringComparison.OrdinalIgnoreCase) >= 0
                || methodName.IndexOf("Calculate", StringComparison.OrdinalIgnoreCase) >= 0
                || methodName.IndexOf("Validate", StringComparison.OrdinalIgnoreCase) >= 0;

            if (!looksLikeRefresh)
                return false;

            return methodName.IndexOf("Hydrophone", StringComparison.OrdinalIgnoreCase) >= 0
                || methodName.IndexOf("Range", StringComparison.OrdinalIgnoreCase) >= 0
                || methodName.IndexOf("Modifier", StringComparison.OrdinalIgnoreCase) >= 0;
        }

        private static bool IsHydrophoneRangeMemberName(string name, bool ownerLooksHydrophone)
        {
            if (string.IsNullOrEmpty(name))
                return false;

            foreach (string key in HydrophoneParameterKeys)
            {
                if (string.Equals(name, key, StringComparison.OrdinalIgnoreCase))
                    return true;
            }

            if (IsExcludedRangeName(name))
                return false;

            bool explicitHydrophone =
                name.IndexOf("Hydrophone", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("PassiveSonar", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Listening", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Horch", StringComparison.OrdinalIgnoreCase) >= 0;

            if (!explicitHydrophone && !ownerLooksHydrophone)
                return false;

            return name.IndexOf("Range", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Distance", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Radius", StringComparison.OrdinalIgnoreCase) >= 0;
        }

        private static bool IsExcludedRangeName(string name)
        {
            return name.IndexOf("Arc", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Fade", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Angle", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Bearing", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Heading", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Direction", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Rotation", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Fov", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Volume", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Frequency", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Noise", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Speed", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Delay", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Time", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Cooldown", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Duration", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Damage", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Explosion", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Torpedo", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Accuracy", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Performance", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Penalty", StringComparison.OrdinalIgnoreCase) >= 0
                || name.IndexOf("Modifier", StringComparison.OrdinalIgnoreCase) >= 0;
        }

        private static Parameter GetParameter(ParameterCollection parameters, string key)
        {
            try
            {
                return parameters.GetParameter(key);
            }
            catch
            {
                return null;
            }
        }

        private static Type[] GetTypesSafely(Assembly assembly)
        {
            try
            {
                return assembly.GetTypes();
            }
            catch (ReflectionTypeLoadException ex)
            {
                return ex.Types;
            }
            catch
            {
                return null;
            }
        }

        private static string GetMemberKey(MemberInfo member)
        {
            if (member == null)
                return string.Empty;

            string declaringType = member.DeclaringType == null ? string.Empty : member.DeclaringType.FullName;
            return declaringType + "." + member.Name;
        }

        private static string SafeObjectName(object target)
        {
            if (target == null)
                return "null";

            UnityEngine.Object unityObject = target as UnityEngine.Object;
            if (unityObject != null)
                return target.GetType().Name + "(" + unityObject.name + ")";

            return target.GetType().Name;
        }

        private static float GetFloatTolerance(float reference)
        {
            return Math.Max(0.0001f, Math.Abs(reference) * 0.0001f);
        }

        private static bool IsFinite(float value)
        {
            return !float.IsNaN(value) && !float.IsInfinity(value);
        }

        private static bool TryEnter(object target)
        {
            if (target == null)
                return false;

            return ApplicationGuardIds.Add(RuntimeHelpers.GetHashCode(target));
        }

        private static void Exit(object target)
        {
            if (target == null)
                return;

            ApplicationGuardIds.Remove(RuntimeHelpers.GetHashCode(target));
        }
    }

    internal sealed class MegaSonarObjectPatchData
    {
        public readonly Dictionary<string, MegaSonarFloatMemberPatchData> FloatMembers =
            new Dictionary<string, MegaSonarFloatMemberPatchData>();
    }

    internal sealed class MegaSonarFloatMemberPatchData
    {
        public float OriginalValue;
        public float LastAppliedFactor;
        public float LastPatchedValue;

        public MegaSonarFloatMemberPatchData(float originalValue, float lastAppliedFactor, float lastPatchedValue)
        {
            OriginalValue = originalValue;
            LastAppliedFactor = lastAppliedFactor;
            LastPatchedValue = lastPatchedValue;
        }
    }

    [HarmonyPatch]
    internal static class MegaSonarHydrophoneRefreshPatch
    {
        private static IEnumerable<MethodBase> TargetMethods()
        {
            return MegaSonarRuntimePatcher.FindHydrophoneTargetMethods();
        }

        private static void Postfix(object __instance)
        {
            MegaSonarRuntimePatcher.ApplyObject(__instance, "hydrophone refresh hook");
        }
    }

    internal sealed class ParameterScalePatchData
    {
        public readonly Modifier ScaleModifier;

        public ParameterScalePatchData(Modifier scaleModifier)
        {
            ScaleModifier = scaleModifier;
        }
    }

    internal sealed class ParameterDeltaPatchData
    {
        public readonly Modifier DeltaModifier;

        public ParameterDeltaPatchData(Modifier deltaModifier)
        {
            DeltaModifier = deltaModifier;
        }
    }

    internal sealed class OxygenModifierPatchData
    {
        public float OriginalValue;
        public float LastAppliedFactor;
        public float LastPatchedValue;

        public OxygenModifierPatchData(float originalValue)
        {
            OriginalValue = originalValue;
            LastAppliedFactor = 1f;
            LastPatchedValue = originalValue;
        }
    }

    internal sealed class TorpedoGuidancePatchData
    {
        public bool HasOriginalValues;
        public bool GuidanceApplied;
        public bool GuidanceLogged;
        public bool ForcingDetonation;
        public float OriginalGyroAngle;
        public Vector3 OriginalTargetPosition;
        public Vector3 OriginalTargetPositionForReports;
    }

    internal static class SuperStealthRuntimePatcher
    {
        private const string SuperStealthScaleModifierName = "LongSubmerged10x Super Stealth Runtime";

        private static readonly ConditionalWeakTable<Parameter, ParameterScalePatchData> StealthScaleData =
            new ConditionalWeakTable<Parameter, ParameterScalePatchData>();

        private static readonly HashSet<int> ApplicationGuardIds = new HashSet<int>();

        public static void ApplyPlayerShip(PlayerShip ship, string reason)
        {
            if (ship == null)
                return;

            if (!TryEnter(ship))
                return;

            try
            {
                float scale = GetEffectiveStealthScale();

                ApplyParameter(ship.CrewNoiseModifier, scale);
                ApplyParameter(ship.StationaryNoise, scale);
                ApplyEntityDetectability(ship, scale);
                ApplySandboxDetectability(ship, scale);
            }
            catch (Exception ex)
            {
                Debug.LogWarning("[LongSubmerged10x] Super discrétion skipped after " + reason + " -> " + ex.GetType().Name + ": " + ex.Message);
            }
            finally
            {
                Exit(ship);
            }
        }

        public static void ApplyEntity(Entity entity, string reason)
        {
            ApplyPlayerShip(entity as PlayerShip, reason);
        }

        public static void ApplyEquipment(Component component, string reason)
        {
            PlayerShip owner = GetPlayerShipOwner(component);
            if (owner == null)
                return;

            // Equipment noise feeds the player's final detectability parameters. Reapply the player-level
            // scale after vanilla component updates instead of scaling the same contribution twice.
            ApplyPlayerShip(owner, reason);
        }

        private static void ApplyEntityDetectability(Entity entity, float scale)
        {
            if (entity == null)
                return;

            ApplyParameter(entity.HydrophoneDetectability, scale);
            ApplyParameter(entity.SonarDetectability, scale);
            ApplyParameter(entity.VisualDetectability, scale);
            ApplyParameter(entity.RadarDetectorDetectability, scale);
        }

        private static void ApplySandboxDetectability(Entity entity, float scale)
        {
            if (entity == null)
                return;

            SandboxEntity sandboxEntity = null;

            try
            {
                sandboxEntity = entity.SandboxEntity;
            }
            catch
            {
                sandboxEntity = null;
            }

            if (sandboxEntity == null)
                return;

            ApplyParameter(sandboxEntity.HydrophoneDetectability, scale);
            ApplyParameter(sandboxEntity.RadarDetectability, scale);
            ApplyParameter(sandboxEntity.IndirectVisualDetectability, scale);
            ApplyParameter(sandboxEntity.SignatureRadius, scale);
        }

        private static void ApplyParameter(Parameter parameter, float scale)
        {
            if (parameter == null)
                return;

            ParameterScalePatchData data;
            if (!StealthScaleData.TryGetValue(parameter, out data))
            {
                data = new ParameterScalePatchData(parameter.AddScaleModifier(SuperStealthScaleModifierName, false));
                StealthScaleData.Add(parameter, data);
            }

            if (data.ScaleModifier == null)
                return;

            if (Math.Abs(data.ScaleModifier.Value - scale) > 0.0001f)
                data.ScaleModifier.Value = scale;
        }

        private static PlayerShip GetPlayerShipOwner(Component component)
        {
            if (component == null)
                return null;

            try
            {
                return component.GetComponentInParent<PlayerShip>();
            }
            catch
            {
                return null;
            }
        }

        private static float GetEffectiveStealthScale()
        {
            if (!LongSubmergedRuntimeSettings.SuperStealth)
                return 1f;

            return 1f / LongSubmergedRuntimeSettings.SuperStealthFactor;
        }

        private static bool TryEnter(object target)
        {
            if (target == null)
                return false;

            return ApplicationGuardIds.Add(RuntimeHelpers.GetHashCode(target));
        }

        private static void Exit(object target)
        {
            if (target == null)
                return;

            ApplicationGuardIds.Remove(RuntimeHelpers.GetHashCode(target));
        }
    }

    internal struct HeavyArmorDamageScaleState
    {
        public bool ScaledDamage;
        public bool PreserveDistributionBudget;
        public float OriginalDamage;

        public HeavyArmorDamageScaleState(bool scaledDamage, bool preserveDistributionBudget, float originalDamage)
        {
            ScaledDamage = scaledDamage;
            PreserveDistributionBudget = preserveDistributionBudget;
            OriginalDamage = originalDamage;
        }
    }

    internal static class HeavyArmorRuntimePatcher
    {
        public static readonly Type[] AddDamageParameterTypes = new Type[]
        {
            typeof(float),
            typeof(Entity),
            typeof(Vector3),
            typeof(Vector3),
            typeof(float),
            typeof(float),
            typeof(DamageType),
            typeof(float),
            typeof(float),
            typeof(bool).MakeByRefType()
        };

        public static readonly Type[] AddWaterDamageParameterTypes = new Type[]
        {
            typeof(float),
            typeof(bool)
        };

        public static readonly Type[] DamageUtilityDoApplyDamageParameterTypes = new Type[]
        {
            typeof(Entity),
            typeof(PlayableCharacterData[]),
            typeof(Entity),
            typeof(Vector3),
            typeof(Vector3),
            typeof(float),
            typeof(float),
            typeof(float),
            typeof(float),
            typeof(float),
            typeof(Entity),
            typeof(DamageType),
            typeof(Action<DamageEvent>),
            typeof(Ship),
            typeof(float)
        };

        public static readonly Type[] DamageUtilityApplyDamageToComponentsParameterTypes = new Type[]
        {
            typeof(Entity),
            typeof(Vector3),
            typeof(Vector3),
            typeof(float),
            typeof(DamageType),
            typeof(float),
            typeof(float).MakeByRefType(),
            typeof(float)
        };

        public static readonly Type[] HullEffectsApplyImpactParameterTypes = new Type[]
        {
            typeof(Vector3),
            typeof(float),
            typeof(float)
        };

        [ThreadStatic]
        private static int damageScaleScopeDepth;

        [ThreadStatic]
        private static int componentDamageDistributionScopeDepth;

        [ThreadStatic]
        private static int pressureWaterDamageScopeDepth;

        public static void ScalePlayerEquipmentDamage(Equipment equipment, ref float damage)
        {
            TryScalePlayerEquipmentDamage(equipment, ref damage);
        }

        public static HeavyArmorDamageScaleState TryScalePlayerEquipmentDamage(Equipment equipment, ref float damage)
        {
            float ignoredFlawProbabilityFactor = 0f;
            float ignoredFireChance = 0f;
            return TryScalePlayerEquipmentDamage(equipment, ref damage, ref ignoredFlawProbabilityFactor, ref ignoredFireChance);
        }

        public static HeavyArmorDamageScaleState TryScalePlayerEquipmentDamage(
            Equipment equipment,
            ref float damage,
            ref float flawProbabilityFactor,
            ref float fireChance
        )
        {
            if (!ShouldScaleDamage(damage) || !IsPlayerShipEquipment(equipment))
                return new HeavyArmorDamageScaleState(false, false, 0f);

            HeavyArmorDamageScaleState state = new HeavyArmorDamageScaleState(
                true,
                IsComponentDamageDistributionScopeActive(),
                damage
            );

            damage = ScaleDamage(damage);

            if (ShouldScaleRawValue(flawProbabilityFactor))
                flawProbabilityFactor = ScaleDamage(flawProbabilityFactor);

            if (ShouldScaleRawValue(fireChance))
                fireChance = ScaleDamage(fireChance);

            return state;
        }

        public static HeavyArmorDamageScaleState TryScalePlayerWaterDamage(Equipment equipment, ref float damage)
        {
            if (IsPressureWaterDamageScopeActive())
                return new HeavyArmorDamageScaleState(false, false, 0f);

            return TryScalePlayerEquipmentDamage(equipment, ref damage);
        }

        public static void ScalePlayerCharacterDamage(PlayableCharacter character, ref float damage)
        {
            if (!ShouldScaleDamage(damage) || !IsPlayerCrewMember(character))
                return;

            damage = ScaleDamage(damage);
        }

        public static void ScalePlayerCrewDamage(Ship target, ref float crewDamage)
        {
            if (!IsHeavyArmorActive() || !IsPlayerEntity(target) || !ShouldScaleRawValue(crewDamage))
                return;

            crewDamage = ScaleDamage(crewDamage);
        }

        public static void ScalePlayerHullImpact(HullEffectsRenderer renderer, ref float intensity)
        {
            if (!IsHeavyArmorActive() || !IsPlayerHullRenderer(renderer) || !ShouldScaleRawValue(intensity))
                return;

            intensity = ScaleDamage(intensity);
        }

        public static bool BeginComponentDamageDistributionScope()
        {
            if (!IsHeavyArmorActive())
                return false;

            componentDamageDistributionScopeDepth++;
            return true;
        }

        public static void EndComponentDamageDistributionScope(bool entered)
        {
            if (!entered)
                return;

            if (componentDamageDistributionScopeDepth > 0)
                componentDamageDistributionScopeDepth--;
        }

        public static bool BeginPressureWaterDamageScope()
        {
            pressureWaterDamageScopeDepth++;
            return true;
        }

        public static void EndPressureWaterDamageScope(bool entered)
        {
            if (!entered)
                return;

            if (pressureWaterDamageScopeDepth > 0)
                pressureWaterDamageScopeDepth--;
        }

        public static void BeginDamageScaleScope(HeavyArmorDamageScaleState state)
        {
            if (state.ScaledDamage)
                damageScaleScopeDepth++;
        }

        public static void EndDamageScaleScope(HeavyArmorDamageScaleState state)
        {
            if (!state.ScaledDamage)
                return;

            if (damageScaleScopeDepth > 0)
                damageScaleScopeDepth--;
        }

        public static void RestoreComponentDistributionBudget(HeavyArmorDamageScaleState state, ref float result)
        {
            if (!state.PreserveDistributionBudget || !ShouldScaleRawValue(result))
                return;

            float restoredResult = result * LongSubmergedRuntimeSettings.HeavyArmorDamageFactor;
            if (ShouldScaleRawValue(state.OriginalDamage))
                result = Mathf.Min(state.OriginalDamage, restoredResult);
            else
                result = restoredResult;
        }

        private static bool ShouldScaleDamage(float damage)
        {
            return IsHeavyArmorActive() && !IsDamageScaleScopeActive() && ShouldScaleRawValue(damage);
        }

        private static bool IsDamageScaleScopeActive()
        {
            return damageScaleScopeDepth > 0;
        }

        private static bool IsComponentDamageDistributionScopeActive()
        {
            return componentDamageDistributionScopeDepth > 0;
        }

        private static bool IsPressureWaterDamageScopeActive()
        {
            return pressureWaterDamageScopeDepth > 0;
        }

        private static bool ShouldScaleRawValue(float value)
        {
            return IsFinite(value) && value > 0f;
        }

        private static float ScaleDamage(float value)
        {
            return value / LongSubmergedRuntimeSettings.HeavyArmorDamageFactor;
        }

        private static bool IsHeavyArmorActive()
        {
            return LongSubmergedRuntimeSettings.HeavyArmor
                && LongSubmergedRuntimeSettings.HeavyArmorDamageFactor > 1.0001f;
        }

        private static bool IsPlayerShipEquipment(Equipment equipment)
        {
            if (equipment == null)
                return false;

            try
            {
                if (IsPlayerEntity(equipment.ParentEntity))
                    return true;
            }
            catch
            {
            }

            try
            {
                return equipment.GetComponentInParent<PlayerShip>() != null;
            }
            catch
            {
                return false;
            }
        }

        private static bool IsPlayerCrewMember(PlayableCharacter character)
        {
            if (character == null)
                return false;

            try
            {
                if (IsPlayerEntity(character.ParentEntity))
                    return true;
            }
            catch
            {
            }

            try
            {
                return character.GetComponentInParent<PlayerShip>() != null;
            }
            catch
            {
                return false;
            }
        }

        private static bool IsPlayerHullRenderer(HullEffectsRenderer renderer)
        {
            if (renderer == null)
                return false;

            try
            {
                if (IsPlayerEntity(renderer.ParentEntity))
                    return true;
            }
            catch
            {
            }

            try
            {
                return renderer.GetComponentInParent<PlayerShip>() != null;
            }
            catch
            {
                return false;
            }
        }

        private static bool IsPlayerEntity(Entity entity)
        {
            if (entity == null)
                return false;

            if (entity is PlayerShip)
                return true;

            try
            {
                SandboxEntity sandboxEntity = entity.SandboxEntity;
                return sandboxEntity != null && sandboxEntity.IsPlayerShip;
            }
            catch
            {
                return false;
            }
        }

        private static bool IsFinite(float value)
        {
            return !float.IsNaN(value) && !float.IsInfinity(value);
        }
    }

    // DonJ : profondeur x2 sans falsifier le profondimetre reel.
    // Les ordres de profondeur > 10 m sont transformes en profondeur reelle :
    // 20->40, 40->80, 150->300, 300->600. Le crush vanilla est neutralise sous 700 m.
    internal static class DeepDiveRuntimePatcher
    {
        public const float DisplayedDepthCommandFactor = 2f;
        public const float ShallowDepthPassthroughMeters = 10f;
        public const float MaxDisplayedCommandDepthMeters = 300f;
        public const float MaxRealCommandDepthMeters = 600f;
        public const float CrushDepthMeters = 700f;

        private const float MetersPerAtmosphere = 10f;
        private const float SeaLevelPressureAtmospheres = 1f;
        private const float Epsilon = 0.01f;
        private const float FullScanIntervalSeconds = 2f;
        private const int MaxObjectPatchLogs = 20;
        private const string HullCrushDepthDeltaModifierName = "LongSubmerged10x DeepDive Crush Depth";

        private static readonly BindingFlags InstanceMemberFlags =
            BindingFlags.Instance | BindingFlags.Public | BindingFlags.NonPublic;

        public static readonly Type[] PlayerShipSetTargetDepthParameterTypes =
            new Type[] { typeof(float), typeof(bool), typeof(bool) };

        private static readonly MethodInfo PlayerShipSetTargetDepthMethod =
            AccessTools.Method(typeof(PlayerShip), "SetTargetDepth", PlayerShipSetTargetDepthParameterTypes);

        private static readonly MethodInfo PlayerShipUpdateStressAndDisciplineGainMethod =
            AccessTools.Method(typeof(PlayerShip), "UpdateStressAndDisciplineGain", new Type[] { });

        private static readonly FieldInfo DepthStressModifierField =
            AccessTools.Field(typeof(PlayerShip), "depthStressModifier");

        private static readonly Dictionary<Type, DepthMemberCache> MemberCache =
            new Dictionary<Type, DepthMemberCache>();

        private static readonly HashSet<string> MissingTypeWarnings =
            new HashSet<string>();

        private static readonly HashSet<int> ObjectPatchLogIds =
            new HashSet<int>();

        private static readonly ConditionalWeakTable<Parameter, ParameterDeltaPatchData> CrushDepthDeltaData =
            new ConditionalWeakTable<Parameter, ParameterDeltaPatchData>();

        private static readonly ConditionalWeakTable<object, DepthObjectPatchData> DepthObjectPatches =
            new ConditionalWeakTable<object, DepthObjectPatchData>();

        private static float nextFullScanTime;
        private static int objectPatchLogCount;

        public static bool IsEnabled()
        {
            return LongSubmergedRuntimeSettings.DeepDive;
        }

        public static void ApplyAll(string reason)
        {
            try
            {
                ApplyPlayerShip(UnityEngine.Object.FindObjectOfType<PlayerShip>(), reason + ".PlayerShip");
                ApplyNearbyDepthObjects(reason + ".Objects");
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }
        }

        public static void ApplyPlayerShip(PlayerShip ship, string reason)
        {
            if (ship == null)
                return;

            try
            {
                if (!IsEnabled())
                {
                    RestoreDepthObject(ship, reason);
                    return;
                }

                ApplyDepthObject(ship, reason);
                ClampPlayerShipTargetDepth(ship, reason);

                if (Time.unscaledTime >= nextFullScanTime)
                {
                    nextFullScanTime = Time.unscaledTime + FullScanIntervalSeconds;
                    ApplyNearbyDepthObjects(reason + ".Scan");
                }
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }
        }

        public static void UpdatePlayerShipRuntime(PlayerShip ship, string reason)
        {
            ApplyPlayerShip(ship, reason);
        }

        public static void ApplyDepthObject(object target, string reason)
        {
            if (target == null)
                return;

            try
            {
                if (!IsEnabled())
                {
                    RestoreDepthObject(target, reason);
                    return;
                }

                Type type = target.GetType();
                DepthMemberCache cache = GetDepthMemberCache(type);
                int changed = 0;

                for (int index = 0; index < cache.Fields.Length; index++)
                {
                    if (TryPatchField(target, cache.Fields[index]))
                        changed++;
                }

                for (int index = 0; index < cache.Properties.Length; index++)
                {
                    if (TryPatchProperty(target, cache.Properties[index]))
                        changed++;
                }

                if (changed > 0)
                    LogPatchedObjectOnce(target, type, changed, reason);
            }
            catch (Exception ex)
            {
                Debug.LogWarning(
                    "[LongSubmerged10x] DeepDive reflection skipped for "
                    + target.GetType().Name
                    + " after "
                    + SafeReason(reason)
                    + ": "
                    + ex.GetType().Name
                    + ": "
                    + ex.Message
                );
            }
        }

        public static void ScaleTargetDepthCommand(ref float value, string reason)
        {
            if (!IsEnabled())
                return;

            if (!IsFinite(value) || value <= ShallowDepthPassthroughMeters)
                return;

            float original = value;
            float scaled = value <= MaxDisplayedCommandDepthMeters + Epsilon
                ? value * DisplayedDepthCommandFactor
                : value;

            scaled = Mathf.Clamp(scaled, 0f, MaxRealCommandDepthMeters);

            if (Mathf.Abs(scaled - original) <= Epsilon)
                return;

            value = scaled;
            Debug.Log(
                "[LongSubmerged10x] DeepDive target depth "
                + original.ToString("0.#")
                + " m -> "
                + scaled.ToString("0.#")
                + " m after "
                + SafeReason(reason)
                + "."
            );
        }

        public static void ApplyDepthStressModifier(PlayerShip ship, string reason)
        {
            if (ship == null || !IsEnabled())
                return;

            try
            {
                if (DepthStressModifierField == null)
                {
                    WarnMissingTypeOnce(
                        "PlayerShip.depthStressModifier",
                        "champ depthStressModifier introuvable, stress de profondeur vanilla conserve."
                    );
                    return;
                }

                Modifier depthStressModifier = DepthStressModifierField.GetValue(ship) as Modifier;
                if (depthStressModifier == null)
                {
                    WarnMissingTypeOnce(
                        "PlayerShip.depthStressModifier.Value",
                        "modificateur stress profondeur introuvable, stress de profondeur vanilla conserve."
                    );
                    return;
                }

                float deckDepth = ship.DeckDepth;
                float targetDepth = ship.TargetDepth;
                if (!IsFinite(deckDepth) && !IsFinite(targetDepth))
                    return;

                if (!IsFinite(deckDepth))
                    deckDepth = 0f;

                if (!IsFinite(targetDepth))
                    targetDepth = 0f;

                float realDepth = Mathf.Max(Mathf.Max(0f, deckDepth), Mathf.Max(0f, targetDepth));
                float vanillaTier = GetDepthStressTier(realDepth);
                if (vanillaTier <= 0f)
                    return;

                float effectiveDepth = realDepth / DisplayedDepthCommandFactor;
                float effectiveTier = GetDepthStressTier(effectiveDepth);
                float currentValue = depthStressModifier.Value;
                if (!IsFinite(currentValue))
                    return;

                float desiredValue = currentValue * (effectiveTier / vanillaTier);
                if (Math.Abs(currentValue - desiredValue) <= 0.000001f)
                    return;

                depthStressModifier.Value = desiredValue;
            }
            catch (Exception ex)
            {
                Debug.LogWarning(
                    "[LongSubmerged10x] DeepDive stress patch skipped after "
                    + SafeReason(reason)
                    + ": "
                    + ex.GetType().Name
                    + ": "
                    + ex.Message
                );
            }
        }

        public static MethodBase FindMethodOnKnownType(string simpleTypeName, string methodName)
        {
            Type type = FindKnownType(simpleTypeName);
            if (type == null)
                return null;

            MethodBase exactNoArgs = AccessTools.Method(type, methodName, new Type[] { });
            if (exactNoArgs != null)
                return exactNoArgs;

            return AccessTools.Method(type, methodName);
        }

        public static MethodBase FindPlayerShipSetTargetDepthMethod()
        {
            return PlayerShipSetTargetDepthMethod;
        }

        public static MethodBase FindPlayerShipUpdateStressAndDisciplineGainMethod()
        {
            return PlayerShipUpdateStressAndDisciplineGainMethod;
        }

        public static bool ShouldRunHullCrushDoUpdate(object controller, ref float result, string reason)
        {
            if (!IsEnabled())
            {
                RestoreDepthObject(controller, reason);
                TryPatchHullCrushControllerData(controller, reason);
                return true;
            }

            ApplyDepthObject(controller, reason);

            if (TryPatchHullCrushControllerData(controller, reason))
                return true;

            float depth = GetPlayerDeckDepthMeters();
            if (!IsFinite(depth))
            {
                WarnMissingTypeOnce(
                    "PlayerShipDepth",
                    "profondeur joueur introuvable, HullCrushController vanilla conserve."
                );
                return true;
            }

            // Sous 700 m, on neutralise le crush vanilla. A 700 m et plus, on laisse le jeu executer
            // son chemin original de destruction pour garder la vraie logique de game over.
            if (depth < CrushDepthMeters - Epsilon)
            {
                result = 1f;
                return false;
            }

            return true;
        }

        public static bool ShouldSkipPressureWaterDamageTick(object controller, string reason)
        {
            if (!IsEnabled())
            {
                RestoreDepthObject(controller, reason);
                return false;
            }

            ApplyDepthObject(controller, reason);

            float depth = GetPlayerDeckDepthMeters();
            if (!IsFinite(depth))
                return false;

            // Jusqu'a 600 m reels, la profondeur demandee est consideree comme operationnelle.
            // Au-dessus, la pression vanilla peut recommencer a infliger des avaries avant le crush a 700 m.
            return depth < MaxRealCommandDepthMeters - Epsilon;
        }

        private static void ApplyNearbyDepthObjects(string reason)
        {
            foreach (DivingPlanesStation station in UnityEngine.Object.FindObjectsOfType<DivingPlanesStation>())
                ApplyDepthObject(station, reason + ".DivingPlanesStation");

            foreach (Equipment equipment in UnityEngine.Object.FindObjectsOfType<Equipment>())
                ApplyDepthObject(equipment, reason + ".Equipment");

            ApplyObjectsByTypeName("HullCrushController", reason + ".HullCrushController");
            ApplyObjectsByTypeName("ApplyWaterDamageToPlayerShip", reason + ".ApplyWaterDamageToPlayerShip");
        }

        private static void ApplyObjectsByTypeName(string simpleTypeName, string reason)
        {
            Type type = FindKnownType(simpleTypeName);
            if (type == null)
            {
                WarnMissingTypeOnce(
                    simpleTypeName,
                    "type " + simpleTypeName + " introuvable, patch profondeur partiel."
                );
                return;
            }

            UnityEngine.Object[] objects = UnityEngine.Object.FindObjectsOfType(type);
            for (int index = 0; index < objects.Length; index++)
            {
                ApplyDepthObject(objects[index], reason);

                if (simpleTypeName == "HullCrushController")
                    TryPatchHullCrushControllerData(objects[index], reason);
            }
        }

        private static bool TryPatchHullCrushControllerData(object controller, string reason)
        {
            if (controller == null)
                return false;

            try
            {
                Type controllerType = controller.GetType();
                MethodInfo parseMethod = AccessTools.Method(controllerType, "ParseNewEntities", new Type[] { });
                if (parseMethod != null)
                    parseMethod.Invoke(controller, null);

                FieldInfo dataField = AccessTools.Field(controllerType, "hullCrushData");
                if (dataField == null)
                    return false;

                Array data = dataField.GetValue(controller) as Array;
                if (data == null)
                    return false;

                int changed = 0;
                for (int index = 0; index < data.Length; index++)
                {
                    object item = data.GetValue(index);
                    if (item == null)
                        continue;

                    Type itemType = item.GetType();
                    FieldInfo entityField = AccessTools.Field(itemType, "Entity");
                    object entity = entityField == null ? null : entityField.GetValue(item);
                    if (!(entity is PlayerShip))
                        continue;

                    FieldInfo crushDepthField = AccessTools.Field(itemType, "CrushDepth");
                    Parameter crushDepth = crushDepthField == null ? null : crushDepthField.GetValue(item) as Parameter;
                    if (crushDepth == null)
                        continue;

                    if (ApplyCrushDepthParameter(crushDepth))
                        changed++;
                }

                if (changed > 0)
                {
                    if (IsEnabled())
                    {
                        Debug.Log(
                            "[LongSubmerged10x] DeepDive raised player crush depth to "
                            + CrushDepthMeters.ToString("0.#")
                            + " m after "
                            + SafeReason(reason)
                            + "."
                        );
                    }
                    else
                    {
                        Debug.Log(
                            "[LongSubmerged10x] DeepDive restored player crush depth after "
                            + SafeReason(reason)
                            + "."
                        );
                    }
                }

                return true;
            }
            catch (Exception ex)
            {
                Debug.LogWarning(
                    "[LongSubmerged10x] DeepDive hull crush data patch skipped after "
                    + SafeReason(reason)
                    + ": "
                    + ex.GetType().Name
                    + ": "
                    + ex.Message
                );
                return false;
            }
        }

        private static bool ApplyCrushDepthParameter(Parameter parameter)
        {
            if (parameter == null)
                return false;

            float baseValue = parameter.GetValueExcludingModifier(HullCrushDepthDeltaModifierName);
            if (!IsFinite(baseValue))
                return false;

            float target = GetHullCrushDepthTarget();
            float desiredDelta = IsEnabled()
                ? (baseValue <= target + Epsilon ? 0f : target - baseValue)
                : 0f;

            ParameterDeltaPatchData data;
            if (!CrushDepthDeltaData.TryGetValue(parameter, out data))
            {
                if (Math.Abs(desiredDelta) <= 0.0001f)
                    return false;

                data = new ParameterDeltaPatchData(parameter.AddDeltaModifier(HullCrushDepthDeltaModifierName, false));
                CrushDepthDeltaData.Add(parameter, data);
            }

            if (data.DeltaModifier == null)
                return false;

            if (Math.Abs(data.DeltaModifier.Value - desiredDelta) <= 0.0001f)
                return false;

            data.DeltaModifier.Value = desiredDelta;
            return true;
        }

        private static float GetHullCrushDepthTarget()
        {
            return -CrushDepthMeters;
        }

        private static Type FindKnownType(string simpleTypeName)
        {
            if (string.IsNullOrEmpty(simpleTypeName))
                return null;

            Type type = AccessTools.TypeByName(simpleTypeName);
            if (type != null)
                return type;

            string[] namespaces = new string[]
            {
                "UBOAT.Game",
                "UBOAT.Game.Scene",
                "UBOAT.Game.Scene.Tasks",
                "UBOAT.Game.Scene.Entities",
                "UBOAT.Game.Scene.Utilities",
                "UBOAT.Game.Core"
            };

            for (int index = 0; index < namespaces.Length; index++)
            {
                type = AccessTools.TypeByName(namespaces[index] + "." + simpleTypeName);
                if (type != null)
                    return type;
            }

            return null;
        }

        private static void ClampPlayerShipTargetDepth(PlayerShip ship, string reason)
        {
            try
            {
                float targetDepth = ship.TargetDepth;
                if (!IsFinite(targetDepth) || targetDepth <= MaxRealCommandDepthMeters + Epsilon)
                    return;

                if (TrySetPlayerShipTargetDepth(ship, MaxRealCommandDepthMeters, reason))
                {
                    Debug.Log(
                        "[LongSubmerged10x] DeepDive target depth clamped from "
                        + targetDepth.ToString("0.#")
                        + " m to "
                        + MaxRealCommandDepthMeters.ToString("0.#")
                        + " m after "
                        + SafeReason(reason)
                        + "."
                    );
                }
            }
            catch
            {
                float targetDepth;
                if (!TryReadFloatMember(ship, "TargetDepth", out targetDepth))
                    return;

                if (!IsFinite(targetDepth) || targetDepth <= MaxRealCommandDepthMeters + Epsilon)
                    return;

                TrySetPlayerShipTargetDepth(ship, MaxRealCommandDepthMeters, reason);
            }
        }

        private static bool TrySetPlayerShipTargetDepth(PlayerShip ship, float value, string reason)
        {
            if (ship == null)
                return false;

            try
            {
                if (PlayerShipSetTargetDepthMethod != null)
                {
                    PlayerShipSetTargetDepthMethod.Invoke(ship, new object[] { value, true, false });
                    return true;
                }
            }
            catch (Exception ex)
            {
                Debug.LogWarning(
                    "[LongSubmerged10x] DeepDive SetTargetDepth invoke skipped after "
                    + SafeReason(reason)
                    + ": "
                    + ex.GetType().Name
                    + ": "
                    + ex.Message
                );
            }

            if (TryWriteFloatMember(ship, "targetDepth", value))
                return true;

            if (TryWriteFloatMember(ship, "TargetDepth", value))
                return true;

            WarnMissingTypeOnce(
                "PlayerShip.SetTargetDepth",
                "PlayerShip.SetTargetDepth introuvable, clamp profondeur impossible."
            );
            return false;
        }

        private static float GetPlayerDeckDepthMeters()
        {
            return GetPlayerDeckDepthMeters(UnityEngine.Object.FindObjectOfType<PlayerShip>());
        }

        private static float GetPlayerDeckDepthMeters(PlayerShip ship)
        {
            if (ship == null)
                return float.NaN;

            try
            {
                if (IsFinite(ship.DeckDepth))
                    return Mathf.Max(0f, ship.DeckDepth);
            }
            catch
            {
                // Fallback reflection ci-dessous.
            }

            float value;
            if (TryReadFloatMember(ship, "DeckDepth", out value))
                return Mathf.Max(0f, value);

            if (TryReadFloatMember(ship, "KeelDepth", out value))
                return Mathf.Max(0f, value);

            if (TryReadFloatMember(ship, "Depth", out value))
                return Mathf.Max(0f, value);

            return float.NaN;
        }

        private static bool TryReadFloatMember(object target, string memberName, out float value)
        {
            value = 0f;
            if (target == null || string.IsNullOrEmpty(memberName))
                return false;

            Type type = target.GetType();

            FieldInfo field = AccessTools.Field(type, memberName);
            if (field != null && TryObjectToFloat(field.GetValue(target), out value))
                return true;

            PropertyInfo property = AccessTools.Property(type, memberName);
            if (property != null && property.CanRead && property.GetIndexParameters().Length == 0)
            {
                try
                {
                    return TryObjectToFloat(property.GetValue(target, null), out value);
                }
                catch
                {
                    return false;
                }
            }

            return false;
        }

        private static bool TryWriteFloatMember(object target, string memberName, float value)
        {
            if (target == null || string.IsNullOrEmpty(memberName))
                return false;

            Type type = target.GetType();

            FieldInfo field = AccessTools.Field(type, memberName);
            if (field != null && !field.IsInitOnly && !field.IsLiteral)
            {
                if (field.FieldType == typeof(float))
                {
                    field.SetValue(target, value);
                    return true;
                }

                if (field.FieldType == typeof(double))
                {
                    field.SetValue(target, (double)value);
                    return true;
                }

                if (field.FieldType == typeof(int))
                {
                    field.SetValue(target, Mathf.RoundToInt(value));
                    return true;
                }
            }

            PropertyInfo property = AccessTools.Property(type, memberName);
            if (
                property != null
                && property.CanWrite
                && property.GetIndexParameters().Length == 0
                && property.GetSetMethod(true) != null
            )
            {
                if (property.PropertyType == typeof(float))
                {
                    property.SetValue(target, value, null);
                    return true;
                }

                if (property.PropertyType == typeof(double))
                {
                    property.SetValue(target, (double)value, null);
                    return true;
                }

                if (property.PropertyType == typeof(int))
                {
                    property.SetValue(target, Mathf.RoundToInt(value), null);
                    return true;
                }
            }

            return false;
        }

        private static DepthMemberCache GetDepthMemberCache(Type type)
        {
            DepthMemberCache cache;
            if (MemberCache.TryGetValue(type, out cache))
                return cache;

            List<FieldInfo> fields = new List<FieldInfo>();
            List<PropertyInfo> properties = new List<PropertyInfo>();

            FieldInfo[] allFields = type.GetFields(InstanceMemberFlags);
            for (int index = 0; index < allFields.Length; index++)
            {
                FieldInfo field = allFields[index];
                if (CanPatchNumericType(field.FieldType) && !field.IsInitOnly && !field.IsLiteral && IsDepthLimitMemberName(field.Name))
                    fields.Add(field);
            }

            PropertyInfo[] allProperties = type.GetProperties(InstanceMemberFlags);
            for (int index = 0; index < allProperties.Length; index++)
            {
                PropertyInfo property = allProperties[index];
                if (
                    CanPatchNumericType(property.PropertyType)
                    && property.CanRead
                    && property.CanWrite
                    && property.GetIndexParameters().Length == 0
                    && property.GetSetMethod(true) != null
                    && IsDepthLimitMemberName(property.Name)
                )
                {
                    properties.Add(property);
                }
            }

            cache = new DepthMemberCache(fields.ToArray(), properties.ToArray());
            MemberCache[type] = cache;
            return cache;
        }

        private static bool TryPatchField(object target, FieldInfo field)
        {
            object rawValue = field.GetValue(target);
            float current;
            if (!TryObjectToFloat(rawValue, out current))
                return false;

            float patched;
            if (!TryBuildPatchedLimit(field.Name, current, out patched))
                return false;

            object patchedValue;
            if (!TryBuildTypedNumericValue(field.FieldType, patched, out patchedValue))
                return false;

            RememberPatchedMember(target, GetMemberPatchKey(field), rawValue, patchedValue);
            field.SetValue(target, patchedValue);
            return true;
        }

        private static bool TryPatchProperty(object target, PropertyInfo property)
        {
            object rawValue = property.GetValue(target, null);
            float current;
            if (!TryObjectToFloat(rawValue, out current))
                return false;

            float patched;
            if (!TryBuildPatchedLimit(property.Name, current, out patched))
                return false;

            object patchedValue;
            if (!TryBuildTypedNumericValue(property.PropertyType, patched, out patchedValue))
                return false;

            RememberPatchedMember(target, GetMemberPatchKey(property), rawValue, patchedValue);
            property.SetValue(target, patchedValue, null);
            return true;
        }

        private static void RestoreDepthObject(object target, string reason)
        {
            if (target == null)
                return;

            DepthObjectPatchData data;
            if (!DepthObjectPatches.TryGetValue(target, out data) || data.Values.Count == 0)
                return;

            try
            {
                Type type = target.GetType();
                DepthMemberCache cache = GetDepthMemberCache(type);
                int restored = 0;

                for (int index = 0; index < cache.Fields.Length; index++)
                {
                    if (TryRestoreField(target, cache.Fields[index], data))
                        restored++;
                }

                for (int index = 0; index < cache.Properties.Length; index++)
                {
                    if (TryRestoreProperty(target, cache.Properties[index], data))
                        restored++;
                }

                if (data.Values.Count == 0)
                    DepthObjectPatches.Remove(target);

                if (restored > 0)
                {
                    Debug.Log(
                        "[LongSubmerged10x] DeepDive restored "
                        + restored
                        + " depth/pressure limits on "
                        + type.Name
                        + " after "
                        + SafeReason(reason)
                        + "."
                    );
                }
            }
            catch (Exception ex)
            {
                Debug.LogWarning(
                    "[LongSubmerged10x] DeepDive restore skipped for "
                    + target.GetType().Name
                    + " after "
                    + SafeReason(reason)
                    + ": "
                    + ex.GetType().Name
                    + ": "
                    + ex.Message
                );
            }
        }

        private static bool TryRestoreField(object target, FieldInfo field, DepthObjectPatchData data)
        {
            string key = GetMemberPatchKey(field);
            DepthOriginalValue originalValue;
            if (!data.Values.TryGetValue(key, out originalValue))
                return false;

            object currentValue = field.GetValue(target);
            if (!ShouldRestoreMemberValue(currentValue, originalValue.PatchedValue))
            {
                data.Values.Remove(key);
                return false;
            }

            field.SetValue(target, originalValue.OriginalValue);
            data.Values.Remove(key);
            return true;
        }

        private static bool TryRestoreProperty(object target, PropertyInfo property, DepthObjectPatchData data)
        {
            string key = GetMemberPatchKey(property);
            DepthOriginalValue originalValue;
            if (!data.Values.TryGetValue(key, out originalValue))
                return false;

            object currentValue = property.GetValue(target, null);
            if (!ShouldRestoreMemberValue(currentValue, originalValue.PatchedValue))
            {
                data.Values.Remove(key);
                return false;
            }

            property.SetValue(target, originalValue.OriginalValue, null);
            data.Values.Remove(key);
            return true;
        }

        private static void RememberPatchedMember(object target, string key, object originalValue, object patchedValue)
        {
            DepthObjectPatchData data;
            if (!DepthObjectPatches.TryGetValue(target, out data))
            {
                data = new DepthObjectPatchData();
                DepthObjectPatches.Add(target, data);
            }

            DepthOriginalValue storedValue;
            if (!data.Values.TryGetValue(key, out storedValue))
            {
                storedValue = new DepthOriginalValue(originalValue, patchedValue);
                data.Values.Add(key, storedValue);
                return;
            }

            storedValue.PatchedValue = patchedValue;
        }

        private static bool ShouldRestoreMemberValue(object currentValue, object patchedValue)
        {
            float current;
            float patched;
            if (TryObjectToFloat(currentValue, out current) && TryObjectToFloat(patchedValue, out patched))
                return Math.Abs(current - patched) <= Epsilon;

            return object.Equals(currentValue, patchedValue);
        }

        private static string GetMemberPatchKey(MemberInfo member)
        {
            return member.MemberType.ToString() + ":" + member.DeclaringType.FullName + ":" + member.Name;
        }

        private static bool TryBuildTypedNumericValue(Type type, float value, out object typedValue)
        {
            typedValue = null;

            if (type == typeof(float))
            {
                typedValue = value;
                return true;
            }

            if (type == typeof(double))
            {
                typedValue = (double)value;
                return true;
            }

            if (type == typeof(int))
            {
                typedValue = Mathf.RoundToInt(value);
                return true;
            }

            return false;
        }

        private static bool TryBuildPatchedLimit(string memberName, float current, out float patched)
        {
            patched = current;

            if (!IsFinite(current) || Math.Abs(current) <= Epsilon)
                return false;

            string name = memberName.ToLowerInvariant();
            float wanted;

            if (IsPressureMemberName(name) && !name.Contains("depth"))
            {
                if (current <= 0f)
                    return false;

                wanted = DepthMetersToAtmospheres(IsCrushMemberName(name) ? CrushDepthMeters : MaxRealCommandDepthMeters);
            }
            else if (IsCrushMemberName(name))
                wanted = GetSignedDepthLimit(current, CrushDepthMeters);
            else
                wanted = GetSignedDepthLimit(current, MaxRealCommandDepthMeters);

            // Never make an existing limit from the game or another mod stricter.
            if (current < 0f)
            {
                if (current <= wanted + Epsilon)
                    return false;
            }
            else if (current >= wanted - Epsilon)
            {
                return false;
            }

            // Avoid changing tiny multipliers/probabilities with ambiguous depth-like names.
            if (Math.Abs(wanted) > 50f && Math.Abs(current) < 1f)
                return false;

            patched = wanted;
            return true;
        }

        private static float GetSignedDepthLimit(float current, float positiveMeters)
        {
            return current < 0f ? -positiveMeters : positiveMeters;
        }

        private static bool IsDepthLimitMemberName(string memberName)
        {
            if (string.IsNullOrEmpty(memberName))
                return false;

            string name = memberName.ToLowerInvariant();

            bool depthLike =
                name.Contains("depth")
                || name.Contains("pressure")
                || name.Contains("atm");

            if (!depthLike)
                return false;

            // Protection : on ne touche pas la profondeur courante du bateau.
            if (
                name.Contains("current")
                || name.Contains("actual")
                || name.Contains("deck")
                || name.Contains("keel")
                || name.Contains("real")
            )
            {
                return false;
            }

            return
                name.Contains("max")
                || name.Contains("maximum")
                || name.Contains("limit")
                || name.Contains("allowed")
                || name.Contains("operational")
                || name.Contains("safe")
                || name.Contains("danger")
                || name.Contains("warning")
                || name.Contains("test")
                || name.Contains("design")
                || IsCrushMemberName(name);
        }

        private static bool IsCrushMemberName(string name)
        {
            return name.Contains("crush")
                || name.Contains("implosion")
                || name.Contains("collapse")
                || name.Contains("destroy")
                || name.Contains("destruct")
                || name.Contains("breakdepth");
        }

        private static bool IsPressureMemberName(string name)
        {
            return name.Contains("pressure") || name.Contains("atm");
        }

        private static float DepthMetersToAtmospheres(float depthMeters)
        {
            return SeaLevelPressureAtmospheres + Mathf.Max(0f, depthMeters) / MetersPerAtmosphere;
        }

        private static float GetDepthStressTier(float depthMeters)
        {
            if (!IsFinite(depthMeters))
                return 0f;

            if (depthMeters > 300f)
                return 6f;

            if (depthMeters > 250f)
                return 5f;

            if (depthMeters > 200f)
                return 4f;

            if (depthMeters > 150f)
                return 3f;

            if (depthMeters > 100f)
                return 2f;

            if (depthMeters > 25f)
                return 1f;

            return 0f;
        }

        private static bool CanPatchNumericType(Type type)
        {
            return type == typeof(float) || type == typeof(double) || type == typeof(int);
        }

        private static bool TryObjectToFloat(object rawValue, out float value)
        {
            value = 0f;

            if (rawValue is float)
            {
                value = (float)rawValue;
                return true;
            }

            if (rawValue is double)
            {
                value = (float)(double)rawValue;
                return true;
            }

            if (rawValue is int)
            {
                value = (int)rawValue;
                return true;
            }

            return false;
        }

        private static bool IsFinite(float value)
        {
            return !float.IsNaN(value) && !float.IsInfinity(value);
        }

        private static void LogPatchedObjectOnce(object target, Type type, int changedCount, string reason)
        {
            if (objectPatchLogCount >= MaxObjectPatchLogs)
                return;

            int id = GetObjectId(target);
            if (ObjectPatchLogIds.Contains(id))
                return;

            ObjectPatchLogIds.Add(id);
            objectPatchLogCount++;

            Debug.Log(
                "[LongSubmerged10x] DeepDive patched "
                + changedCount
                + " depth/pressure limits on "
                + type.Name
                + " after "
                + SafeReason(reason)
                + "."
            );
        }

        private static int GetObjectId(object target)
        {
            UnityEngine.Object unityObject = target as UnityEngine.Object;
            if (unityObject != null)
                return unityObject.GetInstanceID();

            return RuntimeHelpers.GetHashCode(target);
        }

        private static void WarnMissingTypeOnce(string key, string message)
        {
            if (MissingTypeWarnings.Contains(key))
                return;

            MissingTypeWarnings.Add(key);
            Debug.LogWarning("[LongSubmerged10x] DeepDive: " + message);
        }

        private static string SafeReason(string reason)
        {
            return string.IsNullOrEmpty(reason) ? "unknown" : reason;
        }
    }

    internal sealed class DepthMemberCache
    {
        public readonly FieldInfo[] Fields;
        public readonly PropertyInfo[] Properties;

        public DepthMemberCache(FieldInfo[] fields, PropertyInfo[] properties)
        {
            Fields = fields ?? new FieldInfo[0];
            Properties = properties ?? new PropertyInfo[0];
        }
    }

    internal sealed class DepthObjectPatchData
    {
        public readonly Dictionary<string, DepthOriginalValue> Values =
            new Dictionary<string, DepthOriginalValue>();
    }

    internal sealed class DepthOriginalValue
    {
        public readonly object OriginalValue;
        public object PatchedValue;

        public DepthOriginalValue(object originalValue, object patchedValue)
        {
            OriginalValue = originalValue;
            PatchedValue = patchedValue;
        }
    }

    internal static class OxygenBreathRecalculator
    {
        private static readonly MethodInfo ValidateOxygenBreathModifierMethod =
            AccessTools.Method(typeof(PlayerShip), "ValidateOxygenBreathModifier");

        public static void Recalculate(PlayerShip ship, string reason)
        {
            if (ship == null || ValidateOxygenBreathModifierMethod == null)
                return;

            try
            {
                // SurfaceSafe 1.4.7 :
                // UBOAT recalcule d'abord sa respiration vanilla.
                // Ensuite seulement, le mod réduit le drain négatif si Mega Oxygène est actif.
                // On ne touche pas aux valeurs nulles/positives utilisées pendant la surface/recharge.
                ValidateOxygenBreathModifierMethod.Invoke(ship, null);
                LongSubmergedRuntimeApplier.ApplyOxygenBreathModifier(ship, reason);
                Debug.Log("[LongSubmerged10x] Oxygen runtime breath modifier applied after " + reason + ".");
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }
        }
    }

    // DonJ : SuperVitesse ne change pas toutes les allures. Je booste seulement les deux derniers crans avant,
    // le plafond du sous-marin joueur et le multiplicateur de propulseur quand ces crans rapides sont actifs.
    internal static class EngineFastSpeedPatcher
    {
        private const float FastSpeedFactor = __FAST_SPEED_FACTOR__;
        private const float FastSpeedFuelFactor = __FAST_SPEED_FUEL_FACTOR__;
        private const float PlayerSubmarineMaxSpeed = __PLAYER_SUBMARINE_MAX_SPEED__;
        private const float LegacyDataSheetPlayerSubmarineMaxSpeed = __PLAYER_SUBMARINE_MAX_SPEED__;
        private const float VanillaPlayerSubmarineMaxSpeedFallback = 32.8f;
        private const int FastForwardGearCount = __FAST_FORWARD_GEAR_COUNT__;
        private const string RuntimeVelocityModifierName = "LongSubmerged10x Player Speed Cap";

        private static readonly FieldInfo ForwardPresetsField =
            AccessTools.Field(typeof(PlayerShipEngine), "forwardPresets");

        private static readonly FieldInfo ExpectedVelocityPerGearField =
            AccessTools.Field(typeof(PlayerShipEngine), "expectedVelocityPerGear");

        private static readonly FieldInfo ExpectedVelocityPerGearUnderwaterField =
            AccessTools.Field(typeof(PlayerShipEngine), "expectedVelocityPerGearUnderwater");

        private static readonly Type EngineSpeedPresetType =
            typeof(PlayerShipEngine).GetNestedType("EngineSpeedPreset", BindingFlags.Public | BindingFlags.NonPublic);

        private static readonly FieldInfo BasePowerField =
            EngineSpeedPresetType == null ? null : AccessTools.Field(EngineSpeedPresetType, "basePower");

        private static readonly FieldInfo FuelConsumptionField =
            EngineSpeedPresetType == null ? null : AccessTools.Field(EngineSpeedPresetType, "fuelConsumptionInLitersPerHour");

        private static readonly FieldInfo ShipPropellersField =
            AccessTools.Field(typeof(Ship), "propellers");

        private static readonly ConditionalWeakTable<PlayerShipEngine, EngineSpeedPatchData> OriginalData =
            new ConditionalWeakTable<PlayerShipEngine, EngineSpeedPatchData>();

        private static readonly ConditionalWeakTable<PlayerShip, ShipRuntimePatchData> ShipRuntimeData =
            new ConditionalWeakTable<PlayerShip, ShipRuntimePatchData>();

        private static readonly ConditionalWeakTable<Propeller, PropellerPatchData> PropellerRuntimeData =
            new ConditionalWeakTable<Propeller, PropellerPatchData>();

        private static readonly HashSet<int> WarnedEngines = new HashSet<int>();

        public static void PatchPlayerShip(PlayerShip ship, string reason)
        {
            if (ship == null)
                return;

            try
            {
                PatchEngine(ship.DieselEngine, reason + ".DieselEngine");
                PatchEngine(ship.ElectricEngine, reason + ".ElectricEngine");
                PatchShipVelocityCap(ship, reason, true);
                ApplyPropellerSpeedMultiplier(ship, reason, true);
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }
        }

        public static void UpdatePlayerShipRuntime(PlayerShip ship, string reason)
        {
            if (ship == null)
                return;

            try
            {
                // DonJ : appele regulierement par PlayerShip.Update/ValidateTargetVelocity.
                // Les propulseurs sont une valeur runtime, donc il faut la remettre quand le joueur change de cran.
                PatchShipVelocityCap(ship, reason, false);
                ApplyPropellerSpeedMultiplier(ship, reason, false);
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }
        }

        public static void PatchEngine(PlayerShipEngine engine, string reason)
        {
            if (engine == null)
                return;

            try
            {
                // DonJ : les champs moteur sont prives dans UBOAT, donc je passe par reflection.
                // Si une version du jeu renomme un champ, je log une seule alerte et je laisse le moteur vanilla.
                if (!FieldsReady())
                {
                    WarnOnce(engine, "champs moteur introuvables, patch vitesse ignore.");
                    return;
                }

                Array forwardPresets = ForwardPresetsField.GetValue(engine) as Array;
                float[] expectedVelocityPerGear = ExpectedVelocityPerGearField.GetValue(engine) as float[];
                float[] expectedVelocityPerGearUnderwater =
                    ExpectedVelocityPerGearUnderwaterField.GetValue(engine) as float[];

                if (forwardPresets == null || forwardPresets.Length < FastForwardGearCount)
                {
                    WarnOnce(engine, "moins de " + FastForwardGearCount + " crans avant, patch vitesse ignore.");
                    return;
                }

                EngineSpeedPatchData data;
                if (!OriginalData.TryGetValue(engine, out data))
                {
                    data = EngineSpeedPatchData.Capture(
                        forwardPresets,
                        expectedVelocityPerGear,
                        expectedVelocityPerGearUnderwater,
                        BasePowerField,
                        FuelConsumptionField
                    );
                    OriginalData.Add(engine, data);
                }

                float speedFactor = GetEffectiveFastSpeedFactor();
                float fuelFactor = GetEffectiveFastFuelFactor(speedFactor);

                // DonJ : je garde une copie des valeurs originales, puis je recalcule depuis ces bases.
                // Comme ca le slider F10 peut monter/descendre sans empiler les multiplicateurs.
                ApplyTopGearBasePower(forwardPresets, data.ForwardBasePower, speedFactor);
                ApplyTopGearFuelConsumption(forwardPresets, data.ForwardFuelConsumption, fuelFactor);
                ApplyTopGearFloatArray(expectedVelocityPerGear, data.ExpectedVelocityPerGear, speedFactor);
                ApplyTopGearFloatArray(expectedVelocityPerGearUnderwater, data.ExpectedVelocityPerGearUnderwater, speedFactor);

                Debug.Log("[LongSubmerged10x] Fast speed patch applied after " + reason + " with x" + speedFactor + ".");
            }
            catch (Exception ex)
            {
                Debug.LogException(ex);
            }
        }

        private static bool FieldsReady()
        {
            return ForwardPresetsField != null
                && ExpectedVelocityPerGearField != null
                && ExpectedVelocityPerGearUnderwaterField != null
                && BasePowerField != null
                && FuelConsumptionField != null;
        }

        private static void ApplyTopGearBasePower(Array forwardPresets, float[] originalBasePower, float speedFactor)
        {
            if (forwardPresets == null || originalBasePower == null)
                return;

            int patchCount = Math.Min(FastForwardGearCount, Math.Min(forwardPresets.Length, originalBasePower.Length));
            int firstPatchedGear = forwardPresets.Length - patchCount;

            for (int index = firstPatchedGear; index < forwardPresets.Length; index++)
            {
                object preset = forwardPresets.GetValue(index);
                if (preset == null)
                    continue;

                BasePowerField.SetValue(preset, originalBasePower[index] * speedFactor);
                forwardPresets.SetValue(preset, index);
            }
        }

        private static void ApplyTopGearFuelConsumption(Array forwardPresets, float[] originalFuelConsumption, float fuelFactor)
        {
            if (forwardPresets == null || originalFuelConsumption == null)
                return;

            int patchCount = Math.Min(FastForwardGearCount, Math.Min(forwardPresets.Length, originalFuelConsumption.Length));
            int firstPatchedGear = forwardPresets.Length - patchCount;

            for (int index = firstPatchedGear; index < forwardPresets.Length; index++)
            {
                object preset = forwardPresets.GetValue(index);
                if (preset == null)
                    continue;

                FuelConsumptionField.SetValue(preset, originalFuelConsumption[index] * fuelFactor);
                forwardPresets.SetValue(preset, index);
            }
        }

        private static void ApplyTopGearFloatArray(float[] target, float[] original, float speedFactor)
        {
            if (target == null || original == null)
                return;

            int patchCount = Math.Min(FastForwardGearCount, Math.Min(target.Length, original.Length));
            int firstPatchedGear = target.Length - patchCount;

            for (int index = firstPatchedGear; index < target.Length; index++)
                target[index] = original[index] * speedFactor;
        }

        private static void PatchShipVelocityCap(PlayerShip ship, string reason, bool verboseLog)
        {
            if (ship == null || ship.Blueprint == null || ship.Blueprint.Velocity == null)
                return;

            ShipRuntimePatchData data;
            if (!ShipRuntimeData.TryGetValue(ship, out data))
            {
                float originalVelocity = ship.Blueprint.Velocity;
                if (!IsFinite(originalVelocity) || originalVelocity <= 0f)
                    return;

                // DonJ : on cree toujours le modifier, meme si le vieux XLSX a deja mis 45 km/h.
                // Ca permet au slider x1 de revenir a la vitesse normale estimee au lieu de rester a 45.
                Modifier modifier = ship.Blueprint.Velocity.AddDeltaModifier(RuntimeVelocityModifierName, false);
                data = new ShipRuntimePatchData(originalVelocity, modifier);
                ShipRuntimeData.Add(ship, data);
            }

            if (data.VelocityModifier == null)
                return;

            float baseVelocity = GetRuntimeBaseVelocity(data.OriginalVelocity);
            float effectiveSpeedFactor = GetEffectiveFastSpeedFactor();
            float desiredMaxSpeed = baseVelocity * effectiveSpeedFactor;
            float desiredDelta = desiredMaxSpeed - data.OriginalVelocity;

            if (Math.Abs(data.VelocityModifier.Value - desiredDelta) > 0.001f)
                data.VelocityModifier.Value = desiredDelta;

            if (verboseLog)
            {
                Debug.Log(
                    "[LongSubmerged10x] Player ship speed cap patched after "
                    + reason
                    + ": base "
                    + baseVelocity
                    + " km/h, x"
                    + effectiveSpeedFactor
                    + " -> "
                    + desiredMaxSpeed
                    + " km/h."
                );
            }
        }

        private static float GetRuntimeBaseVelocity(float originalVelocity)
        {
            if (!IsFinite(originalVelocity) || originalVelocity <= 0f)
                return 1f;

            // DonJ : compatibilite avec les builds 1.4.7 deja installes.
            // Ces builds ecrivaient 45 km/h dans Entities.xlsx, ce qui empechait x1 d'etre vanilla.
            // Les types joueur vanilla VIIC/IXC utilises par le mod sont a 32.8 km/h dans les Data Sheets.
            if (Math.Abs(originalVelocity - LegacyDataSheetPlayerSubmarineMaxSpeed) <= 0.05f)
                return VanillaPlayerSubmarineMaxSpeedFallback;

            return originalVelocity;
        }

        private static void ApplyPropellerSpeedMultiplier(PlayerShip ship, string reason, bool verboseLog)
        {
            if (ship == null)
                return;

            Propeller[] propellers = ShipPropellersField == null
                ? ship.Propellers
                : ShipPropellersField.GetValue(ship) as Propeller[];

            if (propellers == null || propellers.Length == 0)
                return;

            bool fastForwardGear = IsActiveEngineInFastForwardGear(ship);
            float appliedFactor = fastForwardGear ? GetEffectiveFastSpeedFactor() : 1f;
            int changedCount = 0;

            foreach (Propeller propeller in propellers)
            {
                if (propeller == null)
                    continue;

                PropellerPatchData data;
                if (!PropellerRuntimeData.TryGetValue(propeller, out data))
                {
                    data = new PropellerPatchData(propeller.PowerMultiplier);
                    PropellerRuntimeData.Add(propeller, data);
                }

                float desiredMultiplier = data.OriginalPowerMultiplier * appliedFactor;

                if (Math.Abs(propeller.PowerMultiplier - desiredMultiplier) > 0.001f)
                {
                    propeller.PowerMultiplier = desiredMultiplier;
                    changedCount++;
                }
            }

            if (verboseLog && changedCount > 0)
            {
                Debug.Log(
                    "[LongSubmerged10x] Propeller multiplier "
                    + appliedFactor
                    + " applied after "
                    + reason
                    + "."
                );
            }
        }

        private static bool IsActiveEngineInFastForwardGear(PlayerShip ship)
        {
            PlayerShipEngine engine = ship.ActiveEngine;
            if (engine == null || engine.GearIndex <= 0 || ForwardPresetsField == null)
                return false;

            Array forwardPresets = ForwardPresetsField.GetValue(engine) as Array;
            if (forwardPresets == null || forwardPresets.Length < FastForwardGearCount)
                return false;

            int firstFastGearIndex = forwardPresets.Length - FastForwardGearCount + 1;
            return engine.GearIndex >= firstFastGearIndex;
        }

        private static float GetEffectiveFastSpeedFactor()
        {
            if (!LongSubmergedRuntimeSettings.SuperSpeed)
                return 1f;

            return LongSubmergedRuntimeSettings.ClampSpeedFactor(LongSubmergedRuntimeSettings.SpeedFactor);
        }

        private static float GetEffectiveFastFuelFactor(float speedFactor)
        {
            if (!LongSubmergedRuntimeSettings.SuperSpeed || speedFactor <= 1.0001f)
                return 1f;

            float referenceSpeedFactor = Math.Max(1.0001f, FastSpeedFactor);
            float normalized = (speedFactor - 1f) / (referenceSpeedFactor - 1f);
            return Math.Max(1f, 1f + normalized * (FastSpeedFuelFactor - 1f));
        }

        private static bool IsFinite(float value)
        {
            return !float.IsNaN(value) && !float.IsInfinity(value);
        }

        private static void WarnOnce(PlayerShipEngine engine, string message)
        {
            int instanceId = engine.GetInstanceID();
            if (WarnedEngines.Add(instanceId))
                Debug.LogWarning("[LongSubmerged10x] " + message);
        }
    }

    internal sealed class EngineSpeedPatchData
    {
        public readonly float[] ForwardBasePower;
        public readonly float[] ForwardFuelConsumption;
        public readonly float[] ExpectedVelocityPerGear;
        public readonly float[] ExpectedVelocityPerGearUnderwater;

        private EngineSpeedPatchData(
            float[] forwardBasePower,
            float[] forwardFuelConsumption,
            float[] expectedVelocityPerGear,
            float[] expectedVelocityPerGearUnderwater)
        {
            ForwardBasePower = forwardBasePower;
            ForwardFuelConsumption = forwardFuelConsumption;
            ExpectedVelocityPerGear = expectedVelocityPerGear;
            ExpectedVelocityPerGearUnderwater = expectedVelocityPerGearUnderwater;
        }

        public static EngineSpeedPatchData Capture(
            Array forwardPresets,
            float[] expectedVelocityPerGear,
            float[] expectedVelocityPerGearUnderwater,
            FieldInfo basePowerField,
            FieldInfo fuelConsumptionField)
        {
            float[] basePower = new float[forwardPresets.Length];
            float[] fuelConsumption = new float[forwardPresets.Length];

            for (int index = 0; index < forwardPresets.Length; index++)
            {
                object preset = forwardPresets.GetValue(index);
                if (preset == null)
                    continue;

                object rawValue = basePowerField.GetValue(preset);
                if (rawValue is float)
                    basePower[index] = (float)rawValue;

                object rawFuelConsumption = fuelConsumptionField.GetValue(preset);
                if (rawFuelConsumption is float)
                    fuelConsumption[index] = (float)rawFuelConsumption;
            }

            return new EngineSpeedPatchData(
                basePower,
                fuelConsumption,
                CloneFloatArray(expectedVelocityPerGear),
                CloneFloatArray(expectedVelocityPerGearUnderwater)
            );
        }

        private static float[] CloneFloatArray(float[] source)
        {
            if (source == null)
                return null;

            float[] clone = new float[source.Length];
            Array.Copy(source, clone, source.Length);
            return clone;
        }
    }

    internal sealed class ShipRuntimePatchData
    {
        public readonly float OriginalVelocity;
        public readonly Modifier VelocityModifier;

        public ShipRuntimePatchData(float originalVelocity, Modifier velocityModifier)
        {
            OriginalVelocity = originalVelocity;
            VelocityModifier = velocityModifier;
        }
    }

    internal sealed class PropellerPatchData
    {
        public readonly float OriginalPowerMultiplier;

        public PropellerPatchData(float originalPowerMultiplier)
        {
            OriginalPowerMultiplier = originalPowerMultiplier;
        }
    }

    // DonJ : hooks Harmony courts et delegues. Chaque hook appelle une methode robuste du runtime,
    // ce qui limite le risque de casser UBOAT si un objet arrive partiellement initialise.
    [HarmonyPatch(typeof(PlayerShip), "Awake")]
    internal static class PlayerShipAwakePatch
    {
        private static void Postfix(PlayerShip __instance)
        {
            LongSubmergedRuntimeApplier.ApplyPlayerShip(__instance, "PlayerShip.Awake");
        }
    }

    [HarmonyPatch(typeof(PlayerShip), "OnAfterDeserialize")]
    internal static class PlayerShipOnAfterDeserializePatch
    {
        private static void Postfix(PlayerShip __instance)
        {
            LongSubmergedRuntimeApplier.ApplyPlayerShip(__instance, "PlayerShip.OnAfterDeserialize");
        }
    }

    [HarmonyPatch(typeof(PlayerShip), "Update")]
    internal static class PlayerShipUpdatePatch
    {
        private static void Postfix(PlayerShip __instance)
        {
            // DonJ: I keep the battery full after the submarine frame.
            // Resource.UpdateAmount is also guarded below, but only for PlayerShip.Energy.
            LongSubmergedRuntimeApplier.ApplyBatteryResource(__instance, "PlayerShip.Update");

            // DonJ : la vitesse est un etat runtime du bateau et du cran moteur actif.
            // On la remet ici pour que le slider F10 et les changements de cran soient visibles immediatement.
            EngineFastSpeedPatcher.UpdatePlayerShipRuntime(__instance, "PlayerShip.Update");

            DeepDiveRuntimePatcher.UpdatePlayerShipRuntime(__instance, "PlayerShip.Update");

            SuperStealthRuntimePatcher.ApplyPlayerShip(__instance, "PlayerShip.Update");
        }
    }

    [HarmonyPatch(typeof(Resource), "UpdateAmount")]
    internal static class ResourceUpdateAmountBatteryPatch
    {
        private static bool Prefix(Resource __instance)
        {
            if (LongSubmergedRuntimeApplier.TryFreezeInfiniteBatteryResource(
                __instance,
                "Resource.UpdateAmount.Prefix"
            ))
            {
                return false;
            }

            return true;
        }
    }

    [HarmonyPatch(typeof(ResourceGUI), "GetTooltipContents")]
    internal static class ResourceGuiGetTooltipContentsPatch
    {
        private static bool Prefix(ResourceGUI __instance, ref string __result)
        {
            Resource resource = LongSubmergedRuntimeApplier.GetResourceFromGui(__instance);
            if (!LongSubmergedRuntimeApplier.ShouldSuppressBatteryDepletionUi(resource, "ResourceGUI.GetTooltipContents"))
                return true;

            __result = LongSubmergedRuntimeApplier.BuildInfiniteBatteryTooltip(resource);
            return false;
        }
    }

    [HarmonyPatch(typeof(ResourceGUI), "UpdateDisplayedValue")]
    internal static class ResourceGuiUpdateDisplayedValuePatch
    {
        private static void Prefix(ResourceGUI __instance)
        {
            LongSubmergedRuntimeApplier.TryMaintainBatteryResource(
                LongSubmergedRuntimeApplier.GetResourceFromGui(__instance),
                "ResourceGUI.UpdateDisplayedValue"
            );
        }
    }

    [HarmonyPatch(typeof(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting), "Awake")]
    internal static class InteriorLightingPlayerShipInteriorLightingAwakePatch
    {
        private static void Postfix(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting __instance)
        {
            InteriorLightingColorPatcher.ApplyInteriorLighting(
                __instance,
                "PlayerShipInteriorLighting.Awake",
                false
            );
        }
    }

    [HarmonyPatch(typeof(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting), "Start")]
    internal static class InteriorLightingPlayerShipInteriorLightingStartPatch
    {
        private static void Prefix(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting __instance)
        {
            InteriorLightingColorPatcher.ApplyInteriorLighting(
                __instance,
                "PlayerShipInteriorLighting.Start",
                false
            );
        }
    }

    [HarmonyPatch(typeof(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting), "ApplyLightControllersPresets")]
    internal static class InteriorLightingPlayerShipInteriorLightingApplyPresetsPatch
    {
        private static void Prefix(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting __instance)
        {
            InteriorLightingColorPatcher.ApplyInteriorLighting(
                __instance,
                "PlayerShipInteriorLighting.ApplyLightControllersPresets",
                false
            );
        }
    }

    [HarmonyPatch(typeof(UBOAT.Game.Scene.Effects.LightController), "UpdatePresets", new Type[] { typeof(float[]), typeof(float[]) })]
    internal static class InteriorLightingLightControllerUpdatePresetsPatch
    {
        private static void Prefix(UBOAT.Game.Scene.Effects.LightController __instance)
        {
            InteriorLightingColorPatcher.ApplyLightController(__instance, "LightController.UpdatePresets");
        }
    }

    [HarmonyPatch(typeof(UBOAT.Game.Scene.Effects.FillLight), "UpdatePresets", new Type[] { typeof(float[]) })]
    internal static class InteriorLightingFillLightUpdatePresetsPatch
    {
        private static void Prefix(UBOAT.Game.Scene.Effects.FillLight __instance)
        {
            InteriorLightingColorPatcher.ApplyFillLight(__instance, "FillLight.UpdatePresets");
        }
    }

    [HarmonyPatch(typeof(DepletingResourceNotification), "DoUpdate")]
    internal static class DepletingResourceNotificationDoUpdatePatch
    {
        private static bool Prefix(DepletingResourceNotification __instance, ref float __result)
        {
            Resource resource = LongSubmergedRuntimeApplier.GetResourceFromDepletingNotification(__instance);
            if (!LongSubmergedRuntimeApplier.ShouldSuppressBatteryDepletionUi(resource, "DepletingResourceNotification.DoUpdate"))
                return true;

            __result = 5f;
            return false;
        }
    }

    [HarmonyPatch]
    internal static class DeepDivePlayerShipTargetDepthSetterPatch
    {
        private static MethodBase TargetMethod()
        {
            return DeepDiveRuntimePatcher.FindPlayerShipSetTargetDepthMethod();
        }

        private static void Prefix(ref float __0)
        {
            DeepDiveRuntimePatcher.ScaleTargetDepthCommand(ref __0, "PlayerShip.SetTargetDepth");
        }
    }

    [HarmonyPatch]
    internal static class DeepDiveHullCrushControllerDoUpdatePatch
    {
        private static MethodBase TargetMethod()
        {
            return DeepDiveRuntimePatcher.FindMethodOnKnownType("HullCrushController", "DoUpdate");
        }

        private static bool Prefix(object __instance, ref float __result)
        {
            return DeepDiveRuntimePatcher.ShouldRunHullCrushDoUpdate(__instance, ref __result, "HullCrushController.DoUpdate");
        }
    }

    [HarmonyPatch]
    internal static class DeepDivePlayerShipUpdateStressAndDisciplineGainPatch
    {
        private static MethodBase TargetMethod()
        {
            return DeepDiveRuntimePatcher.FindPlayerShipUpdateStressAndDisciplineGainMethod();
        }

        private static void Postfix(PlayerShip __instance)
        {
            DeepDiveRuntimePatcher.ApplyDepthStressModifier(__instance, "PlayerShip.UpdateStressAndDisciplineGain");
        }
    }

    [HarmonyPatch(typeof(PlayerShip), "ValidateTargetVelocity")]
    internal static class PlayerShipValidateTargetVelocityPatch
    {
        private static void Postfix(PlayerShip __instance)
        {
            EngineFastSpeedPatcher.UpdatePlayerShipRuntime(__instance, "PlayerShip.ValidateTargetVelocity");
        }
    }

    [HarmonyPatch(typeof(PlayerShip), "ValidateOxygenBreathModifier")]
    internal static class PlayerShipValidateOxygenBreathModifierPatch
    {
        private static void Postfix(PlayerShip __instance)
        {
            LongSubmergedRuntimeApplier.ApplyOxygenBreathModifier(__instance, "PlayerShip.ValidateOxygenBreathModifier");
        }
    }

    [HarmonyPatch(typeof(PlayerShip), "SavesManagerOnLoaded")]
    internal static class PlayerShipSavesManagerOnLoadedPatch
    {
        private static void Postfix(PlayerShip __instance, Queue<Action> __0)
        {
            LongSubmergedRuntimeApplier.ApplyPlayerShip(__instance, "SavesManagerOnLoaded");
        }
    }

    [HarmonyPatch(typeof(PlayerShip), "Crew_Added")]
    internal static class PlayerShipCrewAddedPatch
    {
        private static void Postfix(PlayerShip __instance)
        {
            LongSubmergedRuntimeApplier.ApplyPlayerShip(__instance, "Crew_Added");
        }
    }

    [HarmonyPatch(typeof(PlayerShip), "Crew_Removed")]
    internal static class PlayerShipCrewRemovedPatch
    {
        private static void Postfix(PlayerShip __instance)
        {
            LongSubmergedRuntimeApplier.ApplyPlayerShip(__instance, "Crew_Removed");
        }
    }

    [HarmonyPatch(typeof(PlayerShipEngine), "Awake")]
    internal static class PlayerShipEngineAwakePatch
    {
        private static void Postfix(PlayerShipEngine __instance)
        {
            LongSubmergedRuntimeApplier.ApplyBatteryObject(__instance, "PlayerShipEngine.Awake");
            EngineFastSpeedPatcher.PatchEngine(__instance, "PlayerShipEngine.Awake");
        }
    }

    [HarmonyPatch(typeof(PlayerShipEngine), "OnAfterDeserialize")]
    internal static class PlayerShipEngineOnAfterDeserializePatch
    {
        private static void Postfix(PlayerShipEngine __instance)
        {
            LongSubmergedRuntimeApplier.ApplyBatteryObject(__instance, "PlayerShipEngine.OnAfterDeserialize");
            EngineFastSpeedPatcher.PatchEngine(__instance, "PlayerShipEngine.OnAfterDeserialize");
        }
    }

    [HarmonyPatch(typeof(PlayerShipEngine), "SavesManagerOnLoaded")]
    internal static class PlayerShipEngineSavesManagerOnLoadedPatch
    {
        private static void Postfix(PlayerShipEngine __instance, Queue<Action> __0)
        {
            LongSubmergedRuntimeApplier.ApplyBatteryObject(__instance, "PlayerShipEngine.SavesManagerOnLoaded");
            EngineFastSpeedPatcher.PatchEngine(__instance, "PlayerShipEngine.SavesManagerOnLoaded");
        }
    }

    [HarmonyPatch(typeof(AccumulatorsUpgrade), "Start")]
    internal static class AccumulatorsUpgradeStartPatch
    {
        private static void Postfix(AccumulatorsUpgrade __instance)
        {
            LongSubmergedRuntimeApplier.ApplyBatteryObject(__instance, "AccumulatorsUpgrade.Start");
        }
    }

    [HarmonyPatch(typeof(DivingPlanesStation), "Awake")]
    internal static class DivingPlanesStationAwakePatch
    {
        private static void Postfix(DivingPlanesStation __instance)
        {
            LongSubmergedRuntimeApplier.ApplyBatteryObject(__instance, "DivingPlanesStation.Awake");
            DeepDiveRuntimePatcher.ApplyDepthObject(__instance, "DivingPlanesStation.Awake");
        }
    }

    [HarmonyPatch(typeof(DivingPlanesStation), "UpdateModifiers")]
    internal static class DivingPlanesStationUpdateModifiersPatch
    {
        private static void Postfix(DivingPlanesStation __instance)
        {
            LongSubmergedRuntimeApplier.ApplyBatteryObject(__instance, "DivingPlanesStation.UpdateModifiers");
            DeepDiveRuntimePatcher.ApplyDepthObject(__instance, "DivingPlanesStation.UpdateModifiers");
        }
    }

    [HarmonyPatch(typeof(Gyrocompass), "ApplyModifiers")]
    internal static class GyrocompassApplyModifiersPatch
    {
        private static void Postfix(Gyrocompass __instance)
        {
            LongSubmergedRuntimeApplier.ApplyBatteryObject(__instance, "Gyrocompass.ApplyModifiers");
        }
    }

    [HarmonyPatch(typeof(TrimPump), "OnEnable")]
    internal static class TrimPumpOnEnablePatch
    {
        private static void Postfix(TrimPump __instance)
        {
            LongSubmergedRuntimeApplier.ApplyBatteryObject(__instance, "TrimPump.OnEnable");
        }
    }

    [HarmonyPatch(typeof(StoredTorpedo), "Start")]
    internal static class StoredTorpedoStartPatch
    {
        private static void Postfix(StoredTorpedo __instance)
        {
            LongSubmergedRuntimeApplier.ApplyStoredTorpedo(__instance, "StoredTorpedo.Start");
        }
    }

    [HarmonyPatch(typeof(StoredTorpedo), "ApplyWarmUpModifier")]
    internal static class StoredTorpedoApplyWarmUpModifierPatch
    {
        private static void Postfix(StoredTorpedo __instance)
        {
            LongSubmergedRuntimeApplier.ApplyStoredTorpedo(__instance, "StoredTorpedo.ApplyWarmUpModifier");
        }
    }

    [HarmonyPatch(typeof(Torpedo), "Awake")]
    internal static class TorpedoAwakePatch
    {
        private static void Postfix(Torpedo __instance)
        {
            LongSubmergedRuntimeApplier.ApplyLaunchedTorpedo(__instance, "Torpedo.Awake");
        }
    }

    [HarmonyPatch(typeof(Torpedo), "FixedUpdate")]
    internal static class TorpedoFixedUpdatePatch
    {
        private static void Prefix(Torpedo __instance)
        {
            LongSubmergedRuntimeApplier.ApplyLaunchedTorpedo(__instance, "Torpedo.FixedUpdate");
        }
    }

    [HarmonyPatch(typeof(Torpedo), "Detonate")]
    internal static class TorpedoDetonatePatch
    {
        private static void Prefix(Torpedo __instance)
        {
            LongSubmergedRuntimeApplier.ApplyLaunchedTorpedo(__instance, "Torpedo.Detonate");
        }
    }

    [HarmonyPatch(typeof(Entity), "UpdateDetectability")]
    internal static class SuperStealthEntityUpdateDetectabilityPatch
    {
        private static void Postfix(Entity __instance)
        {
            SuperStealthRuntimePatcher.ApplyEntity(__instance, "Entity.UpdateDetectability");
        }
    }

    [HarmonyPatch(typeof(AirCompressor), "OnEnable")]
    internal static class SuperStealthAirCompressorOnEnablePatch
    {
        private static void Postfix(AirCompressor __instance)
        {
            SuperStealthRuntimePatcher.ApplyEquipment(__instance, "AirCompressor.OnEnable");
        }
    }

    [HarmonyPatch(typeof(AirCompressor), "OnDisable")]
    internal static class SuperStealthAirCompressorOnDisablePatch
    {
        private static void Postfix(AirCompressor __instance)
        {
            SuperStealthRuntimePatcher.ApplyEquipment(__instance, "AirCompressor.OnDisable");
        }
    }

    [HarmonyPatch(typeof(Ventilation), "OnEnable")]
    internal static class SuperStealthVentilationOnEnablePatch
    {
        private static void Postfix(Ventilation __instance)
        {
            SuperStealthRuntimePatcher.ApplyEquipment(__instance, "Ventilation.OnEnable");
        }
    }

    [HarmonyPatch(typeof(Ventilation), "OnDisable")]
    internal static class SuperStealthVentilationOnDisablePatch
    {
        private static void Postfix(Ventilation __instance)
        {
            SuperStealthRuntimePatcher.ApplyEquipment(__instance, "Ventilation.OnDisable");
        }
    }

    [HarmonyPatch(typeof(Propeller), "set_Power")]
    internal static class SuperStealthPropellerPowerPatch
    {
        private static void Postfix(Propeller __instance)
        {
            SuperStealthRuntimePatcher.ApplyEquipment(__instance, "Propeller.set_Power");
        }
    }

    [HarmonyPatch(typeof(Propeller), "set_PowerMultiplier")]
    internal static class SuperStealthPropellerPowerMultiplierPatch
    {
        private static void Postfix(Propeller __instance)
        {
            SuperStealthRuntimePatcher.ApplyEquipment(__instance, "Propeller.set_PowerMultiplier");
        }
    }

    [HarmonyPatch(typeof(Snorkel), "Update")]
    internal static class SuperStealthSnorkelUpdatePatch
    {
        private static void Postfix(Snorkel __instance)
        {
            SuperStealthRuntimePatcher.ApplyEquipment(__instance, "Snorkel.Update");
        }
    }

    [HarmonyPatch(typeof(Periscope), "Update")]
    internal static class SuperStealthPeriscopeUpdatePatch
    {
        private static void Postfix(Periscope __instance)
        {
            SuperStealthRuntimePatcher.ApplyEquipment(__instance, "Periscope.Update");
        }
    }

    [HarmonyPatch]
    internal static class HeavyArmorHullAddDamagePatch
    {
        private static MethodBase TargetMethod()
        {
            return AccessTools.Method(typeof(Hull), "AddDamage", HeavyArmorRuntimePatcher.AddDamageParameterTypes);
        }

        private static void Prefix(
            Hull __instance,
            ref float damage,
            ref float flawProbabilityFactor,
            ref float fireChance,
            out HeavyArmorDamageScaleState __state
        )
        {
            __state = HeavyArmorRuntimePatcher.TryScalePlayerEquipmentDamage(
                __instance,
                ref damage,
                ref flawProbabilityFactor,
                ref fireChance
            );
            HeavyArmorRuntimePatcher.BeginDamageScaleScope(__state);
        }

        private static void Postfix(HeavyArmorDamageScaleState __state, ref float __result)
        {
            HeavyArmorRuntimePatcher.RestoreComponentDistributionBudget(__state, ref __result);
        }

        private static void Finalizer(HeavyArmorDamageScaleState __state)
        {
            HeavyArmorRuntimePatcher.EndDamageScaleScope(__state);
        }
    }

    [HarmonyPatch]
    internal static class HeavyArmorEquipmentAddDamagePatch
    {
        private static MethodBase TargetMethod()
        {
            return AccessTools.Method(typeof(Equipment), "AddDamage", HeavyArmorRuntimePatcher.AddDamageParameterTypes);
        }

        private static void Prefix(
            Equipment __instance,
            ref float damage,
            ref float flawProbabilityFactor,
            ref float fireChance,
            out HeavyArmorDamageScaleState __state
        )
        {
            __state = HeavyArmorRuntimePatcher.TryScalePlayerEquipmentDamage(
                __instance,
                ref damage,
                ref flawProbabilityFactor,
                ref fireChance
            );
            HeavyArmorRuntimePatcher.BeginDamageScaleScope(__state);
        }

        private static void Postfix(HeavyArmorDamageScaleState __state, ref float __result)
        {
            HeavyArmorRuntimePatcher.RestoreComponentDistributionBudget(__state, ref __result);
        }

        private static void Finalizer(HeavyArmorDamageScaleState __state)
        {
            HeavyArmorRuntimePatcher.EndDamageScaleScope(__state);
        }
    }

    [HarmonyPatch]
    internal static class HeavyArmorEquipmentAddWaterDamagePatch
    {
        private static MethodBase TargetMethod()
        {
            return AccessTools.Method(typeof(Equipment), "AddWaterDamage", HeavyArmorRuntimePatcher.AddWaterDamageParameterTypes);
        }

        private static void Prefix(Equipment __instance, ref float damage, out HeavyArmorDamageScaleState __state)
        {
            __state = HeavyArmorRuntimePatcher.TryScalePlayerWaterDamage(__instance, ref damage);
            HeavyArmorRuntimePatcher.BeginDamageScaleScope(__state);
        }

        private static void Finalizer(HeavyArmorDamageScaleState __state)
        {
            HeavyArmorRuntimePatcher.EndDamageScaleScope(__state);
        }
    }

    [HarmonyPatch]
    internal static class HeavyArmorPlayableCharacterAddDamagePatch
    {
        private static MethodBase TargetMethod()
        {
            return AccessTools.Method(typeof(PlayableCharacter), "AddDamage", HeavyArmorRuntimePatcher.AddDamageParameterTypes);
        }

        private static void Prefix(PlayableCharacter __instance, ref float damage)
        {
            HeavyArmorRuntimePatcher.ScalePlayerCharacterDamage(__instance, ref damage);
        }
    }

    [HarmonyPatch]
    internal static class HeavyArmorDamageUtilityDoApplyDamagePatch
    {
        private static MethodBase TargetMethod()
        {
            return AccessTools.Method(
                typeof(DamageUtility),
                "DoApplyDamage",
                HeavyArmorRuntimePatcher.DamageUtilityDoApplyDamageParameterTypes
            );
        }

        private static void Prefix(Ship target, ref float crewDamage)
        {
            HeavyArmorRuntimePatcher.ScalePlayerCrewDamage(target, ref crewDamage);
        }
    }

    [HarmonyPatch]
    internal static class HeavyArmorDamageUtilityApplyDamageToComponentsPatch
    {
        private static MethodBase TargetMethod()
        {
            return AccessTools.Method(
                typeof(DamageUtility),
                "ApplyDamageToComponents",
                HeavyArmorRuntimePatcher.DamageUtilityApplyDamageToComponentsParameterTypes
            );
        }

        private static void Prefix(out bool __state)
        {
            __state = HeavyArmorRuntimePatcher.BeginComponentDamageDistributionScope();
        }

        private static void Finalizer(bool __state)
        {
            HeavyArmorRuntimePatcher.EndComponentDamageDistributionScope(__state);
        }
    }

    [HarmonyPatch]
    internal static class HeavyArmorApplyWaterDamageToPlayerShipDoDamageTickPatch
    {
        private static MethodBase TargetMethod()
        {
            return AccessTools.Method(typeof(ApplyWaterDamageToPlayerShip), "DoDamageTick", new Type[] { });
        }

        private static bool Prefix(ApplyWaterDamageToPlayerShip __instance, out bool __state)
        {
            __state = HeavyArmorRuntimePatcher.BeginPressureWaterDamageScope();
            return !DeepDiveRuntimePatcher.ShouldSkipPressureWaterDamageTick(
                __instance,
                "ApplyWaterDamageToPlayerShip.DoDamageTick"
            );
        }

        private static void Finalizer(bool __state)
        {
            HeavyArmorRuntimePatcher.EndPressureWaterDamageScope(__state);
        }
    }

    [HarmonyPatch]
    internal static class HeavyArmorHullEffectsRendererApplyImpactPatch
    {
        private static MethodBase TargetMethod()
        {
            return AccessTools.Method(
                typeof(HullEffectsRenderer),
                "ApplyImpact",
                HeavyArmorRuntimePatcher.HullEffectsApplyImpactParameterTypes
            );
        }

        private static void Prefix(HullEffectsRenderer __instance, ref float intensity)
        {
            HeavyArmorRuntimePatcher.ScalePlayerHullImpact(__instance, ref intensity);
        }
    }
}
'''

    source = source.replace("__FAST_SPEED_FACTOR__", format_csharp_float(args.fast_speed_factor))
    source = source.replace("__FAST_SPEED_FUEL_FACTOR__", format_csharp_float(args.fast_speed_fuel_factor))
    source = source.replace("__PLAYER_SUBMARINE_MAX_SPEED__", format_csharp_float(args.player_submarine_max_speed))
    source = source.replace("__FAST_FORWARD_GEAR_COUNT__", str(args.fast_speed_top_gears))
    source = source.replace("__OXYGEN_CONSUMPTION_FACTOR__", format_csharp_float(args.oxygen_consumption_factor))
    source = source.replace("__BATTERY_CAPACITY_FACTOR__", format_csharp_float(args.battery_capacity_factor))
    source = source.replace("__ENERGY_USAGE_FACTOR__", format_csharp_float(args.energy_usage_factor))
    source = source.replace("__TORPEDO_DAMAGE_FACTOR__", format_csharp_float(args.torpedo_damage_factor))
    source = source.replace("__TORPEDO_CREW_DAMAGE_FACTOR__", format_csharp_float(args.torpedo_crew_damage_factor))
    source = source.replace("__TORPEDO_EXPLOSION_RADIUS_FACTOR__", format_csharp_float(args.torpedo_explosion_radius_factor))
    source = source.replace("__TORPEDO_EXPLOSION_INTENSITY_FACTOR__", format_csharp_float(args.torpedo_explosion_intensity_factor))
    source = source.replace("__PERFECT_TORPEDO_RELIABILITY__", "true" if args.perfect_torpedo_reliability else "false")
    source = source.replace("__DEFAULT_MEGA_TORPEDOES__", "true" if args.mega_torpedoes else "false")
    source = source.replace("__MOD_VERSION__", MOD_VERSION)

    source_path.write_text(source, encoding="utf-8")
    report.file(source_path)
    report.note(
        "Patch runtime Harmony ajoute : recalcul de OxygenBreathModifier apres Awake, "
        "chargement de sauvegarde et changement d'equipage, plus vitesse rapide, "
        "plafond runtime, propulseurs, carburant rapide, menu F10, renforts amis, toggles mega, blindage lourd et super discretion."
    )
    report.note("Mega Batterie : la case F10 rend la batterie infinie et maintient la ressource au maximum.")
    report.note("Mega Oxygene : profil par defaut calibre pour environ 90 jours.")
    report.note("Menu F10 : sliders runtime bornes par profil et bouton Par defaut.")
    report.note("SuperVitesse : runtime F10 reglable 1-20 sur les deux crans rapides avant.")
    report.note("Mega torpilles : runtime F10 reglable 1-10, degats defaut x10, effets visuels bornes x3, aucune ligne torpille XLSX ecrasee.")
    report.note("Mega Sonar : runtime F10 reglable 1-10, defaut x3, applique aux portees hydrophone.")
    report.note("Blindage lourd : case F10 desactivee par defaut, activable manuellement, degats joueur divises par 3.")
    report.note("DeepDive : case F10 activee par defaut, ordres de profondeur x2 au-dessus de 10 m, stress profondeur calcule sur profondeur /2, max operationnel 600 m, crush a 700 m.")
    report.note("Blindage lourd : migration settings v16 sur OFF une seule fois ; le blindage ne masque pas le crush DeepDive.")
    report.note("Super discrétion : case F10 desactivee par defaut, bruit et detectabilite joueur divisibles par 3 sans rendre le sous-marin invisible.")
    report.note("Eclairage interieur : case F10 Couleurs eclairage activee par defaut ; menus Alarm/SilentRun avec palette predefinie, defaut orange ambre/vert sous-marin, decochee restaure les couleurs vanilla.")
    report.note("Appeler renforts : bouton F10 tente les patrouilles vanilla amies si disponibles, puis cree des U-boats amis en fallback manuel plus proche, a portee visuelle raisonnable.")


def write_readme(mod_dir: Path, args: argparse.Namespace, report: PatchReport | None = None) -> None:
    lines = [
        f"{MOD_DISPLAY_NAME} v{MOD_VERSION}",
        "",
        "Paramètres utilisés :",
        "- Oxygene long : applique au runtime sur le drain negatif de respiration",
        "- Recharge surface : vanilla, aucune capacite Air/Oxygen/Atmosphere XLSX modifiee",
        f"- Discipline/fatigue sous l'eau : divisé par {args.discipline_factor:g}",
        f"- Batterie / Accumulators : x{args.battery_capacity_factor:g}",
        f"- EnergyUsage consommateurs hors ventilation/compresseurs : x{args.energy_usage_factor:g} dans les datasheets",
        "- Mega Batterie runtime : case F10 active = batterie infinie, pompe incluse",
        "- EnergyUsage recharge/production batterie : vanilla",
        f"- Deux derniers crans avant : vitesse/propulsion x{args.fast_speed_factor:g}",
        f"- Deux derniers crans avant : carburant x{args.fast_speed_fuel_factor:g}",
        f"- Vitesse max sous-marin joueur : {args.player_submarine_max_speed:g} km/h",
        "- Menu F10 : Batterie 1-100, Oxygene 1-100, SuperVitesse 1-20, Torpilles 1-10, Sonar 1-10, Blindage lourd x3, Super discrétion x3, Plongée x2, Couleurs eclairage avec choix Alarm/SilentRun, Appeler renforts",
        "- Slider Batterie : valeur legacy conservee, l'infini depend seulement de la case Mega Batterie",
        "- Slider Oxygene : 1 = vanilla, 100 = profil environ 90 jours",
        f"- Slider SuperVitesse : 1 = vanilla, {args.fast_speed_factor:g} = defaut actuel, 20 = maximum",
        f"- Slider Torpilles : 1 = vanilla, {args.torpedo_damage_factor:g} = maximum",
        "- Slider Sonar : 1 = vanilla, 3 = defaut actuel, 10 = maximum",
        "- Toggle Plongée x2 : actif par defaut, decoche = profondeur vanilla",
        "- Toggle Couleurs eclairage : actif par defaut, decoche = couleurs vanilla",
        "- Couleurs eclairage : deux listes F10 predefinies pour Alarm et SilentRun, defaut orange ambre / vert sous-marin",
        "- Blindage lourd : case desactivee par defaut, activable dans F10, degats joueur divises par 3 quand activee",
        "- Super discrétion : case desactivee par defaut, bruit et detectabilite joueur divises par 3 quand activee",
        "- DeepDive : case F10 Plongée x2 activee par defaut, ordres de profondeur > 10 m doubles (20->40, 40->80, 150->300, 300->600), stress profondeur calcule sur profondeur /2, max operationnel 600 m, crush 700 m",
        "- Bouton Par defaut : restaure les reglages du profil actuel",
        f"- Mega torpilles : {'oui' if args.mega_torpedoes else 'non'}",
        f"- Mega torpilles degats : x{args.torpedo_damage_factor:g}",
        f"- Mega torpilles effets visuels rayon explosion : x{args.torpedo_explosion_radius_factor:g}",
        f"- Mega torpilles effets visuels intensite explosion : x{args.torpedo_explosion_intensity_factor:g}",
        "- Mega torpilles guidage runtime : desactive pour stabilite surface/alarme",
        f"- Fiabilite parfaite torpilles : {'oui' if args.perfect_torpedo_reliability else 'non'}",
        f"- DudChance torpilles : {args.torpedo_dud_chance:g}",
        f"- Defaillance magnetique torpilles : {args.torpedo_magnetic_failure_chance:g}",
        f"- Explosion magnetique prematuree torpilles : {args.torpedo_premature_magnetic_chance:g}",
        "- Menu en jeu : F10 pour activer/desactiver Mega Batterie, Mega Oxygene, SuperVitesse, Mega Torpilles, Mega Sonar, Blindage lourd, Super discrétion, Plongée x2, Couleurs eclairage et Appeler renforts",
        "- Bouton Appeler renforts : appelle des U-boats amis pres du joueur (10-16 km, minimum 8 km); avions/warships seulement si des spawners amis compatibles existent",
        "- Plongée x2 : case F10 active par defaut, decochee pour revenir au mode profondeur vanilla.",
        "- Couleurs eclairage : case F10 active par defaut, decochee pour restaurer les couleurs vanilla.",
        "- Eclairage interieur : deux listes F10 changent visuellement Alarm et SilentRun ; defaut Alarm orange ambre, SilentRun vert sous-marin, gameplay inchange quand la case est activee",
        "- DLC Type IX officiel : lignes joueur Type IXA/IXC/IXC40 incluses si le DLC est installe",
        f"- Ventilation vanilla : {'non' if args.patch_ventilation else 'oui'}",
        f"- Patch runtime : {MOD_ASSEMBLY_NAME}, air apres chargement, plafond vitesse, carburant rapide, torpilles, sonar, blindage lourd, super discretion, renforts, menu et stabilite surface/alarme",
        "",
        "Installation :",
        "1. Fermer UBOAT.",
        "2. Générer avec --force --clear-cache.",
        "3. Activer le mod dans le launcher.",
        "4. Charger la sauvegarde ou démarrer une nouvelle carrière pour tester les changements d'air.",
        "",
        "Notes :",
        "- La jauge du jeu est une qualité d'air/atmosphère, pas un vrai compteur O2 détaillé.",
        "- La lumiere d'alarme utilise la couleur Alarm choisie uniquement au rendu quand Couleurs eclairage est cochee ; le mode Alarm et ses effets restent vanilla.",
        "- La lumiere SilentRun utilise la couleur SilentRun choisie uniquement au rendu quand Couleurs eclairage est cochee ; le mode SilentRun et ses effets restent vanilla.",
        "- La ventilation reste vanilla par défaut pour éviter les bugs vus dans les essais précédents.",
        "- Le patch runtime recalcule la respiration vanilla puis reduit seulement le drain negatif si Mega Oxygene est actif.",
        "- Le profil air vise environ 90 jours d'immersion avec Mega Oxygene actif, sans toucher a la recharge surface.",
        "- Mega Batterie cochee rend la batterie infinie ; decochee, la batterie revient vanilla.",
        "- Blindage lourd est desactive par defaut ; coche dans F10, il divise les degats joueur par 3 sans rendre le sous-marin immortel.",
        "- Migration settings v16 / v1.4.16 : les anciennes installations repassent Blindage lourd sur OFF une seule fois, puis tes choix F10 sont conserves.",
        "- Super discrétion cochee divise le bruit et les detectabilites joueur par 3 sans supprimer les contacts ennemis.",
        "- DeepDive : coche Plongée x2 dans F10 pour doubler les ordres de profondeur au-dessus de 10 m (20->40, 40->80, 150->300, 300->600).",
        "- DeepDive : active, le stress de profondeur de l'equipage est calcule sur la profondeur /2.",
        "- DeepDive : active, la profondeur operationnelle vise 600 m reels et le crush est repousse a 700 m ; decochee, la profondeur redevient vanilla.",
        "- Les sliders F10 sont persistants et s'appliquent en partie avec un debounce ou Reappliquer maintenant.",
        "- Les vitesses lentes et mi-vitesse restent vanilla ; seuls les deux crans rapides avant sont boostés vers 40/45 km/h.",
        "- Les crans rapides consomment plus de carburant pour garder une autonomie logique.",
        "- Les torpilles gardent leur vitesse/portee vanilla ; les degats, explosions, rates et le guidage verrouille sont geres en runtime.",
        "- Mega Sonar augmente seulement la portee hydrophone ; x1 ou case decochee revient vanilla.",
        "- Le guidage mega met les tirs verrouilles en cible cartésienne dynamique et force l'impact a courte distance.",
        "- La fiabilite parfaite met DudChance, MagneticExplosionFail, MagneticExplosionOnArm et MagneticExplosionAfterArm a 0 quand Mega Torpilles est actif.",
        "- Couper Mega Torpilles remet les torpilles sur les valeurs vanilla, car les XLSX torpilles ne sont pas ecrases.",
        "- La recharge diesel reste vanilla pour éviter une recharge batterie instantanée.",
        "- Le plafond de vitesse inclut les Type IX officiels du DLC Steam quand le DLC est installe.",
        "- Compatible sauvegarde existante après fermeture complète puis relance du jeu.",
        "- Si un autre mod touche l'air, mets Long Submerged 10x+ après lui dans l'ordre de chargement.",
    ]

    if report is not None:
        lines.extend(
            [
                "",
                "Compteurs de génération :",
                f"- Lignes batterie : {report.counters.get('battery_capacity_rows', 0)}",
                f"- Lignes EnergyUsage consommation : {report.counters.get('energy_usage_rows', 0)}",
                f"- Lignes EnergyUsage recharge : {report.counters.get('energy_recharge_rows', 0)}",
                "- Mega Batterie : case F10 active = batterie infinie, pompe incluse",
                "- Menu F10 : sliders runtime bornes par profil, Blindage lourd x3, Super discrétion x3, Plongée x2, Couleurs eclairage avec choix Alarm/SilentRun, Appeler renforts et bouton Par defaut",
                "- SuperVitesse : runtime F10 reglable 1-20 sur les deux crans rapides avant",
                f"- Lignes vitesse sous-marin joueur : {report.counters.get('player_submarine_speed_rows', 0)}",
                "- Mega torpilles : runtime F10 reglable 1-10, degats defaut x10, effets visuels bornes x3, aucune ligne torpille XLSX ecrasee",
                "- Mega Sonar : runtime F10 reglable 1-10, defaut x3, applique aux portees hydrophone",
                "- Blindage lourd : case F10 desactivee par defaut, activable manuellement, degats joueur divises par 3",
                "- Super discrétion : case F10 desactivee par defaut, bruit et detectabilite joueur divisibles par 3",
            ]
        )

    (mod_dir / "README_LongSubmerged10x.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_report(mod_dir: Path, report: PatchReport) -> None:
    lines: list[str] = []

    lines.append("=== Rapport Long Submerged 10x+ ===")
    lines.append("")
    lines.append("Compteurs :")
    for key in sorted(report.counters):
        lines.append(f"- {key}: {report.counters[key]}")

    lines.append("")
    lines.append("Fichiers générés :")
    for path in report.changed_files:
        lines.append(f"- {path}")

    lines.append("")
    lines.append("Changements :")
    for item in report.info:
        lines.append(f"- {item}")

    if report.warnings:
        lines.append("")
        lines.append("Avertissements :")
        for warning in report.warnings:
            lines.append(f"- {warning}")

    (mod_dir / "LongSubmerged10x_generation_report.txt").write_text(
        "\n".join(lines) + "\n",
        encoding="utf-8",
    )


def clear_uboat_cache(local_uboat_dir: Path, report: PatchReport) -> None:
    for folder_name in ("Cache", "Data Sheets", "Temp"):
        folder = local_uboat_dir / folder_name

        if not folder.exists():
            continue

        for child in folder.iterdir():
            try:
                if child.is_dir():
                    shutil.rmtree(child)
                else:
                    child.unlink()
            except OSError as exc:
                report.warn(f"Impossible de supprimer {child}: {exc}")

        report.note(f"Cache vidé : {folder}")


# =============================================================================
# ARGUMENTS
# =============================================================================

def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Génère le mod UBOAT Long Submerged 10x+.",
    )

    parser.add_argument(
        "--uboat",
        type=Path,
        required=True,
        help=r"Chemin d'installation UBOAT, ex: C:\Program Files (x86)\Steam\steamapps\common\UBOAT",
    )

    parser.add_argument(
        "--out",
        type=Path,
        default=default_local_mods_root() / MOD_FOLDER_NAME,
        help=(
            "Dossier de sortie du mod. Par défaut : "
            "AppData\\LocalLow\\Deep Water Studio\\UBOAT\\Mods\\LongSubmerged10x"
        ),
    )

    parser.add_argument(
        "--force",
        action="store_true",
        help="Écrase le dossier de mod existant.",
    )

    parser.add_argument(
        "--clear-cache",
        action="store_true",
        help="Vide Cache, Data Sheets et Temp dans AppData\\LocalLow\\Deep Water Studio\\UBOAT.",
    )

    parser.add_argument(
        "--game-version",
        default=DEFAULT_GAME_VERSION,
        help='Version affichée par le launcher, ex: "2026.1 Patch 20".',
    )

    parser.add_argument(
        "--air-capacity-factor",
        type=float,
        default=DEFAULT_AIR_CAPACITY_FACTOR,
        help="Option conservee pour compatibilite, mais SurfaceSafe 1.4.7 ne modifie plus les capacites d'air XLSX.",
    )

    parser.add_argument(
        "--oxygen-consumption-factor",
        type=float,
        default=DEFAULT_OXYGEN_CONSUMPTION_FACTOR,
        help="Facteur runtime qui divise seulement le drain negatif de respiration. Defaut : 250.",
    )

    parser.add_argument(
        "--discipline-factor",
        type=float,
        default=DEFAULT_DISCIPLINE_FACTOR,
        help="Divise discipline/fatigue sous l'eau par ce facteur. Défaut : 15.",
    )

    parser.add_argument(
        "--battery-capacity-factor",
        type=float,
        default=DEFAULT_BATTERY_CAPACITY_FACTOR,
        help="Multiplicateur capacité batterie / Accumulators. Défaut : 10.",
    )

    parser.add_argument(
        "--energy-usage-factor",
        type=float,
        default=DEFAULT_EQUIPMENT_ENERGY_USAGE_FACTOR,
        help="Multiplicateur EnergyUsage positif datasheet hors ventilation/compresseurs. Défaut : 0.1 ; Mega Batterie runtime force 0.",
    )

    parser.add_argument(
        "--fast-speed-factor",
        type=float,
        default=DEFAULT_FAST_SPEED_FACTOR,
        help="Multiplicateur vitesse/propulsion des derniers crans rapides de marche avant. Defaut : 8.",
    )

    parser.add_argument(
        "--fast-speed-fuel-factor",
        type=float,
        default=DEFAULT_FAST_SPEED_FUEL_FACTOR,
        help="Multiplicateur carburant des derniers crans rapides de marche avant. Défaut : 8.",
    )

    parser.add_argument(
        "--fast-speed-top-gears",
        type=int,
        default=DEFAULT_FAST_SPEED_TOP_GEARS,
        help="Nombre de crans rapides avant à booster. Défaut : 2.",
    )

    parser.add_argument(
        "--player-submarine-max-speed",
        type=float,
        default=DEFAULT_PLAYER_SUBMARINE_MAX_SPEED,
        help="Vitesse max km/h des types de sous-marins joueur. Défaut : 45.",
    )

    parser.add_argument(
        "--patch-player-submarine-speed-rows",
        action="store_true",
        default=DEFAULT_PATCH_PLAYER_SUBMARINE_SPEED_ROWS,
        help="Compatibilite ancienne version : reecrit les lignes Types joueur a --player-submarine-max-speed. Desactive par defaut en 1.4.8.",
    )

    mega_torpedoes_group = parser.add_mutually_exclusive_group()
    mega_torpedoes_group.add_argument(
        "--mega-torpedoes",
        dest="mega_torpedoes",
        action="store_true",
        default=DEFAULT_MEGA_TORPEDOES,
        help="Active les mega torpilles : degats et explosions renforces. Active par defaut.",
    )
    mega_torpedoes_group.add_argument(
        "--no-mega-torpedoes",
        dest="mega_torpedoes",
        action="store_false",
        help="Desactive le patch mega torpilles.",
    )

    parser.add_argument(
        "--torpedo-damage-factor",
        type=float,
        default=DEFAULT_TORPEDO_DAMAGE_FACTOR,
        help="Multiplicateur Damage des torpilles. Défaut : 10.",
    )

    parser.add_argument(
        "--torpedo-crew-damage-factor",
        type=float,
        default=DEFAULT_TORPEDO_CREW_DAMAGE_FACTOR,
        help="Multiplicateur CrewDamage des torpilles. Défaut : 10.",
    )

    parser.add_argument(
        "--torpedo-explosion-radius-factor",
        type=float,
        default=DEFAULT_TORPEDO_EXPLOSION_RADIUS_FACTOR,
        help="Multiplicateur DamageRadius et DamageEffectsRadius des torpilles. Défaut : 3.",
    )

    parser.add_argument(
        "--torpedo-explosion-intensity-factor",
        type=float,
        default=DEFAULT_TORPEDO_EXPLOSION_INTENSITY_FACTOR,
        help="Multiplicateur DamageEffectsIntensity des torpilles. Défaut : 3.",
    )

    torpedo_reliability_group = parser.add_mutually_exclusive_group()
    torpedo_reliability_group.add_argument(
        "--perfect-torpedo-reliability",
        dest="perfect_torpedo_reliability",
        action="store_true",
        default=DEFAULT_PERFECT_TORPEDO_RELIABILITY,
        help="Supprime les rates et defaillances des torpilles. Active par defaut.",
    )
    torpedo_reliability_group.add_argument(
        "--no-perfect-torpedo-reliability",
        dest="perfect_torpedo_reliability",
        action="store_false",
        help="Garde la fiabilite vanilla des torpilles.",
    )

    parser.add_argument(
        "--torpedo-dud-chance",
        type=float,
        default=DEFAULT_TORPEDO_DUD_CHANCE,
        help="Valeur DudChance appliquee aux torpilles si la fiabilite parfaite est active. Défaut : 0.",
    )

    parser.add_argument(
        "--torpedo-magnetic-failure-chance",
        type=float,
        default=DEFAULT_TORPEDO_MAGNETIC_FAILURE_CHANCE,
        help="Valeur MagneticExplosionFail appliquee aux torpilles si la fiabilite parfaite est active. Défaut : 0.",
    )

    parser.add_argument(
        "--torpedo-premature-magnetic-chance",
        type=float,
        default=DEFAULT_TORPEDO_PREMATURE_MAGNETIC_CHANCE,
        help=(
            "Valeur MagneticExplosionOnArm/AfterArm appliquee aux torpilles "
            "si la fiabilite parfaite est active. Défaut : 0."
        ),
    )

    parser.add_argument(
        "--patch-ventilation",
        action="store_true",
        default=DEFAULT_PATCH_VENTILATION,
        help="Optionnel : patche aussi OxygenGain de la ventilation. Désactivé par défaut.",
    )

    parser.add_argument(
        "--ventilation-gain-factor",
        type=float,
        default=DEFAULT_VENTILATION_GAIN_FACTOR,
        help="Si --patch-ventilation est activé, multiplie OxygenGain par ce facteur. Défaut : 1.",
    )

    parser.add_argument(
        "--patch-potassium",
        action="store_true",
        default=DEFAULT_PATCH_POTASSIUM,
        help="Optionnel : allonge les durées des absorbeurs potassium.",
    )

    parser.add_argument(
        "--potassium-duration-factor",
        type=float,
        default=DEFAULT_POTASSIUM_DURATION_FACTOR,
        help="Si --patch-potassium est activé, multiplie les durées par ce facteur. Défaut : 1.",
    )

    parser.add_argument(
        "--patch-energy-base-scale",
        action="store_true",
        default=DEFAULT_PATCH_ENERGY_BASE_SCALE,
        help="Patche aussi General.xlsx / Energy Base Scale. Désactivé par défaut.",
    )

    args = parser.parse_args(argv)

    positive_names = (
        "air_capacity_factor",
        "oxygen_consumption_factor",
        "discipline_factor",
        "battery_capacity_factor",
        "energy_usage_factor",
        "fast_speed_factor",
        "fast_speed_fuel_factor",
        "player_submarine_max_speed",
        "torpedo_damage_factor",
        "torpedo_crew_damage_factor",
        "torpedo_explosion_radius_factor",
        "torpedo_explosion_intensity_factor",
        "ventilation_gain_factor",
        "potassium_duration_factor",
    )

    for name in positive_names:
        value = getattr(args, name)
        if value <= 0:
            raise ValueError(f"--{name.replace('_', '-')} doit être > 0.")

    probability_names = (
        "torpedo_dud_chance",
        "torpedo_magnetic_failure_chance",
        "torpedo_premature_magnetic_chance",
    )

    for name in probability_names:
        value = getattr(args, name)
        if value < 0 or value > 1:
            raise ValueError(f"--{name.replace('_', '-')} doit être entre 0 et 1.")

    if args.fast_speed_top_gears <= 0:
        raise ValueError("--fast-speed-top-gears doit être > 0.")

    return args


# =============================================================================
# MAIN
# =============================================================================

def main(argv: list[str]) -> int:
    try:
        args = parse_args(argv)
    except Exception as exc:
        print(f"Arguments invalides : {exc}", file=sys.stderr)
        return 2

    report = PatchReport()
    out_mod_dir: Path | None = None

    try:
        uboat_root = args.uboat.expanduser().resolve()
        data_sheets_dir = resolve_data_sheets_dir(uboat_root)

        out_mod_dir = args.out.expanduser().resolve()
        out_data_sheets = out_mod_dir / "Data Sheets"

        ensure_clean_directory(out_mod_dir, force=args.force)
        out_data_sheets.mkdir(parents=True, exist_ok=True)

        write_manifest(out_mod_dir, args)
        write_runtime_patch_source(out_mod_dir, args, report)

        # 1) General.xlsx et Realistic Travel/General.xlsx.
        general_patches = build_general_patches(args)
        general_changed = 0

        for vanilla_general in find_datasheet_files(data_sheets_dir, "General.xlsx"):
            relative = vanilla_general.relative_to(data_sheets_dir)
            output_general = out_data_sheets / relative

            if create_general_override(
                vanilla_general,
                output_general,
                general_patches,
                report,
                data_sheets_dir,
            ):
                general_changed += 1

        if general_changed == 0:
            report.warn(
                "Aucun General.xlsx patché. "
                "La discipline ne changera pas."
            )

        # 2) XLSX de gameplay : batterie et EnergyUsage seulement.
        # SurfaceSafe 1.4.7 ne genere plus de ligne Air/Oxygen/Atmosphere pour laisser la recharge surface vanilla.
        # Je ne scanne pas les gros fichiers Locales/Story/Graphics : ils ne portent pas ces systèmes
        # et ralentissent inutilement la génération.
        generic_changed = 0
        generic_targets = (
            "Entities.xlsx",
            "U-boat.xlsx",
            "Sandbox.xlsx",
            "CharacterClasses.xlsx",
        )
        type_ix_dlc_data_sheets = find_type_ix_dlc_data_sheets_dir(uboat_root)

        if type_ix_dlc_data_sheets is not None:
            report.note(f"DLC Type IX detecte : {type_ix_dlc_data_sheets}")

        for target_name in generic_targets:
            for vanilla_xlsx in find_datasheet_files(data_sheets_dir, target_name):
                relative = vanilla_xlsx.relative_to(data_sheets_dir)
                output_xlsx = out_data_sheets / relative

                if create_generic_xlsx_override(
                    vanilla_xlsx,
                    output_xlsx,
                    report,
                    data_sheets_dir,
                    args,
                ):
                    generic_changed += 1

            if type_ix_dlc_data_sheets is None:
                continue

            for dlc_xlsx in find_datasheet_files(type_ix_dlc_data_sheets, target_name):
                relative = dlc_xlsx.relative_to(type_ix_dlc_data_sheets)
                output_xlsx = out_data_sheets / relative

                if create_generic_xlsx_override(
                    dlc_xlsx,
                    output_xlsx,
                    report,
                    type_ix_dlc_data_sheets,
                    args,
                ):
                    generic_changed += 1

        if generic_changed == 0:
            report.warn(
                "Aucun XLSX hors General.xlsx n'a été patché. "
                "La batterie exposée dans Entities/U-boat ne changera peut-être pas."
            )

        if report.counters.get("battery_capacity_rows", 0) == 0:
            report.warn(
                "Aucune ligne Accumulators/Battery patchée. "
                "Si la batterie ne reste pas longue durée, vérifie le nom local de la ligne accumulateurs."
            )

        if args.clear_cache:
            clear_uboat_cache(default_local_uboat_dir(), report)

        write_readme(out_mod_dir, args, report)
        write_report(out_mod_dir, report)

    except Exception as exc:
        print("\nERREUR : génération impossible.", file=sys.stderr)
        print(str(exc), file=sys.stderr)
        return 1

    assert out_mod_dir is not None

    print("\n=== Long Submerged 10x+ généré avec succès ===")
    print(f"Dossier du mod : {out_mod_dir}")
    print(f"Version déclarée : {args.game_version}")

    print("\nProfil appliqué :")
    print("  - Oxygene long : runtime uniquement sur le drain negatif de respiration")
    print("  - Recharge surface : vanilla, aucune capacite Air/Oxygen/Atmosphere XLSX modifiee")
    print("  - Objectif air : environ 90 jours avec Mega Oxygene actif")
    print(f"  - Discipline/fatigue : /{args.discipline_factor:g}")
    print(f"  - Batterie Accumulators : x{args.battery_capacity_factor:g}")
    print(f"  - EnergyUsage consommateurs hors ventilation/compresseurs : x{args.energy_usage_factor:g} dans les datasheets")
    print("  - Mega Batterie runtime : case F10 active = batterie infinie, pompe incluse")
    print("  - EnergyUsage recharge/production batterie : vanilla")
    print(f"  - Deux derniers crans avant : vitesse/propulsion x{args.fast_speed_factor:g}")
    print(f"  - Deux derniers crans avant : carburant x{args.fast_speed_fuel_factor:g}")
    print(f"  - Vitesse max sous-marin joueur : {args.player_submarine_max_speed:g} km/h")
    print("  - Menu F10 : Batterie 1-100, Oxygene 1-100, SuperVitesse 1-20, Torpilles 1-10, Sonar 1-10, Blindage lourd x3, Super discrétion x3, Plongée x2, Couleurs eclairage Alarm/SilentRun, Appeler renforts")
    print("  - Blindage lourd : desactive par defaut, activable dans F10, degats joueur divises par 3")
    print("  - Super discrétion : desactivee par defaut, bruit et detectabilite joueur divisibles par 3")
    print("  - Plongée x2 : activee par defaut, stress profondeur /2, decochee = profondeur vanilla")
    print("  - Couleurs eclairage : actives par defaut, listes Alarm/SilentRun, decochees = couleurs vanilla")
    print(f"  - Mega torpilles : {'OUI' if args.mega_torpedoes else 'NON'}")
    print(f"  - Torpilles degats : x{args.torpedo_damage_factor:g}")
    print(f"  - Torpilles effets visuels rayon explosion : x{args.torpedo_explosion_radius_factor:g}")
    print(f"  - Torpilles effets visuels intensite explosion : x{args.torpedo_explosion_intensity_factor:g}")
    print("  - Torpilles guidage runtime : desactive pour stabilite surface/alarme")
    print(f"  - Fiabilite parfaite torpilles : {'OUI' if args.perfect_torpedo_reliability else 'NON'}")
    print(f"  - DudChance / defaillances torpilles : {args.torpedo_dud_chance:g}")
    print(f"  - Ventilation vanilla : {'NON, patchée' if args.patch_ventilation else 'OUI, laissée normale'}")

    print("\nCompteurs importants :")
    print(f"  - Batterie : {report.counters.get('battery_capacity_rows', 0)}")
    print(f"  - EnergyUsage consommation : {report.counters.get('energy_usage_rows', 0)}")
    print(f"  - EnergyUsage recharge : {report.counters.get('energy_recharge_rows', 0)}")
    print("  - Mega Batterie : case F10 active = batterie infinie, pompe incluse")
    print(f"  - Vitesse sous-marin joueur : {report.counters.get('player_submarine_speed_rows', 0)}")
    print("  - SuperVitesse : runtime F10 reglable 1-20")
    print("  - Mega torpilles : runtime F10 reglable 1-10, degats defaut x10 et effets visuels bornes x3")
    print("  - Mega Sonar : runtime F10 reglable 1-10, defaut x3")
    print("  - Blindage lourd : case F10 desactivee par defaut, activable manuellement, degats joueur divises par 3")
    print("  - Super discrétion : case F10 desactivee par defaut, bruit et detectabilite joueur divisibles par 3")
    print("  - Plongée x2 : case F10 activee par defaut, stress profondeur /2, decochee = profondeur vanilla")
    print("  - Couleurs eclairage : case F10 activee par defaut, choix Alarm/SilentRun, decochee = couleurs vanilla")
    print("  - Fiabilite torpilles : runtime F10")

    print("\nFichiers générés :")
    print(f"  - {out_mod_dir / 'Manifest.json'}")
    print(f"  - {out_mod_dir / 'README_LongSubmerged10x.txt'}")
    print(f"  - {out_mod_dir / 'LongSubmerged10x_generation_report.txt'}")
    for file_path in report.changed_files:
        print(f"  - {file_path}")

    if report.warnings:
        print("\nAvertissements :")
        for warning in report.warnings:
            print(f"  - {warning}")

    print("\nÉtapes obligatoires :")
    print("  1. Ferme complètement UBOAT.")
    print("  2. Lance le script avec --force --clear-cache.")
    print("  3. Active Long Submerged 10x+ dans le launcher.")
    print("  4. Place-le après les autres mods qui touchent General.xlsx, Entities.xlsx ou U-boat.xlsx.")
    print("  5. Charge ta sauvegarde existante ou lance une nouvelle carrière.")
    print("  6. Plonge à 100% air : le tooltip doit viser environ 90 jours avec Mega Oxygene actif.")
    print("  7. Teste pleine vitesse / machines à fond : on vise environ 40/45 km/h.")
    print("  8. Si la ventilation était cassée par une ancienne version, celle-ci doit la remettre vanilla car elle ne copie plus la ligne Ventilation.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
