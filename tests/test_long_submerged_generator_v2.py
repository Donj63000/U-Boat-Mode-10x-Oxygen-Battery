from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

import build_uboat_long_submerged_mod_v2 as generator_v2


def make_general_workbook(path: Path, oxygen_value: float) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Settings"

    # Je garde la structure vanilla importante : une catégorie Discipline et une catégorie Resources.
    rows = [
        ["/Discipline", "Normal", "Hard", "Very Hard"],
        ["Underwater Discipline Loss", 0.000015, 0.000016, 0.000017],
        ["Fatigue - Per Day", -0.0000021, -0.0000022, -0.0000023],
        ["Fatigue - Max Penalty", -0.00012, -0.000155, -0.0002],
        [None, None, None, None],
        ["/Resources", "Value", None, None],
        ["Oxygen Consumption Per Character", oxygen_value, None, None],
        ["Energy Base Scale", 0.32, None, None],
    ]

    for row in rows:
        ws.append(row)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def make_entities_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment"

    headers = [f"Column {i}" for i in range(1, 16)] + ["Parameters"]
    headers[0] = "Id"
    ws.append(headers)

    def append_equipment(row_id: str, row_name: str, parameters: str) -> None:
        row = [None] * 16
        row[0] = row_id
        row[1] = row_name
        row[15] = parameters
        ws.append(row)

    append_equipment("Accumulators I", "Accumulators, Electric Engines Upgrade", "EnergyCapacityGain=25%")
    append_equipment("Electric Engines", "Engines, Electric Engines", "EnergyUsage = 0.000035, Noise = 0.52")
    append_equipment(
        "Ventilation",
        "Ventilation",
        "EnergyUsage = 0.0001, OxygenGain = 0.00011, RegenerationLimit = 0.5",
    )
    append_equipment("Decorative Item", "No patch", "Noise = 0.1")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def find_row_by_id(ws, row_id: str) -> int:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == row_id:
            return row
    raise AssertionError(f"Ligne introuvable : {row_id}")


class LongSubmergedGeneratorV2Tests(unittest.TestCase):
    def test_v2_generates_manifest_nested_general_and_entities_overrides(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            uboat_root = tmp_path / "UBOAT"
            data_sheets = uboat_root / "UBOAT_Data" / "Data Sheets"
            out_mod = tmp_path / "LongSubmerged10x"

            make_general_workbook(data_sheets / "General.xlsx", -0.000009)
            make_general_workbook(data_sheets / "Realistic Travel" / "General.xlsx", -0.00000133)
            make_entities_workbook(data_sheets / "Entities.xlsx")

            exit_code = generator_v2.main(
                [
                    "--uboat",
                    str(uboat_root),
                    "--out",
                    str(out_mod),
                    "--force",
                    "--game-version",
                    "2026.1 Patch 20",
                ]
            )

            self.assertEqual(exit_code, 0)

            manifest = json.loads((out_mod / "Manifest.json").read_text(encoding="utf-8"))
            self.assertEqual(manifest["name"], "Long Submerged 10x+")
            self.assertIn("supportedGameVersions", manifest)
            self.assertIn("2026.1 Patch 20", manifest["supportedGameVersions"])
            self.assertIn("2026.1", manifest["supportedGameVersions"])

            general_root = out_mod / "Data Sheets" / "General.xlsx"
            general_realistic = out_mod / "Data Sheets" / "Realistic Travel" / "General.xlsx"
            entities_path = out_mod / "Data Sheets" / "Entities.xlsx"

            self.assertTrue(general_root.exists())
            self.assertTrue(general_realistic.exists())
            self.assertTrue(entities_path.exists())

            root_ws = load_workbook(general_root, data_only=False)["Settings"]
            realistic_ws = load_workbook(general_realistic, data_only=False)["Settings"]

            root_underwater_row = find_row_by_id(root_ws, "Underwater Discipline Loss")
            root_oxygen_row = find_row_by_id(root_ws, "Oxygen Consumption Per Character")
            realistic_oxygen_row = find_row_by_id(realistic_ws, "Oxygen Consumption Per Character")

            self.assertAlmostEqual(root_ws.cell(root_underwater_row, 2).value, 0.000015 / 15)
            self.assertAlmostEqual(root_ws.cell(root_underwater_row, 3).value, 0.000016 / 15)
            self.assertAlmostEqual(root_ws.cell(root_underwater_row, 4).value, 0.000017 / 15)
            self.assertAlmostEqual(root_ws.cell(root_oxygen_row, 2).value, -0.000009 / 15)
            self.assertAlmostEqual(realistic_ws.cell(realistic_oxygen_row, 2).value, -0.00000133 / 15)

            entities_ws = load_workbook(entities_path, data_only=False)["Equipment"]
            accumulator_row = find_row_by_id(entities_ws, "Accumulators I")
            engine_row = find_row_by_id(entities_ws, "Electric Engines")
            ventilation_row = find_row_by_id(entities_ws, "Ventilation")

            self.assertIn("250%", entities_ws.cell(accumulator_row, 16).value)
            self.assertIn("EnergyUsage = 3.5e-06", entities_ws.cell(engine_row, 16).value)

            ventilation_params = entities_ws.cell(ventilation_row, 16).value
            # La v2 fournie applique EnergyUsage global puis EnergyUsage ventilation sur cette ligne.
            self.assertIn("EnergyUsage = 1e-06", ventilation_params)
            self.assertIn("OxygenGain = 0.0011", ventilation_params)
            self.assertIn("RegenerationLimit = 5", ventilation_params)


if __name__ == "__main__":
    unittest.main()
