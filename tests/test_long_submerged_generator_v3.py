from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

import build_uboat_long_submerged_mod_v3_airfix as generator_v3


def make_general_workbook(path: Path, oxygen_value: float) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Settings"

    rows = [
        ["/Discipline", "Normal", "Hard", "Very Hard"],
        ["Underwater Discipline Loss", 0.000015, 0.000016, 0.000017],
        ["Fatigue - Per Day", -0.0000021, -0.0000022, -0.0000023],
        ["Fatigue - Max Penalty", -0.00012, -0.000155, -0.0002],
        [None, None, None, None],
        ["/Resources", "Value", None, None],
        ["Oxygen Consumption Per Character", oxygen_value, None, None],
        ["Energy Base Scale", 0.32, None, None],
    ]

    for row in rows:
        ws.append(row)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def make_entities_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment"

    headers = [f"Column {i}" for i in range(1, 16)] + ["Parameters"]
    headers[0] = "Id"
    headers[1] = "Name"
    ws.append(headers)

    def append_equipment(row_id: str, row_name: str, parameters: str) -> None:
        row = [None] * 16
        row[0] = row_id
        row[1] = row_name
        row[15] = parameters
        ws.append(row)

    append_equipment("Accumulators I", "Accumulators, Electric Engines Upgrade", "EnergyCapacityGain=25%")
    append_equipment("Electric Engines", "Engines, Electric Engines", "EnergyUsage = 0.000035, Noise = 0.52")
    append_equipment("Atmosphere Tank", "Player U-boat Atmosphere Air", "AirCapacity = 100")
    append_equipment("Ventilation", "Ventilation", "EnergyUsage = 0.0001, OxygenGain = 0.00011, RegenerationLimit = 0.5")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def make_crew_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Statements"

    ws.append(["/", "Tags", "Queue Behaviour", "Queue Priority", "Show on Portrait", "Ignore Injuries", "Spatial Blend", "Spatial Blend (FPP)", "Volume", "Min Distance", "Max Distance"])
    ws.append(["Voice 426", "Air Quality Down To 25%", "Queue", -1, True, False, 0.84, 0.84, 1, 1, 100000])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def find_row_by_id(ws, row_id: str) -> int:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == row_id:
            return row
    raise AssertionError(f"Ligne introuvable : {row_id}")


class LongSubmergedGeneratorV3Tests(unittest.TestCase):
    def test_v3_keeps_ventilation_vanilla_and_patches_air_capacity(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            uboat_root = tmp_path / "UBOAT"
            data_sheets = uboat_root / "UBOAT_Data" / "Data Sheets"
            out_mod = tmp_path / "LongSubmerged10x"

            make_general_workbook(data_sheets / "General.xlsx", -0.000009)
            make_general_workbook(data_sheets / "Realistic Travel" / "General.xlsx", -0.00000133)
            make_entities_workbook(data_sheets / "Entities.xlsx")
            make_crew_workbook(data_sheets / "Crew.xlsx")

            exit_code = generator_v3.main(
                [
                    "--uboat",
                    str(uboat_root),
                    "--out",
                    str(out_mod),
                    "--force",
                    "--game-version",
                    "2026.1 Patch 20",
                ]
            )

            self.assertEqual(exit_code, 0)

            manifest = json.loads((out_mod / "Manifest.json").read_text(encoding="utf-8"))
            self.assertEqual(manifest["version"], "1.2.1")
            self.assertEqual(manifest["assemblyName"], "LongSubmerged10xPatch")
            self.assertIn("Reflection", manifest["permissions"])
            self.assertIn("2026.1 Patch 20", manifest["supportedGameVersions"])
            self.assertFalse((out_mod / "Data Sheets" / "Crew.xlsx").exists())

            runtime_patch = out_mod / "Source" / "LongSubmergedRuntimePatch.cs"
            runtime_patch_text = runtime_patch.read_text(encoding="utf-8")
            self.assertIn("IUserMod", runtime_patch_text)
            self.assertIn("ValidateOxygenBreathModifier", runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(PlayerShip), "SavesManagerOnLoaded")', runtime_patch_text)

            root_ws = load_workbook(out_mod / "Data Sheets" / "General.xlsx", data_only=False)["Settings"]
            realistic_ws = load_workbook(out_mod / "Data Sheets" / "Realistic Travel" / "General.xlsx", data_only=False)["Settings"]

            root_oxygen_row = find_row_by_id(root_ws, "Oxygen Consumption Per Character")
            realistic_oxygen_row = find_row_by_id(realistic_ws, "Oxygen Consumption Per Character")
            discipline_row = find_row_by_id(root_ws, "Underwater Discipline Loss")

            self.assertAlmostEqual(root_ws.cell(root_oxygen_row, 2).value, -0.000009 / 125)
            self.assertAlmostEqual(realistic_ws.cell(realistic_oxygen_row, 2).value, -0.00000133 / 125)
            self.assertAlmostEqual(root_ws.cell(discipline_row, 2).value, 0.000015 / 15)
            self.assertAlmostEqual(root_ws.cell(discipline_row, 3).value, 0.000016 / 15)
            self.assertAlmostEqual(root_ws.cell(discipline_row, 4).value, 0.000017 / 15)

            entities_ws = load_workbook(out_mod / "Data Sheets" / "Entities.xlsx", data_only=False)["Equipment"]
            ids = [entities_ws.cell(row, 1).value for row in range(1, entities_ws.max_row + 1)]

            self.assertIn("Accumulators I", ids)
            self.assertIn("Electric Engines", ids)
            self.assertIn("Atmosphere Tank", ids)
            self.assertNotIn("Ventilation", ids)

            accumulator_row = find_row_by_id(entities_ws, "Accumulators I")
            engine_row = find_row_by_id(entities_ws, "Electric Engines")
            air_row = find_row_by_id(entities_ws, "Atmosphere Tank")

            self.assertIn("250%", entities_ws.cell(accumulator_row, 16).value)
            self.assertIn("EnergyUsage = 3.5e-06", entities_ws.cell(engine_row, 16).value)
            self.assertIn("AirCapacity = 12500", entities_ws.cell(air_row, 16).value)

            report_text = (out_mod / "LongSubmerged10x_generation_report.txt").read_text(encoding="utf-8")
            self.assertIn("air_capacity_parameter_rows: 1", report_text)


if __name__ == "__main__":
    unittest.main()
