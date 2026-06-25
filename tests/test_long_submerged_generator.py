from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

import build_uboat_long_submerged_mod as generator


def extract_csharp_block(text: str, declaration: str) -> str:
    start = text.index(declaration)
    brace_start = text.index("{", start)
    depth = 0
    for index in range(brace_start, len(text)):
        char = text[index]
        if char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                return text[start : index + 1]

    raise AssertionError(f"C# block not closed: {declaration}")


def make_general_workbook(path: Path, oxygen_value: float) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Settings"

    rows = [
        ["/Discipline", "Normal", "Hard", "Very Hard"],
        ["Underwater Discipline Loss", 0.000015, 0.000016, 0.000017],
        ["Fatigue - Per Day", -0.0000021, -0.0000022, -0.0000023],
        ["Fatigue - Max Penalty", -0.00012, -0.000155, -0.0002],
        [None, None, None, None],
        ["/Resources", "Value", None, None],
        ["Oxygen Consumption Per Character", oxygen_value, None, None],
        ["Energy Base Scale", 0.32, None, None],
    ]

    for row in rows:
        ws.append(row)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def make_entities_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment"

    headers = [f"Column {i}" for i in range(1, 16)] + ["Parameters"]
    headers[0] = "Id"
    headers[1] = "Name"
    ws.append(headers)

    def append_equipment(row_id: str, row_name: str, parameters: str) -> None:
        row = [None] * 16
        row[0] = row_id
        row[1] = row_name
        row[15] = parameters
        ws.append(row)

    append_equipment("Accumulators I", "Accumulators, Electric Engines Upgrade", "EnergyCapacityGain=25%")
    append_equipment("Electric Engines", "Engines, Electric Engines", "EnergyUsage = 0.000035, Noise = 0.52")
    append_equipment("Diesel Engines", "Engines, Diesel Engines", "EnergyUsage = -1.0, Noise = 0.7")
    append_equipment("Atmosphere Tank", "Player U-boat Atmosphere Air", "AirCapacity = 100")
    append_equipment("Ventilation", "Ventilation", "EnergyUsage = 0.0001, OxygenGain = 0.00011, RegenerationLimit = 0.5")
    append_equipment(
        "G7a Torpedo T1 - Pi1",
        "Torpedo",
        (
            "Range1 = 5000, Speed1 = 22.63555, DudChance = 0.19, "
            "Damage = 7.0, CrewDamage = 0.8, DamageRadius = 6.5, "
            "DamageEffectsRadius = 7.0, DamageEffectsIntensity = 1.0, "
            "MagneticExplosionOnArm = 0.1, MagneticExplosionAfterArm = 0.005, "
            "MagneticExplosionFail = 0.1"
        ),
    )
    append_equipment("Bow Torpedo Launcher", "Torpedo Launcher", "ReloadTime = 180, AimPerformance = 0.55")

    ws_types = wb.create_sheet("Types")
    ws_types.append([None, None, None, "Displacement (t)"])
    ws_types.append(["/", "Category", "Speed (km/h)", "Standard", "Full"])
    ws_types.append(["Type VIIC", "Submarine", 32.8, 769, 871])
    ws_types.append(["Type VIIC (Player)", "Submarine", 32.8, 769, 871])
    ws_types.append(["F-Class", "Destroyer", 65.7, 1428, 1970])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def make_type_ix_dlc_entities_workbook(path: Path) -> None:
    wb = Workbook()
    ws_types = wb.active
    ws_types.title = "Types"
    ws_types.append([None, None, None, "Displacement (t)"])
    ws_types.append(["/", "Category", "Speed (km/h)", "Standard", "Full"])
    ws_types.append(["Type IXC", "Submarine", 32.8, 1120, 1232])
    ws_types.append(["Type IXC (Player)", "Submarine", 32.8, 1120, 1232])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def make_crew_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Statements"

    ws.append(["/", "Tags", "Queue Behaviour", "Queue Priority", "Show on Portrait", "Ignore Injuries", "Spatial Blend", "Spatial Blend (FPP)", "Volume", "Min Distance", "Max Distance"])
    ws.append(["Voice 426", "Air Quality Down To 25%", "Queue", -1, True, False, 0.84, 0.84, 1, 1, 100000])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def find_row_by_id(ws, row_id: str) -> int:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == row_id:
            return row
    raise AssertionError(f"Ligne introuvable : {row_id}")


class LongSubmergedGeneratorTests(unittest.TestCase):
    def test_github_presentation_matches_generator_defaults(self) -> None:
        repo_root = Path(__file__).resolve().parents[1]
        readme_text = (repo_root / "README.md").read_text(encoding="utf-8")
        install_text = (repo_root / "mod-here" / "install.txt").read_text(encoding="utf-8")

        expected_defaults = {
            "--oxygen-consumption-factor": generator.DEFAULT_OXYGEN_CONSUMPTION_FACTOR,
            "--battery-capacity-factor": generator.DEFAULT_BATTERY_CAPACITY_FACTOR,
            "--energy-usage-factor": generator.DEFAULT_EQUIPMENT_ENERGY_USAGE_FACTOR,
            "--fast-speed-factor": generator.DEFAULT_FAST_SPEED_FACTOR,
            "--fast-speed-fuel-factor": generator.DEFAULT_FAST_SPEED_FUEL_FACTOR,
            "--player-submarine-max-speed": generator.DEFAULT_PLAYER_SUBMARINE_MAX_SPEED,
            "--torpedo-damage-factor": generator.DEFAULT_TORPEDO_DAMAGE_FACTOR,
            "--torpedo-crew-damage-factor": generator.DEFAULT_TORPEDO_CREW_DAMAGE_FACTOR,
            "--torpedo-explosion-radius-factor": generator.DEFAULT_TORPEDO_EXPLOSION_RADIUS_FACTOR,
            "--torpedo-explosion-intensity-factor": generator.DEFAULT_TORPEDO_EXPLOSION_INTENSITY_FACTOR,
        }

        for option, value in expected_defaults.items():
            self.assertIn(f"| `{option}` | `{value:g}` |", readme_text)

        self.assertIn(f"- SuperSpeed: {generator.DEFAULT_FAST_SPEED_FACTOR:g}, boosted fast gears.", install_text)
        self.assertIn("- Sonar: 3, boosted hydrophone range.", install_text)
        self.assertIn("- Heavy Armor: disabled by default, can reduce player submarine damage to about one third when enabled in F10.", install_text)
        self.assertIn("- Heavy Armor migration: settings v16 / v1.4.16 turns old saved Heavy Armor settings off once, then keeps your F10 choices.", install_text)
        self.assertIn("- Interior lighting colors: enabled by default, with F10 dropdowns for Alarm and SilentRun. Defaults are amber orange and submarine green.", install_text)
        self.assertNotIn("v1.4.13 turns old saved Heavy Armor settings off", install_text)
        self.assertIn("- Super discrétion / Super Stealth: disabled by default, can reduce player noise and detectability to about one third.", install_text)
        self.assertIn("| **Heavy Armor x3** | `Heavy Armor = off` |", readme_text)
        self.assertIn("| **Heavy Armor x3** | `off` | `off` | `on`, x3 protection |", readme_text)
        self.assertIn("## Heavy Armor x3", readme_text)
        self.assertIn("Reduces flaw and fire chances together with player equipment damage.", readme_text)
        self.assertIn("| **DeepDive** | `Plongée x2 = on` |", readme_text)
        self.assertIn("evaluates crew depth stress on half real depth", readme_text)
        self.assertIn("stress thresholds, and crush handling to vanilla", readme_text)
        self.assertIn("| **Interior Lighting** | `Couleurs eclairage = on` |", readme_text)
        self.assertIn("F10 color dropdowns for Alarm and SilentRun", readme_text)
        self.assertIn("Defaults remain amber orange and submarine green.", readme_text)
        self.assertIn("Disabled restores vanilla lighting colors.", readme_text)
        self.assertIn("| **Super discrétion / Super Stealth** | `Super discrétion = off` |", readme_text)
        self.assertIn("| **Super discrétion / Super Stealth** | `off` | `off` | `on`, x3 stealth |", readme_text)
        self.assertIn("migration settings v16 / v1.4.16 turns old saved Heavy Armor settings off once", readme_text)
        self.assertIn("`Plongée x2` is enabled by default in F10", readme_text)
        self.assertIn("Enemy contacts and detection mechanics still exist.", readme_text)

    def test_generator_keeps_surface_oxygen_vanilla_and_runtime_safe(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            uboat_root = tmp_path / "UBOAT"
            data_sheets = uboat_root / "UBOAT_Data" / "Data Sheets"
            out_mod = tmp_path / "LongSubmerged10x"

            make_general_workbook(data_sheets / "General.xlsx", -0.000009)
            make_general_workbook(data_sheets / "Realistic Travel" / "General.xlsx", -0.00000133)
            make_entities_workbook(data_sheets / "Entities.xlsx")
            make_type_ix_dlc_entities_workbook(
                uboat_root
                / "UBOAT_Data"
                / "StreamingAssets"
                / "Packages"
                / "uboat.dlc.type-ix"
                / "Data Sheets"
                / "Entities.xlsx"
            )
            make_crew_workbook(data_sheets / "Crew.xlsx")

            exit_code = generator.main(
                [
                    "--uboat",
                    str(uboat_root),
                    "--out",
                    str(out_mod),
                    "--force",
                    "--game-version",
                    "2026.1 Patch 20",
                ]
            )

            self.assertEqual(exit_code, 0)

            manifest = json.loads((out_mod / "Manifest.json").read_text(encoding="utf-8"))
            self.assertEqual(manifest["version"], "1.4.16")
            self.assertEqual(manifest["assemblyName"], "LongSubmerged10xPatch_1_4_16")
            self.assertIn("Blindage lourd F10 desactive par defaut", manifest["description"])
            self.assertIn("DeepDive F10 active par defaut : ordres de profondeur x2, stress profondeur calcule sur profondeur /2", manifest["description"])
            self.assertIn("Couleurs eclairage F10 personnalisables par listes predefinies", manifest["description"])
            self.assertIn("defaut Alarm orange ambre et SilentRun vert sous-marin", manifest["description"])
            self.assertIn("Super discrétion F10 desactivee par defaut", manifest["description"])
            self.assertIn("Reflection", manifest["permissions"])
            self.assertIn("2026.1 Patch 20", manifest["supportedGameVersions"])
            self.assertFalse((out_mod / "Data Sheets" / "Crew.xlsx").exists())

            runtime_patch = out_mod / "Source" / "LongSubmergedRuntimePatch.cs"
            runtime_patch_text = runtime_patch.read_text(encoding="utf-8")
            self.assertIn("IUserMod", runtime_patch_text)
            self.assertIn("DonJ : point d'entree du runtime UBOAT", runtime_patch_text)
            self.assertIn("DonJ : vrai menu Unity UI", runtime_patch_text)
            self.assertIn("DonJ : coeur gameplay du mod", runtime_patch_text)
            self.assertIn("Mega Batterie now means a fully infinite battery", runtime_patch_text)
            self.assertIn("DonJ : les torpilles sont reglees au runtime", runtime_patch_text)
            self.assertIn("DonJ : SuperVitesse ne change pas toutes les allures", runtime_patch_text)
            self.assertIn("DonJ : hooks Harmony courts et delegues", runtime_patch_text)
            self.assertIn("ValidateOxygenBreathModifier", runtime_patch_text)
            self.assertIn("internal static class DeepDiveRuntimePatcher", runtime_patch_text)
            self.assertIn("public const float DisplayedDepthCommandFactor = 2f;", runtime_patch_text)
            self.assertIn("public const float ShallowDepthPassthroughMeters = 10f;", runtime_patch_text)
            self.assertIn("public const float MaxRealCommandDepthMeters = 600f;", runtime_patch_text)
            self.assertIn("public const float CrushDepthMeters = 700f;", runtime_patch_text)
            self.assertIn('AccessTools.Method(typeof(PlayerShip), "UpdateStressAndDisciplineGain", new Type[] { })', runtime_patch_text)
            self.assertIn('AccessTools.Field(typeof(PlayerShip), "depthStressModifier")', runtime_patch_text)
            self.assertIn("public static void ApplyDepthStressModifier(PlayerShip ship, string reason)", runtime_patch_text)
            self.assertIn("float effectiveDepth = realDepth / DisplayedDepthCommandFactor;", runtime_patch_text)
            for threshold, tier in [("300", "6"), ("250", "5"), ("200", "4"), ("150", "3"), ("100", "2"), ("25", "1")]:
                self.assertIn(f"if (depthMeters > {threshold}f)\n                return {tier}f;", runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(PlayerShip), "SavesManagerOnLoaded")', runtime_patch_text)
            self.assertIn("private const float FastSpeedFactor = 8f;", runtime_patch_text)
            self.assertIn("private const float FastSpeedFuelFactor = 8f;", runtime_patch_text)
            self.assertIn("private const float PlayerSubmarineMaxSpeed = 45f;", runtime_patch_text)
            self.assertIn("private const int FastForwardGearCount = 2;", runtime_patch_text)
            self.assertIn("EngineFastSpeedPatcher", runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(PlayerShip), "OnAfterDeserialize")', runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(PlayerShip), "ValidateTargetVelocity")', runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(PlayerShipEngine), "Awake")', runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(PlayerShipEngine), "OnAfterDeserialize")', runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(PlayerShipEngine), "SavesManagerOnLoaded")', runtime_patch_text)
            self.assertIn("expectedVelocityPerGear", runtime_patch_text)
            self.assertIn("expectedVelocityPerGearUnderwater", runtime_patch_text)
            self.assertIn("basePower", runtime_patch_text)
            self.assertIn("fuelConsumptionInLitersPerHour", runtime_patch_text)
            self.assertIn("PowerMultiplier", runtime_patch_text)
            self.assertIn("AddDeltaModifier", runtime_patch_text)
            self.assertIn("LongSubmergedMenuController", runtime_patch_text)
            self.assertIn("KeyCode.F10", runtime_patch_text)
            self.assertNotIn("KeyCode.F11", runtime_patch_text)
            self.assertIn("MenuKey = KeyCode.F10", runtime_patch_text)
            self.assertIn('private const string RuntimeVersion = "1.4.16";', runtime_patch_text)
            self.assertIn("Runtime patch loaded v", runtime_patch_text)
            self.assertIn('new Harmony("donj.longsubmerged10x.runtimefix1416")', runtime_patch_text)
            self.assertIn("LongSubmergedRuntimePatcher.PatchSafely", runtime_patch_text)
            self.assertIn("Harmony patch active", runtime_patch_text)
            self.assertIn("Harmony patch skipped", runtime_patch_text)
            self.assertNotIn("PatchAll()", runtime_patch_text)
            self.assertIn("InteriorLightingColorPatcher", runtime_patch_text)
            self.assertIn("internal static class InteriorLightingColorPatcher", runtime_patch_text)
            self.assertIn("private static readonly InteriorLightingColorPreset[] LightingColorPresets", runtime_patch_text)
            self.assertIn('new InteriorLightingColorPreset("Orange ambre", new Color(1f, 0.55f, 0.12f, 1f))', runtime_patch_text)
            self.assertIn('new InteriorLightingColorPreset("Vert sous-marin", new Color(0.12f, 0.78f, 0.28f, 1f))', runtime_patch_text)
            self.assertIn('new InteriorLightingColorPreset("Rouge", new Color(0.95f, 0.12f, 0.10f, 1f))', runtime_patch_text)
            self.assertIn('new InteriorLightingColorPreset("Blanc froid", new Color(0.78f, 0.90f, 1f, 1f))', runtime_patch_text)
            self.assertIn("public static int LightingColorPresetCount", runtime_patch_text)
            self.assertIn("public static List<string> GetLightingColorPresetNames()", runtime_patch_text)
            self.assertIn("public static Color GetLightingColorPresetColor(int presetIndex)", runtime_patch_text)
            self.assertIn("private static Color AlarmColor", runtime_patch_text)
            self.assertIn("private static Color SilentRunColor", runtime_patch_text)
            self.assertNotIn("AlarmOrangeColor", runtime_patch_text)
            self.assertNotIn("SilentRunGreenColor", runtime_patch_text)
            self.assertIn("public static bool IsEnabled()", runtime_patch_text)
            self.assertIn("return LongSubmergedRuntimeSettings.InteriorLightingColors;", runtime_patch_text)
            self.assertIn("public static void ApplyAll(string reason)", runtime_patch_text)
            self.assertIn("InteriorLightingColorPatcher.ApplyAll", runtime_patch_text)
            self.assertIn("typeof(InteriorLightingPlayerShipInteriorLightingAwakePatch)", runtime_patch_text)
            self.assertIn("typeof(InteriorLightingPlayerShipInteriorLightingStartPatch)", runtime_patch_text)
            self.assertIn("typeof(InteriorLightingPlayerShipInteriorLightingApplyPresetsPatch)", runtime_patch_text)
            self.assertIn("typeof(InteriorLightingLightControllerUpdatePresetsPatch)", runtime_patch_text)
            self.assertIn("typeof(InteriorLightingFillLightUpdatePresetsPatch)", runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting), "Awake")', runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting), "Start")', runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting), "ApplyLightControllersPresets")', runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(UBOAT.Game.Scene.Effects.LightController), "UpdatePresets", new Type[] { typeof(float[]), typeof(float[]) })', runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(UBOAT.Game.Scene.Effects.FillLight), "UpdatePresets", new Type[] { typeof(float[]) })', runtime_patch_text)
            self.assertIn('AccessTools.Field(typeof(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting), "alarmInteriorFogColor")', runtime_patch_text)
            self.assertIn('AccessTools.Field(typeof(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting), "silentRunInteriorFogColor")', runtime_patch_text)
            self.assertIn('AccessTools.Field(typeof(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting), "alarmLightsColorMultiplier")', runtime_patch_text)
            self.assertIn('AccessTools.Field(typeof(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting), "silentRunLightsColorMultiplier")', runtime_patch_text)
            self.assertIn('AccessTools.Method(typeof(UBOAT.Game.Scene.Effects.PlayerShipInteriorLighting), "ApplyColorMultiplier")', runtime_patch_text)
            self.assertIn('SetColorProperty(controller, "AlarmColor", controller.AlarmColor, AlarmColor);', runtime_patch_text)
            self.assertIn('SetColorProperty(controller, "BlueColor", controller.BlueColor, SilentRunColor);', runtime_patch_text)
            self.assertIn('SetColorProperty(fillLight, "RedColor", fillLight.RedColor, AlarmColor);', runtime_patch_text)
            self.assertIn('SetColorProperty(fillLight, "BlueColor", fillLight.BlueColor, SilentRunColor);', runtime_patch_text)
            self.assertIn('SetColorField(lighting, AlarmInteriorFogColorField, "alarmInteriorFogColor", AlarmColor);', runtime_patch_text)
            self.assertIn('SetColorField(lighting, SilentRunInteriorFogColorField, "silentRunInteriorFogColor", SilentRunColor);', runtime_patch_text)
            self.assertIn("RestoreInteriorLighting", runtime_patch_text)
            self.assertIn("RestoreLightController", runtime_patch_text)
            self.assertIn("RestoreFillLight", runtime_patch_text)
            self.assertIn("RememberColorPatch", runtime_patch_text)
            self.assertIn("TryConsumeColorPatch", runtime_patch_text)
            consume_color_patch_text = extract_csharp_block(runtime_patch_text, "private static bool TryConsumeColorPatch")
            self.assertLess(
                consume_color_patch_text.index("if (!ColorsEqual(current, stored.PatchedValue))"),
                consume_color_patch_text.index("data.Values.Remove(memberName);"),
            )
            self.assertIn("ObjectColorPatches", runtime_patch_text)
            self.assertIn("WarnMissingMember", runtime_patch_text)
            self.assertNotIn("OrangeAlarmLightingPatcher", runtime_patch_text)
            self.assertNotIn("SwitchLightAction", runtime_patch_text)
            self.assertNotIn("HarmonyPatch(typeof(SwitchLightAction)", runtime_patch_text)
            self.assertIn("Runtime Unity UI menu controller ready on F10", runtime_patch_text)
            self.assertIn("using UnityEngine.UI;", runtime_patch_text)
            self.assertIn("using UBOAT.Game.Core.AI;", runtime_patch_text)
            self.assertIn("using UBOAT.Game.Core.AI.GroupBehaviours;", runtime_patch_text)
            self.assertIn("using UBOAT.Game.Scene.Characters;", runtime_patch_text)
            self.assertIn("using UBOAT.Game.Scene.Utilities;", runtime_patch_text)
            self.assertIn("using UBOAT.Game.Sandbox;", runtime_patch_text)
            self.assertIn("Canvas", runtime_patch_text)
            self.assertIn("RenderMode.ScreenSpaceOverlay", runtime_patch_text)
            self.assertIn("CanvasScaler", runtime_patch_text)
            self.assertIn("GraphicRaycaster", runtime_patch_text)
            self.assertIn("Button", runtime_patch_text)
            self.assertIn("Toggle", runtime_patch_text)
            self.assertIn("Slider", runtime_patch_text)
            self.assertIn("CreateFactorSlider", runtime_patch_text)
            self.assertIn("OnFactorSliderChanged", runtime_patch_text)
            self.assertIn("Par defaut", runtime_patch_text)
            self.assertIn("SuperVitesse", runtime_patch_text)
            self.assertIn("Vitesses rapides", runtime_patch_text)
            self.assertIn("private Button callReinforcementsButton;", runtime_patch_text)
            self.assertIn('callReinforcementsButton = CreateButton(panelObject.transform, "Appeler renforts"', runtime_patch_text)
            self.assertIn("callReinforcementsButton.onClick.AddListener(OnCallReinforcementsClicked);", runtime_patch_text)
            self.assertIn("private void OnCallReinforcementsClicked()", runtime_patch_text)
            self.assertIn('ReinforcementRuntimeController.CallReinforcements("unity ui call reinforcements")', runtime_patch_text)
            self.assertIn("private Text reinforcementsStatusText;", runtime_patch_text)
            self.assertIn("RefreshReinforcementsStatus", runtime_patch_text)
            self.assertIn('callReinforcementsButton.interactable = availabilityStatus == "Pret";', runtime_patch_text)
            self.assertIn("internal static class ReinforcementRuntimeController", runtime_patch_text)
            self.assertIn("private const float ReinforcementCooldownSeconds = 300f;", runtime_patch_text)
            self.assertIn("private static bool reinforcementCallInProgress;", runtime_patch_text)
            self.assertIn("private static float nextAllowedReinforcementCallTime;", runtime_patch_text)
            self.assertIn("if (reinforcementCallInProgress)", runtime_patch_text)
            self.assertIn("nextAllowedReinforcementCallTime = Time.unscaledTime + ReinforcementCooldownSeconds;", runtime_patch_text)
            self.assertIn("ActiveReinforcementGroups", runtime_patch_text)
            self.assertIn("private const float ReinforcementActiveTrackingSeconds = 900f;", runtime_patch_text)
            self.assertIn("ActiveReinforcementGroupTrackedAt", runtime_patch_text)
            self.assertIn("Time.unscaledTime - trackedAt >= ReinforcementActiveTrackingSeconds", runtime_patch_text)
            self.assertIn("Reinforcement group tracking expired", runtime_patch_text)
            self.assertIn("DestroyCreatedGroups(primaryGroups, \"primary incomplete\")", runtime_patch_text)
            self.assertIn("Reinforcement call requested", runtime_patch_text)
            self.assertIn("Reinforcement call skipped: cooldown active", runtime_patch_text)
            self.assertIn("Reinforcement call skipped: already running", runtime_patch_text)
            self.assertIn("Reinforcement call spawned primary groups", runtime_patch_text)
            self.assertIn("Reinforcement call spawned manual fallback U-boats", runtime_patch_text)
            self.assertIn("Reinforcement call failed", runtime_patch_text)
            self.assertIn("MissionUtility.SpawnPatrol", runtime_patch_text)
            self.assertIn("Country.Relation.Ally", runtime_patch_text)
            self.assertIn('"Air Patrol"', runtime_patch_text)
            self.assertIn('"Warships"', runtime_patch_text)
            self.assertIn('"Submarine"', runtime_patch_text)
            self.assertIn("private const int DesiredFallbackUboatCount = 2;", runtime_patch_text)
            self.assertIn('"Type VIIC"', runtime_patch_text)
            self.assertIn('"Type VIIB"', runtime_patch_text)
            self.assertIn('"Type VIIC41"', runtime_patch_text)
            self.assertIn('"Type IID"', runtime_patch_text)
            self.assertIn('"Type IIB"', runtime_patch_text)
            self.assertIn('"Type IIA"', runtime_patch_text)
            self.assertIn("CreateManualFriendlyUboats", runtime_patch_text)
            self.assertIn("SandboxGroup.Create<SandboxMobileGroup>", runtime_patch_text)
            self.assertIn("SandboxEntity.Create(submarineTypeName, country)", runtime_patch_text)
            self.assertIn("group.AddEntity(entity);", runtime_patch_text)
            self.assertIn("bool entityAttachedToGroup = false;", runtime_patch_text)
            self.assertIn("entityAttachedToGroup = true;", runtime_patch_text)
            self.assertIn("if (!entityAttachedToGroup && entity != null)", runtime_patch_text)
            self.assertIn("sandbox.AddGroup(group);", runtime_patch_text)
            self.assertIn("group.gameObject.AddComponent<CharacterAI>()", runtime_patch_text)
            self.assertIn("new SailToBehaviour(ai, 1.5f, rallyPoint)", runtime_patch_text)
            self.assertIn("AIBehaviourFlags.OneShot", runtime_patch_text)
            self.assertIn("GetGroupsInRange(position, FallbackGroupClearance, false)", runtime_patch_text)
            self.assertIn('AccessTools.Field(typeof(SandboxGroup), "worldNavMesh")', runtime_patch_text)
            self.assertIn("ResolveWorldNavMesh", runtime_patch_text)
            self.assertIn("worldNavMesh.SnapWorld(position)", runtime_patch_text)
            self.assertIn("worldNavMesh.IsOnNavMesh(position)", runtime_patch_text)
            self.assertIn("worldNavMesh.RaycastLandsNavMesh(position, playerGroup.Position, out hit)", runtime_patch_text)
            self.assertIn("Manual U-boat fallback navmesh check skipped", runtime_patch_text)
            self.assertIn('"1 U-boat appele"', runtime_patch_text)
            self.assertIn('" U-boats appeles"', runtime_patch_text)
            self.assertIn('"Aucun U-boat ami disponible"', runtime_patch_text)
            self.assertNotIn('"Air Raid"', runtime_patch_text)
            self.assertNotIn("AirRaid", runtime_patch_text)
            self.assertNotIn("EngageBehaviour", runtime_patch_text)
            self.assertNotIn("RelationsManager", runtime_patch_text)
            self.assertNotRegex(runtime_patch_text, r"\b(?:Set|Force|Override)\w*Relation")
            self.assertNotRegex(runtime_patch_text, r"PlayerPrefs\.(?:Get|Set)\w+\([^;\n]*(?:Reinforcement|Reinforcements|Renfort)")
            self.assertNotIn('CreateToggle(panelObject.transform, "Appeler renforts"', runtime_patch_text)
            self.assertNotIn('GUILayout.Button("Appeler renforts"', runtime_patch_text)
            self.assertNotIn("private void OnGUI", runtime_patch_text)
            self.assertNotIn("GUI.Window", runtime_patch_text)
            self.assertNotIn("GUILayout", runtime_patch_text)
            reinforcement_controller_text = extract_csharp_block(runtime_patch_text, "internal static class ReinforcementRuntimeController")
            manual_fallback_text = reinforcement_controller_text[
                reinforcement_controller_text.index("private static int CreateManualFriendlyUboats") :
                reinforcement_controller_text.index("private static SandboxGroup ResolvePlayerGroup")
            ]
            call_reinforcements_text = extract_csharp_block(reinforcement_controller_text, "public static string CallReinforcements")
            self.assertNotIn("MissionUtility.SpawnPatrol", manual_fallback_text)
            self.assertNotIn("GroupSpawner.SpawnInstantly", manual_fallback_text)
            self.assertNotIn("UnityEngine.Random", manual_fallback_text)
            self.assertNotIn('"(Player)"', manual_fallback_text)
            self.assertIn("StartReinforcementCooldown();", call_reinforcements_text)
            self.assertNotIn("nextAllowedReinforcementCallTime = Time.unscaledTime + ReinforcementCooldownSeconds;", call_reinforcements_text)
            self.assertIn("PlayerPrefs", runtime_patch_text)
            self.assertIn("MegaBattery", runtime_patch_text)
            self.assertIn("MegaOxygen", runtime_patch_text)
            self.assertIn("SuperSpeed", runtime_patch_text)
            self.assertIn("MegaTorpedoes", runtime_patch_text)
            self.assertIn("MegaSonar", runtime_patch_text)
            self.assertIn("HeavyArmor", runtime_patch_text)
            self.assertIn("SuperStealth", runtime_patch_text)
            self.assertIn("BatteryFactor", runtime_patch_text)
            self.assertIn("OxygenFactor", runtime_patch_text)
            self.assertIn("SpeedFactor", runtime_patch_text)
            self.assertIn("TorpedoFactor", runtime_patch_text)
            self.assertIn("SonarFactor", runtime_patch_text)
            self.assertIn("MinRuntimeFactor = 1f", runtime_patch_text)
            self.assertIn("MaxRuntimeFactor = BatteryMaxFactor", runtime_patch_text)
            self.assertIn("ResetToDefaults", runtime_patch_text)
            self.assertIn("RuntimeSettingsVersion = 19", runtime_patch_text)
            self.assertIn("BatteryMaxFactor = 100f", runtime_patch_text)
            self.assertIn("OxygenMaxFactor = 100f", runtime_patch_text)
            self.assertIn("SpeedMaxFactor = 20f", runtime_patch_text)
            self.assertIn("TorpedoMaxFactor = 10f", runtime_patch_text)
            self.assertIn("SonarMaxFactor = 10f", runtime_patch_text)
            self.assertIn("DefaultSonarFactor = 3f", runtime_patch_text)
            self.assertIn("HeavyArmorDamageFactor = 3f", runtime_patch_text)
            self.assertIn("SuperStealthFactor = 3f", runtime_patch_text)
            self.assertIn("DefaultHeavyArmor = false", runtime_patch_text)
            self.assertIn("DefaultSuperStealth = false", runtime_patch_text)
            self.assertIn("DefaultDeepDive = true", runtime_patch_text)
            self.assertIn("DefaultInteriorLightingColors = true", runtime_patch_text)
            self.assertIn("DefaultInteriorLightingAlarmColorPresetIndex = 0", runtime_patch_text)
            self.assertIn("DefaultInteriorLightingSilentRunColorPresetIndex = 1", runtime_patch_text)
            self.assertIn("public static int InteriorLightingAlarmColorPresetIndex = DefaultInteriorLightingAlarmColorPresetIndex", runtime_patch_text)
            self.assertIn("public static int InteriorLightingSilentRunColorPresetIndex = DefaultInteriorLightingSilentRunColorPresetIndex", runtime_patch_text)
            self.assertIn("public static Color InteriorLightingAlarmColor", runtime_patch_text)
            self.assertIn("public static Color InteriorLightingSilentRunColor", runtime_patch_text)
            self.assertIn('ReadBool("HeavyArmor", DefaultHeavyArmor)', runtime_patch_text)
            self.assertIn('ReadBool("SuperStealth", DefaultSuperStealth)', runtime_patch_text)
            self.assertIn('ReadBool("DeepDive", DefaultDeepDive)', runtime_patch_text)
            self.assertIn('ReadBool("InteriorLightingColors", DefaultInteriorLightingColors)', runtime_patch_text)
            self.assertIn('ReadInteriorLightingColorPresetIndex(\n                "InteriorLightingAlarmColorPresetIndex"', runtime_patch_text)
            self.assertIn('ReadInteriorLightingColorPresetIndex(\n                "InteriorLightingSilentRunColorPresetIndex"', runtime_patch_text)
            self.assertIn("ClampInteriorLightingColorPresetIndex", runtime_patch_text)
            self.assertIn('PlayerPrefs.SetInt(PrefPrefix + "HeavyArmor", HeavyArmor ? 1 : 0);', runtime_patch_text)
            self.assertIn('PlayerPrefs.SetInt(PrefPrefix + "SuperStealth", SuperStealth ? 1 : 0);', runtime_patch_text)
            self.assertIn('PlayerPrefs.SetInt(PrefPrefix + "DeepDive", DeepDive ? 1 : 0);', runtime_patch_text)
            self.assertIn('PlayerPrefs.SetInt(PrefPrefix + "InteriorLightingColors", InteriorLightingColors ? 1 : 0);', runtime_patch_text)
            self.assertIn('PlayerPrefs.SetInt(PrefPrefix + "InteriorLightingAlarmColorPresetIndex", InteriorLightingAlarmColorPresetIndex);', runtime_patch_text)
            self.assertIn('PlayerPrefs.SetInt(PrefPrefix + "InteriorLightingSilentRunColorPresetIndex", InteriorLightingSilentRunColorPresetIndex);', runtime_patch_text)
            self.assertIn("HeavyArmor = DefaultHeavyArmor", runtime_patch_text)
            self.assertIn("if (savedVersion < 16)", runtime_patch_text)
            self.assertIn("HeavyArmor = false;", runtime_patch_text)
            self.assertIn("SuperStealth = DefaultSuperStealth", runtime_patch_text)
            self.assertIn("DeepDive = DefaultDeepDive", runtime_patch_text)
            self.assertIn("InteriorLightingColors = DefaultInteriorLightingColors", runtime_patch_text)
            self.assertIn("InteriorLightingAlarmColorPresetIndex = DefaultInteriorLightingAlarmColorPresetIndex", runtime_patch_text)
            self.assertIn("InteriorLightingSilentRunColorPresetIndex = DefaultInteriorLightingSilentRunColorPresetIndex", runtime_patch_text)
            settings_load_text = extract_csharp_block(runtime_patch_text, "public static void Load()")
            settings_save_text = extract_csharp_block(runtime_patch_text, "public static void Save()")
            settings_reset_text = extract_csharp_block(runtime_patch_text, "public static void ResetToDefaults()")
            self.assertLess(
                settings_load_text.index('HeavyArmor = ReadBool("HeavyArmor", DefaultHeavyArmor);'),
                settings_load_text.index("if (savedVersion < RuntimeSettingsVersion)"),
            )
            migration_text = settings_load_text[
                settings_load_text.index("if (savedVersion < RuntimeSettingsVersion)") :
                settings_load_text.index("Debug.Log")
            ]
            self.assertIn("if (savedVersion < 16)", migration_text)
            self.assertIn("HeavyArmor = false;", migration_text)
            self.assertIn("Save();", migration_text)
            self.assertNotIn("MegaBattery =", migration_text)
            self.assertNotIn("SuperStealth =", migration_text)
            self.assertNotIn("DeepDive =", migration_text)
            self.assertNotIn("InteriorLightingColors =", migration_text)
            self.assertNotIn("InteriorLightingAlarmColorPresetIndex =", migration_text)
            self.assertNotIn("InteriorLightingSilentRunColorPresetIndex =", migration_text)
            self.assertIn('PlayerPrefs.SetInt(PrefPrefix + "HeavyArmor", HeavyArmor ? 1 : 0);', settings_save_text)
            self.assertIn('PlayerPrefs.SetInt(PrefPrefix + "DeepDive", DeepDive ? 1 : 0);', settings_save_text)
            self.assertIn('PlayerPrefs.SetInt(PrefPrefix + "InteriorLightingColors", InteriorLightingColors ? 1 : 0);', settings_save_text)
            self.assertIn('PlayerPrefs.SetInt(PrefPrefix + "InteriorLightingAlarmColorPresetIndex", InteriorLightingAlarmColorPresetIndex);', settings_save_text)
            self.assertIn('PlayerPrefs.SetInt(PrefPrefix + "InteriorLightingSilentRunColorPresetIndex", InteriorLightingSilentRunColorPresetIndex);', settings_save_text)
            self.assertIn('PlayerPrefs.SetInt(PrefPrefix + "RuntimeSettingsVersion", RuntimeSettingsVersion);', settings_save_text)
            self.assertIn("HeavyArmor = DefaultHeavyArmor;", settings_reset_text)
            self.assertIn("DeepDive = DefaultDeepDive;", settings_reset_text)
            self.assertIn("InteriorLightingColors = DefaultInteriorLightingColors;", settings_reset_text)
            self.assertIn("InteriorLightingAlarmColorPresetIndex = DefaultInteriorLightingAlarmColorPresetIndex;", settings_reset_text)
            self.assertIn("InteriorLightingSilentRunColorPresetIndex = DefaultInteriorLightingSilentRunColorPresetIndex;", settings_reset_text)
            self.assertIn('CreateToggle(panelObject.transform, "Blindage lourd x3"', runtime_patch_text)
            self.assertIn('CreateToggle(panelObject.transform, "Plongée x2"', runtime_patch_text)
            self.assertIn('CreateToggle(panelObject.transform, "Couleurs eclairage"', runtime_patch_text)
            self.assertNotIn('CreateToggle(panelObject.transform, "Lumières orange/vert"', runtime_patch_text)
            self.assertIn('CreateToggle(panelObject.transform, "Super discrétion"', runtime_patch_text)
            self.assertIn("private Dropdown alarmColorDropdown;", runtime_patch_text)
            self.assertIn("private Dropdown silentRunColorDropdown;", runtime_patch_text)
            self.assertIn("private Image alarmColorSwatch;", runtime_patch_text)
            self.assertIn("private Image silentRunColorSwatch;", runtime_patch_text)
            self.assertIn('alarmColorDropdown = CreateColorDropdown(panelObject.transform, "Alarme"', runtime_patch_text)
            self.assertIn('silentRunColorDropdown = CreateColorDropdown(panelObject.transform, "Silencieux"', runtime_patch_text)
            self.assertIn("private Dropdown CreateColorDropdown", runtime_patch_text)
            self.assertIn("private RectTransform CreateDropdownTemplate", runtime_patch_text)
            self.assertIn("dropdown.AddOptions(InteriorLightingColorPatcher.GetLightingColorPresetNames());", runtime_patch_text)
            self.assertIn("dropdown.onValueChanged.AddListener(OnColorDropdownChanged);", runtime_patch_text)
            self.assertIn("private void OnColorDropdownChanged(int ignored)", runtime_patch_text)
            self.assertIn('SaveAndApplyCurrentControlsNow("unity ui color dropdown")', runtime_patch_text)
            self.assertIn("SetColorDropdownValue(alarmColorDropdown, LongSubmergedRuntimeSettings.InteriorLightingAlarmColorPresetIndex);", runtime_patch_text)
            self.assertIn("SetColorSwatch(alarmColorSwatch, LongSubmergedRuntimeSettings.InteriorLightingAlarmColor);", runtime_patch_text)
            self.assertIn("ReadColorDropdownPresetIndex(alarmColorDropdown, LongSubmergedRuntimeSettings.InteriorLightingAlarmColorPresetIndex)", runtime_patch_text)
            self.assertIn("Dropdown dropdown = dropdownObject.AddComponent<Dropdown>();", runtime_patch_text)
            self.assertIn("ScrollRect scrollRect = template.AddComponent<ScrollRect>();", runtime_patch_text)
            self.assertIn("Mask viewportMask = viewport.AddComponent<Mask>();", runtime_patch_text)
            self.assertIn("heavyArmorToggle.isOn = LongSubmergedRuntimeSettings.HeavyArmor", runtime_patch_text)
            self.assertIn("superStealthToggle.isOn = LongSubmergedRuntimeSettings.SuperStealth", runtime_patch_text)
            self.assertIn("deepDiveToggle.isOn = LongSubmergedRuntimeSettings.DeepDive", runtime_patch_text)
            self.assertIn("interiorLightingToggle.isOn = LongSubmergedRuntimeSettings.InteriorLightingColors", runtime_patch_text)
            self.assertIn("LongSubmergedRuntimeSettings.HeavyArmor = heavyArmorToggle != null && heavyArmorToggle.isOn", runtime_patch_text)
            self.assertIn("LongSubmergedRuntimeSettings.SuperStealth = superStealthToggle != null && superStealthToggle.isOn", runtime_patch_text)
            self.assertIn("LongSubmergedRuntimeSettings.DeepDive = deepDiveToggle != null && deepDiveToggle.isOn", runtime_patch_text)
            self.assertIn("LongSubmergedRuntimeSettings.InteriorLightingColors = interiorLightingToggle != null && interiorLightingToggle.isOn", runtime_patch_text)
            self.assertIn("LongSubmergedRuntimeSettings.InteriorLightingAlarmColorPresetIndex =", runtime_patch_text)
            self.assertIn("LongSubmergedRuntimeSettings.InteriorLightingSilentRunColorPresetIndex =", runtime_patch_text)
            self.assertIn("ClampSonarFactor", runtime_patch_text)
            self.assertIn("SaveAndApplyCurrentControlsNow", runtime_patch_text)
            self.assertIn('SaveAndApplyCurrentControlsNow("unity ui toggle")', runtime_patch_text)
            self.assertIn('SaveAndApplyCurrentControlsNow("unity ui slider")', runtime_patch_text)
            self.assertIn('SaveAndApplyCurrentControlsNow("unity ui refresh")', runtime_patch_text)
            self.assertIn('SaveAndApplyCurrentControlsNow("unity ui call reinforcements")', runtime_patch_text)
            self.assertIn('LongSubmergedRuntimeApplier.ApplyAll(string.IsNullOrEmpty(reason) ? "unity ui change" : reason);', runtime_patch_text)
            self.assertIn("rowHitImage.color = new Color(1f, 1f, 1f, 0f);", runtime_patch_text)
            self.assertIn("rowHitImage.raycastTarget = true;", runtime_patch_text)
            self.assertIn("try\n            {\n                if (megaBatteryToggle != null)", runtime_patch_text)
            self.assertIn("finally\n            {\n                suppressToggleEvents = false;", runtime_patch_text)
            self.assertNotIn("RuntimeApplyDebounceSeconds", runtime_patch_text)
            self.assertNotIn("QueueRuntimeApply", runtime_patch_text)
            self.assertNotIn("FlushPendingRuntimeApply", runtime_patch_text)
            self.assertNotIn("pendingRuntimeApply", runtime_patch_text)
            self.assertIn("RunBatteryMaintenanceTick", runtime_patch_text)
            self.assertIn("MegaSonarMaintenanceIntervalSeconds = 1.00f", runtime_patch_text)
            self.assertIn("RunMegaSonarMaintenanceTick", runtime_patch_text)
            self.assertIn("MegaSonarRuntimePatcher.ApplyAll", runtime_patch_text)
            self.assertIn("ApplyBatteryObject", runtime_patch_text)
            self.assertIn("ApplyBatteryResource", runtime_patch_text)
            self.assertIn("MaintainBatteryRuntime", runtime_patch_text)
            self.assertIn("ApplyBatteryConsumers", runtime_patch_text)
            self.assertIn("ApplyGenericBatteryConsumers", runtime_patch_text)
            self.assertIn("GenericEnergyUsageFieldCache", runtime_patch_text)
            self.assertIn("IsEnergyUsageMemberName", runtime_patch_text)
            self.assertIn("ResourceGuiGetTooltipContentsPatch", runtime_patch_text)
            self.assertIn("ResourceGuiUpdateDisplayedValuePatch", runtime_patch_text)
            self.assertIn("MegaSonarHydrophoneRefreshPatch", runtime_patch_text)
            self.assertIn("typeof(SuperStealthEntityUpdateDetectabilityPatch)", runtime_patch_text)
            self.assertIn("typeof(SuperStealthAirCompressorOnEnablePatch)", runtime_patch_text)
            self.assertIn("typeof(SuperStealthAirCompressorOnDisablePatch)", runtime_patch_text)
            self.assertIn("typeof(SuperStealthVentilationOnEnablePatch)", runtime_patch_text)
            self.assertIn("typeof(SuperStealthVentilationOnDisablePatch)", runtime_patch_text)
            self.assertIn("typeof(SuperStealthPropellerPowerPatch)", runtime_patch_text)
            self.assertIn("typeof(SuperStealthPropellerPowerMultiplierPatch)", runtime_patch_text)
            self.assertIn("typeof(SuperStealthSnorkelUpdatePatch)", runtime_patch_text)
            self.assertIn("typeof(SuperStealthPeriscopeUpdatePatch)", runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(Entity), "UpdateDetectability")', runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(AirCompressor), "OnEnable")', runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(AirCompressor), "OnDisable")', runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(Ventilation), "OnEnable")', runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(Ventilation), "OnDisable")', runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(Propeller), "set_Power")', runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(Propeller), "set_PowerMultiplier")', runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(Snorkel), "Update")', runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(Periscope), "Update")', runtime_patch_text)
            self.assertIn("SuperStealthRuntimePatcher", runtime_patch_text)
            self.assertIn("LongSubmerged10x Super Stealth Runtime", runtime_patch_text)
            self.assertIn("ApplyEntityDetectability", runtime_patch_text)
            self.assertIn("ApplySandboxDetectability", runtime_patch_text)
            self.assertIn("ApplyEquipment(Component component", runtime_patch_text)
            self.assertIn("entity.HydrophoneDetectability", runtime_patch_text)
            self.assertIn("entity.SonarDetectability", runtime_patch_text)
            self.assertIn("entity.VisualDetectability", runtime_patch_text)
            self.assertIn("entity.RadarDetectorDetectability", runtime_patch_text)
            self.assertIn("sandboxEntity.HydrophoneDetectability", runtime_patch_text)
            self.assertIn("sandboxEntity.RadarDetectability", runtime_patch_text)
            self.assertIn("sandboxEntity.IndirectVisualDetectability", runtime_patch_text)
            self.assertIn("sandboxEntity.SignatureRadius", runtime_patch_text)
            self.assertIn("ship.CrewNoiseModifier", runtime_patch_text)
            self.assertIn("ship.StationaryNoise", runtime_patch_text)
            self.assertIn("return 1f / LongSubmergedRuntimeSettings.SuperStealthFactor;", runtime_patch_text)
            self.assertIn("return 1f;", runtime_patch_text)
            self.assertIn("typeof(HeavyArmorHullAddDamagePatch)", runtime_patch_text)
            self.assertIn("typeof(HeavyArmorEquipmentAddDamagePatch)", runtime_patch_text)
            self.assertIn("typeof(HeavyArmorEquipmentAddWaterDamagePatch)", runtime_patch_text)
            self.assertIn("typeof(HeavyArmorPlayableCharacterAddDamagePatch)", runtime_patch_text)
            self.assertIn("typeof(HeavyArmorDamageUtilityDoApplyDamagePatch)", runtime_patch_text)
            self.assertIn("typeof(HeavyArmorDamageUtilityApplyDamageToComponentsPatch)", runtime_patch_text)
            self.assertIn("typeof(HeavyArmorApplyWaterDamageToPlayerShipDoDamageTickPatch)", runtime_patch_text)
            self.assertIn("typeof(HeavyArmorHullEffectsRendererApplyImpactPatch)", runtime_patch_text)
            self.assertIn("typeof(DeepDivePlayerShipTargetDepthSetterPatch)", runtime_patch_text)
            self.assertIn("typeof(DeepDiveHullCrushControllerDoUpdatePatch)", runtime_patch_text)
            self.assertIn("typeof(DeepDivePlayerShipUpdateStressAndDisciplineGainPatch)", runtime_patch_text)
            self.assertIn('AccessTools.Method(typeof(PlayerShip), "SetTargetDepth", PlayerShipSetTargetDepthParameterTypes)', runtime_patch_text)
            self.assertIn('FindMethodOnKnownType("HullCrushController", "DoUpdate")', runtime_patch_text)
            self.assertIn("FindPlayerShipUpdateStressAndDisciplineGainMethod", runtime_patch_text)
            self.assertIn("DeepDiveRuntimePatcher.ScaleTargetDepthCommand", runtime_patch_text)
            self.assertIn("DeepDiveRuntimePatcher.ShouldRunHullCrushDoUpdate", runtime_patch_text)
            self.assertIn("DeepDiveRuntimePatcher.ShouldSkipPressureWaterDamageTick", runtime_patch_text)
            self.assertIn("DeepDiveRuntimePatcher.ApplyDepthStressModifier", runtime_patch_text)
            self.assertIn('DeepDiveRuntimePatcher.ApplyDepthObject(__instance, "DivingPlanesStation.Awake")', runtime_patch_text)
            self.assertIn("public static bool IsEnabled()", runtime_patch_text)
            self.assertIn("return LongSubmergedRuntimeSettings.DeepDive;", runtime_patch_text)
            self.assertIn("if (!IsEnabled())\n                return;", runtime_patch_text)
            self.assertIn("RestoreDepthObject(ship, reason);", runtime_patch_text)
            self.assertIn("RestoreDepthObject(controller, reason);", runtime_patch_text)
            self.assertIn("return false;\n            }\n\n            ApplyDepthObject(controller, reason);", runtime_patch_text)
            self.assertIn("DepthObjectPatches", runtime_patch_text)
            self.assertIn("RememberPatchedMember", runtime_patch_text)
            self.assertIn("ShouldRestoreMemberValue", runtime_patch_text)
            self.assertIn("TryPatchHullCrushControllerData", runtime_patch_text)
            self.assertIn("GetHullCrushDepthTarget", runtime_patch_text)
            self.assertIn("return -CrushDepthMeters;", runtime_patch_text)
            self.assertIn("target - baseValue", runtime_patch_text)
            self.assertIn("return current < 0f ? -positiveMeters : positiveMeters;", runtime_patch_text)
            self.assertNotIn("CrushDepthMeters - baseValue", runtime_patch_text)
            self.assertIn('AccessTools.Method(typeof(Hull), "AddDamage", HeavyArmorRuntimePatcher.AddDamageParameterTypes)', runtime_patch_text)
            self.assertIn('AccessTools.Method(typeof(Equipment), "AddDamage", HeavyArmorRuntimePatcher.AddDamageParameterTypes)', runtime_patch_text)
            self.assertIn('AccessTools.Method(typeof(Equipment), "AddWaterDamage", HeavyArmorRuntimePatcher.AddWaterDamageParameterTypes)', runtime_patch_text)
            self.assertIn('AccessTools.Method(typeof(PlayableCharacter), "AddDamage", HeavyArmorRuntimePatcher.AddDamageParameterTypes)', runtime_patch_text)
            self.assertIn("DamageUtilityDoApplyDamageParameterTypes", runtime_patch_text)
            self.assertIn("DamageUtilityApplyDamageToComponentsParameterTypes", runtime_patch_text)
            self.assertIn('AccessTools.Method(typeof(ApplyWaterDamageToPlayerShip), "DoDamageTick", new Type[] { })', runtime_patch_text)
            self.assertIn("HullEffectsApplyImpactParameterTypes", runtime_patch_text)
            self.assertIn("HeavyArmorRuntimePatcher", runtime_patch_text)
            self.assertIn("HeavyArmorDamageScaleState", runtime_patch_text)
            self.assertIn("ScalePlayerEquipmentDamage", runtime_patch_text)
            self.assertIn("TryScalePlayerEquipmentDamage", runtime_patch_text)
            self.assertIn("TryScalePlayerWaterDamage", runtime_patch_text)
            self.assertIn("ScalePlayerCharacterDamage", runtime_patch_text)
            self.assertIn("ScalePlayerCrewDamage", runtime_patch_text)
            self.assertIn("ScalePlayerHullImpact", runtime_patch_text)
            self.assertIn("BeginDamageScaleScope", runtime_patch_text)
            self.assertIn("EndDamageScaleScope", runtime_patch_text)
            self.assertIn("BeginComponentDamageDistributionScope", runtime_patch_text)
            self.assertIn("EndComponentDamageDistributionScope", runtime_patch_text)
            self.assertIn("BeginPressureWaterDamageScope", runtime_patch_text)
            self.assertIn("EndPressureWaterDamageScope", runtime_patch_text)
            self.assertIn("RestoreComponentDistributionBudget", runtime_patch_text)
            self.assertIn("damageScaleScopeDepth", runtime_patch_text)
            self.assertIn("componentDamageDistributionScopeDepth", runtime_patch_text)
            self.assertIn("pressureWaterDamageScopeDepth", runtime_patch_text)
            self.assertIn("Finalizer(HeavyArmorDamageScaleState __state)", runtime_patch_text)
            self.assertIn("Postfix(HeavyArmorDamageScaleState __state, ref float __result)", runtime_patch_text)
            self.assertIn("value / LongSubmergedRuntimeSettings.HeavyArmorDamageFactor", runtime_patch_text)
            self.assertIn("return LongSubmergedRuntimeSettings.HeavyArmor", runtime_patch_text)
            self.assertIn("if (!ShouldScaleDamage(damage) || !IsPlayerShipEquipment(equipment))", runtime_patch_text)
            self.assertIn("ref float flawProbabilityFactor", runtime_patch_text)
            self.assertIn("ref float fireChance", runtime_patch_text)
            self.assertIn("flawProbabilityFactor = ScaleDamage(flawProbabilityFactor);", runtime_patch_text)
            self.assertIn("fireChance = ScaleDamage(fireChance);", runtime_patch_text)
            self.assertIn("!IsDamageScaleScopeActive()", runtime_patch_text)
            self.assertIn("equipment.ParentEntity", runtime_patch_text)
            self.assertIn("character.ParentEntity", runtime_patch_text)
            self.assertIn("renderer.ParentEntity", runtime_patch_text)
            self.assertIn("sandboxEntity.IsPlayerShip", runtime_patch_text)
            self.assertIn("if (!IsHeavyArmorActive() || !IsPlayerEntity(target) || !ShouldScaleRawValue(crewDamage))", runtime_patch_text)
            self.assertNotIn("ScalePlayerCrewAndEffects", runtime_patch_text)
            self.assertNotIn("CouldAffectPlayerShip", runtime_patch_text)
            self.assertNotIn("HarmonyPatch(typeof(HullCrushController)", runtime_patch_text)
            self.assertIn("HydrophoneRange", runtime_patch_text)
            self.assertIn("GroupHydrophoneRange", runtime_patch_text)
            self.assertIn("Hydrophone portee", runtime_patch_text)
            self.assertIn("LongSubmerged10x Mega Sonar Runtime", runtime_patch_text)
            self.assertIn("FindHydrophoneTargetMethods", runtime_patch_text)
            self.assertIn("NoiseHydrophoneRange", runtime_patch_text)
            self.assertIn("IsExcludedRangeName", runtime_patch_text)
            self.assertIn("new MegaSonarFloatMemberPatchData(currentValue, 1f, currentValue)", runtime_patch_text)
            self.assertIn("memberData.LastPatchedValue", runtime_patch_text)
            self.assertIn("currentValue - memberData.LastPatchedValue", runtime_patch_text)
            self.assertIn("memberData.LastAppliedFactor > 1.0001f", runtime_patch_text)
            self.assertIn("DepletingResourceNotificationDoUpdatePatch", runtime_patch_text)
            self.assertNotIn("typeof(DepletingResourceNotificationDoUpdatePatch)", runtime_patch_text)
            self.assertIn("using UBOAT.Game.Scene.Effects;", runtime_patch_text)
            self.assertIn("using UBOAT.Game.Scene.Tasks;", runtime_patch_text)
            self.assertNotIn("FlamesRegistrySpawnParticlesPatch", runtime_patch_text)
            self.assertNotIn("typeof(FlamesRegistrySpawnParticlesPatch)", runtime_patch_text)
            self.assertNotIn('HarmonyPatch(typeof(FlamesRegistry), "SpawnParticles", new Type[] { })', runtime_patch_text)
            self.assertNotIn("ShouldRunFlamesRegistrySpawnParticles", runtime_patch_text)
            self.assertNotIn("HandleFlamesRegistrySpawnParticlesException", runtime_patch_text)
            self.assertNotIn("FlamesRegistry particle guard prevented crash", runtime_patch_text)
            self.assertNotIn("flamesEffect2", runtime_patch_text)
            self.assertNotIn("NullReferenceException absorbe", runtime_patch_text)
            self.assertIn("ShouldSuppressBatteryDepletionUi", runtime_patch_text)
            self.assertIn("ApplyBatteryRuntimeToResource", runtime_patch_text)
            self.assertIn("ApplyNuclearBatteryCapacityOverride", runtime_patch_text)
            self.assertIn("RuntimeNuclearBatteryCapacityModifierName", runtime_patch_text)
            self.assertIn("NuclearBatteryCapacityFloor = 100000f", runtime_patch_text)
            self.assertIn("NuclearBatteryCapacityDeltaData", runtime_patch_text)
            self.assertIn("NuclearBatteryCapacityLoggedResourceIds", runtime_patch_text)
            self.assertIn("Mega Batterie nuclear capacity active", runtime_patch_text)
            self.assertIn("ClampBatteryAmountToCapacity", runtime_patch_text)
            self.assertIn("SetDelta", runtime_patch_text)
            self.assertIn("BuildInfiniteBatteryTooltip", runtime_patch_text)
            self.assertIn("Mega Batterie : batterie nucleaire active.", runtime_patch_text)
            self.assertIn("Decoche Mega Batterie dans F10 pour revenir a la batterie vanilla.", runtime_patch_text)
            self.assertIn('LongSubmergedRuntimeSettings.MegaBattery ? "inf" : null', runtime_patch_text)
            self.assertIn("MaintainInfiniteBatteryCharge", runtime_patch_text)
            self.assertIn("IsInfiniteBatteryRuntimeActive", runtime_patch_text)
            self.assertIn("return LongSubmergedRuntimeSettings.MegaBattery;", runtime_patch_text)
            self.assertNotIn("ClampBatteryFactor(LongSubmergedRuntimeSettings.BatteryFactor)\n                    >= LongSubmergedRuntimeSettings.BatteryMaxFactor", runtime_patch_text)
            self.assertNotIn("Passe le slider Batterie sous 100", runtime_patch_text)
            self.assertNotIn("UpdateBatteryResourceAfterVanilla", runtime_patch_text)
            self.assertIn("ResourceUpdateAmountBatteryPatch", runtime_patch_text)
            self.assertIn("typeof(ResourceUpdateAmountBatteryPatch)", runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(Resource), "UpdateAmount")', runtime_patch_text)
            self.assertIn("TryFreezeInfiniteBatteryResource", runtime_patch_text)
            self.assertIn("if (!IsInfiniteBatteryRuntimeActive())\n                    return false;", runtime_patch_text)
            self.assertIn("if (!IsPlayerShipEnergyResource(resource))\n                    return false;", runtime_patch_text)
            self.assertIn('"Resource.UpdateAmount.Prefix"', runtime_patch_text)
            self.assertNotIn("Postfix(Resource __instance)", runtime_patch_text)
            self.assertNotIn("return !LongSubmergedRuntimeApplier.TryUpdateBatteryResourceAmount", runtime_patch_text)
            self.assertIn("ResourcePlayerShipField", runtime_patch_text)
            self.assertIn("return object.ReferenceEquals(owner.Energy, resource);", runtime_patch_text)
            self.assertIn("energy.Amount = capacity", runtime_patch_text)
            self.assertIn("RuntimeBatteryGainModifierName", runtime_patch_text)
            self.assertIn("ApplyBatteryGainModifiers", runtime_patch_text)
            self.assertIn("ApplyBatteryGainParameter", runtime_patch_text)
            self.assertIn("BatteryGainDeltaData", runtime_patch_text)
            self.assertIn("ParameterDeltaPatchData", runtime_patch_text)
            self.assertIn("GetValueExcludingModifier", runtime_patch_text)
            self.assertIn("energy.GainSandboxTimeScale", runtime_patch_text)
            self.assertIn("Mega Batterie infinite gain guard active", runtime_patch_text)
            self.assertIn("return LongSubmergedRuntimeSettings.BatteryMaxFactor;", runtime_patch_text)
            self.assertNotIn("return LongSubmergedRuntimeSettings.ClampBatteryFactor(LongSubmergedRuntimeSettings.BatteryFactor);", runtime_patch_text)
            self.assertIn("baseValue = parameter.GetValueExcludingModifier(RuntimeScaleModifierName)", runtime_patch_text)
            self.assertNotIn("SetAmountQuiet", runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(PlayerShip), "Update")', runtime_patch_text)
            self.assertNotIn('HarmonyPatch(typeof(AirCompressor), "EnergyUsage_Changed")', runtime_patch_text)
            self.assertNotIn("typeof(AirCompressorOnEnablePatch)", runtime_patch_text)
            self.assertNotIn("typeof(AirCompressorEnergyUsageChangedPatch)", runtime_patch_text)
            self.assertNotIn("typeof(VentilationOnEnablePatch)", runtime_patch_text)
            self.assertNotIn('ApplyBatteryObject(__instance, "AirCompressor', runtime_patch_text)
            self.assertNotIn('ApplyBatteryObject(__instance, "Ventilation', runtime_patch_text)
            self.assertIn("BatteryObjectApplicationGuardIds", runtime_patch_text)
            self.assertIn("IsSurfaceAirRuntimeObject", runtime_patch_text)
            self.assertIn("IsSurfaceAirName", runtime_patch_text)
            self.assertIn("ApplyDirectEnergyGainModifier", runtime_patch_text)
            self.assertIn("private const float OxygenRuntimeMaxFactor = 250f;", runtime_patch_text)
            self.assertIn("GetEffectiveOxygenRuntimeFactor", runtime_patch_text)
            self.assertIn("ApplyOxygenBreathModifier", runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(PlayerShip), "ValidateOxygenBreathModifier")', runtime_patch_text)
            self.assertIn("typeof(PlayerShipValidateOxygenBreathModifierPatch)", runtime_patch_text)
            self.assertIn("OxygenModifierData", runtime_patch_text)
            self.assertIn("internal sealed class OxygenModifierPatchData", runtime_patch_text)
            self.assertIn("currentValue >= 0f", runtime_patch_text)
            self.assertIn("data.LastAppliedFactor > 1.0001f", runtime_patch_text)
            self.assertIn("currentValue - data.LastPatchedValue", runtime_patch_text)
            self.assertIn("data.OriginalValue / factor", runtime_patch_text)
            self.assertNotIn("OxygenVanillaRestoreFactor", runtime_patch_text)
            self.assertNotIn("GetEffectiveOxygenDataFactor", runtime_patch_text)
            self.assertNotIn("RestoreVanillaOxygenIfNeeded", runtime_patch_text)
            self.assertIn("GetEffectiveBatteryCapacityScale", runtime_patch_text)
            self.assertIn("GetEffectiveBatteryEnergyUsageScale", runtime_patch_text)
            self.assertIn("ApplyStoredTorpedo", runtime_patch_text)
            self.assertIn("ApplyLaunchedTorpedo", runtime_patch_text)
            self.assertIn("LongSubmergedRuntimeSettings.TorpedoFactor = ReadSliderFactor(torpedoFactorSlider, LongSubmergedRuntimeSettings.TorpedoMaxFactor);", runtime_patch_text)
            self.assertIn("ApplyLockedTargetGuidance", runtime_patch_text)
            self.assertIn("RestoreLockedTargetGuidance", runtime_patch_text)
            self.assertIn("RestoreLockedTargetGuidance(torpedo);", runtime_patch_text)
            self.assertNotIn("ApplyLockedTargetGuidance(torpedo, reason);", runtime_patch_text)
            self.assertIn("Mathf.Min(torpedoFactor, 3f)", runtime_patch_text)
            self.assertIn("PredictLockedTargetPoint", runtime_patch_text)
            self.assertIn("TryForceLockedTargetDetonation", runtime_patch_text)
            self.assertIn("TorpedoGuidanceLeadSeconds = 4f", runtime_patch_text)
            self.assertIn("TorpedoGuidanceMinimumDetonationDistance = 20f", runtime_patch_text)
            self.assertIn("TorpedoGuidanceMaximumDetonationDistance = 80f", runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(Torpedo), "FixedUpdate")', runtime_patch_text)
            self.assertIn("torpedo.GyroAngle = float.NaN", runtime_patch_text)
            self.assertIn("torpedo.TargetPosition = targetPoint", runtime_patch_text)
            self.assertIn("TorpedoHomingTargetField", runtime_patch_text)
            self.assertIn("TorpedoDoExplosionHitMethod.Invoke", runtime_patch_text)
            self.assertIn("TorpedoDetonateMethod.Invoke", runtime_patch_text)
            self.assertIn("ForcingDetonation", runtime_patch_text)
            self.assertIn("GetEffectiveTorpedoFactor", runtime_patch_text)
            self.assertIn("IsMegaTorpedoRuntimeActive", runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(AccumulatorsUpgrade), "Start")', runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(StoredTorpedo), "Start")', runtime_patch_text)
            self.assertIn('HarmonyPatch(typeof(Torpedo), "Detonate")', runtime_patch_text)
            self.assertIn("private const float TorpedoDamageScale = 10f;", runtime_patch_text)
            self.assertIn("private const float TorpedoCrewDamageScale = 10f;", runtime_patch_text)
            self.assertIn("private const float TorpedoExplosionRadiusScale = 3f;", runtime_patch_text)
            self.assertIn("private const float TorpedoExplosionIntensityScale = 3f;", runtime_patch_text)
            self.assertIn("private const bool PerfectTorpedoReliability = true;", runtime_patch_text)
            self.assertIn("private const bool DefaultMegaTorpedoes = true;", runtime_patch_text)
            self.assertIn("GetEffectiveFastSpeedFactor", runtime_patch_text)
            self.assertIn("GetEffectiveFastFuelFactor", runtime_patch_text)

            root_ws = load_workbook(out_mod / "Data Sheets" / "General.xlsx", data_only=False)["Settings"]
            realistic_ws = load_workbook(out_mod / "Data Sheets" / "Realistic Travel" / "General.xlsx", data_only=False)["Settings"]

            root_ids = [root_ws.cell(row, 1).value for row in range(1, root_ws.max_row + 1)]
            realistic_ids = [realistic_ws.cell(row, 1).value for row in range(1, realistic_ws.max_row + 1)]
            discipline_row = find_row_by_id(root_ws, "Underwater Discipline Loss")

            self.assertNotIn("Oxygen Consumption Per Character", root_ids)
            self.assertNotIn("Oxygen Consumption Per Character", realistic_ids)
            self.assertAlmostEqual(root_ws.cell(discipline_row, 2).value, 0.000015 / 15)
            self.assertAlmostEqual(root_ws.cell(discipline_row, 3).value, 0.000016 / 15)
            self.assertAlmostEqual(root_ws.cell(discipline_row, 4).value, 0.000017 / 15)

            entities_ws = load_workbook(out_mod / "Data Sheets" / "Entities.xlsx", data_only=False)["Equipment"]
            ids = [entities_ws.cell(row, 1).value for row in range(1, entities_ws.max_row + 1)]

            self.assertIn("Accumulators I", ids)
            self.assertIn("Electric Engines", ids)
            self.assertNotIn("Diesel Engines", ids)
            self.assertNotIn("Atmosphere Tank", ids)
            self.assertNotIn("Ventilation", ids)
            self.assertNotIn("G7a Torpedo T1 - Pi1", ids)
            self.assertNotIn("Bow Torpedo Launcher", ids)

            accumulator_row = find_row_by_id(entities_ws, "Accumulators I")
            engine_row = find_row_by_id(entities_ws, "Electric Engines")

            self.assertIn("250%", entities_ws.cell(accumulator_row, 16).value)
            self.assertIn("EnergyUsage = 3.5e-06", entities_ws.cell(engine_row, 16).value)

            report_text = (out_mod / "LongSubmerged10x_generation_report.txt").read_text(encoding="utf-8")
            generated_readme_text = (out_mod / "README_LongSubmerged10x.txt").read_text(encoding="utf-8")
            self.assertNotIn("air_capacity_parameter_rows:", report_text)
            self.assertNotIn("air_capacity_cell_rows:", report_text)
            self.assertNotIn("general_oxygen_consumption_rows:", report_text)
            self.assertIn("Mega Batterie : la case F10 rend la batterie infinie", report_text)
            self.assertIn("SuperVitesse : runtime F10 reglable 1-20", report_text)
            self.assertIn("Mega torpilles : runtime F10 reglable 1-10, degats defaut x10, effets visuels bornes x3", report_text)
            self.assertIn("Mega Sonar : runtime F10 reglable 1-10, defaut x3", report_text)
            self.assertIn("Appeler renforts : bouton F10 tente les patrouilles vanilla amies si disponibles, puis cree des U-boats amis en fallback manuel plus proche, a portee visuelle raisonnable.", report_text)
            self.assertIn("Blindage lourd : case F10 desactivee par defaut, activable manuellement, degats joueur divises par 3", report_text)
            self.assertIn("DeepDive : case F10 activee par defaut, ordres de profondeur x2 au-dessus de 10 m, stress profondeur calcule sur profondeur /2", report_text)
            self.assertIn("le blindage ne masque pas le crush DeepDive", report_text)
            self.assertIn("Eclairage interieur : case F10 Couleurs eclairage activee par defaut ; menus Alarm/SilentRun avec palette predefinie, defaut orange ambre/vert sous-marin, decochee restaure les couleurs vanilla", report_text)
            self.assertIn("Super discrétion : case F10 desactivee par defaut, bruit et detectabilite joueur divisibles par 3", report_text)
            self.assertIn("Bouton Appeler renforts : appelle des U-boats amis pres du joueur (10-16 km, minimum 8 km); avions/warships seulement si des spawners amis compatibles existent", generated_readme_text)
            self.assertIn("Toggle Couleurs eclairage : actif par defaut, decoche = couleurs vanilla", generated_readme_text)
            self.assertIn("Couleurs eclairage : deux listes F10 predefinies pour Alarm et SilentRun, defaut orange ambre / vert sous-marin", generated_readme_text)
            self.assertIn("Couleurs eclairage : case F10 active par defaut, decochee pour restaurer les couleurs vanilla", generated_readme_text)
            self.assertIn("Eclairage interieur : deux listes F10 changent visuellement Alarm et SilentRun ; defaut Alarm orange ambre, SilentRun vert sous-marin, gameplay inchange quand la case est activee", generated_readme_text)
            self.assertIn("La lumiere SilentRun utilise la couleur SilentRun choisie uniquement au rendu quand Couleurs eclairage est cochee", generated_readme_text)
            self.assertNotIn("Lumières orange/vert", generated_readme_text)
            self.assertNotIn("bleue reste vanilla", generated_readme_text)
            self.assertNotIn("tente 2 patrouilles avions + 2 patrouilles warships", generated_readme_text)
            self.assertIn("Blindage lourd : case desactivee par defaut, activable dans F10, degats joueur divises par 3 quand activee", generated_readme_text)
            self.assertIn("Migration settings v16 / v1.4.16 : les anciennes installations repassent Blindage lourd sur OFF une seule fois", generated_readme_text)
            self.assertIn("Super discrétion : case desactivee par defaut, bruit et detectabilite joueur divises par 3 quand activee", generated_readme_text)
            self.assertIn("sans supprimer les contacts ennemis", generated_readme_text)
            self.assertIn("Toggle Plongée x2 : actif par defaut, decoche = profondeur vanilla", generated_readme_text)
            self.assertIn("DeepDive : case F10 Plongée x2 activee par defaut, ordres de profondeur > 10 m doubles (20->40, 40->80, 150->300, 300->600), stress profondeur calcule sur profondeur /2", generated_readme_text)
            self.assertIn("DeepDive : active, le stress de profondeur de l'equipage est calcule sur la profondeur /2.", generated_readme_text)
            self.assertIn("DeepDive : active, la profondeur operationnelle vise 600 m reels et le crush est repousse a 700 m ; decochee, la profondeur redevient vanilla", generated_readme_text)
            self.assertNotIn("mega_torpedo_rows:", report_text)
            self.assertNotIn("perfect_torpedo_reliability_rows:", report_text)
            self.assertNotIn("energy_recharge_rows:", report_text)
            self.assertNotIn("player_submarine_speed_rows:", report_text)
            self.assertIn("DLC Type IX detecte", report_text)

            entities_wb = load_workbook(out_mod / "Data Sheets" / "Entities.xlsx", data_only=False)
            self.assertNotIn("Types", entities_wb.sheetnames)

    def test_reinforcement_fallback_spawn_distances_are_close_but_visible(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            uboat_root = tmp_path / "UBOAT"
            data_sheets = uboat_root / "UBOAT_Data" / "Data Sheets"
            out_mod = tmp_path / "LongSubmerged10x"

            make_general_workbook(data_sheets / "General.xlsx", -0.000009)
            make_general_workbook(data_sheets / "Realistic Travel" / "General.xlsx", -0.00000133)
            make_entities_workbook(data_sheets / "Entities.xlsx")

            exit_code = generator.main(
                [
                    "--uboat",
                    str(uboat_root),
                    "--out",
                    str(out_mod),
                    "--force",
                    "--game-version",
                    "2026.1 Patch 20",
                ]
            )

            self.assertEqual(exit_code, 0)

            runtime_patch_text = (out_mod / "Source" / "LongSubmergedRuntimePatch.cs").read_text(encoding="utf-8")
            reinforcement_controller_text = extract_csharp_block(runtime_patch_text, "internal static class ReinforcementRuntimeController")

            self.assertIn("private const float FallbackMinimumPlayerDistance = 8f;", reinforcement_controller_text)
            self.assertIn("private const float FallbackRallyDistance = 6f;", reinforcement_controller_text)
            self.assertIn("private static readonly float[] FallbackSpawnDistances = new float[] { 10f, 12f, 14f, 16f };", reinforcement_controller_text)
            self.assertIn("private static readonly float[] FallbackSpawnAngles = new float[] { 110f, -110f, 130f, -130f, 150f, -150f, 90f, -90f };", reinforcement_controller_text)
            self.assertNotIn("private const float FallbackMinimumPlayerDistance = 15f;", reinforcement_controller_text)
            self.assertNotIn("new float[] { 18f, 22f, 26f, 16f }", reinforcement_controller_text)
            self.assertIn("fromPlayer.sqrMagnitude < FallbackMinimumPlayerDistance * FallbackMinimumPlayerDistance", reinforcement_controller_text)
            self.assertIn("private const float FallbackGroupClearance = 2.5f;", reinforcement_controller_text)
            self.assertIn("GetGroupsInRange(position, FallbackGroupClearance, false)", reinforcement_controller_text)
            self.assertIn("IsFallbackSpawnPositionOnNavMesh(worldNavMesh, playerGroup, position)", reinforcement_controller_text)
            self.assertIn("worldNavMesh.RaycastLandsNavMesh(position, playerGroup.Position, out hit)", reinforcement_controller_text)


if __name__ == "__main__":
    unittest.main()
